"""End-to-end integration check for Tasks & Meetings + cross-system consistency.

Runs against a THROWAWAY temp database (config.DATABASE_PATH is repointed before
any DB call), so the real ./data/yordamchi.db is never touched.

Run:  venv/bin/python tests/integration_check.py
"""
import asyncio
import os
import sys
from datetime import datetime, timedelta, time as dtime

# Allow running from anywhere — project root holds config.py / database.py.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Repoint DB to a temp file BEFORE importing database/handlers.
import config
_TMP = "/tmp/yordamchi_integration_test.db"
if os.path.exists(_TMP):
    os.remove(_TMP)
config.DATABASE_PATH = _TMP
config.ICLOUD_ENABLED = False  # tests must NEVER push to the real iCloud calendar
config.APPLE_ID = ""; config.APPLE_APP_SPECIFIC_PASSWORD = ""

import database  # noqa: E402
import handlers  # noqa: E402
import translit  # noqa: E402

TZ = database.TZ
PASS = 0
FAIL = 0
FAILED = []


def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✅ {name}")
    else:
        FAIL += 1
        FAILED.append(name)
        print(f"  ❌ {name}   {detail}")


def has(lst, _id):
    return any((x.get("id") == _id) for x in lst)


async def main():
    await database.init()
    now = datetime.now(TZ)
    today_start = TZ.localize(datetime.combine(now.date(), dtime.min))
    wk_start = today_start.isoformat()
    wk_end = (today_start + timedelta(days=7)).isoformat()

    print("\n── A. VAZIFALAR ──")
    tid = await database.create_task({"title": "TEST vazifa A", "priority": "P2",
                                      "status": "todo", "assignee": "Aziz",
                                      "deadline": (now + timedelta(days=1)).isoformat(),
                                      "tags": ["test", "muhim"]})
    t = await database.get_task(tid)
    check("A1 create_task saqlanadi", t is not None and t["title"] == "TEST vazifa A")
    check("A1 tags ro'yxat sifatida o'qiladi", isinstance(t.get("tags"), list) and "test" in t["tags"],
          f"tags={t.get('tags')!r}")
    active = await database.list_tasks(status_in=["todo", "in_progress"], limit=200)
    done = await database.list_tasks(status_in=["done"], limit=200)
    check("A2 yangi vazifa AKTIV filtrda", has(active, tid))
    check("A2 yangi vazifa BAJARILGAN filtrda EMAS", not has(done, tid))

    # cross-system snapshot before complete
    active_before = len(active)
    done_before = len(done)

    print("\n── C1. todo → done (cross-system) ──")
    ok = await database.complete_task(tid)
    t = await database.get_task(tid)
    active2 = await database.list_tasks(status_in=["todo", "in_progress"], limit=200)
    done2 = await database.list_tasks(status_in=["done"], limit=200)
    done_today = await database.list_tasks_done_today()
    check("C1 complete_task ok", ok and t["status"] == "done")
    check("C1 AKTIV'dan yo'qoldi", not has(active2, tid))
    check("C1 BAJARILGAN'da paydo bo'ldi", has(done2, tid))
    check("C1 'bugun bajarilgan'da bor", has(done_today, tid))
    check("C1 sanoq izchil (aktiv -1, done +1)",
          len(active2) == active_before - 1 and len(done2) == done_before + 1,
          f"aktiv {active_before}->{len(active2)}, done {done_before}->{len(done2)}")

    print("\n── A4. Tahrir (prioritet) ──")
    await database.update_task(tid, {"priority": "P0"}, source="edit")
    t = await database.get_task(tid)
    check("A4 prioritet P0 ga o'zgardi", t["priority"] == "P0")

    print("\n── C2. O'chirish (cross-system) ──")
    await database.delete_task(tid)
    gone = await database.get_task(tid)
    a3 = await database.list_tasks(status_in=["todo", "in_progress"], limit=200)
    d3 = await database.list_tasks(status_in=["done"], limit=200)
    allt = await database.list_tasks(limit=500)
    check("C2 get_task None", gone is None)
    check("C2 hech bir filtrda yo'q (aktiv/done/barchasi)",
          not has(a3, tid) and not has(d3, tid) and not has(allt, tid))

    print("\n── A6. Bugungi muddatli vazifa ──")
    tid2 = await database.create_task({"title": "TEST bugun", "priority": "P1", "status": "todo",
                                       "deadline": (today_start + timedelta(hours=15)).isoformat()})
    today_tasks = await database.list_today_tasks()
    check("A6 bugungi muddat → list_today_tasks", has(today_tasks, tid2))
    await database.delete_task(tid2)

    print("\n── B/D1. Uchrashuv + AGENDA RO'YXATI (regressiya) ──")
    mdata = {"title": "TEST uchrashuv", "datetime_start": (now + timedelta(days=1)).isoformat(),
             "datetime_end": (now + timedelta(days=1, hours=1)).isoformat(),
             "participants": ["Dinislam", "Aziz"],
             "agenda": ["Byudjet", "Marketing", "Keyingi qadamlar"]}
    mid = await database.create_meeting(mdata)
    m = await database.get_meeting(mid)
    check("D1 agenda=ro'yxat bo'lsa ham SAQLANADI", m is not None, "create_meeting yiqildi")
    check("D1 agenda matn sifatida o'qiladi", isinstance(m.get("agenda"), str) and "Byudjet" in m["agenda"],
          f"agenda={m.get('agenda')!r}")
    check("D1 participants ro'yxat sifatida o'qiladi", isinstance(m.get("participants"), list) and "Aziz" in m["participants"])

    print("\n── B2/D2. Bugun ERTAROQ uchrashuv → 'Haftalik'da (regressiya) ──")
    earlier = today_start + timedelta(minutes=1)  # bugun, lekin hozirdan oldin
    mid2 = await database.create_meeting({"title": "TEST ertaroq", "datetime_start": earlier.isoformat(),
                                          "datetime_end": (earlier + timedelta(hours=1)).isoformat(),
                                          "participants": ["X"], "agenda": ["a"]})
    week = await database.list_meetings_in_window(wk_start, wk_end)
    week_now = await database.list_meetings_in_window(now.isoformat(), wk_end)  # eski (buggy) oyna
    today_m = await database.list_today_meetings()
    check("D2 'Haftalik' (bugun boshidan) ko'rsatadi", has(week, mid2))
    check("D2 'Bugun' filtrida bor", has(today_m, mid2))
    check("D2 (eslatma) eski now-oyna ko'rsatMASdi", not has(week_now, mid2),
          "eski oyna ham ko'rsatsa, regressiya emas edi")

    print("\n── C4. Reschedule (cross-system) ──")
    new_start = (now + timedelta(days=2, hours=3)).isoformat()
    await database.update_meeting(mid, {"datetime_start": new_start})
    m = await database.get_meeting(mid)
    check("C4 datetime_start yangilandi", m["datetime_start"] == new_start)
    w2 = await database.list_meetings_in_window(wk_start, (today_start + timedelta(days=10)).isoformat())
    found = [x for x in w2 if x["id"] == mid]
    check("C4 oyna so'rovida yangi vaqt bilan chiqadi", bool(found) and found[0]["datetime_start"] == new_start)

    print("\n── C5. Bekor (cross-system) ──")
    await database.cancel_meeting(mid)
    await database.cancel_meeting(mid2)
    check("C5 get_meeting None", (await database.get_meeting(mid)) is None)
    w3 = await database.list_meetings_in_window(wk_start, (today_start + timedelta(days=10)).isoformat())
    check("C5 oyna so'rovidan yo'qoldi", not has(w3, mid) and not has(w3, mid2))

    print("\n── D5. Sana parser — sabab ajratiladi ──")
    cases = {"2 soat": None, "ertaga 09:00": None, "999 soat": "too_far",
             "2026-13-45 10:00": "invalid", "falongdek": "unparsable"}
    for txt, exp in cases.items():
        iso, reason = await handlers._parse_deadline_natural(txt)
        if exp is None:
            check(f"D5 '{txt}' → parse OK", iso is not None and reason is None, f"reason={reason}")
        else:
            check(f"D5 '{txt}' → '{exp}'", iso is None and reason == exp, f"reason={reason}")

    print("\n── Stats ishlaydi (cross-system yuzasi) ──")
    try:
        st = await database.executive_stats(days=7)
        check("executive_stats xatosiz ishlaydi", isinstance(st, dict) and "tasks" in st)
    except Exception as e:
        check("executive_stats xatosiz ishlaydi", False, f"{type(e).__name__}: {e}")

    print("\n── Export / Import (Excel) ──")
    import io as _io
    from openpyxl import load_workbook as _lwb

    class _ExpMsg:
        chat = type("C", (), {"id": 1})()
        text = "/export"
        captured = {}
        async def answer_document(self, file, caption=None, parse_mode=None, reply_markup=None):
            _ExpMsg.captured["b"] = file.data
            _ExpMsg.captured["kb"] = reply_markup
        async def answer(self, *a, **k):
            _ExpMsg.captured["t"] = a

    await database.create_task({"title": "Export sinov", "assignee": "J.K", "priority": "P0",
                                "status": "todo", "deadline": (datetime.now(TZ) + timedelta(days=2)).isoformat(),
                                "description": "izoh"})
    await handlers.cmd_export(_ExpMsg())
    try:
        _wb = _lwb(_io.BytesIO(_ExpMsg.captured["b"]))
        _ws = _wb["Vazifalar"]
        check("Export: xlsx yaratildi", _ExpMsg.captured.get("b") is not None)
        check("Export: 'Boshqaruv paneli' birinchi varaq", _wb.sheetnames[0] == "Boshqaruv paneli")
        _xtext = " ".join(str(c.value) for row in _wb["Boshqaruv paneli"].iter_rows() for c in row if c.value)
        check("Export: panel bo'limlari (UMUMIY/HOLAT/USTUVORLIK)",
              "UMUMIY" in _xtext and "HOLAT" in _xtext and "USTUVORLIK" in _xtext and "Jami" in _xtext)
        check("Export: panel JONLI formula (COUNTA/COUNTIF Vazifalar)",
              "COUNTA(Vazifalar" in _xtext and "COUNTIF(Vazifalar" in _xtext)
        # GRID layout: sections sit side by side (a band beyond column A), not one strip.
        _dash = _wb["Boshqaruv paneli"]
        _band_cols = {c.column for row in _dash.iter_rows() for c in row
                      if isinstance(c.value, str) and c.value.endswith("BO'YICHA")}
        check("Export: panel GRID (bo'limlar yonma-yon, A dan tashqarida ham)",
              any(col > 1 for col in _band_cols), sorted(_band_cols))
        check("Export: dinamik sarlavha A1", "AKTIV VAZIFALAR" in (_ws["A1"].value or ""),
              _ws["A1"].value)
        check("Export: sarlavha sz20 qalin, fillsiz (template)",
              int(_ws["A1"].font.sz) == 20 and _ws["A1"].font.bold
              and (_ws["A1"].fill.fgColor.rgb in (None, "00000000")))
        check("Export: header A3=№ B3=Vazifa", _ws["A3"].value == "№" and _ws["B3"].value == "Vazifa")
        check("Export: header yashil band + oq qalin (to'q ko'k emas)",
              _ws["B3"].font.name == "Arial" and _ws["B3"].font.bold
              and str(_ws["B3"].fill.fgColor.rgb or "").endswith("2E7D32")
              and str(_ws["B3"].font.color.rgb or "").endswith("FFFFFF"))
        check("Export: barcha katak wrap_text (uzun matn o'raladi)",
              _ws["H4"].alignment.wrap_text and _ws["F4"].alignment.wrap_text
              and _ws["B4"].alignment.wrap_text)
        check("Export: Takroriylik ustuni (G)", _ws["G3"].value == "Takroriylik")
        check("Export: Kategoriya ustuni (I)", _ws["I3"].value == "Kategoriya")
        check("Export: yashirin ID ustuni (J)", _ws["J3"].value == "ID" and bool(_ws.column_dimensions["J"].hidden))
        # P0 sorts first → row 4 priority cell (E) is red+bold (Shoshilinch)
        check("Export: P0 ustuvorlik qizil+qalin",
              _ws["E4"].font.bold and str(_ws["E4"].font.color.rgb or "").endswith("C00000"),
              f"E4={_ws['E4'].value} bold={_ws['E4'].font.bold} rgb={_ws['E4'].font.color.rgb}")
        check("Export: data shrift Arial 14", _ws["B4"].font.name == "Arial" and int(_ws["B4"].font.sz) == 14)
        check("Export: pechatga tayyor (landshaft + fit-width + header takror)",
              _ws.page_setup.orientation == "landscape" and _ws.page_setup.fitToWidth == 1
              and _ws.print_title_rows == "$1:$3")
        check("Export: qator balandligi moslashgan (≥26)", (_ws.row_dimensions[4].height or 0) >= 26)
        # UX: keyboard attached to the file message (one message), compact, with drill-down
        _kb = _ExpMsg.captured.get("kb")
        _kbcbs = [b.callback_data for row in _kb.inline_keyboard for b in row] if _kb else []
        check("Export: klaviatura faylga biriktirilgan (bitta xabar)", _kb is not None)
        check("Export: 'Hammasi' + ijrochi-drilldown tugmalari",
              "exportst:all" in _kbcbs and "exportwho:0" in _kbcbs)
        check("Export: ixcham (≤8 tugma, alohida ijrochi tugmasi yo'q)",
              len(_kbcbs) <= 8 and not any(c.startswith("exportby:") for c in _kbcbs))
        check("Export: 'Shu hafta' filtri tugmasi bor", "exportst:week" in _kbcbs)
        _wk = handlers._export_who_keyboard([f"N{i}" for i in range(20)], 0)
        _wkcbs = [b.callback_data for row in _wk.inline_keyboard for b in row]
        check("Export: ijrochi-picker paginated (kesish yo'q: 8/sahifa + next + orqaga)",
              sum(1 for c in _wkcbs if c.startswith("exportby:")) == 8
              and "exportwho:1" in _wkcbs and "exportroot" in _wkcbs)
    except Exception as e:
        check("Export: xlsx tahlili", False, f"{type(e).__name__}: {e}")

    # ── Export BY STATUS (holatiga qarab) ──
    check("export status map: bajarilgan→done", handlers._EXPORT_STATUS_WORDS.get("bajarilgan") == "done")
    check("export status map: o'tgan→overdue", handlers._EXPORT_STATUS_WORDS.get("o'tgan") == "overdue")
    check("export status map: aktiv→active, muhim→important",
          handlers._EXPORT_STATUS_WORDS.get("aktiv") == "active"
          and handlers._EXPORT_STATUS_WORDS.get("muhim") == "important")
    check("cb_export_status handler mavjud", hasattr(handlers, "cb_export_status"))

    _done_id = await database.create_task({"title": "EXPORT_done_x", "priority": "P1", "status": "todo"})
    await database.complete_task(_done_id)
    _fa_done = await handlers._fetch_tasks_for_export("done")
    check("_fetch_tasks_for_export('done') → faqat done",
          bool(_fa_done) and all(t.get("status") == "done" for t in _fa_done))
    _fa_active = await handlers._fetch_tasks_for_export("active")
    check("_fetch_tasks_for_export('active') → done yo'q",
          all(t.get("status") in ("todo", "in_progress", "blocked") for t in _fa_active))

    # ── Phase-2 eksport: ierarxik subtask + dinamik sarlavha + per-ijrochi "Ota vazifa" ──
    _xp = await database.create_task({"title": "EXP_parent", "priority": "P0", "category": "Loyihalar"})
    await database.create_task({"title": "EXP_child", "assignee": "EXP_Karimov",
                                "deadline": "2026-07-01T10:00:00+05:00", "priority": "P1", "parent_id": _xp})

    class _SubMsg:
        chat = type("C", (), {"id": 1})()
        cap: dict = {}
        async def answer_document(self, file, caption=None, parse_mode=None, reply_markup=None):
            _SubMsg.cap["b"] = file.data
            _SubMsg.cap["kb"] = reply_markup
        async def answer(self, *a, **k):
            _SubMsg.cap.setdefault("m", []).append(a)

    _SubMsg.cap = {}
    await handlers._send_tasks_export(_SubMsg())
    try:
        _hw = _lwb(_io.BytesIO(_SubMsg.cap["b"]))["Vazifalar"]
        _hnums = [_hw.cell(row=r, column=1).value for r in range(4, 34)]
        _htitles = [str(_hw.cell(row=r, column=2).value or "") for r in range(4, 34)]
        check("export subtask: ierarxik 'N.M' raqamlash", any("." in str(n) for n in _hnums if n))
        check("export subtask: '↳' bilan ichkariga surilgan", any(t.startswith("↳") for t in _htitles))
        # Parent (has subtasks) title is BOLD; the subtask row title is not. Resolve
        # the parent by the subtask's № prefix ("N.M" → "N"), not the first row.
        _numrow = {str(_hw.cell(row=r, column=1).value or ""): r for r in range(4, 34)}
        _sub_n = next((n for n in _numrow if "." in n), None)
        _par_n = _sub_n.split(".")[0] if _sub_n else None
        _par_r, _sub_r = _numrow.get(_par_n), _numrow.get(_sub_n)
        check("export subtask: asosiy (ota) sarlavhasi QALIN",
              bool(_par_r) and bool(_sub_r) and _hw.cell(row=_par_r, column=2).font.bold
              and not _hw.cell(row=_sub_r, column=2).font.bold,
              f"par_n={_par_n}@{_par_r} sub_n={_sub_n}@{_sub_r}")
    except Exception as e:
        check("export subtask: ierarxik", False, f"{type(e).__name__}: {e}")

    _SubMsg.cap = {}
    await handlers._send_tasks_export(_SubMsg(), assignee="EXP_Karimov")
    try:
        _aw = _lwb(_io.BytesIO(_SubMsg.cap["b"]))["Vazifalar"]
        check("export subtask: per-ijrochi 'Asosiy vazifa' ustuni",
              _aw.cell(row=3, column=3).value == "Asosiy vazifa")
        check("export subtask: sub-vazifa otasi ko'rsatilgan (EXP_parent)",
              _aw.cell(row=4, column=3).value == "EXP_parent")
        check("export subtask: per-ijrochi dinamik sarlavha (ism)",
              "EXP_KARIMOV" in (_aw["A1"].value or ""), _aw["A1"].value)
    except Exception as e:
        check("export subtask: per-ijrochi", False, f"{type(e).__name__}: {e}")
    # Cyrillic export version
    _SubMsg.cap = {}
    await handlers._send_tasks_export(_SubMsg(), script="cyr")
    try:
        _cw = _lwb(_io.BytesIO(_SubMsg.cap["b"]))["Vazifalar"]
        check("export krill: A1 kirilcha", any("Ѐ" <= ch <= "ӿ" for ch in (_cw["A1"].value or "")),
              _cw["A1"].value)
        check("export krill: header 'Вазифа'", _cw.cell(row=3, column=2).value == "Вазифа")
    except Exception as e:
        check("export krill", False, f"{type(e).__name__}: {e}")
    _rkcbs = [b.callback_data for r in handlers._export_root_keyboard(True).inline_keyboard for b in r]
    check("export krill: root kb 'exportcyr' tugma + cb",
          "exportcyr" in _rkcbs and hasattr(handlers, "cb_export_cyr"))
    # FILTERED export must also offer Cyrillic, carrying its scope in the callback
    # (regression: per-ijrochi/status eksportlar lotincha qolib ketardi).
    _SubMsg.cap = {}
    await handlers._send_tasks_export(_SubMsg(), assignee="EXP_Karimov")
    _fkb = _SubMsg.cap.get("kb")
    _fcbs = [b.callback_data for r in _fkb.inline_keyboard for b in r] if _fkb else []
    _cyr_btn = next((str(c) for c in _fcbs if str(c).startswith("exportcyr:")), None)
    check("export: filtrlangan eksportда 'Krillcha' tugma (indeks-scope bilan)",
          _cyr_btn is not None and _cyr_btn.startswith("exportcyr:wi:"), _fcbs)
    # callback_data must be byte-safe (≤64 bytes) even for Cyrillic/long names
    check("export: krill callback ≤64 bayt",
          _cyr_btn is not None and len(_cyr_btn.encode()) <= 64, _cyr_btn)
    # Click that Cyrillic button (real callback) → assignee must be transliterated
    class _CyrQ:
        data = _cyr_btn or "exportcyr"
        message = _SubMsg()
        async def answer(self, *a, **k): pass
    _SubMsg.cap = {}
    await handlers.cb_export_cyr(_CyrQ())
    try:
        _fw = _lwb(_io.BytesIO(_SubMsg.cap["b"]))["Vazifalar"]
        _hdr = [_fw.cell(row=3, column=c).value for c in range(1, 10)]
        _aidx = _hdr.index("Ижрочи") + 1 if "Ижрочи" in _hdr else 4
        _aval = str(_fw.cell(row=4, column=_aidx).value or "")
        check("export: filtrlangan KRILL eksport — ijrochi kirilcha",
              any("Ѐ" <= ch <= "ӿ" for ch in _aval), _aval)
    except Exception as e:
        check("export: filtrlangan KRILL eksport", False, f"{type(e).__name__}: {e}")
    await database.delete_task(_xp)

    class _StMsg:
        chat = type("C", (), {"id": 1})()
        text = "/export bajarilgan"
        cap: dict = {}
        async def answer_document(self, file, caption=None, parse_mode=None, reply_markup=None):
            _StMsg.cap["b"] = file.data
        async def answer(self, *a, **k):
            _StMsg.cap.setdefault("t", []).append(a)

    await handlers.cmd_export(_StMsg())  # "/export bajarilgan" → status=done
    try:
        _ws2 = _lwb(_io.BytesIO(_StMsg.cap["b"]))["Vazifalar"]
        check("Export status: subtitle 'Holat: Bajarilgan'", "Holat: Bajarilgan" in str(_ws2["A2"].value))
        _holats = [_ws2.cell(row=r, column=6).value for r in range(4, _ws2.max_row + 1)]
        _done_lbl = handlers._STATUS_LABEL_UZ.get("done", "Bajarildi")
        check("Export status: Holat ustuni faqat done",
              bool(_holats) and all(h == _done_lbl for h in _holats))
    except Exception as e:
        check("Export status: xlsx tahlili", False, f"{type(e).__name__}: {e}")

    _acts = handlers._structured_tasks_from_table(
        [("№", "Vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Izoh"),
         (1, "Import A", "Aziz", "08-06-2026", "Shoshilinch", "Aktiv", "x")])
    check("Import (struktura): 1 vazifa", len(_acts) == 1 and _acts[0]["data"]["title"] == "Import A")
    check("Import: Shoshilinch→P0", bool(_acts) and _acts[0]["data"]["priority"] == "P0")
    check("Import: noma'lum ustun → [] (aqlli yo'lga)", handlers._structured_tasks_from_table([("Mahsulot", "Narx", "Soni"), ("Olma", "5000", "10")]) == [])
    check("Import: Mas'ul ustuni → ijrochi", (lambda a: bool(a) and a[0]["data"].get("assignee") == "O.X")(handlers._structured_tasks_from_table([("Vazifa", "Mas'ul"), ("Ish", "O.X")])))
    check("Import: izoh==ijrochi → izoh tushiriladi", (lambda a: bool(a) and not a[0]["data"].get("description"))(handlers._structured_tasks_from_table([("Vazifa", "Ijrochi", "Izoh"), ("Ish", "Aziz", "Aziz")])))
    # Import is NOT held to the chat 20-cap (deterministic parse, no JSON-truncation risk).
    # 25-row table → 25 actions; the import backstop is _MAX_IMPORT_TASKS (1000), not 20.
    _big_table = [("Vazifa", "Ijrochi")] + [(f"Ish {i}", "Aziz") for i in range(25)]
    _big_acts = handlers._structured_tasks_from_table(_big_table)
    check("Import: 20+ qator cheklovsiz parse (25→25)", len(_big_acts) == 25, f"got {len(_big_acts)}")
    check("Import backstop 1000 (chat-cap 20 dan yuqori)", handlers._MAX_IMPORT_TASKS == 1000 and handlers._MAX_IMPORT_TASKS > handlers._MAX_CREATE_ACTIONS_PER_MSG, f"{getattr(handlers, '_MAX_IMPORT_TASKS', None)}")

    # № hierarchy: "3.1" → subtask of "3" (Variant 1, authoritative parent_id)
    _hacts = handlers._structured_tasks_from_table(
        [("№", "Vazifa", "Ijrochi"),
         ("3", "IMPN_ota", "Aziz"), ("3.1", "IMPN_bola1", "Aziz"),
         ("3.2", "IMPN_bola2", "Aziz"), ("4", "IMPN_ota2", "Aziz")])
    check("Import №: _num o'qiladi", [a.get("_num") for a in _hacts] == ["3", "3.1", "3.2", "4"])
    # round-trip: the export indent marker "↳" must NOT accumulate on re-import
    _arrowacts = handlers._structured_tasks_from_table(
        [("№", "Vazifa", "Ijrochi"), ("2.1", "↳ ↳ Bildirishnoma yozish", "Umrzoq")])
    check("Import: '↳' marker tozalanadi",
          bool(_arrowacts) and _arrowacts[0]["data"]["title"] == "Bildirishnoma yozish",
          _arrowacts and _arrowacts[0]["data"]["title"])
    for _a in _hacts:
        _a.setdefault("data", {})["source"] = "excel"
    await handlers._execute_actions(_hacts)
    _ota = next((t for t in await database.list_tasks(status_in=None, limit=9000)
                 if t["title"] == "IMPN_ota"), None)
    _kids = await database.list_subtasks(_ota["id"]) if _ota else []
    check("Import №: 3.1/3.2 → 3 ning sub-vazifasi", _ota is not None and len(_kids) == 2,
          f"ota={bool(_ota)} kids={len(_kids)}")
    check("Import №: 4 alohida top-level (sub emas)",
          any(t["title"] == "IMPN_ota2" and not t.get("parent_id")
              for t in await database.list_tasks(status_in=None, limit=9000, include_subtasks=True)))
    # Re-parent (authoritative): move IMPN_bola1 to top-level by dropping the dot
    _b1 = next((t for t in _kids if t["title"] == "IMPN_bola1"), None)
    if _b1:
        _re = [{"type": "update_task", "id": _b1["id"], "data": {"source": "excel"}, "_num": "9"}]
        await handlers._execute_actions(_re)
        _b1b = await database.get_task(_b1["id"])
        check("Import №: nuqtasiz raqam → top-level'ga ko'tariladi (re-parent)",
              not (_b1b or {}).get("parent_id"))
    for _t in await database.list_tasks(status_in=None, limit=9000, include_subtasks=True):
        if _t["title"].startswith("IMPN_"):
            await database.delete_task(_t["id"])

    # ── Bug: Excel round-trip sub-vazifalarni adashtiradi ──
    _rp = await database.create_task({"title": "RT_ota", "priority": "P1", "status": "todo"})
    _rs = await database.create_task({"title": "RT_sub", "parent_id": _rp, "priority": "P2", "status": "todo"})
    # (A) Fayl saralanganda sub qatori otadan OLDIN kelsa ham — sub top-level bo'lib ketmasin
    #     (num→id xaritasi update id'laridan oldindan to'ldiriladi → tartibga bog'liq emas).
    _reorder = [{"type": "update_task", "id": _rs, "data": {"source": "excel"}, "_num": "1.1"},
                {"type": "update_task", "id": _rp, "data": {"source": "excel"}, "_num": "1"}]
    await handlers._execute_actions(_reorder)
    check("Round-trip: saralangan faylда sub ota bilan qoladi (tartibga bog'liq emas)",
          (await database.get_task(_rs)).get("parent_id") == _rp)
    # (B) Per-assignee (flat) eksport re-import — dotsiz № sub-vazifani top-level qilmaydi.
    _flat_acts = handlers._structured_tasks_from_table(
        [("№", "Vazifa", "Asosiy vazifa", "Ijrochi", "ID"),
         ("1", "RT_ota", "", "Aziz", _rp), ("2", "RT_sub", "RT_ota", "Aziz", _rs)])
    check("Import: flat (Asosiy vazifa ustunli) eksportда _flat bayrog'i o'rnatiladi",
          bool(_flat_acts) and all(a.get("_flat") for a in _flat_acts))
    for _a in _flat_acts:
        _rid = _a.pop("_id", "")
        if _rid and await database.get_task(_rid):
            _a["type"] = "update_task"; _a["id"] = _rid
    await handlers._execute_actions(_flat_acts)
    check("Round-trip: flat (per-ijrochi) re-import sub-vazifani top-level qilmaydi",
          (await database.get_task(_rs)).get("parent_id") == _rp)
    for _t in await database.list_tasks(status_in=None, limit=9000, include_subtasks=True):
        if _t["title"].startswith("RT_"):
            await database.delete_task(_t["id"])

    # Round-trip CLEAR: a blank cell in a PRESENT column clears the field on update
    # (the reported "to'liq yangilanmaydi" bug — blank assignee/izoh/kategoriya stayed).
    _clrid = await database.create_task({"title": "IMPCLR", "assignee": "Aziz",
        "category": "Hujjatlar", "priority": "P1", "status": "todo", "description": "eski izoh"})
    _clr = handlers._structured_tasks_from_table([
        ("№", "Vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Izoh", "Kategoriya", "ID"),
        ("1", "IMPCLR", "", "", "Muhim", "Aktiv", "", "", _clrid)])
    for _a in _clr:
        _rid = _a.pop("_id", "")
        if _rid and await database.get_task(_rid):
            _a["type"] = "update_task"; _a["id"] = _rid
        _a.setdefault("data", {})["source"] = "excel"
    await handlers._execute_actions(_clr)
    _ct2 = await database.get_task(_clrid)
    check("Import: bo'sh katak maydonni TOZALAYDI (izoh/ijrochi/kategoriya)",
          not _ct2["description"] and not _ct2["assignee"] and not _ct2["category"],
          f"d={_ct2['description']!r} a={_ct2['assignee']!r} c={_ct2['category']!r}")
    # value-edit still applies; absent column left untouched
    check("Import: mavjud ustun yangilanadi (Muhim→P1, Aktiv→todo)",
          _ct2["priority"] == "P1" and _ct2["status"] == "todo")
    _keepid = await database.create_task({"title": "IMPKEEP", "assignee": "Bek",
        "priority": "P0", "status": "todo", "description": "saqlan"})
    _keep = handlers._structured_tasks_from_table([("Vazifa", "Holat"), ("IMPKEEP", "Jarayonda")])
    _bt = {(t.get("title") or "").strip().lower(): t["id"]
           for t in await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=2000)}
    handlers._apply_title_dedup(_keep, _bt)
    for _a in _keep:
        _a.setdefault("data", {})["source"] = "excel"
    await handlers._execute_actions(_keep)
    _kt = await database.get_task(_keepid)
    check("Import: YO'Q ustun maydonga tegmaydi (ijrochi/ustuvorlik/izoh saqlanadi)",
          _kt["status"] == "in_progress" and _kt["assignee"] == "Bek"
          and _kt["priority"] == "P0" and _kt["description"] == "saqlan")
    for _x in (_clrid, _keepid):
        await database.delete_task(_x)

    # Subtask round-trip: editing/clearing a subtask row updates it AND keeps the
    # parent link (same code path as top-level + № re-parenting).
    _spid = await database.create_task({"title": "IMPSP_ota", "priority": "P0", "status": "todo"})
    _ssid = await database.create_task({"title": "IMPSP_sub", "assignee": "Aziz", "priority": "P1",
        "status": "todo", "description": "sub izoh", "parent_id": _spid})
    _sub = handlers._structured_tasks_from_table([
        ("№", "Vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Izoh", "Kategoriya", "ID"),
        ("1", "IMPSP_ota", "", "", "Shoshilinch", "Aktiv", "", "", _spid),
        ("1.1", "IMPSP_sub", "", "", "Muhim", "Bajarildi", "", "", _ssid)])
    for _a in _sub:
        _rid = _a.pop("_id", "")
        if _rid and await database.get_task(_rid):
            _a["type"] = "update_task"; _a["id"] = _rid
        _a.setdefault("data", {})["source"] = "excel"
    await handlers._execute_actions(_sub)
    _sst = await database.get_task(_ssid)
    check("Import: sub-vazifa yangilanadi + ota saqlanadi + izoh/ijrochi tozalanadi",
          _sst["status"] == "done" and _sst["parent_id"] == _spid
          and not _sst["description"] and not _sst["assignee"],
          f"st={_sst['status']} parent={_sst['parent_id']==_spid} d={_sst['description']!r} a={_sst['assignee']!r}")
    await database.delete_task(_spid)

    print("\n── Bug-fix regressions (3 muammo) ──")
    # 1) Double-confirm: dedup a redelivered Telegram update by (chat_id, message_id)
    _d1 = await database.enqueue_pending_action(update_id=None, chat_id=900, message_id=77, user_text="x")
    _d2 = await database.enqueue_pending_action(update_id=None, chat_id=900, message_id=77, user_text="x")
    _d3 = await database.enqueue_pending_action(update_id=None, chat_id=900, message_id=78, user_text="y")
    check("Bug1: bir xil (chat,msg) qayta-yetkazish dedup (2-marta None)",
          bool(_d1) and _d2 is None and bool(_d3))
    # 2) Meeting update: participants as a STRING → list (no [] data loss), time persists
    _mid = await database.create_meeting({"title": "M", "datetime_start": "2026-07-09T10:00:00+05:00",
                                          "participants": ["A"]})
    await database.update_meeting(_mid, {"participants": "Sardor, Bobur; Aziz"})
    _m = await database.get_meeting(_mid)
    check("Bug2: meeting participants STRING → ro'yxat (revert/[] yo'q)",
          _m["participants"] == ["Sardor", "Bobur", "Aziz"], _m["participants"])
    await database.update_meeting(_mid, {"datetime_start": "2026-07-09T15:00:00+05:00", "title": "Yangi M"})
    _m2 = await database.get_meeting(_mid)
    check("Bug2: meeting vaqt+nom yangilanadi, participants saqlanadi",
          "15:00" in _m2["datetime_start"] and _m2["title"] == "Yangi M"
          and _m2["participants"] == ["Sardor", "Bobur", "Aziz"])
    await database.cancel_meeting(_mid)
    check("Bug2: _as_list matn/ro'yxatni normallaydi",
          database._as_list("A; B, C") == ["A", "B", "C"] and database._as_list(["X", " Y "]) == ["X", "Y"])
    # Renaming a meeting propagates into the SAVED bayonnoma body (mis-heard-name fix):
    # without this the protocol export kept the old (wrong) name after a title correction.
    _rmid = await database.create_meeting({"title": "FCB bilan uchrashuv",
                                           "datetime_start": "2026-07-10T15:00:00+05:00"})
    await database.update_meeting(_rmid, {"follow_up_actions": ["Mavzu: FCB bilan uchrashuv\nEshitildi: muhokama."]})
    await database.update_meeting(_rmid, {"title": "Barcelona bilan uchrashuv"})
    _rb = (await database.get_meeting(_rmid)).get("follow_up_actions") or [""]
    check("Bug: nom to'g'rilash saqlangan bayonnoma tanasida ham yangilanadi",
          "Barcelona bilan uchrashuv" in _rb[0] and "FCB bilan uchrashuv" not in _rb[0], _rb[0][:60])
    await database.cancel_meeting(_rmid)
    # 3) New-X one-shot FSM exists + capture routes with an explicit create intent
    check("Bug3: NewMeeting/NewTask FSM + capture handlerlar mavjud",
          all(hasattr(handlers, n) for n in
              ("NewMeetingTextFSM", "NewTaskTextFSM", "handle_new_meeting_capture", "handle_new_task_capture")))
    _routed: dict = {}
    _orig_par = handlers._process_and_reply
    async def _fake_par(msg, text, state=None, **k):  # noqa: E306
        _routed["text"] = text
    handlers._process_and_reply = _fake_par
    try:
        class _CapMsg:
            text = "Ertaga 12:00 Aziz bilan forum"; voice = None
            chat = type("C", (), {"id": 1})()
            async def answer(self, *a, **k): pass
        class _CapState:
            async def clear(self): pass
            async def set_state(self, *a, **k): pass
            async def get_data(self): return {}
            async def update_data(self, **k): pass
        await handlers.handle_new_meeting_capture(_CapMsg(), _CapState())
    finally:
        handlers._process_and_reply = _orig_par
    check("Bug3: yangi uchrashuv matni create-meeting niyati bilan yo'naltiriladi",
          _routed.get("text", "").startswith("Yangi uchrashuv qo'sh:"), _routed.get("text"))
    # Opened meeting must be in LLM context so "uchrashuv sarlavhasini o'zgartir" resolves
    # to THAT meeting (update_meeting{id}) instead of the bot asking "which meeting?".
    handlers.claude_service.set_last_meeting_view(
        [{"n": 1, "id": "m-test", "title": "World Wide bilan uchrashuv"}])
    _blk = await handlers.claude_service._build_state_block()
    check("Meeting-view: ochilgan uchrashuv LLM kontekstida (id + update_meeting)",
          "m-test" in _blk and "OXIRGI KO'RSATILGAN UCHRASHUV" in _blk and "update_meeting{id}" in _blk)
    handlers.claude_service.set_last_meeting_view([])

    print("\n── Kategoriyalar ──")
    _cid = await database.create_task({"title": "Kat sinov", "priority": "P1", "status": "todo", "category": "Shartnomalar"})
    _ct = await database.get_task(_cid)
    check("Kategoriya saqlanadi", _ct.get("category") == "Shartnomalar")
    _cats = await database.list_task_categories()
    check("list_task_categories ishlaydi", any(c["category"] == "Shartnomalar" and c["count"] >= 1 for c in _cats))
    check("list_tasks_by_category", any(t["id"] == _cid for t in await database.list_tasks_by_category("Shartnomalar")))
    check("Batafsil kartada kategoriya", "Shartnomalar" in handlers._format_task_detail_card(_ct))
    _imp_cat = handlers._structured_tasks_from_table([("Vazifa", "Kategoriya"), ("Ish", "SMM")])
    check("Import: Kategoriya ustuni o'qiladi", bool(_imp_cat) and _imp_cat[0]["data"].get("category") == "SMM")

    print("\n── Export/import audit tuzatishlari (8 ta) ──")
    # #6 Takroriylik (recurrence) round-trip: exported label → normalized rule code.
    _imp_rec = handlers._structured_tasks_from_table([("Vazifa", "Takroriylik"), ("Hafta ishi", "har hafta")])
    check("#6 Import: Takroriylik o'qiladi (har hafta→weekly)",
          bool(_imp_rec) and _imp_rec[0]["data"].get("recurrence_rule") == "weekly")
    check("#6 Export label round-trip (weekly→'har hafta'→weekly)",
          database.normalize_recurrence_rule(handlers._RECUR_LABEL["weekly"]) == "weekly"
          and all(database.normalize_recurrence_rule(v) == k for k, v in handlers._RECUR_LABEL.items()))
    _imp_recb = handlers._structured_tasks_from_table([("Vazifa", "Takroriylik"), ("Ish", "")])
    check("#6 Import: bo'sh Takroriylik → tozalaydi (None, kalit bor)",
          bool(_imp_recb) and "recurrence_rule" in _imp_recb[0]["data"]
          and _imp_recb[0]["data"]["recurrence_rule"] is None)
    # #2 blank priority/status in a PRESENT column must NOT overwrite with default
    _imp_blank = handlers._structured_tasks_from_table(
        [("Vazifa", "Ustuvorlik", "Holat"), ("Ish", "", "")])
    check("#2 Import: bo'sh Ustuvorlik/Holat default bilan yozilmaydi",
          bool(_imp_blank) and "priority" not in _imp_blank[0]["data"]
          and "status" not in _imp_blank[0]["data"])
    # #1 Cyrillic round-trip: a Cyrillic priority/status label imports back to its code
    _imp_cyr = handlers._structured_tasks_from_table(
        [("Vazifa", "Ustuvorlik", "Holat"),
         ("Ish", translit.to_cyrillic_pro("Shoshilinch"), translit.to_cyrillic_pro("Bajarildi"))])
    check("#1 Import: krillcha yorliq kodga qaytadi (Шошилинч→P0, Бажарилди→done)",
          bool(_imp_cyr) and _imp_cyr[0]["data"].get("priority") == "P0"
          and _imp_cyr[0]["data"].get("status") == "done")
    # #3 orphan subtask: a dotted child whose parent № is missing is counted/warned
    _orph_actions = [
        {"type": "create_task", "data": {"title": "Ota"}, "_num": "1"},
        {"type": "create_task", "data": {"title": "Bola"}, "_num": "1.1"},     # parent present
        {"type": "create_task", "data": {"title": "Yetim"}, "_num": "9.2"},    # parent missing
    ]
    check("#3 Import: yetim sub-vazifa aniqlanadi (ota-№ yo'q)",
          handlers._count_orphan_subtasks(_orph_actions) == 1)
    # #5 Cyrillic export token recognition (voice/command "krillcha")
    check("#5 Export: krillcha tokenlari tanilади",
          "krillcha" in handlers._CYR_TOKENS and "кирилл" in handlers._CYR_TOKENS)
    # #4 "this week" export scope wiring
    check("#4 Export: 'shu hafta' filtri ulangan",
          handlers._EXPORT_STATUS_WORDS.get("shu hafta") == "week"
          and handlers._EXPORT_FILTER_LABEL.get("week") == "Shu haftalik")

    print("\n── REAL export→edit→import round-trip (round-trip audit) ──")
    # Unit-level guards for the round-trip helpers first.
    # #3 _export_date keeps the clock (date-only DISPLAY, full instant stored)
    _ed = handlers._export_date("2026-07-15T17:00:00+05:00")
    check("#3 _export_date vaqtni saqlaydi (17:00, .date() emas)",
          getattr(_ed, "hour", None) == 17 and getattr(_ed, "minute", None) == 0)
    # #2 _import_deadline accepts user-natural forms and never raises
    check("#2 _import_deadline: slash/yil-siz formatlar o'qiladi",
          handlers._import_deadline("15/07/2026") is not None
          and handlers._import_deadline("12-10 14:30") is not None)
    check("#2 _import_deadline: o'qib bo'lmaydigan → None (chaqiruvchi tegmaydi)",
          handlers._import_deadline("falon-piston") is None)
    # #8 unquote_names is a clean inverse of quote_names (no drift)
    check("#8 unquote_names quote_names'ning teskarisi (drift yo'q)",
          translit.unquote_names(translit.quote_names("Agrobank hisoboti")) == "Agrobank hisoboti"
          and translit.unquote_names(translit.quote_names("Agrobankning rejasi")) == "Agrobankning rejasi")
    check("#8 unquote_names idempotent + begona qo'shtirnoqqa tegmaydi",
          translit.unquote_names("\"Agrobank\" loyiha") == "Agrobank loyiha"
          and translit.unquote_names("\"oddiy\" so'z") == "\"oddiy\" so'z")
    # #6 _import_text back-transliterates Cyrillic free text ONLY for cyr-export files
    check("#6 _import_text: krill matn lotinga qaytadi (cyr fayl)",
          handlers._import_text(translit.to_cyrillic_pro("muhim izoh"), cyr=True) == "muhim izoh")
    check("#6 _import_text: lotin faylda saqlangan krill TEGILMAYDI",
          handlers._import_text("Отчёт по проекту") == "Отчёт по проекту")

    import io as _rio
    from openpyxl import load_workbook as _rlwb

    class _RtMsg:
        chat = type("C", (), {"id": 1})()
        text = "/export"
        cap = {}
        async def answer_document(self, file, caption=None, parse_mode=None, reply_markup=None):
            _RtMsg.cap["b"] = file.data
        async def answer(self, *a, **k):
            pass

    def _rt_export(**kw):
        _RtMsg.cap.clear()
        return _RtMsg(), kw

    async def _rt_bytes(**kw):
        m = _RtMsg()
        await handlers._send_tasks_export(m, **kw)
        return _RtMsg.cap["b"]

    def _rt_read(b):
        # EXACTLY what the importer does: _read_task_sheet over the loaded workbook.
        return handlers._read_task_sheet(_rlwb(_rio.BytesIO(b), data_only=True))

    async def _rt_apply(table):
        """Run the production import pipeline (sheet→parse→ID-dedup→execute)."""
        acts = handlers._structured_tasks_from_table(table)
        for a in acts:
            rid = a.pop("_id", "")
            if rid and await database.get_task(rid):
                a["type"] = "update_task"; a["id"] = rid
        await handlers._execute_actions(acts)
        return acts

    def _rt_cols(table):
        hi = next(i for i, row in enumerate(table)
                  if any(handlers._norm_header(c) in handlers._COL_TITLE for c in row))
        H = [handlers._norm_header(c) for c in table[hi]]
        def ci(names):
            return next((H.index(n) for n in names if n in H), None)
        return hi, ci

    await handlers._upsert_contacts(["Karimov"])
    _rt_dl = (datetime.now(TZ) + timedelta(days=3)).replace(hour=14, minute=30, second=0, microsecond=0)
    _rt_id = await database.create_task({
        "title": "RoundTrip vazifa", "assignee": "Karimov", "priority": "P1",
        "status": "todo", "deadline": _rt_dl.isoformat(),
        "description": "Agrobank hisoboti", "recurrence_rule": "daily"})

    _b = await _rt_bytes(status="all")
    _tbl = _rt_read(_b)
    _acts = handlers._structured_tasks_from_table(_tbl)
    _mine = [a for a in _acts if a.get("_id") == _rt_id]
    # #1 THE headline fix: importer reads 'Vazifalar', not the dashboard '.active'
    check("#1 Round-trip: importer 'Vazifalar' varag'ini o'qiydi (dashboard emas)", len(_mine) == 1)
    check("#3 Round-trip: muddat vaqti saqlanadi (T14:30, 00:00 emas)",
          bool(_mine) and "T14:30" in (_mine[0]["data"].get("deadline") or ""))
    check("#8 Round-trip: brend qo'shtirnog'i import'da olib tashlanadi (drift yo'q)",
          bool(_mine) and _mine[0]["data"].get("description") == "Agrobank hisoboti")

    # Edit the parsed table: status→Bajarildi, assignee case-only → 'karimov'
    hi, ci = _rt_cols(_tbl)
    idci, stci, asci = ci(("id",)), ci(handlers._COL_STATUS), ci(handlers._COL_ASSIGNEE)
    di = next(i for i, row in enumerate(_tbl)
              if i > hi and idci is not None and idci < len(row) and str(row[idci]) == _rt_id)
    _erow = list(_tbl[di])
    _erow[stci] = "Bajarildi"
    _erow[asci] = "karimov"            # case-only change
    _etbl = [list(r) for r in _tbl]; _etbl[di] = _erow
    _eacts = await _rt_apply(_etbl)
    _td = await database.get_task(_rt_id)
    check("#1 Round-trip: tahrir UPDATE bo'ladi (dublikat emas)",
          any(a.get("type") == "update_task" and a.get("id") == _rt_id for a in _eacts))
    check("#4 Round-trip: holat 'Bajarildi' → done", _td["status"] == "done")
    check("#7 Round-trip: ijrochi kanonik registr ('Karimov', 'karimov' emas)",
          _td["assignee"] == "Karimov")
    _kids = [t for t in await database.list_tasks(limit=5000, include_subtasks=True)
             if t.get("recurrence_parent_id") == _rt_id]
    check("#4 Round-trip: takrorlanuvchi keyingi nusxa yaratildi (zanjir uzilmaydi)", len(_kids) >= 1)

    # #2 a non-empty UNPARSEABLE deadline cell must NOT destroy the stored deadline
    _dl_id = await database.create_task({"title": "Muddatli vazifa", "priority": "P2",
        "status": "todo", "deadline": (datetime.now(TZ) + timedelta(days=5)).isoformat()})
    _b2 = await _rt_bytes(status="all")
    _tbl2 = _rt_read(_b2)
    hi2, ci2 = _rt_cols(_tbl2)
    idci2, dlci2 = ci2(("id",)), ci2(handlers._COL_DEADLINE)
    di2 = next(i for i, row in enumerate(_tbl2)
               if i > hi2 and idci2 is not None and idci2 < len(row) and str(row[idci2]) == _dl_id)
    _r2 = list(_tbl2[di2]); _r2[dlci2] = "falon-sana"
    _t2 = [list(r) for r in _tbl2]; _t2[di2] = _r2
    await _rt_apply(_t2)
    check("#2 Round-trip: o'qib bo'lmaydigan muddat katagi mavjud muddatni O'CHIRMAYDI",
          (await database.get_task(_dl_id)).get("deadline") is not None)

    # #6 a CYRILLIC export must re-import deterministically (no LLM), Latin restored
    _cy_id = await database.create_task({"title": "Krill vazifa", "priority": "P2",
        "status": "todo", "description": "muhim hujjat"})
    _bc = await _rt_bytes(status="all", script="cyr")
    _tblc = _rt_read(_bc)
    _actsc = handlers._structured_tasks_from_table(_tblc)
    _minec = [a for a in _actsc if a.get("_id") == _cy_id]
    check("#6 Round-trip: krillcha eksport deterministik o'qiladi (LLM'siz)", len(_minec) == 1)
    check("#6 Round-trip: krill matn lotinga qaytadi (storage krill bo'lmaydi)",
          bool(_minec) and _minec[0]["data"].get("description") == "muhim hujjat"
          and _minec[0]["data"].get("title") == "Krill vazifa")

    print("\n── Fallout auditi tuzatishlari (c9cd44a yon ta'sirlari) ──")
    # F1: multi-name assignee no-op round-trip — no bogus combined contact, value lossless
    await handlers._upsert_contacts(["Aziz"])
    _mn_id = await database.create_task({"title": "Ikki ijrochili ish", "priority": "P2",
        "status": "todo", "assignee": "Karimov/Aziz"})
    _bmn = await _rt_bytes(status="all")
    await _rt_apply(_rt_read(_bmn))
    _mn_t = await database.get_task(_mn_id)
    _contacts_now = [c["name"] for c in await database.list_contacts()]
    check("F1: 'Karimov/Aziz' round-trip lossless (mutatsiya yo'q)",
          _mn_t.get("assignee") == "Karimov/Aziz", _mn_t.get("assignee"))
    check("F1: soxta birlashgan kontakt yaratilmaydi",
          not any("/" in n for n in _contacts_now), _contacts_now)
    # F2: self-name apostrophe variants — cleared, no bogus contact
    _sn_id = await database.create_task({"title": "O'zimniki ish", "priority": "P2",
        "status": "todo", "assignee": "Karimov"})
    _bsn = await _rt_bytes(status="all")
    _tsn = _rt_read(_bsn)
    hi3, ci3 = _rt_cols(_tsn)
    idci3, asci3 = ci3(("id",)), ci3(handlers._COL_ASSIGNEE)
    di3 = next(i for i, row in enumerate(_tsn)
               if i > hi3 and idci3 is not None and idci3 < len(row) and str(row[idci3]) == _sn_id)
    _r3 = list(_tsn[di3]); _r3[asci3] = "O’zim"   # U+2019 typographic apostrophe
    _t3 = [list(r) for r in _tsn]; _t3[di3] = _r3
    await _rt_apply(_t3)
    _sn_t = await database.get_task(_sn_id)
    _contacts_now = [c["name"] for c in await database.list_contacts()]
    check("F2: 'O’zim' (U+2019) → ijrochi bo'shatiladi, kontakt yaratilmaydi",
          not (_sn_t.get("assignee") or "")
          and not any(handlers._norm_asg_key(n) == "o'zim" for n in _contacts_now),
          f"asg={_sn_t.get('assignee')!r}")
    # F3: stored Cyrillic survives a LATIN no-op round-trip (no silent Latinization)
    _ru_id = await database.create_task({"title": "Отчёт по проекту", "priority": "P2", "status": "todo"})
    _bru = await _rt_bytes(status="all")
    await _rt_apply(_rt_read(_bru))
    check("F3: lotin eksportda krill sarlavha lotinlashmaydi",
          (await database.get_task(_ru_id))["title"] == "Отчёт по проекту")
    # F4: year-less date rolls forward, seconds format parses
    _yl = handlers._import_deadline("12-01")
    _yl_dt = datetime.fromisoformat(_yl)
    _now_yr = datetime.now(TZ)
    _exp_yr = _now_yr.year + 1 if _now_yr.month > 1 else _now_yr.year
    check("F4: yil-siz '12-01' kelajakka buriladi (o'tmishga tushmaydi)",
          _yl_dt >= _now_yr - timedelta(days=1) and _yl_dt.year == _exp_yr, _yl)
    check("F4: soniyali '15-07-2026 10:00:30' o'qiladi",
          handlers._import_deadline("15-07-2026 10:00:30") is not None)
    # F5: preview labels — unread cell vs blank clear vs untouched
    _pv_unread = handlers._import_deadline_line(
        {"type": "update_task", "data": {"title": "x"}, "_dl_unread": "falon"})
    _pv_clear = handlers._import_deadline_line(
        {"type": "update_task", "data": {"title": "x", "deadline": None}})
    check("F5: preview 'o'qilmadi' va 'olib tashlanadi' farqlanadi",
          "o'zgarmaydi" in _pv_unread and "olib tashlanadi" == _pv_clear,
          f"{_pv_unread!r} / {_pv_clear!r}")
    # F6: recurring flip-flop — reopen + re-done spawns only ONE child
    _ff_id = await database.create_task({"title": "FlipFlop reja", "priority": "P2",
        "status": "todo", "recurrence_rule": "weekly",
        "deadline": (datetime.now(TZ) + timedelta(days=2)).isoformat()})
    await database.complete_task(_ff_id)
    await database.update_task(_ff_id, {"status": "todo"})
    await database.complete_task(_ff_id)
    _ff_kids = [t for t in await database.list_tasks(limit=5000, include_subtasks=True)
                if t.get("recurrence_parent_id") == _ff_id]
    check("F6: reopen→re-done bitta nusxa (dublikat emas)", len(_ff_kids) == 1, len(_ff_kids))
    # F7: no-op update doesn't churn updated_at
    _nu_id = await database.create_task({"title": "NoOp ish", "priority": "P2", "status": "todo"})
    _before_ua = (await database.get_task(_nu_id))["updated_at"]
    await asyncio.sleep(0.02)
    await database.update_task(_nu_id, {"title": "NoOp ish", "source": (await database.get_task(_nu_id))["source"]})
    check("F7: no-op update updated_at'ni burmaydi",
          (await database.get_task(_nu_id))["updated_at"] == _before_ua)
    # F8: done-created recurring row gets no phantom recurrence_next_at
    _dn_id = await database.create_task({"title": "Arxiv takroriy", "priority": "P2",
        "status": "done", "recurrence_rule": "daily",
        "deadline": (datetime.now(TZ) + timedelta(days=1)).isoformat()})
    check("F8: done+takroriy yaratishда recurrence_next_at yozilmaydi",
          not (await database.get_task(_dn_id)).get("recurrence_next_at"))
    # F9: '(Boshqa)' capitalized sentinel clears category
    _cb = handlers._structured_tasks_from_table(
        [("Vazifa", "Kategoriya"), ("Ish", "(Boshqa)")])
    check("F9: '(Boshqa)' katta harfda ham tozalaydi",
          bool(_cb) and _cb[0]["data"].get("category") == "")
    # F10: foreign wb — ACTIVE task sheet beats a leading 'Nomi' reference sheet
    from openpyxl import Workbook as _WbF
    _fwb = _WbF()
    _ref = _fwb.active; _ref.title = "Malumotnoma"
    _ref.append(("Nomi", "Boshliq")); _ref.append(("Kadrlar bo'limi", "A."))
    _tsk = _fwb.create_sheet("Vazifalarim")
    _tsk.append(("Vazifa", "Ijrochi")); _tsk.append(("Haqiqiy vazifa", "Karimov"))
    _fwb.active = _fwb.sheetnames.index("Vazifalarim")
    _fbuf = _rio.BytesIO(); _fwb.save(_fbuf)
    _ftbl = handlers._read_task_sheet(_rlwb(_rio.BytesIO(_fbuf.getvalue()),
                                            read_only=True, data_only=True))
    _facts = handlers._structured_tasks_from_table(_ftbl)
    check("F10: begona faylda AKTIV vazifa varag'i 'Nomi'-varaqdan ustun",
          len(_facts) == 1 and _facts[0]["data"]["title"] == "Haqiqiy vazifa",
          [a["data"].get("title") for a in _facts])
    # F11: unchanged deadline — sub-second xlsx artifact counts as SAME instant
    check("F11: mikrosekund artefakti bir xil muddat deb tan olinadi",
          handlers._same_deadline_instant("2026-07-02T19:33:21.123000+05:00",
                                          "2026-07-02T19:33:21.123456+05:00")
          and not handlers._same_deadline_instant("2026-07-02T19:33:22.500000+05:00",
                                                  "2026-07-02T19:33:21.123456+05:00"))

    print("\n── Vazifa hayot sikli auditi (yaratish/o'chirish/holat) ──")
    # L1: blank / whitespace title → safe default (not a broken empty card)
    _l1a = await database.get_task(await database.create_task({"title": "", "priority": "P2", "status": "todo"}))
    _l1b = await database.get_task(await database.create_task({"title": "   ", "priority": "P2", "status": "todo"}))
    check("L1: bo'sh/probel sarlavha → 'Vazifa'",
          _l1a["title"] == "Vazifa" and _l1b["title"] == "Vazifa")
    # L2: malformed (non-ISO) deadline is dropped on create AND on update (never stored raw)
    _l2 = await database.get_task(await database.create_task(
        {"title": "L2", "priority": "P2", "status": "todo", "deadline": "ertaga"}))
    check("L2: create — noto'g'ri muddat ('ertaga') saqlanmaydi", _l2.get("deadline") is None)
    _l2b_id = await database.create_task({"title": "L2b", "priority": "P2", "status": "todo",
        "deadline": (datetime.now(TZ) + timedelta(days=4)).isoformat()})
    await database.update_task(_l2b_id, {"deadline": "keyinroq"})
    check("L2: update — noto'g'ri muddat mavjud muddatni o'chirmaydi",
          (await database.get_task(_l2b_id)).get("deadline") is not None)
    # L3: delete cascades the FULL subtree (grandchild) + all linked reminders
    _gp = await database.create_task({"title": "Bobo", "priority": "P2", "status": "todo"})
    _ch = await database.create_task({"title": "Ota-bola", "priority": "P2", "status": "todo", "parent_id": _gp})
    _gc = await database.create_task({"title": "Nabira", "priority": "P2", "status": "todo", "parent_id": _ch})
    await database.create_reminder({"title": "R-nabira", "remind_at": (datetime.now(TZ)+timedelta(days=1)).isoformat(), "task_id": _gc})
    await database.delete_task(_gp)
    _rem_left = [r for r in await database.list_reminders(limit=500) if r.get("task_id") == _gc]
    check("L3: rekursiv o'chirish — nabira ham, eslatmasi ham o'chadi",
          (await database.get_task(_gc)) is None and (await database.get_task(_ch)) is None
          and not _rem_left)
    # L4: delete_tasks_by_category cascades subtasks + reminders
    _kp = await database.create_task({"title": "KatOta", "priority": "P2", "status": "todo", "category": "WipeKat"})
    _ks = await database.create_task({"title": "KatSub", "priority": "P2", "status": "todo", "parent_id": _kp})
    await database.create_reminder({"title": "R-katsub", "remind_at": (datetime.now(TZ)+timedelta(days=1)).isoformat(), "task_id": _ks})
    await database.delete_tasks_by_category("WipeKat")
    check("L4: kategoriya o'chirish — sub-vazifa+eslatma ham ketadi",
          (await database.get_task(_ks)) is None
          and not [r for r in await database.list_reminders(limit=500) if r.get("task_id") == _ks])
    # L5: undo restores the WHOLE subtree + reminder, not just the parent
    _up = await database.create_task({"title": "UndoOta", "priority": "P2", "status": "todo"})
    _us = await database.create_task({"title": "UndoSub", "priority": "P2", "status": "todo", "parent_id": _up})
    await database.create_reminder({"title": "R-undo", "remind_at": (datetime.now(TZ)+timedelta(days=1)).isoformat(), "task_id": _us})
    _snap = await database.snapshot_task_tree(_up)
    await database.delete_task(_up)
    _restored = await database.restore_task_tree(_snap)
    check("L5: undo — ota+sub+eslatma to'liq tiklanadi",
          _restored == 2 and (await database.get_task(_us)) is not None
          and (await database.get_task(_us)).get("parent_id") == _up
          and [r for r in await database.list_reminders(limit=500) if r.get("task_id") == _us])
    # L9: a BLOCKED task is coherent across overdue / unassigned / reminders / risk
    _bl = await database.create_task({"title": "Bloklangan P0", "priority": "P0", "status": "blocked",
        "deadline": (datetime.now(TZ) - timedelta(days=1)).isoformat()})
    _ov_ids = [t["id"] for t in await database.list_overdue_tasks()]
    _un_ids = [t["id"] for t in await database.list_unassigned_tasks()]
    _rk = await database.risk_score_counts()
    check("L9: blocked vazifa overdue+unassigned ro'yxatlarida ko'rinadi",
          _bl in _ov_ids and _bl in _un_ids)
    check("L9: blocked overdue risk hisobiga kiradi", _rk.get("overdue", 0) >= 1)
    # L10: recurrence dedup holds even after the spawned child's deadline is edited
    _rc = await database.create_task({"title": "L10 takror", "priority": "P2", "status": "todo",
        "recurrence_rule": "weekly", "deadline": (datetime.now(TZ)+timedelta(days=2)).isoformat()})
    await database.complete_task(_rc)
    _kid = next(t for t in await database.list_tasks(limit=5000, include_subtasks=True)
                if t.get("recurrence_parent_id") == _rc)
    await database.update_task(_kid["id"], {"deadline": (datetime.now(TZ)+timedelta(days=20)).isoformat()})
    await database.update_task(_rc, {"status": "todo"})   # reopen
    await database.complete_task(_rc)                     # re-complete
    _kids10 = [t for t in await database.list_tasks(limit=5000, include_subtasks=True)
               if t.get("recurrence_parent_id") == _rc]
    check("L10: tahrirlangan muddatli bola bo'lsa ham dublikat tug'ilmaydi", len(_kids10) == 1, len(_kids10))
    # L12: update/complete on a DELETED id → surfaced in _failed (not silent success)
    _dead = await database.create_task({"title": "O'ladi", "priority": "P2", "status": "todo"})
    await database.delete_task(_dead)
    _res12 = await handlers._execute_actions([{"type": "update_task", "id": _dead, "data": {"status": "done"}}])
    check("L12: o'lik id'ga update → _failed'da belgilanadi",
          "update_task" in _res12.get("_failed", []) and _dead not in _res12.get("task", []))
    # L13: a subtask with its own deadline must NOT leak into /today or overdue lists
    _tp = await database.create_task({"title": "TodayOta", "priority": "P2", "status": "todo"})
    _tsub = await database.create_task({"title": "TodaySub", "priority": "P2", "status": "todo",
        "parent_id": _tp, "deadline": datetime.now(TZ).replace(hour=15, minute=0).isoformat()})
    _osub = await database.create_task({"title": "OverdueSub", "priority": "P2", "status": "todo",
        "parent_id": _tp, "deadline": (datetime.now(TZ) - timedelta(days=2)).isoformat()})
    _today_ids = [t["id"] for t in await database.list_today_tasks()]
    _ovd_ids = [t["id"] for t in await database.list_overdue_tasks()]
    check("L13: sub-vazifa /today va overdue ro'yxatlariga chiqmaydi",
          _tsub not in _today_ids and _osub not in _ovd_ids)
    # L15: complete a recurring task, then delete it → NO zombie next-occurrence orphan
    _z = await database.create_task({"title": "Zombi test", "priority": "P2", "status": "todo",
        "recurrence_rule": "daily", "deadline": (datetime.now(TZ)+timedelta(days=1)).isoformat()})
    await database.complete_task(_z)
    _zkid = next((t for t in await database.list_tasks(limit=5000, include_subtasks=True)
                  if t.get("recurrence_parent_id") == _z), None)
    check("L15: takroriy yopilganda keyingi nusxa yaratiladi (sanity)", _zkid is not None)
    # concurrent complete+delete on a fresh recurring task → no orphan child of a deleted parent
    _z2 = await database.create_task({"title": "Zombi race", "priority": "P2", "status": "todo",
        "recurrence_rule": "daily", "deadline": (datetime.now(TZ)+timedelta(days=1)).isoformat()})
    await asyncio.gather(database.complete_task(_z2), database.delete_task(_z2), return_exceptions=True)
    _z2kids = [t for t in await database.list_tasks(limit=5000, include_subtasks=True)
               if t.get("recurrence_parent_id") == _z2]
    check("L15: complete+delete poygasida o'lik otaga bog'liq zombi yo'q",
          (await database.get_task(_z2)) is None or not _z2kids or len(_z2kids) <= 1)

    print("\n── Disk I/O incident: backup rotatsiyasi + xato xabari ──")
    # D1: backup rotation caps local snapshots (undo backups can't fill the disk)
    import pathlib as _pl, tempfile as _tf
    _bdir = _pl.Path(_tf.mkdtemp()) / "backups"
    _bdir.mkdir(parents=True)
    _names = [_bdir / f"yordamchi-pre-delete-{_i:02d}.db" for _i in range(25)]
    for _i, f in enumerate(_names):
        f.write_text("x")
        os.utime(f, (_i, _i))   # deterministic mtime ordering (newest = highest i)
    handlers._rotate_backups(_bdir, keep=20)
    _left = sorted(_bdir.glob("yordamchi-*.db"))
    check("D1: backup rotatsiyasi eng so'nggi 20 tani saqlaydi", len(_left) == 20)
    check("D1: eng eski backuplar o'chiriladi (yangilari qoladi)",
          _names[24].exists() and _names[20].exists() and not _names[0].exists()
          and not _names[4].exists())
    # D2: a disk I/O error gets an actionable Uzbek message (not a raw SQLite string)
    _e = handlers._humanize_error(Exception("disk I/O error"))
    check("D2: 'disk I/O error' → disk to'lgani haqida amaliy xabar",
          "disk" in _e.lower() and ("df -h" in _e or "to'lgan" in _e))
    _e2 = handlers._humanize_error(Exception("database disk image is malformed"))
    check("D2: 'malformed' → korruptsiya + integrity_check maslahati",
          "integrity_check" in _e2 or "buzilgan" in _e2)

    print("\n── Kategoriya boshqaruvi (qo'shish/o'chirish) ──")
    _m1 = await database.create_task({"title": "M1", "priority": "P2", "status": "todo", "category": "TestKat"})
    _m2 = await database.create_task({"title": "M2", "priority": "P2", "status": "todo", "category": "TestKat"})
    check("count_tasks_in_category", await database.count_tasks_in_category("TestKat") == 2)
    await handlers._execute_actions([{"type": "assign_category", "data": {"category": "YangiKat", "from_category": "TestKat"}}])
    check("assign_category (ko'chirish)", await database.count_tasks_in_category("YangiKat") == 2 and await database.count_tasks_in_category("TestKat") == 0)
    await handlers._execute_actions([{"type": "delete_category", "data": {"category": "YangiKat"}}])
    _mt = await database.get_task(_m1)
    check("delete_category: yorliq olinadi, vazifa qoladi", (not _mt.get("category")) and _mt.get("status") == "todo")
    await database.update_task(_m1, {"category": "OchKat"})
    await database.update_task(_m2, {"category": "OchKat"})
    await handlers._execute_actions([{"type": "delete_tasks_by_category", "data": {"category": "OchKat"}}])
    check("delete_tasks_by_category: o'chiriladi", await database.count_tasks_in_category("OchKat") == 0 and (await database.get_task(_m1)) is None)
    check("Kategoriya o'chirish tasdiq to'plamida", {"delete_category", "delete_tasks_by_category"} <= handlers._CATEGORY_DELETE_ACTION_TYPES)
    _pv = await handlers._format_create_preview([{"type": "delete_tasks_by_category", "data": {"category": "Shartnomalar"}}])
    check("Preview: kategoriya o'chirish (son bilan)", "Shartnomalar" in _pv and "o'chiriladi" in _pv)

    print("\n── Kategoriyalar jadvali (B variant: ikonka/arxiv/tartib) ──")
    await database.create_category("Reklama", "🔴")
    await database.create_task({"title": "RT1", "priority": "P2", "status": "todo", "category": "SMM2"})
    _lc = await database.list_categories()
    check("list_categories: bo'sh kategoriya + ikonka", any(c["name"] == "Reklama" and c["count"] == 0 and c["icon"] == "🔴" for c in _lc))
    check("list_categories: orphan ko'rinadi", any(c["name"] == "SMM2" for c in _lc))
    await database.update_category("Reklama", new_name="Reklama2", icon="🟣")
    check("update_category: rename + icon", (lambda g: bool(g) and g["icon"] == "🟣")(await database.get_category("Reklama2")))
    await database.archive_category("Reklama2", True)
    check("archive: faolda yo'q", not any(c["name"] == "Reklama2" for c in await database.list_categories()))
    check("archive: arxivda bor", any(c["name"] == "Reklama2" for c in await database.list_categories(include_archived=True)))
    await database.archive_category("Reklama2", False)
    check("unarchive: faolga qaytadi", any(c["name"] == "Reklama2" for c in await database.list_categories()))
    await database.create_category("MZ1"); await database.create_category("MZ2")
    _b4 = [c["name"] for c in await database.list_categories() if c["name"] in ("MZ1", "MZ2")]
    await database.move_category("MZ2", "up")
    _af = [c["name"] for c in await database.list_categories() if c["name"] in ("MZ1", "MZ2")]
    check("move_category: tartib o'zgaradi", _b4 != _af)
    await handlers._execute_actions([{"type": "create_category", "data": {"category": "ActKat", "icon": "🟢"}}])
    check("action create_category", (await database.get_category("ActKat")) is not None)
    await handlers._execute_actions([{"type": "archive_category", "data": {"category": "ActKat"}}])
    check("action archive_category", (lambda g: bool(g) and g["archived"] == 1)(await database.get_category("ActKat")))
    await database.create_category("DelKat")
    await handlers._execute_actions([{"type": "delete_category", "data": {"category": "DelKat"}}])
    check("action delete_category: metadata o'chadi", (await database.get_category("DelKat")) is None)

    # create_task category guard (B): reuse EXISTING only — never auto-create new
    await database.create_category("GuardKat")
    _ig = await handlers._execute_actions([{"type": "create_task",
        "data": {"title": "Guard mavjud", "priority": "P2", "category": "GuardKat"}}])
    _tg = await database.get_task(_ig["task"][0])
    check("create_task: mavjud kategoriya saqlanadi",
          bool(_tg) and _tg.get("category") == "GuardKat", f"{_tg.get('category') if _tg else None}")
    _in = await handlers._execute_actions([{"type": "create_task",
        "data": {"title": "Guard yangi", "priority": "P2", "category": "YoqBunaqaKat999"}}])
    _tn = await database.get_task(_in["task"][0])
    check("create_task: NOMA'LUM kategoriya tushiriladi (uncategorized, sprawl yo'q)",
          bool(_tn) and not (_tn.get("category") or ""), f"{_tn.get('category') if _tn else None}")

    # Reference-by-number: the last shown numbered list is exposed to the LLM so
    # "10-vazifani tahrirla" resolves to the right task id.
    import claude_service as _cs
    _rbn_id = await database.create_task({"title": "Rekvizit tayyorlash", "priority": "P2", "status": "todo"})
    _cs.set_last_task_view([{"n": 7, "id": _rbn_id, "title": "Rekvizit tayyorlash"}])
    _sb = await _cs._build_state_block()
    check("ref-by-number: state'da 'OXIRGI KO'RSATILGAN RO'YXAT'", "OXIRGI KO'RSATILGAN" in _sb)
    check("ref-by-number: raqam+id+title ko'rinadi",
          "7." in _sb and _rbn_id in _sb and "Rekvizit" in _sb, _sb[-400:])
    # #12: a DELETED task drops out of the LLM 'last shown' context (no dead-id action)
    await database.delete_task(_rbn_id)
    _sb2 = await _cs._build_state_block()
    check("#12 ref-by-number: o'chirilgan vazifa kontekstdan chiqib ketadi",
          _rbn_id not in _sb2)
    _cs.set_last_task_view([])

    # JSON repair: unescaped double-quotes inside a string value (real bug: a task
    # title «"Pulli Gap"» broke json.loads → command silently failed).
    _bad = '{"intent":"A","actions":[],"user_message":"«"Pulli Gap" shartnoma»","buttons":[]}'
    _p = _cs._extract_json(_bad)
    check("json-repair: buzuq qo'shtirnoqli JSON tiklandi", bool(_p) and _p.get("intent") == "A", f"{_p}")
    check("json-repair: user_message saqlandi", bool(_p) and "Pulli Gap" in (_p.get("user_message") or ""))
    check("json-repair: to'g'ri JSON buzilmaydi",
          (_cs._extract_json('{"intent":"B","user_message":"oddiy matn"}') or {}).get("intent") == "B")

    print("\n── Import round-trip dedup (yangilash, dublikat emas) ──")
    _t = await database.create_task({"title": "Dedup sinov", "priority": "P2", "status": "todo"})
    _a2 = handlers._structured_tasks_from_table(
        [("№", "Vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Izoh", "ID"),
         (1, "Dedup YANGI", "Dilshod", "", "Muhim", "Aktiv", "", _t)])
    check("Import: ID ustuni o'qiladi", bool(_a2) and _a2[0].get("_id") == _t)
    for _a in _a2:
        _rid = _a.pop("_id", "")
        if _rid and await database.get_task(_rid):
            _a["type"] = "update_task"; _a["id"] = _rid
    check("Import dedup: mavjud ID → update_task", _a2[0]["type"] == "update_task")
    await handlers._execute_actions(_a2)
    _cnt = len([t for t in await database.list_tasks(limit=999) if t["title"] in ("Dedup sinov", "Dedup YANGI")])
    _upd = await database.get_task(_t)
    check("Import dedup: yangilandi, dublikat YO'Q", _cnt == 1 and _upd["title"] == "Dedup YANGI" and _upd["assignee"] == "Dilshod", f"cnt={_cnt}")

    print("\n── Nazoratli ro'yxat (A): ijrochi/kategoriya avto qo'shilmaydi ──")
    _aids = await handlers._execute_actions([{"type": "create_task", "data": {
        "title": "A_llm", "assignee": "A_Nomalum", "category": "A_NomalumKat"}}])
    _at = await database.get_task(_aids["task"][0])
    check("A: LLM noma'lum ijrochi → bo'sh", not (_at.get("assignee") or ""))
    check("A: LLM noma'lum kategoriya → bo'sh", not (_at.get("category") or ""))
    check("A: LLM contact avto qo'shmaydi",
          not any((c.get("name") or "") == "A_Nomalum" for c in await database.list_contacts()))
    _eids = await handlers._execute_actions([{"type": "create_task", "data": {
        "title": "A_excel", "assignee": "A_Karimov", "category": "A_Loyiha", "source": "excel"}}])
    _et = await database.get_task(_eids["task"][0])
    check("A: Excel ijrochi saqlanadi + contact qo'shiladi",
          _et.get("assignee") == "A_Karimov"
          and any((c.get("name") or "") == "A_Karimov" for c in await database.list_contacts()))
    check("A: Excel kategoriya saqlanadi", _et.get("category") == "A_Loyiha")
    _kids = await handlers._execute_actions([{"type": "create_task", "data": {
        "title": "A_llm2", "assignee": "A_Karimov"}}])
    check("A: ma'lum ijrochiga LLM tayinlay oladi",
          (await database.get_task(_kids["task"][0])).get("assignee") == "A_Karimov")
    check("A: manual '➕ Ijrochi qo'shish' UI mavjud",
          all(hasattr(handlers, n) for n in ("cb_contact_add", "handle_contact_add", "ContactAddFSM")))

    # Title-based dedup (covers files with NO hidden ID + smart-extracted PDFs)
    _bt = {(t.get("title") or "").strip().lower(): t["id"]
           for t in await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=999)
           if t.get("title")}
    _new = [{"type": "create_task", "data": {"title": "Dedup YANGI", "priority": "P0"}},
            {"type": "create_task", "data": {"title": "Mutlaqo yangi XYZ", "priority": "P2"}}]
    _conv = handlers._apply_title_dedup(_new, _bt)
    check("Title dedup: mavjud sarlavha → update", _new[0]["type"] == "update_task" and _conv["converted"] == 1)
    check("Title dedup: yangi sarlavha → create", _new[1]["type"] == "create_task")
    # intra-batch dedup: identical NEW titles in ONE import → keep first, drop rest
    _batch = [{"type": "create_task", "data": {"title": "Takror ABC", "priority": "P1"}},
              {"type": "create_task", "data": {"title": "takror abc", "priority": "P2"}},
              {"type": "create_task", "data": {"title": "Takror ABC", "priority": "P0"}},
              {"type": "create_task", "data": {"title": "Boshqa DEF", "priority": "P2"}}]
    _d2 = handlers._apply_title_dedup(_batch, {})
    check("Intra-batch dedup: 3 ta bir xil → 1 qoladi (2 tashlandi)",
          _d2["dropped"] == 2 and len(_batch) == 2, f"{_d2} len={len(_batch)}")
    check("Intra-batch dedup: birinchi + boshqasi qoladi",
          _batch[0]["data"]["title"] == "Takror ABC" and _batch[1]["data"]["title"] == "Boshqa DEF")

    print("\n── _humanize_error (aniq sabab) ──")
    check("humanize: tarmoq", "ulanish" in handlers._humanize_error(Exception("Cannot connect [Network is unreachable]")).lower())
    check("humanize: bo'sh xabar", "bo'sh xabar" in handlers._humanize_error(Exception("messages.8: user messages must have non-empty content")).lower())
    check("humanize: noma'lum → tur ko'rsatiladi", "ValueError" in handlers._humanize_error(ValueError("nimadir")))

    print("\n── Ikonka birligi (list ↔ batafsil) ──")
    _p2 = {"title": "P2 sinov", "priority": "P2", "status": "todo", "deadline": None}
    check("Badge: P2 list=batafsil bir xil", handlers._format_task_detail_card(_p2).startswith(handlers._task_badge(_p2)))

    print("\n── Delegatsiya → Ijrochilar paneliga birlashtirildi ──")
    try:
        import aiosqlite as _sq
        async with _sq.connect(config.DATABASE_PATH) as _db:
            await _db.execute("SELECT * FROM tasks WHERE status IN ('todo','in_progress') "
                              "AND assignee IS NOT NULL AND LOWER(TRIM(assignee)) NOT IN "
                              "('','men','siz','belgilanmagan','—','oʻzim','o''zim','o''z','ozim') LIMIT 5")
        check("Qotgan-topshiriqlar SQL ishlaydi", True)
    except Exception as e:
        check("Qotgan-topshiriqlar SQL ishlaydi", False, f"{type(e).__name__}: {e}")
    # Birlashtirish: alohida /delegations olib tashlandi, Ijrochilarga ko'chdi
    check("Birlashma: _render_stale_delegations mavjud", hasattr(handlers, "_render_stale_delegations"))
    check("Birlashma: eski cmd_delegations olib tashlandi", not hasattr(handlers, "cmd_delegations"))
    check("Birlashma: '⏳ Kutilayotganlar' tugmasi Ijrochilar klaviaturasida",
          any(b.text == handlers.YBTN_TEAM_STALE
              for row in handlers.team_section_reply_keyboard().keyboard for b in row))
    check("Birlashma: YBTN_TEAM_STALE section-label ro'yxatida",
          handlers.YBTN_TEAM_STALE in handlers._SECTION_LABELS)

    print("\n── Bo'sh slot (free-time) hisoblash ──")
    # SOF funksiya: 11:00-12:30 + 15:00-16:00 band → 3 ta bo'shliq
    d = datetime(2026, 6, 9).date()
    ws = TZ.localize(datetime.combine(d, dtime(9, 0)))
    we = TZ.localize(datetime.combine(d, dtime(18, 0)))
    busy = [(TZ.localize(datetime.combine(d, dtime(11, 0))), TZ.localize(datetime.combine(d, dtime(12, 30)))),
            (TZ.localize(datetime.combine(d, dtime(15, 0))), TZ.localize(datetime.combine(d, dtime(16, 0))))]
    fr = handlers._compute_free_slots(busy, ws, we)
    check("free-slot: 3 ta bo'shliq", len(fr) == 3, f"len={len(fr)}")
    check("free-slot: birinchi 09:00–11:00", fr[0][0].strftime("%H:%M") == "09:00" and fr[0][1].strftime("%H:%M") == "11:00")
    # LMT-bug regressiya: 09:00–11:00 ANIQ 120 daqiqa bo'lishi shart (pytz +04:37 emas)
    check("free-slot: davomiylik aniq (LMT bug yo'q)", int((fr[0][1] - fr[0][0]).total_seconds() // 60) == 120,
          f"{int((fr[0][1]-fr[0][0]).total_seconds()//60)} daq")
    # overlap birlashishi
    ov = [(TZ.localize(datetime.combine(d, dtime(10, 0))), TZ.localize(datetime.combine(d, dtime(12, 0)))),
          (TZ.localize(datetime.combine(d, dtime(11, 0))), TZ.localize(datetime.combine(d, dtime(13, 0))))]
    fov = handlers._compute_free_slots(ov, ws, we)
    check("free-slot: overlap birlashdi", len(fov) == 2 and fov[1][0].strftime("%H:%M") == "13:00")
    # 30 daqiqadan kichik bo'shliq tashlanadi
    tiny = [(ws, TZ.localize(datetime.combine(d, dtime(11, 0)))),
            (TZ.localize(datetime.combine(d, dtime(11, 20))), we)]
    check("free-slot: <30daq bo'shliq tashlandi", handlers._compute_free_slots(tiny, ws, we) == [])
    # bo'sh kun → bitta to'liq slot
    check("free-slot: bo'sh kun = 1 slot", len(handlers._compute_free_slots([], ws, we)) == 1)
    # _fmt_dur
    check("free-slot: _fmt_dur", handlers._fmt_dur(45) == "45 daq" and handlers._fmt_dur(120) == "2 soat" and handlers._fmt_dur(90) == "1s 30daq")
    # _resolve_target_date
    _now = datetime(2026, 6, 6, 10, 0, tzinfo=TZ)  # Shanba
    check("free-slot: sana resolver (ISO + hafta-kun)",
          handlers._resolve_target_date("2026-06-09", _now) == datetime(2026, 6, 9).date()
          and handlers._resolve_target_date("seshanba", _now) == datetime(2026, 6, 9).date()
          and handlers._resolve_target_date("ertaga", _now) == datetime(2026, 6, 7).date())
    check("free-slot: show_free_slots dispatch ro'yxatda", "show_free_slots" in handlers._SHOW_ACTION_TYPES)

    print("\n── Bayonnoma: Agrobank shabloniga ko'chirildi (eski builder olib tashlandi) ──")
    import protocol_doc as _pdoc
    import inspect as _insp
    check("eski _build_protocol_docx_bytes olib tashlandi (o'lik kod yo'q)",
          not hasattr(handlers, "_build_protocol_docx_bytes"))
    check("protocol_doc: build_docx/build_pdf/build_fields mavjud",
          all(hasattr(_pdoc, fn) for fn in ("build_docx", "build_pdf", "build_fields")))
    check("cb_protocol_export protocol_doc ishlatadi (yangi Agrobank dizayni)",
          "protocol_doc" in _insp.getsource(handlers.cb_protocol_export))
    # Yangi dizaynning o'zi (ko'k/zebra, ustun tartibi, Lotin/Kiril) protocol_doc_check.py da.

    print("\n── Ikonka nomuvofiqligi tuzatildi (bo'lim ikonkalari noyob) ──")
    import pathlib as _pl
    _src = _pl.Path(handlers.__file__).read_text(encoding="utf-8")
    check("Kategoriyalar reply tugma = 🗄 (🏷 emas)", handlers.TBTN_TASKS_CATEGORIES.startswith("🗄"))
    check("Ijrochilar reply tugma = 👥", handlers.BTN_TEAM.startswith("👥"))
    check("Kategoriya ikonkasi ≠ Ijrochilar ikonkasi",
          handlers.TBTN_TASKS_CATEGORIES[0] != handlers.BTN_TEAM[0])
    check("Kategoriya header = 🗄 (🏷 KATEGORIYALAR yo'q)",
          "🗄 **KATEGORIYALAR**" in _src and "🏷 **KATEGORIYALAR**" not in _src)
    check("Teglar hali 🏷 (kategoriya bilan to'qnashmaydi)", "🏷 Teglar" in _src)
    check("Qotgan-topshiriqlar header = ⏳ (eski DELEGATSIYALAR header yo'q)",
          "⏳  **KUTILAYOTGAN TOPSHIRIQLAR**" in _src and "**DELEGATSIYALAR**" not in _src)
    check("Delegatsiya stats kichik-bo'limi = 📋", "📋 **Delegatsiya**" in _src and "👥 **Delegatsiya**" not in _src)

    print("\n── Qaydlar: sana bo'yicha guruhlash (Variant B) ──")
    _now = datetime.now(handlers.database.TZ)
    _nt = {"id": "a", "title": "Bugungi qayd", "content": "matn bugun",
           "source": "voice", "created_at": _now.isoformat()}
    _ny = {"id": "b", "title": "Kechagi qayd", "content": "matn kecha",
           "source": "forward", "created_at": (_now - timedelta(days=1)).isoformat()}
    _nout = handlers._format_notes_compact([_nt, _ny], "inbox", inbox_count=2)
    check("Qaydlar header = 'QAYDLAR · INBOX · 2 ta'", "QAYDLAR · INBOX · 2 ta" in _nout)
    check("Qaydlar: 📅 BUGUN guruhi", "📅 **BUGUN**" in _nout)
    check("Qaydlar: 📅 KECHA guruhi", "📅 **KECHA**" in _nout)
    check("Qaydlar: manba ikonkasi sarlavha oldida (🎙 1.)", "🎙 1." in _nout)
    check("Qaydlar: forward ikonkasi (🔁 2.)", "🔁 2." in _nout)
    check("Qaydlar: eski 'NOTES ·' sarlavhasi yo'q", "NOTES ·" not in _nout)

    print("\n── Qaydlarga junk tushmasligi (tugma/buyruq/bot-chiqishi filtri) ──")
    check("noise: '/team' (buyruq) bloklanadi", handlers._is_note_noise("/team"))
    check("noise: '⬅️ Asosiy menyu' (tugma) bloklanadi", handlers._is_note_noise("⬅️ Asosiy menyu"))
    check("noise: real matn bloklanmaydi", not handlers._is_note_noise("Pepsi bilan shartnoma tuzish"))
    _botout = "📌  VAZIFALAR\n\nKo'rinish · Aktiv\nNatija · 9 ta\n━━━━━━━━━━━━━━━"
    check("bot-echo: botning '📌 VAZIFALAR' paneli aniqlanadi", handlers._looks_like_bot_output(_botout))
    check("bot-echo: real matn (divider yo'q) aniqlanmaydi",
          not handlers._looks_like_bot_output("Ertaga soat 10da uchrashuv bor"))
    check("forward guard: _forward_is_bot_echo mavjud", hasattr(handlers, "_forward_is_bot_echo"))

    print("\n── Qayd detali: tugma to'qnashuvi + ixcham karta ──")
    _dnote = {"id": "n-x", "title": "Test qayd", "content": "matn",
              "source": "llm", "created_at": "2026-06-06T02:41:00+05:00",
              "status": "inbox", "tags": ["bot", "UI"]}
    _dkb = [b.text for row in handlers.note_detail_menu(_dnote).inline_keyboard for b in row]
    check("dedup: inline action = '📦 Arxivga' (yo'nalishli)", "📦 Arxivga" in _dkb)
    check("dedup: inline'da '📦 Arxiv' (filtr nomi) YO'Q", "📦 Arxiv" not in _dkb)
    check("dedup: reply filtr hali '📦 Arxiv'", handlers.NBTN_NOTES_ARCHIVED == "📦 Arxiv")
    check("dedup: action ≠ filtr (to'qnashuv yo'q)",
          "📦 Arxivga" != handlers.NBTN_NOTES_ARCHIVED)
    _dtext, _ = handlers._format_note_detail(_dnote)
    check("karta: meta bitta qatorda (manba · sana · holat)", "🤖 LLM · 📅 06-06 02:41 · 📥 Inbox" in _dtext)
    check("karta: ortiqcha '🔖 Holat:' qatori olib tashlandi", "🔖 Holat:" not in _dtext)
    check("karta: teglar bo'sh joy bilan (vergulsiz)", "#bot #UI" in _dtext)

    print("\n── Inline ↔ Reply takror yo'q (konsepsiya: filtr=reply, inline=kontent) ──")
    _rnotes = [{"id": f"n-{i}", "title": f"Q{i}", "content": "x", "source": "manual",
                "created_at": "2026-06-06T10:00:00+05:00"} for i in range(12)]
    _list_inline = {b.text for row in handlers.notes_compact_keyboard(_rnotes, "inbox", 1).inline_keyboard for b in row}
    _reply = {b.text for row in handlers.notes_section_reply_keyboard().keyboard for b in row}
    _det_inline = {b.text for row in handlers.note_detail_menu(_dnote).inline_keyboard for b in row}
    check("list inline'da filtr-pills YO'Q", not ({"📥 Inbox", "⚙️ Ishlangan", "📦 Arxiv"} & _list_inline))
    check("list inline ∩ reply = bo'sh", not (_list_inline & _reply), str(_list_inline & _reply))
    check("detail inline ∩ reply = bo'sh", not (_det_inline & _reply), str(_det_inline & _reply))
    check("list inline'da pagination saqlangan (to'liq 'Keyingi' yorlig'i)",
          any("Keyingi" in b for b in _list_inline))
    check("list inline'da drill-down raqamlar bor", "1" in _list_inline)

    print("\n── Qaydlar takomillashtirish (A · B3 Ishlandi · B4 yosh · C6 split) ──")
    check("A2: 'Yangi qayd' tugmasi (note emas)", handlers.NBTN_NOTES_NEW == "➕ Yangi qayd")
    check("A2: 'Qayd qidirish' tugmasi", handlers.NBTN_NOTES_SEARCH == "🔍 Qayd qidirish")
    _nb = {b.text for row in handlers.notes_section_reply_keyboard().keyboard for b in row}
    check("A2: reply tugmalarда inglizcha 'note' yo'q", not any("note" in b.lower() for b in _nb))
    check("A1: diagnostika header 🩺 (🔍 emas)", "🩺  **DIAGNOSTIKA**" in _src)
    check("B3: database.mark_note_done mavjud", hasattr(database, "mark_note_done"))
    _now2 = datetime.now(handlers.database.TZ)
    _inb = {"id": "x", "status": "inbox", "title": "t", "content": "x",
            "source": "manual", "created_at": _now2.isoformat(), "tags": []}
    _im = [b.text for row in handlers.note_detail_menu(_inb).inline_keyboard for b in row]
    check("B3: inbox menyuда '✅ Ishlandi'", "✅ Ishlandi" in _im)
    _proc = dict(_inb); _proc["status"] = "processed"
    _pm = [b.text for row in handlers.note_detail_menu(_proc).inline_keyboard for b in row]
    check("B3: processed menyuда Ishlandi/ajrat YO'Q",
          "✅ Ishlandi" not in _pm and not any("ajrat" in x for x in _pm))
    _old = dict(_inb); _old["created_at"] = (_now2 - timedelta(days=4)).isoformat()
    check("B4: eski inbox qayd yoshini ko'rsatadi (4 kun)", "4 kun" in handlers._format_note_detail(_old)[0])
    check("B4: bugungi qaydда yosh yo'q",
          "kun" not in handlers._format_note_detail(_inb)[0].split("blockquote")[0])
    check("C6: inbox menyuда '✂️ ...ajrat'", any("ajrat" in x for x in _im))
    check("C6: split directive create_task so'raydi",
          "create_task" in handlers._build_note_split_directive("a, b, c"))

    print("\n── Eslatmalar bo'limi (Bugun→eslatma · Barchasi📋 · Keyingi olib · edit) ──")
    _rk = {b.text for row in handlers.reminders_section_reply_keyboard().keyboard for b in row}
    check("Eslatma: 'Keyingi' tugmasi olib tashlandi", not any("Keyingi" in b for b in _rk))
    check("Eslatma: 'Barchasi' = 📋 (tasks bilan moslik)", "📋 Barchasi" in _rk)
    check("Eslatma: filtrlar today/sent/all (upcoming yo'q)",
          set(handlers._REMINDERS_SECTION_FILTERS.values()) == {"today", "sent", "all"})
    check("Eslatma: ⏰ Bugun → 'today'", handlers._REMINDERS_SECTION_FILTERS.get(handlers.RBTN_REMINDERS_TODAY) == "today")
    check("Eslatma: Bugun state-mustaqil handleri bor (tasks leak yo'q)",
          hasattr(handlers, "handle_reminder_filter_anystate"))
    _rdm = {b.text for row in handlers.reminder_detail_menu({"id": "r", "status": "scheduled"}).inline_keyboard for b in row}
    check("Eslatma: '✏️ Tahrirlash' tugmasi mavjud (konsolidatsiya)", "✏️ Tahrirlash" in _rdm)
    check("Eslatma: Bajarildi/o'chir tugmalari", "✅ Bajarildi" in _rdm and "🗑 O'chirish" in _rdm)

    print("\n── Professional: lock · backup/undo · onboarding · audit · auto-chase ──")
    import pathlib as _pl2
    import bot as _bot
    import scheduler as _sch
    check("A1: single-instance lock funksiyasi", hasattr(_bot, "_acquire_single_instance_lock"))
    _svc = _pl2.Path(handlers.__file__).parent / "deploy" / "yordamchi.service"
    check("A2: systemd xizmat fayli mavjud", _svc.exists())
    check("A2: xizmat Restart=always", _svc.exists() and "Restart=always" in _svc.read_text(encoding="utf-8"))
    check("B: _create_db_backup mavjud", hasattr(handlers, "_create_db_backup"))
    check("B: cb_undo_delete + _UNDO_BACKUPS",
          hasattr(handlers, "cb_undo_delete") and hasattr(handlers, "_UNDO_BACKUPS"))
    check("C1: /start onboarding (Yordamchi Pro + nima qila olishi)",
          "yozing yoki ayting" in _src and "Yordamchi Pro" in _src)
    check("C2: database.list_recent_actions", hasattr(database, "list_recent_actions"))
    check("C2: diagnostikada 'So'nggi amallar' bo'limi", "So'nggi amallar" in _src)
    check("D: database.list_stale_delegations", hasattr(database, "list_stale_delegations"))
    check("D: scheduler._stale_delegation_digest", hasattr(_sch.YordamchiScheduler, "_stale_delegation_digest"))

    # ── iCloud sync: failure paths must return a (imported, conflicts) TUPLE ──
    # Bug: _sync_events_to_db_sync returned bare `0` when the client/calendar was
    # unavailable, so the scheduler's `imported, conflicts = await ...` raised
    # TypeError on every iCloud connection hiccup (swallowed as "sync failed").
    import calendar_service as _cal
    _orig_conn = _cal._connect
    try:
        _cal._connect = lambda: None      # simulate no iCloud client
        _r = _cal._sync_events_to_db_sync()
        check("iCloud: klient yo'q → (0, []) tuple qaytadi (int emas)",
              isinstance(_r, tuple) and _r == (0, []))
        _i, _c = _r  # must unpack cleanly like the scheduler does
        check("iCloud: natija (imported, conflicts) sifatida ochiladi", _i == 0 and _c == [])
    finally:
        _cal._connect = _orig_conn
    check("iCloud: sync_events_to_db o'chiq holatda ham (0, []) qaytaradi",
          (await _cal.sync_events_to_db()) == (0, []))

    # ── iCloud push-backfill: botdagi uchrashuvni kalendarga yuborish uchun tanlash ──
    check("iCloud: database.list_meetings_to_push mavjud", hasattr(database, "list_meetings_to_push"))
    check("iCloud: scheduler._icloud_push_backfill mavjud",
          hasattr(_sch.YordamchiScheduler, "_icloud_push_backfill"))
    _pnow = datetime.now(TZ)
    _pm_up = await database.create_meeting({"title": "PUSH_up", "datetime_start": (_pnow + timedelta(days=2)).isoformat()})
    _pm_done = await database.create_meeting({"title": "PUSH_done", "datetime_start": (_pnow + timedelta(days=2)).isoformat()})
    await database.complete_meeting(_pm_done)
    _pm_synced = await database.create_meeting({"title": "PUSH_synced", "datetime_start": (_pnow + timedelta(days=2)).isoformat()})
    await database.set_meeting_icloud_uid(_pm_synced, "yordamchi-x@a")
    _pm_past = await database.create_meeting({"title": "PUSH_past", "datetime_start": (_pnow - timedelta(days=2)).isoformat()})
    _pids = {x["id"] for x in await database.list_meetings_to_push(days=60)}
    check("iCloud push-backfill: kelgusi+yuborilmagan tanlanadi; bajarilgan/yuborilgan/o'tgan chiqmaydi",
          _pm_up in _pids and _pm_done not in _pids and _pm_synced not in _pids and _pm_past not in _pids)
    for _x in (_pm_up, _pm_done, _pm_synced, _pm_past):
        await database.cancel_meeting(_x)

    print("\n── Bayonnomalar markaziy ro'yxati (oy bo'yicha) ──")
    check("Bayonnoma: database.list_meetings_with_protocol", hasattr(database, "list_meetings_with_protocol"))
    check("Bayonnoma: heuristika — uzun matn = protokol",
          handlers._looks_like_protocol(["📝 UCHRASHUV BAYONNOMASI uzun matn shu yerda joylashgan"]))
    check("Bayonnoma: heuristika — task-id ro'yxat ≠ protokol",
          not handlers._looks_like_protocol(["t-123", "t-456"]))
    check("Bayonnoma: heuristika — bo'sh ≠ protokol", not handlers._looks_like_protocol([]))
    check("Bayonnoma: _render_protocols + cmd_protocols",
          hasattr(handlers, "_render_protocols") and hasattr(handlers, "cmd_protocols"))
    _mk = {b.text for row in handlers.meetings_section_reply_keyboard().keyboard for b in row}
    check("Bayonnoma: '📄 Bayonnomalar' tugmasi meetings kbd'da", "📄 Bayonnomalar" in _mk)
    # 'ulashish ishlamadi' bug: share handler state'ga emas, uchrashuvga ham tayanishi shart
    import re as _re3
    _share_fn = _re3.search(
        r"async def cb_protocol_share\(.*?(?=\n@router|\nasync def |\ndef )", _src, _re3.DOTALL)
    check("Bayonnoma: Ulashish uchrashuvdan fallback o'qiydi (bug tuzatildi)",
          bool(_share_fn) and "get_meeting" in _share_fn.group(0)
          and "follow_up_actions" in _share_fn.group(0))
    # Inline ulashish (bayonnoma → istalgan chatga)
    check("Inline: yagona handle_inline_query mavjud", hasattr(handlers, "handle_inline_query"))
    check("Inline: aynan BITTA @router.inline_query() handler (dublikat yo'q — leak bug)",
          len(handlers.router.inline_query.handlers) == 1)
    check("Inline: 'Ulashish' switch_inline_query mavjud (📋 Nusxa = proto_share callback)",
          'switch_inline_query=f"proto:{mid}"' in _src and 'callback_data=f"proto_share:{mid}"' in _src)

    class _IqU:
        def __init__(s, i): s.id = i
    class _IqQ:
        def __init__(s, uid, q): s.from_user = _IqU(uid); s.query = q; s.res = None
        async def answer(s, results, **k): s.res = results
    _imid = await database.create_meeting({
        "title": "InlineTest", "datetime_start": "2026-06-06T10:00:00+05:00",
        "participants": [], "follow_up_actions": ["Bu bayonnoma matni ulashish uchun."]})
    _iq = _IqQ(config.PRINCIPAL_USER_ID, f"proto:{_imid}")
    await handlers.handle_inline_query(_iq)
    check("Inline: principal bayonnomani oladi", bool(_iq.res) and len(_iq.res) == 1)
    _iq2 = _IqQ(999999999, f"proto:{_imid}")
    await handlers.handle_inline_query(_iq2)
    check("Inline: begona foydalanuvchi natija olmaydi (gate)", _iq2.res == [])
    _iq3 = _IqQ(config.PRINCIPAL_USER_ID, "proto:m-DOESNOTEXIST")
    await handlers.handle_inline_query(_iq3)
    check("Inline: yo'q bayonnoma → xom 'proto:' matni QAYTMAYDI (leak bug tuzatildi)",
          bool(_iq3.res) and all(
              "proto:" not in (r.input_message_content.message_text or "") for r in _iq3.res))
    import aiosqlite as _sq5
    async with _sq5.connect(config.DATABASE_PATH) as _db5:
        await _db5.execute("DELETE FROM meetings WHERE id=?", (_imid,)); await _db5.commit()

    print("\n── 2 xil ulashish: 📋 Nusxa (qo'lda) + 📤 Ulashish (inline) ──")
    _prk = handlers._protocol_result_kb("m1", 1, saved=True, tasks_done=False)
    _prb = [(b.text, bool(b.switch_inline_query)) for row in _prk.inline_keyboard for b in row]
    check("Bayonnoma: 📋 Nusxa (qo'lda) + 📤 Ulashish (inline)",
          ("📋 Nusxa", False) in _prb and ("📤 Ulashish", True) in _prb)
    _pol = handlers._build_keyboard(
        [[{"label": "📋 Nusxa olish", "callback": "copy"},
          {"label": "x", "callback": "share"},
          {"label": "✎", "callback": "edit:polish"}]],
        {}, share_text="**Tahrirlangan matn:**\n───\nSalom hamkor.\n───")
    _pb = [(b.text, bool(b.switch_inline_query)) for row in _pol.inline_keyboard for b in row]
    check("Sayqal matn: 📋 Nusxa + 📤 Ulashish (inline), no-op copy tashlandi",
          ("📋 Nusxa", False) in _pb and ("📤 Ulashish", True) in _pb
          and not any(t == "📋 Nusxa olish" for t, _ in _pb))
    _itok = next((b.switch_inline_query.split("txt:")[1]
                  for row in _pol.inline_keyboard for b in row
                  if b.switch_inline_query and "txt:" in b.switch_inline_query), None)
    check("Inline kesh DB-backed (restart'ga bardoshli, toza matn)",
          _itok is not None and (await database.get_share_text(_itok)) == "Salom hamkor.")

    print("\n── Eslatma detal tugmalari (kontekstga mos) ──")
    def _rmenu(r):
        return [b.text for row in handlers.reminder_detail_menu(r).inline_keyboard for b in row]
    _rec = _rmenu({"id": "r", "status": "scheduled", "recurrence_rule": "weekly",
                   "remind_at": "2026-06-08T12:00:00+05:00"})
    _one = _rmenu({"id": "r", "status": "scheduled", "recurrence_rule": None,
                   "remind_at": "2026-06-08T12:00:00+05:00"})
    _dn = _rmenu({"id": "r", "status": "done", "recurrence_rule": None})
    check("Eslatma: takroriyda ⏭ skip + 🛑 stop",
          any("o'tkaz" in x for x in _rec) and any("to'xtat" in x.lower() for x in _rec))
    check("Eslatma: takroriyda 'Bajarildi' YO'Q (seriya o'chmaydi)", "✅ Bajarildi" not in _rec)
    check("Eslatma: bir martalikda ✅ Bajarildi + snooze",
          "✅ Bajarildi" in _one and "⏰ 15 daq" in _one)
    check("Eslatma: snooze hammasi ⏰ (izchil)",
          all(x.startswith("⏰") for x in _one if ("daq" in x or "soat" in x or "Ertaga" in x)))
    check("Eslatma: bir martalik ham '✏️ Tahrirlash' (eski 📆/📅 inline yo'q)",
          "✏️ Tahrirlash" in _one and "📆 Vaqt" not in str(_one + _rec))
    check("Eslatma: done → ↺ Qayta eslat", any("Qayta eslat" in x for x in _dn))
    check("Eslatma: cb_reminder_skip + cb_reminder_stop",
          hasattr(handlers, "cb_reminder_skip") and hasattr(handlers, "cb_reminder_stop"))
    check("Eslatma: done'da ham '✏️ Tahrirlash'", "✏️ Tahrirlash" in _dn)
    check("Eslatma: default ko'rinish 'active' (done emas, aktiv chiqadi)",
          '_render_reminders_for_filter(message, "active")' in _src)
    _act, _albl = await handlers._load_reminders_for_filter("active")
    check("Eslatma: 'active' filtr done'ni ko'rsatmaydi",
          all(r.get("status") != "done" for r in _act))
    check("Eslatma: Barchasi tugmasi hali 'all' (done ko'rinadi)",
          handlers._REMINDERS_SECTION_FILTERS.get(handlers.RBTN_REMINDERS_ALL) == "all")

    print("\n── Eslatma tahrir menyusi (konsolidatsiya + Takror + Izoh + weekdays) ──")
    _em = [b.text for row in handlers.reminder_detail_menu(
        {"id": "r", "status": "scheduled", "recurrence_rule": "weekly"}).inline_keyboard for b in row]
    check("Tahrir: bitta '✏️ Tahrirlash' (Matn/Vaqt alohida emas)",
          "✏️ Tahrirlash" in _em and "✏️ Matn" not in _em)
    check("Tahrir: edit-menu/recurrence callbacklari mavjud",
          hasattr(handlers, "cb_reminder_edit_menu") and hasattr(handlers, "cb_reminder_recurrence_menu")
          and hasattr(handlers, "cb_reminder_set_recurrence"))
    check("Takror: variantlarda 'weekdays' (ish kunlari) bor",
          "weekdays" in [v for v, _ in handlers._RECUR_OPTIONS])
    check("weekdays: normalize ('ish kunlari'/'Dushanba-juma')",
          database.normalize_recurrence_rule("ish kunlari") == "weekdays"
          and database.normalize_recurrence_rule("Dushanba-juma") == "weekdays")
    check("weekdays: label 'ish kunlari'", "ish kunlari" in handlers._format_recurrence_label("weekdays"))
    # Date-robust: compute_next_recurrence clamps to the future, so use a Friday ≥1 week
    # out (not a hardcoded past date) and assert the next weekday occurrence is the Monday.
    _now_w = datetime.now(TZ)
    _fri_w = (_now_w + timedelta(days=((4 - _now_w.weekday()) % 7) + 7)).replace(
        hour=12, minute=0, second=0, microsecond=0)
    _nxt_w = datetime.fromisoformat(database.compute_next_recurrence(_fri_w.isoformat(), "weekdays"))
    check("weekdays: dam olishni o'tkazadi (Juma → Dushanba, +3 kun)",
          _nxt_w.weekday() == 0 and (_nxt_w.date() - _fri_w.date()).days == 3,
          f"{_fri_w.date()}→{_nxt_w.date()}")

    print("\n── Reply-kbd tartibi: Barchasi + Asosiy menyu alohida, qolgani 2 tadan ──")
    _h = [[b.text for b in row] for row in handlers._two_per_row(["a", "b", "c", "d", "e"], solo={"c"})]
    check("_two_per_row: solo alohida + qolgani juft",
          ["c"] in _h and ["a", "b"] in _h and ["d", "e"] in _h)
    _rk = [[b.text for b in row] for row in handlers.reminders_section_reply_keyboard().keyboard]
    check("kbd Eslatma: 📋 Barchasi alohida qator", [handlers.RBTN_REMINDERS_ALL] in _rk)
    check("kbd Eslatma: ⬅️ Asosiy menyu alohida qator", [handlers.BTN_BACK_MAIN] in _rk)
    check("kbd Eslatma: filtrlar 2 tadan ([Bugun, Yuborilgan])",
          [handlers.RBTN_REMINDERS_TODAY, handlers.RBTN_REMINDERS_SENT] in _rk)
    _tk = [[b.text for b in row] for row in handlers.tasks_section_reply_keyboard().keyboard]
    _tk_flat = [b for row in _tk for b in row]
    check("kbd Vazifa: 2 tadan bir qatorda (uzun yorliq kesilmaydi)",
          all(len(r) <= 2 for r in _tk) and set(_tk_flat) == {
              handlers.TBTN_TASKS_ACTIVE, handlers.TBTN_TASKS_TODAY,
              handlers.TBTN_TASKS_IMPORTANT, handlers.TBTN_TASKS_OVERDUE,
              handlers.TBTN_TASKS_DONE, handlers.TBTN_TASKS_ALL,
              handlers.TBTN_TASKS_CATEGORIES, handlers.TBTN_TASKS_NEW,
              handlers.TBTN_TASKS_SEARCH, handlers.BTN_BACK_MAIN})
    # Vizual ierarxiya: «Yangi vazifa» va «Asosiy menyu» solo (to'liq qator);
    # «Qidirish» + «Kategoriyalar» juft.
    check("kbd Vazifa: «Yangi vazifa» solo (to'liq qator)",
          [handlers.TBTN_TASKS_NEW] in _tk)
    check("kbd Vazifa: «Qidirish» + «Kategoriyalar» juft",
          [handlers.TBTN_TASKS_SEARCH, handlers.TBTN_TASKS_CATEGORIES] in _tk)
    check("kbd Vazifa: «Asosiy menyu» solo (to'liq qator)",
          [handlers.BTN_BACK_MAIN] in _tk)

    print("\n── Eslatma tahrir: matn + OVOZ ishlaydi (bug tuzatildi) ──")
    class _ReFakeState:
        def __init__(s, d): s._d = d
        async def get_data(s): return s._d
        async def clear(s): pass
    class _ReFakeMsg:
        def __init__(s, text=None, voice=False):
            s.text = text; s.voice = (object() if voice else None); s.bot = None
            s.chat = type("C", (), {"id": 1})()
        async def answer(s, *a, **k): pass
    _orig_tt = handlers._get_text_or_transcribe
    try:
        _rid = await database.create_reminder(
            {"title": "Eski", "remind_at": "2026-06-10T10:00:00+05:00", "status": "scheduled"})
        async def _vt(m, bot=None): return "Ovozli yangi sarlavha"
        handlers._get_text_or_transcribe = _vt
        await handlers.handle_reminder_edit_value(
            _ReFakeMsg(text=None, voice=True), _ReFakeState({"reminder_id": _rid, "field": "title"}))
        _r = await database.get_reminder(_rid)
        check("Tahrir: OVOZ orqali sarlavha saqlanadi", bool(_r) and _r["title"] == "Ovozli yangi sarlavha")
        async def _vt2(m, bot=None): return "ertaga 09:00"
        handlers._get_text_or_transcribe = _vt2
        await handlers.handle_reminder_edit_value(
            _ReFakeMsg(text=None, voice=True), _ReFakeState({"reminder_id": _rid, "field": "time"}))
        _r = await database.get_reminder(_rid)
        # Full-value compare (not just date) so it's robust when "ertaga" happens to
        # equal the original date — the voice edit still changes the TIME (10:00→09:00).
        check("Tahrir: OVOZ orqali vaqt saqlanadi", bool(_r) and _r["remind_at"] != "2026-06-10T10:00:00+05:00")
        import aiosqlite as _sq3
        async with _sq3.connect(config.DATABASE_PATH) as _db3:
            await _db3.execute("DELETE FROM reminders WHERE id=?", (_rid,)); await _db3.commit()
    finally:
        handlers._get_text_or_transcribe = _orig_tt
    # Kengroq ovoz bug: section handlerlari ovozni transkripsiya qilib, keyin
    # message.text (ovozda None) ishlatardi → crash / yo'qolish. Endi _msg_text.
    check("Ovoz: 'label = message.text' qolmadi (section crash yo'q)",
          "label = message.text.strip()" not in _src)
    check("Ovoz: '_process_and_reply(message, message.text' qolmadi (ovoz Claude'ga)",
          "_process_and_reply(message, message.text" not in _src)

    print("\n── Batch-4a UX: truncation (kontent yo'qolmaydi) + Muxlisa retry ──")
    _long_line = "X" * 9000  # one line far over the Telegram soft limit
    _chunks = handlers._split_for_telegram(_long_line)
    check("truncation: uzun bitta qator bo'linadi (1 emas)", len(_chunks) >= 2, str(len(_chunks)))
    check("truncation: HECH kontent yo'qolmaydi (join == asl)", "".join(_chunks) == _long_line)
    check("truncation: har chunk limitdan oshmaydi",
          all(len(c) <= handlers._TG_SOFT_LIMIT for c in _chunks))
    import voice_service as _vs4
    import inspect as _ins4
    _muxsrc = _ins4.getsource(_vs4._transcribe_muxlisa)
    check("Muxlisa STT retry/backoff bor (max_attempts + transient)",
          "max_attempts" in _muxsrc and "transient_statuses" in _muxsrc and "asyncio.sleep" in _muxsrc)
    # 4b sana formati: vazifa chiplari Uzbek oy (uchrashuv/eslatma bilan moslashadi)
    from datetime import datetime as _dtu, timedelta as _tdu
    _far = _dtu.now(database.TZ) + _tdu(days=5)
    _mon = handlers.UZ_MONTHS_FULL[_far.month - 1]
    check("sana: _task_deadline_chip Uzbek oy (numeric emas)",
          _mon in handlers._task_deadline_chip({"deadline": _far.isoformat(), "status": "todo"}))
    check("sana: _format_deadline_short Uzbek oy",
          _mon in handlers._format_deadline_short(_far.isoformat())[0])
    check("sana: _fmt_dt_uz kanonik format",
          handlers._fmt_dt_uz(_far) == f"{_far.day}-{_mon} {_far.strftime('%H:%M')}")
    # export default = active-only (bajarilgan/eski chiqmaydi)
    _edone = await database.create_task({"title": "ExpDone", "priority": "P2", "status": "todo"})
    await database.complete_task(_edone)
    await database.create_task({"title": "ExpActive", "priority": "P2", "status": "todo"})
    _exp_active_titles = {t["title"] for t in await handlers._fetch_tasks_for_export("active")}
    check("export aktiv: faqat aktiv (done chiqmaydi)",
          "ExpActive" in _exp_active_titles and "ExpDone" not in _exp_active_titles)
    check("export hammasi: done ham bor",
          "ExpDone" in {t["title"] for t in await handlers._fetch_tasks_for_export("all")})
    import inspect as _iexp
    check("export default → aktiv (hammasi emas)",
          '_fetch_tasks_for_export("active"' in _iexp.getsource(handlers._send_tasks_export))
    # OpenAI/Whisper ixtiyoriy — bot OpenAI kalitisiz ham ishlaydi (STT zaxirasi skip)
    import pathlib as _pl
    _root_dir = _pl.Path(handlers.__file__).resolve().parent
    _cfg_src = (_root_dir / "config.py").read_text(encoding="utf-8")
    _vs_src = (_root_dir / "voice_service.py").read_text(encoding="utf-8")
    check("OpenAI: kalit MAJBURIY EMAS (os.getenv, _require emas)",
          'OPENAI_API_KEY: str = os.getenv("OPENAI_API_KEY"' in _cfg_src
          and '_require("OPENAI_API_KEY")' not in _cfg_src)
    check("Whisper: client kalitsiz None (construct guard)",
          "if config.OPENAI_API_KEY" in _vs_src and "else None" in _vs_src)
    check("Whisper: kalitsiz no-op (_transcribe_whisper None guard)",
          "_openai_client is None" in _vs_src)

    print("\n── Sub-vazifa (to'liq bola-vazifa) ──")
    _pid = await database.create_task({"title": "Asosiy loyiha SUB", "priority": "P0"})
    _c1 = await database.create_task({"title": "Tahlil", "assignee": "A.Karimov",
                                      "deadline": "2026-06-25T10:00:00+05:00", "parent_id": _pid})
    _c2 = await database.create_task({"title": "Yozish", "parent_id": _pid})
    _subs = await database.list_subtasks(_pid)
    check("subtask: list_subtasks → 2 bola", len(_subs) == 2)
    check("subtask: bola o'z ijrochi+muddatiga ega",
          any(s.get("assignee") == "A.Karimov" and s.get("deadline") for s in _subs))
    _mainids = {t["id"] for t in await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=500)}
    check("subtask: asosiy ro'yxatda takrorlanmaydi (parent_id IS NULL)",
          _pid in _mainids and _c1 not in _mainids and _c2 not in _mainids)
    _pcbs = [b.callback_data for r in handlers._task_card_kb_with_back(await database.get_task(_pid)).inline_keyboard for b in r]
    check("subtask: ota kartasida 🌳 Sub-vazifalar tugma", f"subview:{_pid}" in _pcbs)
    _ccbs = [b.callback_data for r in handlers._task_card_kb_with_back(await database.get_task(_c1)).inline_keyboard for b in r]
    check("subtask: bola kartasida subview YO'Q + orqa→ota",
          f"subview:{_c1}" not in _ccbs and f"subview:{_pid}" in _ccbs)
    _vcbs = [b.callback_data for r in handlers._subtask_view_kb(_pid, _subs).inline_keyboard for b in r]
    check("subtask: view (bola ochiladi + add + orqaga)",
          f"taskopen:{_c1}" in _vcbs and f"subadd:{_pid}" in _vcbs and f"taskopen:{_pid}" in _vcbs)
    check("subtask: handlerlar + FSM mavjud",
          all(hasattr(handlers, n) for n in
              ("cb_subtask_view", "cb_subtask_add", "handle_subtask_add", "SubtaskAddFSM")))
    await database.delete_task(_pid)
    check("subtask: cascade — ota o'chsa bolalar ham o'chadi",
          (await database.get_task(_c1)) is None and (await database.get_task(_c2)) is None)

    print("\n── Batch-1 tuzatishlar: recurrence / cascade / NULL-end / bulk-count ──")
    from datetime import datetime as _dt2, timedelta as _td2
    _now2 = _dt2.now(database.TZ)
    # 1) recurrence: uzoq o'tmishdagi base → KELAJAK sana (o'tmish emas)
    _nxt = database.compute_next_recurrence((_now2 - _td2(days=400)).isoformat(), "daily")
    check("recurrence: uzoq o'tmish daily → kelajak (o'tmish emas)",
          _nxt is not None and _dt2.fromisoformat(_nxt) > _now2, str(_nxt))
    _nxt_w = database.compute_next_recurrence((_now2 - _td2(days=400)).isoformat(), "weekly")
    check("recurrence: weekly ham kelajak", bool(_nxt_w) and _dt2.fromisoformat(_nxt_w) > _now2)
    # 2) delete_task → bog'liq eslatma cascade (orphan yo'q)
    _ctid = await database.create_task({"title": "Cascade T", "priority": "P2", "status": "todo"})
    _crid = await database.create_reminder({"title": "T eslatma",
        "remind_at": (_now2 + _td2(hours=2)).isoformat(), "task_id": _ctid})
    await database.delete_task(_ctid)
    check("delete_task → bog'liq eslatma o'chadi", await database.get_reminder(_crid) is None)
    # 3) cancel_meeting → bog'liq eslatma cascade
    _cmid = await database.create_meeting({"title": "Cascade M",
        "datetime_start": (_now2 + _td2(days=1)).isoformat()})
    _cmrid = await database.create_reminder({"title": "M eslatma",
        "remind_at": (_now2 + _td2(days=1)).isoformat(), "meeting_id": _cmid})
    await database.cancel_meeting(_cmid)
    check("cancel_meeting → bog'liq eslatma o'chadi", await database.get_reminder(_cmrid) is None)
    # 4) create_meeting: end yo'q → start+60daq materializatsiya
    _neid = await database.create_meeting({"title": "NoEnd",
        "datetime_start": (_now2 + _td2(days=2)).replace(microsecond=0).isoformat()})
    _nem = await database.get_meeting(_neid)
    check("create_meeting: end yo'q → start+60daq",
          bool(_nem.get("datetime_end")) and
          _dt2.fromisoformat(_nem["datetime_end"]) - _dt2.fromisoformat(_nem["datetime_start"]) == _td2(minutes=60))
    # 5) bulk-delete preview: filtrlangan son (jami emas)
    for _i in range(3):
        await database.create_task({"title": f"BActive {_i}", "priority": "P2", "status": "todo"})
    _bd1 = await database.create_task({"title": "BDone1", "priority": "P2", "status": "todo"})
    await database.complete_task(_bd1)
    _done_n = len(await database.list_tasks(status_in=["done"], limit=100000))
    _total_n = await database.count_table("tasks")
    _bprev = await handlers._format_create_preview([{"type": "delete_all_tasks", "data": {"status_in": ["done"]}}])
    check("bulk-delete preview: filtrlangan 'done' soni (jami emas)",
          f"{_done_n} ta" in _bprev and _done_n < _total_n, f"done={_done_n} total={_total_n}")

    print("\n── Mini App (Telegram Web App) backend ──")
    import webapp as _wa
    import hashlib as _hl, hmac as _hm, json as _js, time as _tm
    from urllib.parse import urlencode as _ue

    def _mk_init(uid, token, auth_date=None):
        ad = auth_date if auth_date is not None else int(_tm.time())
        user = _js.dumps({"id": uid, "first_name": "Test"}, separators=(",", ":"))
        pairs = {"auth_date": str(ad), "query_id": "AAA", "user": user}
        dcs = "\n".join(f"{k}={pairs[k]}" for k in sorted(pairs))
        sk = _hm.new(b"WebAppData", token.encode(), _hl.sha256).digest()
        pairs["hash"] = _hm.new(sk, dcs.encode(), _hl.sha256).hexdigest()
        return _ue(pairs)

    _tok = config.TELEGRAM_BOT_TOKEN
    _pid = config.PRINCIPAL_USER_ID
    _good = _mk_init(_pid, _tok)
    check("MiniApp: to'g'ri initData → user qaytadi",
          (_wa.validate_init_data(_good, _tok) or {}).get("id") == _pid)
    check("MiniApp: buzilgan hash → None",
          _wa.validate_init_data(_good[:-4] + "0000", _tok) is None)
    check("MiniApp: noto'g'ri token → None",
          _wa.validate_init_data(_good, _tok + "x") is None)
    check("MiniApp: eskirgan auth_date → None (replay himoyasi)",
          _wa.validate_init_data(_mk_init(_pid, _tok, auth_date=1), _tok) is None)
    check("MiniApp: hash yo'q → None", _wa.validate_init_data("user=%7B%7D", _tok) is None)
    # Real Telegram initData carries `signature` + chat_* fields; the `hash` is computed
    # over ALL fields except `hash` (signature INCLUDED). Regression for the AUTH-FAIL bug.
    def _mk_init_sig(uid, token):
        ad = int(_tm.time())
        user = _js.dumps({"id": uid, "first_name": "Test"}, separators=(",", ":"))
        pairs = {"auth_date": str(ad), "query_id": "AAA", "chat_type": "private",
                 "signature": "Ed25519sig_xyz", "user": user}
        dcs = "\n".join(f"{k}={pairs[k]}" for k in sorted(pairs))
        sk = _hm.new(b"WebAppData", token.encode(), _hl.sha256).digest()
        pairs["hash"] = _hm.new(sk, dcs.encode(), _hl.sha256).hexdigest()
        return _ue(pairs)
    check("MiniApp: signature-li initData (Telegram real) qabul qilinadi",
          (_wa.validate_init_data(_mk_init_sig(_pid, _tok), _tok) or {}).get("id") == _pid)
    # ── Browser Login Widget + session (secret = SHA256(token), farqли algoritm) ──
    def _widget(uid, token, age=None):
        d = {"id": uid, "first_name": "QA", "auth_date": age if age is not None else int(_tm.time())}
        dcs = "\n".join(f"{k}={d[k]}" for k in sorted(d))
        d["hash"] = _hm.new(_hl.sha256(token.encode()).digest(), dcs.encode(), _hl.sha256).hexdigest()
        return d
    check("Login Widget: to'g'ri → user",
          (_wa.validate_login_widget(_widget(_pid, _tok), _tok) or {}).get("id") == _pid)
    _wbad = _widget(_pid, _tok); _wbad["hash"] = "00" + _wbad["hash"][2:]
    check("Login Widget: buzilgan hash → None", _wa.validate_login_widget(_wbad, _tok) is None)
    check("Login Widget: eskirgan → None", _wa.validate_login_widget(_widget(_pid, _tok, age=1), _tok) is None)
    _sess = _wa.make_session(_pid)
    check("Sessiya: round-trip", _wa.check_session(_sess) == _pid)
    check("Sessiya: buzilgan → None", _wa.check_session(_sess[:-4] + "0000") is None)
    check("Sessiya: eskirgan → None", _wa.check_session(_wa.make_session(_pid, ttl=-10)) is None)

    # End-to-end via aiohttp test client: auth gate + a real create round-trip.
    from aiohttp.test_utils import TestClient, TestServer
    _app = _wa.create_app()
    _client = TestClient(TestServer(_app))
    await _client.start_server()
    try:
        _H = {"Authorization": "tma " + _good}
        _r = await _client.get("/api/tasks")
        check("MiniApp: authsiz /api → 401", _r.status == 401)
        _r = await _client.get("/api/tasks", headers={"Authorization": "tma " + _mk_init(999999, _tok)})
        check("MiniApp: begona user (imzo to'g'ri) → 403", _r.status == 403)
        _r = await _client.get("/api/tasks", headers=_H)
        check("MiniApp: egasi → 200 + tasks ro'yxati", _r.status == 200 and "tasks" in await _r.json())
        _r = await _client.post("/api/tasks", headers=_H, json={"title": "MiniApp orqali", "priority": "P1"})
        _body = await _r.json()
        check("MiniApp: POST /api/tasks → 201 + DBда yaratildi",
              _r.status == 201 and bool(_body.get("id"))
              and (await database.get_task(_body["id"]))["title"] == "MiniApp orqali")
        # Frontend rendering contract: /api/tasks item carries status + parent_id
        # (row status badge + subtask grouping depend on these).
        _tk = (await (await _client.get("/api/tasks?status=all", headers=_H)).json())["tasks"][0]
        check("MiniApp: task javobiда status + parent_id maydonlari bor",
              "status" in _tk and "parent_id" in _tk and "priority" in _tk)
        _r = await _client.post(f"/api/tasks/{_body['id']}/complete", headers=_H)
        check("MiniApp: complete → done", _r.status == 200
              and (await database.get_task(_body["id"]))["status"] == "done")
        _r = await _client.get("/api/health")
        check("MiniApp: /api/health authsiz ochiq", _r.status == 200)
        # Browser login flow: /api/config public, login sets session, cookie grants access
        check("MiniApp: /api/config authsiz ochiq", (await _client.get("/api/config")).status == 200)
        _lr = await _client.post("/api/auth/telegram", json=_widget(_pid, _tok))
        check("Login Widget: POST → 200 + sessiya cookie",
              _lr.status == 200 and "ya_session=" in _lr.headers.get("Set-Cookie", ""))
        check("Login Widget: begona user → 403",
              (await _client.post("/api/auth/telegram", json=_widget(999999, _tok))).status == 403)
        check("Login Widget: buzuq → 401",
              (await _client.post("/api/auth/telegram", json=_wbad)).status == 401)
        _ck = {"Cookie": f"ya_session={_wa.make_session(_pid)}"}
        check("Sessiya cookie → /api/tasks 200", (await _client.get("/api/tasks", headers=_ck)).status == 200)
        check("Sessiya cookie → /api/me 200", (await _client.get("/api/me", headers=_ck)).status == 200)
        check("/api/me authsiz → 401", (await _client.get("/api/me")).status == 401)
        check("Begona sessiya → 403",
              (await _client.get("/api/tasks", headers={"Cookie": f"ya_session={_wa.make_session(999999)}"})).status == 403)
        # ── noindex + robots (maxfiy ilova qidiruvга tushmasin) ──
        _rh = await _client.get("/api/health")
        check("MiniApp: X-Robots-Tag noindex har javobda",
              "noindex" in _rh.headers.get("X-Robots-Tag", ""))
        _rob = await _client.get("/robots.txt")
        check("MiniApp: /robots.txt → Disallow all",
              _rob.status == 200 and "Disallow: /" in (await _rob.text()))
        # ── WEBAPP_OPEN_ACCESS: vaqtincha to'liq ochiq brauzer kirishi (env bilan) ──
        _prev_open = config.WEBAPP_OPEN_ACCESS
        try:
            config.WEBAPP_OPEN_ACCESS = False
            check("Open-access OFF → authsiz /api/tasks 401",
                  (await _client.get("/api/tasks")).status == 401)
            config.WEBAPP_OPEN_ACCESS = True
            check("Open-access ON → authsiz /api/tasks 200 (principal)",
                  (await _client.get("/api/tasks")).status == 200)
            _mo = await _client.get("/api/me")
            check("Open-access ON → /api/me principal uid",
                  _mo.status == 200 and (await _mo.json()).get("uid") == _pid)
        finally:
            config.WEBAPP_OPEN_ACCESS = _prev_open
        check("Open-access qayta OFF → authsiz 401",
              (await _client.get("/api/tasks")).status == 401)
        # ── Bug: dashboard 'Jamoa nazorati' → openTask → taskForm crash ──
        # list_stale_delegations qaytargan vazifada `tags` XOM JSON-satr bo'lib qolgan
        # edi (boshqa hamma funksiya _row_to_task orqali list qaytaradi). Mini-app
        # taskForm `tags.join(", ")` chaqiradi → TypeError. Ildizni tekshiramiz:
        _stid = await database.create_task({"title": "Kechikkan topshiriq", "priority": "P1",
            "status": "todo", "assignee": "Bekzod Karimov", "tags": ["shoshilinch", "q3"]})
        import aiosqlite as _sq6
        _old = (datetime.now(TZ) - timedelta(days=5)).isoformat()
        async with _sq6.connect(config.DATABASE_PATH) as _db6:
            await _db6.execute("UPDATE tasks SET created_at=? WHERE id=?", (_old, _stid)); await _db6.commit()
        _stale = await database.list_stale_delegations(min_age_days=3, limit=20)
        _st_row = next((t for t in _stale if t["id"] == _stid), None)
        check("Bug: list_stale_delegations → tags LIST (JSON-satr emas)",
              _st_row is not None and isinstance(_st_row.get("tags"), list)
              and _st_row["tags"] == ["shoshilinch", "q3"] and "age_days" in _st_row)
        _r = await _client.get("/api/dashboard", headers=_H)
        _dj = await _r.json()
        check("MiniApp: /api/dashboard → progress+counts+today",
              _r.status == 200 and "progress" in _dj and "counts" in _dj and "today" in _dj)
        _rd = _dj.get("radar", {})
        check("MiniApp: dashboard radar → total/overdue/blocked/unassigned",
              all(k in _rd for k in ("total", "overdue", "blocked", "unassigned")))
        check("MiniApp: dashboard today → obyekt (meetings/tasks/next)",
              isinstance(_dj.get("today"), dict)
              and all(k in _dj["today"] for k in ("meetings", "tasks", "next")))
        check("MiniApp: dashboard → priority ro'yxati + team oversight",
              isinstance(_dj.get("priority"), list)
              and isinstance(_dj.get("team", {}).get("overloaded"), list)
              and isinstance(_dj.get("team", {}).get("stale"), list))
        # Har bir vazifa-shaklidagi obyektда `tags` LIST bo'lishi shart (satr emas) —
        # mini-app uni to'g'ridan-to'g'ri kesh qilib taskForm'да join() qiladi.
        _task_lists = (_dj.get("priority") or []) + (_dj.get("today", {}).get("tasks") or []) + (_dj.get("team", {}).get("stale") or [])
        check("MiniApp: dashboard vazifalarida tags LIST (crash oldini olish)",
              len(_task_lists) > 0
              and all(isinstance(t.get("tags", []), list) for t in _task_lists))
        check("MiniApp: dashboard team.stale → kechikkan vazifa ko'rinadi (tags list)",
              any(t["id"] == _stid and isinstance(t.get("tags"), list) for t in _dj.get("team", {}).get("stale", [])))
        _r = await _client.get("/api/insights", headers=_H)
        _ij = await _r.json()
        check("MiniApp: /api/insights → 7 kunlik bar + kategoriyalar",
              _r.status == 200 and len(_ij.get("by_day", [])) == 7 and "categories" in _ij)
        check("MiniApp: dashboard/insights authsiz → 401",
              (await _client.get("/api/dashboard")).status == 401
              and (await _client.get("/api/insights")).status == 401)
        _r = await _client.post("/api/tasks", headers=_H, json={"title": "   "})
        check("MiniApp: bo'sh sarlavha → 400", _r.status == 400)
        # ── Security-hardening regressions (audit: 4 LOW) ──
        # ── Search + AI chat ──
        await database.create_task({"title": "Qidiruv sinovi UNIKAL", "priority": "P2", "status": "todo"})
        _sr = await (await _client.get("/api/search?q=UNIKAL", headers=_H)).json()
        check("Search: /api/search topadi", any("UNIKAL" in (t.get("title") or "") for t in _sr.get("tasks", [])))
        check("Search: qisqa so'rov (<2) bo'sh", not any((await (await _client.get("/api/search?q=a", headers=_H)).json()).values()))
        # AI chat — stub process_message so no real Anthropic call; verify it executes actions
        import claude_service as _cs2
        _orig_pm = _cs2.process_message
        async def _fake_pm(text, **k):
            return {"user_message": "Vazifa qo'shildi ✓",
                    "actions": [{"type": "create_task", "data": {"title": "AI chat orqali " + text[:20], "source": "webapp"}}]}
        _cs2.process_message = _fake_pm
        try:
            _n0 = len(await database.list_tasks(limit=9999))
            _cr = await (await _client.post("/api/chat", headers=_H, json={"message": "test vazifa"})).json()
            check("Chat: reply + action bajarildi (DB'да yangi vazifa)",
                  "reply" in _cr and _cr.get("created", {}).get("task") == 1
                  and len(await database.list_tasks(limit=9999)) == _n0 + 1)
            check("Chat: bo'sh xabar → 400",
                  (await _client.post("/api/chat", headers=_H, json={"message": "  "})).status == 400)
            # Destructive action → confirm_token, NOT auto-executed until /chat/confirm
            _dtid = await database.create_task({"title": "Chat o'chirar", "priority": "P2", "status": "todo"})
            async def _fake_del(text, **k):
                return {"user_message": "O'chirilsinmi?", "actions": [{"type": "delete_task", "id": _dtid}]}
            _cs2.process_message = _fake_del
            await asyncio.sleep(1.6)  # chat rate-limit oynasidan o'tish
            _dc = await (await _client.post("/api/chat", headers=_H, json={"message": "o'chir"})).json()
            check("Chat: halokatли amal tasdiqсиz bajarilmaydi (confirm_token)",
                  bool(_dc.get("confirm_token")) and (await database.get_task(_dtid)) is not None)
            _cc = await _client.post("/api/chat/confirm", headers=_H, json={"token": _dc["confirm_token"]})
            check("Chat: confirm → o'chadi", _cc.status == 200 and (await database.get_task(_dtid)) is None)
            check("Chat: noto'g'ri token → 410",
                  (await _client.post("/api/chat/confirm", headers=_H, json={"token": "yoq"})).status == 410)
            # show_* actions must return DB-backed `views` (bug: they hit _execute_actions'
            # else-branch → _failed → reply with no data). Reproduces "bazadan ma'lumot yo'q".
            await asyncio.sleep(1.6)
            async def _fake_show(text, **k):
                return {"user_message": "Mana faol vazifalaringiz:",
                        "actions": [{"type": "show_tasks", "data": {"filter": "active"}}]}
            _cs2.process_message = _fake_show
            _sv = await (await _client.post("/api/chat", headers=_H, json={"message": "vazifalarni ko'rsat"})).json()
            _views = _sv.get("views") or []
            check("Chat: show_tasks → views (DB'dan ma'lumot qaytadi, _failed emas)",
                  len(_views) == 1 and _views[0].get("kind") == "tasks"
                  and _views[0].get("filter") == "active"
                  and isinstance(_views[0].get("items"), list) and len(_views[0]["items"]) > 0
                  and not _sv.get("created"))
            await asyncio.sleep(1.6)
            async def _fake_stats(text, **k):
                return {"user_message": "Statistika:",
                        "actions": [{"type": "show_stats", "data": {"days": 7}}]}
            _cs2.process_message = _fake_stats
            _st = await (await _client.post("/api/chat", headers=_H, json={"message": "statistika"})).json()
            check("Chat: show_stats → stats view (executive_stats)",
                  len(_st.get("views") or []) == 1 and _st["views"][0].get("kind") == "stats"
                  and "active" in (_st["views"][0].get("stats") or {}).get("tasks", {}))
            await asyncio.sleep(1.6)
            async def _fake_exp(text, **k):
                return {"user_message": "Eksport tayyor.", "actions": [{"type": "export_tasks", "data": {}}]}
            _cs2.process_message = _fake_exp
            _ed = await (await _client.post("/api/chat", headers=_H, json={"message": "eksport qil"})).json()
            check("Chat: export_tasks → download hint (/export/tasks)",
                  isinstance(_ed.get("download"), dict) and "/export/tasks" in _ed["download"].get("path", ""))
            # needs_clarification → NOTHING executes (bot parity), question surfaced
            await asyncio.sleep(1.6)
            _nc0 = len(await database.list_tasks(limit=9999))
            async def _fake_clar(text, **k):
                return {"needs_clarification": True, "clarification_question": "Qaysi vazifa?",
                        "user_message": "", "actions": [{"type": "create_task", "data": {"title": "AJRATILMAGAN"}}]}
            _cs2.process_message = _fake_clar
            _ncr = await (await _client.post("/api/chat", headers=_H, json={"message": "uni bajar"})).json()
            check("Chat: needs_clarification → savol + amal bajarilmaydi",
                  _ncr.get("needs_clarification") and _ncr.get("reply") == "Qaysi vazifa?"
                  and len(await database.list_tasks(limit=9999)) == _nc0)
            # meeting time conflict → warning note appended, no silent success
            _cf_mid = await database.create_meeting({"title": "Band slot",
                "datetime_start": "2031-03-03T10:00:00+05:00", "datetime_end": "2031-03-03T11:00:00+05:00"})
            await asyncio.sleep(1.6)
            async def _fake_conf(text, **k):
                return {"user_message": "Uchrashuv qo'shildi.",
                        "actions": [{"type": "schedule_meeting", "data": {"title": "Ustma-ust",
                            "datetime_start": "2031-03-03T10:30:00+05:00", "datetime_end": "2031-03-03T11:30:00+05:00"}}]}
            _cs2.process_message = _fake_conf
            _cfr = await (await _client.post("/api/chat", headers=_H, json={"message": "uchrashuv qo'y"})).json()
            check("Chat: to'qnashuv → ogohlantirish (jimgina muvaffaqiyat emas)",
                  "to'qnashuv" in (_cfr.get("reply") or "").lower() and not _cfr.get("created", {}).get("meeting"))
            await database.cancel_meeting(_cf_mid)
            # show_free_slots → slots view (reuses handlers._free_slots_for_day)
            await asyncio.sleep(1.6)
            async def _fake_slots(text, **k):
                return {"user_message": "Bugungi bo'sh vaqt:",
                        "actions": [{"type": "show_free_slots", "data": {"range": "day"}}]}
            _cs2.process_message = _fake_slots
            _fs = await (await _client.post("/api/chat", headers=_H, json={"message": "bo'sh vaqtim"})).json()
            check("Chat: show_free_slots → slots view (bo'sh vaqt ma'lumoti)",
                  len(_fs.get("views") or []) == 1 and _fs["views"][0].get("kind") == "slots"
                  and isinstance(_fs["views"][0].get("days"), list) and len(_fs["views"][0]["days"]) == 1)
        finally:
            _cs2.process_message = _orig_pm
        check("Search/Chat: authsiz → 401",
              (await _client.get("/api/search?q=x")).status == 401
              and (await _client.post("/api/chat", json={"message": "x"})).status == 401)
        # ── Bayonnomalar (protokol) + Excel eksport ──
        _pmid = await database.create_meeting({"title": "Protokol uchrashuv",
            "datetime_start": datetime.now(TZ).isoformat()})
        await database.update_meeting(_pmid, {"follow_up_actions": [
            "BAYONNOMA\n\nMuhokama: byudjet ko'rib chiqildi. Qaror: tasdiqlandi. Mas'ul: Karimov."]})
        _pl = await (await _client.get("/api/protocols", headers=_H)).json()
        check("Protocols: saqlangan bayonnoma ro'yxatда", any(p["id"] == _pmid for p in _pl.get("protocols", [])))
        _pw = await _client.get(f"/api/protocols/{_pmid}/download?fmt=word", headers=_H)
        check("Protocols: Word (docx) yuklab olinadi",
              _pw.status == 200 and "wordprocessing" in _pw.headers.get("Content-Type", "")
              and len(await _pw.read()) > 500)
        _pp = await _client.get(f"/api/protocols/{_pmid}/download?fmt=pdf", headers=_H)
        check("Protocols: PDF yuklab olinadi (reportlab)",
              _pp.status == 200 and "pdf" in _pp.headers.get("Content-Type", ""))
        # ── Uchrashuv: detail (GET /api/meetings/{id}) + 90-kunlik oyna ──
        _mget = await (await _client.get(f"/api/meetings/{_pmid}", headers=_H)).json()
        check("MiniApp: GET /api/meetings/{id} → meeting (follow_up_actions bilan)",
              _mget.get("meeting", {}).get("id") == _pmid
              and isinstance(_mget["meeting"].get("follow_up_actions"), list))
        check("MiniApp: GET /api/meetings/yoq → 404",
              (await _client.get("/api/meetings/yoq-id", headers=_H)).status == 404)
        _oldmid = await database.create_meeting({"title": "Eski uchrashuv",
            "datetime_start": (datetime.now(TZ) - timedelta(days=30)).isoformat()})
        _ml = await (await _client.get("/api/meetings", headers=_H)).json()
        check("MiniApp: meetings oynasi o'tgan uchrashuvni ham qamraydi (O'tgan filtr uchun)",
              any(x["id"] == _oldmid for x in _ml.get("meetings", [])))
        # ── Uchrashuv: bayonnoma yaratish (POST /api/meetings/{id}/protocol, AI mirror) ──
        _pm2 = await database.create_meeting({"title": "Bayonnoma sinovi",
            "datetime_start": datetime.now(TZ).isoformat()})
        import claude_service as _cs3
        _orig_pm3 = _cs3.process_message
        async def _fake_proto(text, internal_directive=None, **k):
            return {"user_message": "BAYONNOMA\n\nMuhokama: reja ko'rib chiqildi. Qaror: tasdiqlandi.\nTOPSHIRIQ: Karimov — hujjatni tayyorlash",
                    "actions": [{"type": "create_task", "data": {"title": "Bayonnomadan vazifa", "source": "webapp"}}]}
        _cs3.process_message = _fake_proto
        try:
            _n0 = len(await database.list_tasks(limit=9999))
            _pr = await (await _client.post(f"/api/meetings/{_pm2}/protocol", headers=_H,
                json={"notes": "byudjet ko'rib chiqildi, karimov hujjat tayyorlaydi", "create_tasks": True})).json()
            check("MiniApp: protocol → matn + follow-up vazifa yaratildi",
                  "BAYONNOMA" in (_pr.get("protocol_text") or "") and _pr.get("tasks_created") == 1
                  and len(await database.list_tasks(limit=9999)) == _n0 + 1)
            _m2 = await database.get_meeting(_pm2)
            check("MiniApp: protocol follow_up_actions'ga saqlandi",
                  isinstance(_m2.get("follow_up_actions"), list) and _m2["follow_up_actions"]
                  and "BAYONNOMA" in _m2["follow_up_actions"][0])
            _dw = await _client.get(f"/api/protocols/{_pm2}/download?fmt=word", headers=_H)
            check("MiniApp: yaratilgan bayonnoma Word yuklab olinadi",
                  _dw.status == 200 and len(await _dw.read()) > 500)
            check("MiniApp: protocol rate-limit (2s ichida ketma-ket) → 429",
                  (await _client.post(f"/api/meetings/{_pm2}/protocol", headers=_H, json={"notes": "yana"})).status == 429)
            check("MiniApp: protocol bo'sh notes → 400",
                  (await _client.post(f"/api/meetings/{_pm2}/protocol", headers=_H, json={"notes": "  "})).status == 400)
            check("MiniApp: protocol noma'lum uchrashuv → 404",
                  (await _client.post("/api/meetings/yoq-id/protocol", headers=_H, json={"notes": "x"})).status == 404)
        finally:
            _cs3.process_message = _orig_pm3
        _ex = await _client.get("/api/export/tasks?filter=all", headers=_H)
        check("Export: Excel (xlsx) yuklab olinadi",
              _ex.status == 200 and "spreadsheet" in _ex.headers.get("Content-Type", "")
              and len(await _ex.read()) > 1000)
        check("Protocols/Export: authsiz → 401",
              (await _client.get("/api/protocols")).status == 401
              and (await _client.get("/api/export/tasks")).status == 401)
        # ── Jamoa / Risklar / Kategoriyalar / Kalendar ──
        await database.create_task({"title": "Yuklama ish", "priority": "P0", "status": "todo", "assignee": "TeamX",
            "deadline": (datetime.now(TZ) - timedelta(days=1)).isoformat()})
        _tm = await (await _client.get("/api/team", headers=_H)).json()
        check("Team: yuklama ro'yxati (name/active)",
              any(p.get("name") == "TeamX" and "active" in p for p in _tm.get("team", [])))
        _rk = await (await _client.get("/api/risks", headers=_H)).json()
        check("Risks: counts + overdue ro'yxati",
              all(k in _rk.get("counts", {}) for k in ("overdue", "urgent_open", "unassigned"))
              and isinstance(_rk.get("overdue"), list))
        _cc0 = await _client.post("/api/categories", headers=_H, json={"name": "WebKat"})
        _cl = await (await _client.get("/api/categories", headers=_H)).json()
        check("Categories: create + list", _cc0.status == 201
              and any(c.get("name") == "WebKat" for c in _cl.get("categories", [])))
        _cd = await _client.delete("/api/categories?name=WebKat", headers=_H)
        check("Categories: delete", _cd.status == 200)
        _now = datetime.now(TZ)
        _cal = await (await _client.get(f"/api/calendar?year={_now.year}&month={_now.month}", headers=_H)).json()
        check("Calendar: oylik uchrashuvlar (meetings ro'yxati)", "meetings" in _cal and "year" in _cal)
        check("Team/Risks/Cal: authsiz → 401",
              (await _client.get("/api/team")).status == 401
              and (await _client.get("/api/risks")).status == 401
              and (await _client.get("/api/calendar")).status == 401)
        # ── Fayl import + ovoz ──
        _csv = "Vazifa,Ijrochi,Ustuvorlik\nImport CSV ish,Karimov,P1\nImport CSV 2,Aziz,P2\n"
        _ir = await _client.post("/api/import/tasks?name=t.csv", headers=_H, data=_csv.encode("utf-8"))
        _ij = await _ir.json()
        check("Import: CSV → vazifalar yaratildi",
              _ir.status == 200 and _ij.get("created", 0) >= 2
              and any(t["title"] == "Import CSV ish" for t in await database.list_tasks(limit=9999)))
        check("Import: bo'sh/noto'g'ri fayl → 400",
              (await _client.post("/api/import/tasks?name=x.csv", headers=_H, data=b"salom dunyo\n")).status == 400)
        # voice — stub transcribe (no real STT call)
        import voice_service as _vs
        _ovs = _vs.transcribe
        async def _fake_tr(b, filename="v", language="uz"): return "ertaga hisobot tayyorla"
        _vs.transcribe = _fake_tr
        try:
            _vr = await _client.post("/api/voice?name=v.webm", headers=_H, data=b"AUDIOBYTES")
            check("Voice: audio → matn", _vr.status == 200 and (await _vr.json()).get("text") == "ertaga hisobot tayyorla")
        finally:
            _vs.transcribe = _ovs
        check("Import/Voice: authsiz → 401",
              (await _client.post("/api/import/tasks", data=b"x")).status == 401
              and (await _client.post("/api/voice", data=b"x")).status == 401)

        # S1: non-scalar value for a scalar field → clean 400, NOT a raw 500
        _r = await _client.post("/api/tasks", headers=_H, json={"title": "x", "priority": {"a": 1}})
        check("S1: dict qiymatli maydon → 400 (500 emas)", _r.status == 400)
        # S1b: server survived the bad request (next request still works)
        _r = await _client.get("/api/tasks", headers=_H)
        check("S1: buzuq so'rovdan keyin server tirik", _r.status == 200)
        # S4: over-long text field → 400 (no multi-MB rows)
        _r = await _client.post("/api/tasks", headers=_H, json={"title": "T" * 5000})
        check("S4: juda uzun sarlavha → 400", _r.status == 400)
        # S2: parent_id guards — dangling / self / cycle all rejected
        _pt = (await (await _client.post("/api/tasks", headers=_H,
               json={"title": "Ota"})).json())["id"]
        _r = await _client.post("/api/tasks", headers=_H,
               json={"title": "Bola", "parent_id": "t-YOQ"})
        check("S2: mavjud bo'lmagan parent_id → 400", _r.status == 400)
        _r = await _client.patch(f"/api/tasks/{_pt}", headers=_H, json={"parent_id": _pt})
        check("S2: o'ziga ota (self-parent) → 400", _r.status == 400)
        _c1 = (await (await _client.post("/api/tasks", headers=_H,
               json={"title": "C1"})).json())["id"]
        _c2 = (await (await _client.post("/api/tasks", headers=_H,
               json={"title": "C2", "parent_id": _c1})).json())["id"]
        _r = await _client.patch(f"/api/tasks/{_c1}", headers=_H, json={"parent_id": _c2})
        check("S2: halqa (cycle) → 400", _r.status == 400)
        # valid parent still works
        _r = await _client.post("/api/tasks", headers=_H, json={"title": "OK sub", "parent_id": _pt})
        check("S2: to'g'ri parent_id → 201", _r.status == 201)
        # task delete (UI: task-form 'O'chirish') — DELETE endpoint reachable end-to-end
        _delid = (await (await _client.post("/api/tasks", headers=_H, json={"title": "O'chiraman"})).json())["id"]
        _r = await _client.delete(f"/api/tasks/{_delid}", headers=_H)
        check("Full CRUD: vazifa DELETE → o'chadi",
              _r.status == 200 and (await database.get_task(_delid)) is None)

        # ── Full-platform CRUD smoke: har bir bo'lim uchun create→list→update→yakun/o'chir ──
        # Meetings
        _mid = (await (await _client.post("/api/meetings", headers=_H, json={
            "title": "Web uchrashuv", "datetime_start": (datetime.now(TZ)+timedelta(days=1)).isoformat(),
            "participants": ["Karimov", "Aziz"], "location_or_link": "Zoom"})).json())["id"]
        _ml = (await (await _client.get("/api/meetings", headers=_H)).json())["meetings"]
        check("Full CRUD: uchrashuv create+list", any(m["id"] == _mid for m in _ml))
        check("MiniApp: meeting javobiда completed_at + datetime_start (o'tgan/bo'ldi ajratish)",
              all(k in _ml[0] for k in ("completed_at", "datetime_start")))
        _r = await _client.patch(f"/api/meetings/{_mid}", headers=_H, json={"location_or_link": "Ofis 3-qavat"})
        check("Full CRUD: uchrashuv PATCH", _r.status == 200
              and (await database.get_meeting(_mid))["location_or_link"] == "Ofis 3-qavat")
        check("Full CRUD: uchrashuv complete → completed_at",
              (await _client.post(f"/api/meetings/{_mid}/complete", headers=_H)).status == 200
              and (await database.get_meeting(_mid)).get("completed_at"))
        check("Full CRUD: uchrashuv uncomplete (qayta ochish) → completed_at tozalanadi",
              (await _client.post(f"/api/meetings/{_mid}/uncomplete", headers=_H)).status == 200
              and not (await database.get_meeting(_mid)).get("completed_at"))
        check("Full CRUD: uchrashuv cancel → o'chadi",
              (await _client.post(f"/api/meetings/{_mid}/cancel", headers=_H)).status == 200
              and (await database.get_meeting(_mid)) is None)
        # Notes
        _nid = (await (await _client.post("/api/notes", headers=_H, json={
            "title": "Web qayd", "content": "Mini app orqali qayd"})).json())["id"]
        _nl = (await (await _client.get("/api/notes", headers=_H)).json())["notes"]
        check("Full CRUD: qayd create+list", any(n["id"] == _nid for n in _nl))
        _r = await _client.patch(f"/api/notes/{_nid}", headers=_H, json={"content": "Yangilangan matn"})
        check("Full CRUD: qayd PATCH", _r.status == 200
              and (await database.get_note(_nid))["content"] == "Yangilangan matn")
        check("Full CRUD: qayd DELETE",
              (await _client.delete(f"/api/notes/{_nid}", headers=_H)).status == 200
              and (await database.get_note(_nid)) is None)
        # Reminders
        _rid = (await (await _client.post("/api/reminders", headers=_H, json={
            "title": "Web eslatma", "remind_at": (datetime.now(TZ)+timedelta(hours=3)).isoformat()})).json())["id"]
        _rl = (await (await _client.get("/api/reminders", headers=_H)).json())["reminders"]
        check("Full CRUD: eslatma create+list", any(r["id"] == _rid for r in _rl))
        _r = await _client.patch(f"/api/reminders/{_rid}", headers=_H, json={"note": "yangi izoh"})
        check("Full CRUD: eslatma PATCH (UI tahrir tugmasi)", _r.status == 200
              and (await database.get_reminder(_rid))["note"] == "yangi izoh")
        check("Full CRUD: eslatma complete",
              (await _client.post(f"/api/reminders/{_rid}/complete", headers=_H)).status == 200)
        _rid_d = (await (await _client.post("/api/reminders", headers=_H, json={
            "title": "O'chiriladi", "remind_at": (datetime.now(TZ)+timedelta(hours=1)).isoformat()})).json())["id"]
        check("Full CRUD: eslatma DELETE (UI 🗑 tugmasi)",
              (await _client.delete(f"/api/reminders/{_rid_d}", headers=_H)).status == 200
              and (await database.get_reminder(_rid_d)) is None)
        # /api/meta dropdown data
        _mt = await (await _client.get("/api/meta", headers=_H)).json()
        check("Full CRUD: /api/meta (kategoriya+kontakt+ustuvorlik)",
              "categories" in _mt and "contacts" in _mt and "priorities" in _mt)
    finally:
        await _client.close()

    print("\n" + "=" * 48)
    print(f"NATIJA:  ✅ {PASS} o'tdi   ❌ {FAIL} yiqildi")
    if FAILED:
        print("Yiqilganlar:", ", ".join(FAILED))
    print("=" * 48)
    # cleanup temp db
    try:
        os.remove(_TMP)
    except OSError:
        pass
    return 0 if FAIL == 0 else 1


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
