"""Telegram message and callback handlers."""

import asyncio
import json
import logging
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

from aiogram import Bot, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup, default_state
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InlineQuery,
    InlineQueryResultArticle,
    InputTextMessageContent,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)

import calendar_service
import claude_service
import config
import database
import document_service
import scheduler as scheduler_module
import voice_service


_TG_SOFT_LIMIT = 3500  # split outgoing messages above this to preserve readability


def _split_for_telegram(text: str, limit: int = _TG_SOFT_LIMIT) -> list[str]:
    """Split a long message at clean boundaries (blank lines, then single lines).
    Returns the original text as a 1-item list when under the limit.
    """
    if len(text) <= limit:
        return [text]
    parts: list[str] = []
    buf = ""
    # Prefer splitting at blank-line boundaries to keep sections intact.
    for block in text.split("\n\n"):
        chunk = block if not buf else f"{buf}\n\n{block}"
        if len(chunk) <= limit:
            buf = chunk
            continue
        if buf:
            parts.append(buf)
            buf = ""
        # Block alone exceeds the limit — fall back to per-line splitting.
        if len(block) <= limit:
            buf = block
            continue
        line_buf = ""
        for line in block.split("\n"):
            cand = line if not line_buf else f"{line_buf}\n{line}"
            if len(cand) <= limit:
                line_buf = cand
            else:
                if line_buf:
                    parts.append(line_buf)
                    line_buf = ""
                # A single line longer than the limit — HARD-SPLIT into limit-sized
                # pieces so no content is silently dropped (was: line[:limit], which
                # lost the tail — e.g. long compliance/contract notes in a task).
                while len(line) > limit:
                    parts.append(line[:limit])
                    line = line[limit:]
                line_buf = line
        if line_buf:
            buf = line_buf
    if buf:
        parts.append(buf)
    return parts


async def _safe_answer(message: Message, text: str, **kwargs) -> None:
    """Send a message, falling back to plain text if Markdown parsing fails.
    Auto-splits messages longer than _TG_SOFT_LIMIT at section boundaries.
    Reply markup is attached to the last chunk only.
    """
    chunks = _split_for_telegram(text)
    reply_markup = kwargs.pop("reply_markup", None)
    last_idx = len(chunks) - 1
    for i, chunk in enumerate(chunks):
        chunk_kwargs = dict(kwargs)
        if i == last_idx and reply_markup is not None:
            chunk_kwargs["reply_markup"] = reply_markup
        try:
            await message.answer(chunk, **chunk_kwargs)
        except TelegramBadRequest as e:
            if "can't parse entities" in str(e).lower() or "parse" in str(e).lower():
                # Explicit None overrides the bot's DEFAULT parse_mode (Markdown);
                # merely popping the kwarg would still fall back to that default.
                chunk_kwargs["parse_mode"] = None
                await message.answer(chunk, **chunk_kwargs)
            else:
                raise


# ─────────────────────── PERSISTENT REPLY KEYBOARD ───────────────────────


# Main menu — Executive Task Management Assistant layout
BTN_COCKPIT = "🎛 Boshqaruv paneli"
BTN_TODAY = "📅 Bugun"
BTN_TASKS = "📌 Vazifalar"
BTN_REMINDERS = "⏰ Eslatmalar"
# Legacy reply-keyboard labels still cached on user devices — keep accepting them
# so the button works even if the keyboard hasn't refreshed yet.
_LEGACY_BTN_TASKS = {"📋 Vazifalar"}
BTN_TEAM = "👥 Ijrochilar"
BTN_RISKS = "🚨 Risklar"
BTN_NEW = "➕ Yangi"
BTN_STATS = "📊 Statistika"
BTN_SEARCH = "🔍 Qidiruv"
BTN_MEETINGS = "🤝 Uchrashuvlar"
BTN_SETTINGS = "⚙️ Sozlamalar"


def main_reply_keyboard() -> ReplyKeyboardMarkup:
    """Persistent main menu — daily executive workflow.

    Discoverability: Ijrochilar / Risklar / Statistika endi bosh menyuda
    (avval faqat Cockpit drill-down orqali topilardi). Routing o'zgармaydi —
    handle_main_reply_button allaqachon bu labellarni dispatch qiladi.
    """
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=BTN_COCKPIT), KeyboardButton(text=BTN_TODAY)],
            [KeyboardButton(text=BTN_TASKS), KeyboardButton(text=BTN_MEETINGS)],
            [KeyboardButton(text=BTN_TEAM), KeyboardButton(text=BTN_RISKS)],
            [KeyboardButton(text=BTN_STATS), KeyboardButton(text=BTN_REMINDERS)],
            [KeyboardButton(text=BTN_NEW), KeyboardButton(text=BTN_SEARCH),
             KeyboardButton(text=BTN_SETTINGS)],
        ],
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="Matn yoki ovoz yuboring...",
    )


# ─────────────────────── SECTION REPLY KEYBOARDS ───────────────────────
# Bo'limga kirgan foydalanuvchi avtomatik o'sha bo'limning filter/amal
# tugmalarini reply kbd sifatida ko'radi. Har label noyob prefiks bilan
# (📌 = Vazifalar, 🤝 = Uchrashuvlar) — FSM state shart emas, label
# o'zi qaysi bo'limga tegishliligini bildiradi.

BTN_BACK_MAIN = "⬅️ Asosiy menyu"

# Section reply labels — prefiks yo'q (toza ko'rinish uchun).
# Kontekst (qaysi bo'lim) SectionFSM state orqali aniqlanadi.
TBTN_TASKS_ACTIVE = "🔵 Aktiv"
TBTN_TASKS_TODAY = "📅 Bugun"
TBTN_TASKS_OVERDUE = "⌛ O'tgan"
TBTN_TASKS_IMPORTANT = "⭐ Muhim"
TBTN_TASKS_DONE = "✅ Bajarilgan"
TBTN_TASKS_ALL = "📋 Barchasi"
TBTN_TASKS_NEW = "➕ Yangi vazifa"
TBTN_TASKS_SEARCH = "🔍 Vazifa qidirish"
TBTN_TASKS_CATEGORIES = "🗄 Kategoriyalar"

_TASKS_SECTION_FILTERS = {
    TBTN_TASKS_ACTIVE:   "active",
    TBTN_TASKS_TODAY:    "today",
    TBTN_TASKS_OVERDUE:  "overdue",
    TBTN_TASKS_IMPORTANT: "important",
    TBTN_TASKS_DONE:     "done",
    TBTN_TASKS_ALL:      "all",
}

# Last task filter the principal viewed — so an opened task's "⬅️ Ro'yxatga"
# returns to THAT filter, not always "active". Single-user bot → module state ok.
_last_task_filter: str = "active"
# Same idea for meetings — an opened meeting's "⬅️ Ro'yxatga" returns to the
# filter it was opened from (was hard-coded to "week", which jarred when you'd
# filtered by "tomorrow"/"today").
_last_meeting_filter: str = "week"

MBTN_MEETINGS_WEEK = "Haftalik"
MBTN_MEETINGS_TODAY = "Bugun"
MBTN_MEETINGS_TOMORROW = "Ertaga"
MBTN_MEETINGS_ALL = "Barchasi"
MBTN_MEETINGS_PAST = "O'tgan"
MBTN_MEETINGS_NEW = "➕ Yangi uchrashuv"
MBTN_MEETINGS_SEARCH = "🔍 Uchrashuv qidirish"
MBTN_MEETINGS_PROTOCOLS = "📄 Bayonnomalar"

_MEETINGS_SECTION_FILTERS = {
    MBTN_MEETINGS_WEEK:     "week",
    MBTN_MEETINGS_TODAY:    "today",
    MBTN_MEETINGS_TOMORROW: "tomorrow",
    MBTN_MEETINGS_ALL:      "all",
    MBTN_MEETINGS_PAST:     "past",
}

RBTN_REMINDERS_TODAY = "⏰ Bugun"
RBTN_REMINDERS_UPCOMING = "⏭ Keyingi"
RBTN_REMINDERS_SENT = "📤 Yuborilgan"
RBTN_REMINDERS_ALL = "📋 Barchasi"
RBTN_REMINDERS_NEW = "➕ Yangi eslatma"
RBTN_REMINDERS_SEARCH = "🔍 Eslatma qidirish"

_REMINDERS_SECTION_FILTERS = {
    RBTN_REMINDERS_TODAY: "today",
    RBTN_REMINDERS_SENT: "sent",
    RBTN_REMINDERS_ALL: "all",
}

# ── Qaydlar (Notes) section labels ──
NBTN_NOTES_INBOX = "📥 Inbox"
NBTN_NOTES_PROCESSED = "⚙️ Ishlangan"
NBTN_NOTES_ARCHIVED = "📦 Arxiv"
NBTN_NOTES_NEW = "➕ Yangi qayd"
NBTN_NOTES_SEARCH = "🔍 Qayd qidirish"

_NOTES_SECTION_FILTERS = {
    NBTN_NOTES_INBOX:     "inbox",
    NBTN_NOTES_PROCESSED: "processed",
    NBTN_NOTES_ARCHIVED:  "archived",
}

# Per-note source labels (short, used in card meta line)
_NOTES_SOURCE_BADGE = {
    "forward": "🔁 forward",
    "command": "⚡ buyruq",
    "voice":   "🎙 ovoz",
    "manual":  "✍️ qo'lda",
    "llm":     "🤖 LLM",
}
# Emoji-only manba ikonkasi — sana bo'yicha guruhlangan ro'yxatda (Variant B)
# sarlavha oldida turadi.
_NOTES_SOURCE_ICON = {
    "forward": "🔁",
    "command": "⚡",
    "voice":   "🎙",
    "manual":  "✍️",
    "llm":     "🤖",
}
_NOTES_PER_PAGE = 10


class SectionFSM(StatesGroup):
    """Aktiv bo'limni eslab qolish — reply kbd labellari noyob emas,
    shuning uchun "Bugun"/"Barchasi"/"O'tgan" qaysi bo'limga tegishliligini
    state bilan aniqlanadi. State faqat shu maqsad uchun ishlatiladi —
    boshqa FSM oqimlar bilan to'qnashmaydi."""
    in_tasks = State()
    in_reminders = State()
    in_meetings = State()
    in_stats = State()
    in_team = State()
    in_risks = State()
    in_today = State()
    in_new = State()
    in_search = State()
    in_settings = State()
    in_notes = State()


class NoteCaptureFSM(StatesGroup):
    """One-shot FSM for `/qayd` with no body or "➕ Yangi qayd" button —
    next text/voice message becomes the note content."""
    awaiting_text = State()


class NewMeetingTextFSM(StatesGroup):
    """One-shot capture after "➕ Yangi uchrashuv" — the next text/voice is routed
    to create-meeting (with an explicit intent) instead of being reinterpreted as
    an unrelated command. Without this the button only printed a prompt and left
    the user in the section state, so the reply was treated as a fresh command."""
    awaiting_text = State()


class NewTaskTextFSM(StatesGroup):
    """One-shot capture after "➕ Yangi vazifa" — next text/voice → create-task."""
    awaiting_text = State()


# ── Statistika section labels ──
SBTN_STATS_TODAY = "Bugun"
SBTN_STATS_WEEK = "7 kun"
SBTN_STATS_MONTH = "30 kun"
SBTN_STATS_REPORT_WEEK = "📄 Hisobot 7 kun"
SBTN_STATS_REPORT_MONTH = "📄 Hisobot 30 kun"

_STATS_SECTION_PERIODS = {
    SBTN_STATS_TODAY: 1,
    SBTN_STATS_WEEK:  7,
    SBTN_STATS_MONTH: 30,
}

# ── Ijrochilar section labels ──
YBTN_TEAM_REFRESH = "🔄 Yangilash"
YBTN_TEAM_UNASSIGNED = "👤 Ijrochisiz vazifalar"
YBTN_TEAM_REASSIGN = "🔄 Qayta taqsimlash"
# Eski "Delegatsiya trekeri" shu yerga ko'chirildi — boshqalarga berilgan
# vazifalarni eng uzoq kutilgani bo'yicha ko'rsatadi (qotgan topshiriqlar).
YBTN_TEAM_STALE = "⏳ Kutilayotganlar"

# ── Risklar section labels ──
RBTN_RISKS_REFRESH = "🔄 Yangilash"

# ── Bugun section labels ──
DBTN_TODAY_EVENING = "🌙 Kechki yakun"
DBTN_TODAY_ALL_TASKS = "📋 Hamma vazifalar"
DBTN_TODAY_NEW_TASK = "➕ Yangi vazifa (Bugun)"  # main "➕ Yangi" bilan to'qnashmasin
DBTN_TODAY_MEETINGS = "🤝 Bugungi uchrashuvlar"

# ── Yangi section labels ──
NBTN_NEW_TASK = "📝 Yangi vazifa"
NBTN_NEW_MEETING = "🤝 Yangi uchrashuv"
NBTN_NEW_REMINDER = "⏰ Yangi eslatma"
NBTN_NEW_NOTE = "📥 Yangi qayd"
NBTN_NEW_VOICE = "🎙 Ovozli vazifa"
NBTN_NEW_POLISH = "✏️ Matn tahrirlash"

# ── Qidiruv section labels ──
QBTN_SEARCH_TASKS = "📌 Faqat vazifalar"
QBTN_SEARCH_MEETINGS = "🤝 Faqat uchrashuvlar"
QBTN_SEARCH_CONTACTS = "👥 Faqat kontaktlar"
QBTN_SEARCH_ALL = "🗂 Hammasi"

# ── Sozlamalar section labels ──
GBTN_SETTINGS_NOTIFY = "🔔 Bildirishnoma"
GBTN_SETTINGS_BRIEFING = "⏰ Brifing vaqti"
GBTN_SETTINGS_EVENING = "🌙 Kechki vaqt"  # "yakun" is the action in Today section; here we set the time
GBTN_SETTINGS_REMINDER = "📲 Eslatma parametrlari"
GBTN_SETTINGS_VOICE = "🎙 Ovoz tasdig'i"
GBTN_SETTINGS_CREATE_CONFIRM = "✅ Yaratish tasdig'i"
GBTN_SETTINGS_CALENDAR = "📅 Kalendar holati"


def _two_per_row(labels: list[str], solo: set[str]) -> list[list[KeyboardButton]]:
    """Reply-kbd tartibi: `solo` yorliqlari ALOHIDA to'liq qator oladi; qolgan
    tugmalar 2 tadan yonma-yon joylashadi (tartib saqlanadi). Toq son bo'lsa,
    oxirgisi yakka qoladi."""
    rows: list[list[KeyboardButton]] = []
    pending: list[str] = []

    def _flush():
        for i in range(0, len(pending), 2):
            rows.append([KeyboardButton(text=t) for t in pending[i:i + 2]])
        pending.clear()

    for lbl in labels:
        if lbl in solo:
            _flush()
            rows.append([KeyboardButton(text=lbl)])
        else:
            pending.append(lbl)
    _flush()
    return rows


def tasks_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Vazifalar bo'limida — 2 tadan bir qatorda (uzun yorliqlar
    kesilmasin). Vizual ierarxiya: 6 filtr 2 tadan; primary amal «➕ Yangi
    vazifa» to'liq qator (solo) — ko'zga tashlanadi; «🔍 Qidirish» + «🗄
    Kategoriyalar» juft; «⬅️ Asosiy menyu» to'liq qator (solo)."""
    return ReplyKeyboardMarkup(
        keyboard=_two_per_row(
            [TBTN_TASKS_ACTIVE, TBTN_TASKS_TODAY,        # filtr · hozir: aktiv / bugun
             TBTN_TASKS_IMPORTANT, TBTN_TASKS_OVERDUE,   # filtr · diqqat: muhim / o'tgan
             TBTN_TASKS_DONE, TBTN_TASKS_ALL,            # filtr · ko'rib chiqish: bajarilgan / barchasi
             TBTN_TASKS_NEW,                             # primary amal — solo (to'liq qator)
             TBTN_TASKS_SEARCH, TBTN_TASKS_CATEGORIES,   # qidirish + kategoriyalar (juft)
             BTN_BACK_MAIN],                             # navigatsiya — solo (to'liq qator)
            solo={TBTN_TASKS_NEW, BTN_BACK_MAIN}),
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Filter tanlang yoki yangi vazifa...",
    )


def meetings_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Uchrashuvlar bo'limida — filterlar va sub-amallar."""
    return ReplyKeyboardMarkup(
        keyboard=_two_per_row(
            [MBTN_MEETINGS_WEEK, MBTN_MEETINGS_TODAY, MBTN_MEETINGS_TOMORROW,
             MBTN_MEETINGS_PAST, MBTN_MEETINGS_ALL,
             MBTN_MEETINGS_NEW, MBTN_MEETINGS_SEARCH, MBTN_MEETINGS_PROTOCOLS, BTN_BACK_MAIN],
            solo={MBTN_MEETINGS_ALL, BTN_BACK_MAIN}),
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Filter tanlang yoki yangi uchrashuv...",
    )


def reminders_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Eslatmalar bo'limida — filterlar va asosiy amallar."""
    return ReplyKeyboardMarkup(
        keyboard=_two_per_row(
            [RBTN_REMINDERS_TODAY, RBTN_REMINDERS_SENT, RBTN_REMINDERS_ALL,
             RBTN_REMINDERS_NEW, RBTN_REMINDERS_SEARCH, BTN_BACK_MAIN],
            solo={RBTN_REMINDERS_ALL, BTN_BACK_MAIN}),
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Eslatma tanlang yoki yangi eslatma...",
    )


def notes_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Notes bo'limida — Inbox / Ishlangan / Arxiv + amallar."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=NBTN_NOTES_INBOX),
             KeyboardButton(text=NBTN_NOTES_PROCESSED),
             KeyboardButton(text=NBTN_NOTES_ARCHIVED)],
            [KeyboardButton(text=NBTN_NOTES_NEW),
             KeyboardButton(text=NBTN_NOTES_SEARCH)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Qayd tanlang yoki yangisini yozing...",
    )


def stats_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Statistika bo'limida — davrlar va hisobotlar."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=SBTN_STATS_TODAY),
             KeyboardButton(text=SBTN_STATS_WEEK),
             KeyboardButton(text=SBTN_STATS_MONTH)],
            [KeyboardButton(text=SBTN_STATS_REPORT_WEEK),
             KeyboardButton(text=SBTN_STATS_REPORT_MONTH)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Davr tanlang yoki hisobot...",
    )


def team_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Ijrochilar bo'limida — tezkor amallar."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=YBTN_TEAM_STALE),
             KeyboardButton(text=YBTN_TEAM_UNASSIGNED)],
            [KeyboardButton(text=YBTN_TEAM_REFRESH),
             KeyboardButton(text=YBTN_TEAM_REASSIGN)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Ijrochi raqamini yoki amal tanlang...",
    )


def risks_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Risklar bo'limida — minimal."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=RBTN_RISKS_REFRESH)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Risk raqamini bosing yoki yangilang...",
    )


def today_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Bugun bo'limida — tezkor amallar."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=DBTN_TODAY_EVENING),
             KeyboardButton(text=DBTN_TODAY_ALL_TASKS)],
            [KeyboardButton(text=DBTN_TODAY_NEW_TASK),
             KeyboardButton(text=DBTN_TODAY_MEETINGS)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Vazifa raqamini yoki amal tanlang...",
    )


def new_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Yangi bo'limida — turlar."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=NBTN_NEW_TASK),
             KeyboardButton(text=NBTN_NEW_MEETING)],
            [KeyboardButton(text=NBTN_NEW_REMINDER),
             KeyboardButton(text=NBTN_NEW_NOTE)],
            [KeyboardButton(text=NBTN_NEW_VOICE),
             KeyboardButton(text=NBTN_NEW_POLISH)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Yangi item turini tanlang...",
    )


def search_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Qidiruv bo'limida — scope filterlari."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=QBTN_SEARCH_TASKS),
             KeyboardButton(text=QBTN_SEARCH_MEETINGS)],
            [KeyboardButton(text=QBTN_SEARCH_CONTACTS),
             KeyboardButton(text=QBTN_SEARCH_ALL)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Scope tanlang yoki so'z yuboring...",
    )


def settings_section_reply_keyboard() -> ReplyKeyboardMarkup:
    """Reply kbd Sozlamalar bo'limida — parametr toifalari."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=GBTN_SETTINGS_NOTIFY),
             KeyboardButton(text=GBTN_SETTINGS_BRIEFING)],
            [KeyboardButton(text=GBTN_SETTINGS_EVENING),
             KeyboardButton(text=GBTN_SETTINGS_REMINDER)],
            [KeyboardButton(text=GBTN_SETTINGS_VOICE),
             KeyboardButton(text=GBTN_SETTINGS_CREATE_CONFIRM)],
            [KeyboardButton(text=GBTN_SETTINGS_CALENDAR)],
            [KeyboardButton(text=BTN_BACK_MAIN)],
        ],
        resize_keyboard=True, is_persistent=True,
        input_field_placeholder="Sozlama tanlang...",
    )


_SECTION_LABELS: set[str] = (
    set(_TASKS_SECTION_FILTERS)
    | set(_REMINDERS_SECTION_FILTERS)
    | set(_MEETINGS_SECTION_FILTERS)
    | set(_STATS_SECTION_PERIODS)
    | {TBTN_TASKS_NEW, TBTN_TASKS_SEARCH,
       MBTN_MEETINGS_NEW, MBTN_MEETINGS_SEARCH, MBTN_MEETINGS_PROTOCOLS,
       RBTN_REMINDERS_NEW, RBTN_REMINDERS_SEARCH,
       SBTN_STATS_REPORT_WEEK, SBTN_STATS_REPORT_MONTH,
       YBTN_TEAM_REFRESH, YBTN_TEAM_UNASSIGNED, YBTN_TEAM_REASSIGN, YBTN_TEAM_STALE,
       RBTN_RISKS_REFRESH,
       DBTN_TODAY_EVENING, DBTN_TODAY_ALL_TASKS, DBTN_TODAY_NEW_TASK, DBTN_TODAY_MEETINGS,
       NBTN_NEW_TASK, NBTN_NEW_MEETING, NBTN_NEW_REMINDER, NBTN_NEW_VOICE, NBTN_NEW_POLISH,
       QBTN_SEARCH_TASKS, QBTN_SEARCH_MEETINGS, QBTN_SEARCH_CONTACTS, QBTN_SEARCH_ALL,
       GBTN_SETTINGS_NOTIFY, GBTN_SETTINGS_BRIEFING, GBTN_SETTINGS_EVENING,
       GBTN_SETTINGS_REMINDER, GBTN_SETTINGS_VOICE, GBTN_SETTINGS_CREATE_CONFIRM,
       GBTN_SETTINGS_CALENDAR,
       BTN_BACK_MAIN}
)


# Bot restart'dan keyin section labellarni state'ga avtomatik tiklash
# (E1 edge case). Foydalanuvchi cache'dagi section reply kbd dan tugma
# bossa, mos state'ga o'rnatamiz va to'g'ri handler ishlay oladi.

def back_button(callback_data: str = "nav_cockpit", text: str = "⬅️ Orqaga") -> InlineKeyboardButton:
    return InlineKeyboardButton(text=text, callback_data=callback_data)


def single_back_keyboard(callback_data: str = "nav_cockpit", text: str = "⬅️ Orqaga") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[back_button(callback_data, text)]])


def task_inline_actions(task: dict) -> InlineKeyboardMarkup:
    """Per-task quick action row: 3 most-used buttons.

    [👤 Ijrochi] [✅ Bajarildi] [⋯ Batafsil]

    Batafsil opens task_detail_menu with the full action set.
    """
    tid = task["id"]
    if task.get("status") == "done":
        return InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="↺ Qaytarish", callback_data=f"reopen:{tid}"),
            InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"task_del:{tid}"),
            InlineKeyboardButton(text="⋯ Batafsil", callback_data=f"task_detail:{tid}"),
        ]])
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="👤 Ijrochi", callback_data=f"set_assignee:{tid}"),
        InlineKeyboardButton(text="✅ Bajarildi", callback_data=f"complete:{tid}"),
        InlineKeyboardButton(text="⋯ Batafsil", callback_data=f"task_detail:{tid}"),
    ]])


def task_detail_menu(task: dict) -> InlineKeyboardMarkup:
    """Full action menu — compact 2-column layout from ⋯ Batafsil."""
    tid = task["id"]
    if task.get("status") == "done":
        return InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="↺ Qaytarish", callback_data=f"reopen:{tid}"),
                InlineKeyboardButton(text="✏️ Tahrir", callback_data=f"edit:{tid}"),
            ],
            [
                InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"task_del:{tid}"),
                InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"taskopen:{tid}"),
            ],
        ])
    # Quick lifecycle actions only. Field edits (Muddat, Prioritet, Status, …)
    # all live in ✏️ Tahrir → task_edit_menu, so we don't duplicate them here.
    # 👤 Ijrochi stays (it has no field-editor equivalent and is a frequent action).
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Bajarildi", callback_data=f"complete:{tid}"),
            InlineKeyboardButton(text="👤 Ijrochi", callback_data=f"set_assignee:{tid}"),
        ],
        [
            InlineKeyboardButton(text="✏️ Tahrir", callback_data=f"edit:{tid}"),
            InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"task_del:{tid}"),
        ],
        [
            InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"taskopen:{tid}"),
        ],
    ])


class TaskEditFSM(StatesGroup):
    awaiting_value = State()


class AssigneeFSM(StatesGroup):
    awaiting_name = State()


class TaskSearchFSM(StatesGroup):
    awaiting_query = State()


class MeetingSearchFSM(StatesGroup):
    awaiting_query = State()


class MeetingFollowupFSM(StatesGroup):
    awaiting_notes = State()


class MeetingEditFSM(StatesGroup):
    awaiting_value = State()


class MeetingProtocolFSM(StatesGroup):
    awaiting_notes = State()
    awaiting_revision = State()


class VoiceConfirmFSM(StatesGroup):
    """Free-form ovoz xabari konfirmatsiyasi. Foydalanuvchi ovoz yuboradi →
    bot transkripsiyani ko'rsatadi → foydalanuvchi tasdiqlaydi/tahrirlaydi/bekor.
    Faqat FSM holatisiz holatlarda faollashadi (boshqa flow ichida emas)."""
    awaiting_action = State()
    awaiting_revision = State()


class PolishRevisionFSM(StatesGroup):
    """✎ Yana tahrir — polished matnni qayta ishlash. Foydalanuvchi ko'rsatma
    yuboradi (masalan «qisqartir»), bot original + ko'rsatmani qayta polish qiladi."""
    awaiting = State()


class CreateActionConfirmFSM(StatesGroup):
    """Tasdiq state — yangi vazifa/uchrashuv yaratishdan oldin foydalanuvchi
    "Tasdiqlayman/Bekor qilish" bossin uchun. Claude'ning to'liq javobi
    state.data['pending_response']'da saqlanadi; tasdiqdan keyin bajariladi."""
    awaiting = State()


class DocReviseFSM(StatesGroup):
    """✏️ Tahrirla — hujjat tahlilini ko'rsatma bo'yicha qayta ishlash (polish
    revizatsiyasi bilan bir xil konsepsiya). Cache'dagi fayl uid'i
    state.data['revise_uid']'da; bot SHU faylni yangi ko'rsatma bilan qayta tahlil qiladi."""
    awaiting = State()


class GlobalSearchFSM(StatesGroup):
    """Global search across tasks + meetings."""
    awaiting_query = State()


class NewTaskFSM(StatesGroup):
    """Step-by-step guided form for creating a task."""
    awaiting_title = State()
    awaiting_priority = State()
    awaiting_deadline = State()           # step 1: pick a day
    awaiting_deadline_time = State()      # step 2: pick a time for the chosen day
    awaiting_deadline_manual = State()    # power path: type a full date+time
    awaiting_assignee = State()
    awaiting_confirm = State()


class NewReminderFSM(StatesGroup):
    """Step-by-step guided form for creating a standalone reminder."""
    awaiting_title = State()
    awaiting_time = State()
    awaiting_time_manual = State()
    awaiting_repeat = State()
    awaiting_confirm = State()


class ReminderSearchFSM(StatesGroup):
    awaiting_query = State()


class ReminderEditFSM(StatesGroup):
    awaiting_value = State()


def task_edit_menu(task: dict) -> InlineKeyboardMarkup:
    """Menu of editable fields for a task — used by 📝 Tahrir flow. Fields are laid
    out 2-per-row (matching the Tasks section reply-kbd convention) so the menu
    stays compact (5 rows) instead of sprawling one-field-per-row (was 9 rows)."""
    tid = task["id"]
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📝 Sarlavha", callback_data=f"editfield:{tid}:title"),
            InlineKeyboardButton(text="📄 Tavsif", callback_data=f"editfield:{tid}:description"),
        ],
        [
            InlineKeyboardButton(text="⚡ Prioritet", callback_data=f"editfield:{tid}:priority"),
            InlineKeyboardButton(text="📅 Deadline", callback_data=f"editfield:{tid}:deadline"),
        ],
        [
            InlineKeyboardButton(text="👤 Ijrochi", callback_data=f"set_assignee:{tid}"),
            InlineKeyboardButton(text="📊 Status", callback_data=f"editfield:{tid}:status"),
        ],
        [
            InlineKeyboardButton(text="🏷 Teglar", callback_data=f"editfield:{tid}:tags"),
            InlineKeyboardButton(text="📁 Kategoriya", callback_data=f"editfield:{tid}:category"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"taskopen:{tid}")],
    ])


def priority_picker(task_id: str) -> InlineKeyboardMarkup:
    # Uzbek labels — NewTaskFSM bilan bir xil; raw P0/P1/P2/P3 kodlari foydalanuvchiga ko'rinmasin.
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔴 Shoshilinch", callback_data=f"setfield:{task_id}:priority:P0"),
            InlineKeyboardButton(text="🟠 Muhim",      callback_data=f"setfield:{task_id}:priority:P1"),
        ],
        [
            InlineKeyboardButton(text="🔵 Rejadagi",   callback_data=f"setfield:{task_id}:priority:P2"),
            InlineKeyboardButton(text="⚪ Oddiy",      callback_data=f"setfield:{task_id}:priority:P3"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"edit:{task_id}")],
    ])


def status_picker(task_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⏳ Bajariladi", callback_data=f"setfield:{task_id}:status:todo"),
            InlineKeyboardButton(text="🔄 Jarayonda", callback_data=f"setfield:{task_id}:status:in_progress"),
        ],
        [
            InlineKeyboardButton(text="⚠️ To'silgan", callback_data=f"setfield:{task_id}:status:blocked"),
            InlineKeyboardButton(text="✅ Bajarildi", callback_data=f"setfield:{task_id}:status:done"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"edit:{task_id}")],
    ])


def deadline_picker(task_id: str) -> InlineKeyboardMarkup:
    """Quick deadline presets + manual entry."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Bugun 17:00", callback_data=f"deadline_preset:{task_id}:today"),
            InlineKeyboardButton(text="📅 Ertaga 09:00", callback_data=f"deadline_preset:{task_id}:tomorrow"),
        ],
        [
            InlineKeyboardButton(text="📅 +3 kun", callback_data=f"deadline_preset:{task_id}:plus3"),
            InlineKeyboardButton(text="📅 Hafta oxiri", callback_data=f"deadline_preset:{task_id}:weekend"),
        ],
        [
            InlineKeyboardButton(text="✏️ Qo'lda kiritish", callback_data=f"deadline_manual:{task_id}"),
            InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"setfield:{task_id}:deadline:none"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"edit:{task_id}")],
    ])


def meeting_inline_actions(meeting: dict) -> InlineKeyboardMarkup:
    """Drill-down actions for a single meeting — STATE-AWARE.

    Before the meeting happens (not completed): ✅ Bo'ldi · reschedule · edit ·
    cancel. NO "Bayonnoma yaratish" — minutes only make sense after the meeting,
    and a cancelled meeting needs none.
    After ✅ Bo'ldi (completed): 📝 Bayonnoma yaratish · ↺ undo · ⬅️ Ro'yxatga.
    """
    mid = meeting["id"]
    if meeting.get("completed_at"):
        fu = meeting.get("follow_up_actions") or []
        has_proto = bool(fu) if isinstance(fu, list) else bool(str(fu).strip())
        rows: list = []
        if has_proto:  # protocol exists → let the principal VIEW it (was unfindable)
            rows.append([InlineKeyboardButton(text="📄 Bayonnomani ko'rish", callback_data=f"viewproto:{mid}")])
            rows.append([InlineKeyboardButton(text="✏️ Qayta tuzish", callback_data=f"protocol:{mid}")])
        else:
            rows.append([InlineKeyboardButton(text="📝 Bayonnoma yaratish", callback_data=f"protocol:{mid}")])
        rows.append([InlineKeyboardButton(text="↺ Bo'ldi'ni bekor qilish", callback_data=f"meeting_undone:{mid}")])
        rows.append([back_button(f"meetingfilter:{_last_meeting_filter or 'week'}", "⬅️ Ro'yxatga")])
        return InlineKeyboardMarkup(inline_keyboard=rows)
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Bo'ldi", callback_data=f"meeting_done:{mid}")],
        # One-tap reschedule (like task snooze). Reuses the existing resched_preset
        # handler (DB update + iCloud sync + re-render). "🔄 Vaqtni o'zgartirish"
        # still opens the full 6-preset + manual picker for other times.
        [
            InlineKeyboardButton(text="📅 Ertaga 09:00", callback_data=f"resched_preset:{mid}:tomorrow_9"),
            InlineKeyboardButton(text="📅 Keyingi hafta", callback_data=f"resched_preset:{mid}:next_week"),
        ],
        [InlineKeyboardButton(text="🔄 Boshqa vaqt", callback_data=f"reschedule:{mid}")],
        [
            InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"meeting_edit:{mid}"),
            InlineKeyboardButton(text="✕ Bekor qilish", callback_data=f"meeting_cancel:{mid}"),
        ],
        [back_button(f"meetingfilter:{_last_meeting_filter or 'week'}")],
    ])


def settings_keyboard(settings: dict) -> InlineKeyboardMarkup:
    """Settings menu — actions + Back."""
    notif_on = settings.get("notifications_enabled", True)
    notif_label = "🔔 Bildirishnomalar: YOQ" if notif_on else "🔕 Bildirishnomalar: O'CHIQ"
    morning_time = settings.get("morning_briefing_time", "09:00")
    evening_time = settings.get("evening_summary_time", "18:00")
    quiet_on = settings.get("quiet_hours_enabled", False)
    qh_start = settings.get("quiet_hours_start", "22:00")
    qh_end = settings.get("quiet_hours_end", "07:00")
    quiet_label = (f"🌙 Sukunat: {qh_start}–{qh_end}" if quiet_on
                    else "🌙 Sukunat: o'chiq")
    voice_auto = settings.get("voice_auto_confirm", True)
    voice_label = ("🎙 Ovoz: AVTO (tasdiqsiz)" if voice_auto
                    else "🎙 Ovoz: tasdiq so'rash")
    confirm_create = settings.get("confirm_create_actions", True)
    create_label = ("⚠️ Yaratish tasdig'i: YOQ" if confirm_create
                     else "⚠️ Yaratish tasdig'i: O'CHIQ")
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=notif_label, callback_data="setting:notifications_toggle")],
        [InlineKeyboardButton(text=f"⏰ Brifing vaqti: {morning_time}", callback_data="setting:briefing_time")],
        [InlineKeyboardButton(text=f"🌙 Kechki yakun: {evening_time}", callback_data="setting:evening_time")],
        [InlineKeyboardButton(text=quiet_label, callback_data="setting:quiet_hours")],
        [InlineKeyboardButton(text=voice_label, callback_data="setting:voice_auto_toggle")],
        [InlineKeyboardButton(text=create_label, callback_data="setting:confirm_create_toggle")],
        [InlineKeyboardButton(text="📲 Eslatma parametrlari", callback_data="setting:reminders")],
        [InlineKeyboardButton(text="📅 Kalendar holati", callback_data="setting:calendar")],
        [back_button()],
    ])


async def _keep_typing(bot: Bot, chat_id: int) -> None:
    """Persistent typing indicator. Telegram's indicator lasts ~5s, so we refresh every 4s."""
    try:
        while True:
            await bot.send_chat_action(chat_id, "typing")
            await asyncio.sleep(4)
    except asyncio.CancelledError:
        return
    except Exception as e:
        # Network blip, bot blocked, chat deleted — don't crash the caller, but
        # leave a breadcrumb so silent typing-failures are diagnosable.
        logger.debug("_keep_typing stopped on chat %s: %s", chat_id, e)

logger = logging.getLogger(__name__)
router = Router()


# Strong refs to fire-and-forget tasks. Without these, Python's GC can collect
# the task object mid-flight (PEP 3156 only holds weak refs), causing the
# coroutine to be silently cancelled and any exception swallowed.
_background_tasks: set[asyncio.Task] = set()


def _log_background_exception(task: asyncio.Task) -> None:
    if task.cancelled():
        return
    exc = task.exception()
    if exc is not None:
        logger.error("Background task %r failed: %s", task.get_name(), exc, exc_info=exc)


def _spawn_background(coro, *, name: str) -> asyncio.Task:
    """Schedule a fire-and-forget coroutine while keeping a strong reference and
    logging any unhandled exception. Use instead of bare asyncio.create_task()
    for work that runs past the current handler's response."""
    task = asyncio.create_task(coro, name=name)
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)
    task.add_done_callback(_log_background_exception)
    return task


def _cb_int(data: str | None, *, default: int = 0) -> int:
    """Extract an int from callback_data of the form 'prefix:N'. Returns
    `default` on any parse failure (missing data, no colon, non-numeric tail)
    so a malformed or stale callback can't crash the handler with
    IndexError/ValueError. The auth middleware blocks non-principal users
    so this is mainly defence against stale buttons from older bot versions."""
    if not data:
        return default
    parts = data.split(":", 1)
    if len(parts) < 2:
        return default
    try:
        return int(parts[1])
    except (TypeError, ValueError):
        return default


def _cb_part(data: str | None, index: int, *, default: str = "") -> str:
    """Safe positional access into a colon-separated callback_data string.
    `_cb_part('a:b:c', 2)` → 'c'; `_cb_part('a:b', 2)` → ''. Use in place of
    `query.data.split(':')[index]` when the tail is optional."""
    if not data:
        return default
    parts = data.split(":")
    if 0 <= index < len(parts):
        return parts[index]
    return default


async def _get_text_or_transcribe(message: Message, bot: Bot | None = None) -> str | None:
    """Universal text extractor for state handlers that accept F.text | F.voice.
    Returns the message text (for text messages) OR the STT transcript (for
    voice). Returns None if the voice download/transcribe failed — caller
    should respond with a retry prompt in that case.

    This lets every FSM state accept voice without each handler re-implementing
    download + transcribe + error messaging."""
    if message.text:
        return message.text
    if message.voice and (bot or message.bot):
        b = bot or message.bot
        if message.voice.file_size and message.voice.file_size > voice_service.MAX_AUDIO_BYTES:
            await message.answer(
                f"Ovoz xabari juda katta ({message.voice.file_size // 1024} KB). "
                f"Iltimos, {voice_service.MAX_AUDIO_BYTES // (1024 * 1024)} MB dan kichikroq yuboring."
            )
            return None
        try:
            file = await b.get_file(message.voice.file_id)
            audio_io = await b.download_file(file.file_path)
            if hasattr(audio_io, "getvalue"):
                audio_bytes = audio_io.getvalue()
            elif hasattr(audio_io, "read"):
                audio_bytes = audio_io.read()
            else:
                audio_bytes = bytes(audio_io)
        except Exception:
            logger.exception("Voice download failed")
            await message.answer("Ovozni yuklab ololmadim. Iltimos, qaytadan yuboring.")
            return None
        transcript = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not transcript:
            await message.answer("Ovozni o'qiy olmadim. Matn yozib yuboring yoki qaytadan urinib ko'ring.")
            return None
        # Patch message.text in-place so handler bodies using `message.text`
        # transparently see the transcript. Pydantic v2 forbids regular
        # assignment on Message; object.__setattr__ bypasses the descriptor.
        try:
            object.__setattr__(message, "text", transcript)
        except Exception:
            logger.debug("Could not patch message.text in-place; caller must use return value")
        return transcript
    return None


# ─────────────────────── AUTH MIDDLEWARE ───────────────────────


def _is_principal(user_id: int | None) -> bool:
    return user_id == config.PRINCIPAL_USER_ID


def _is_private_chat(chat) -> bool:
    """True if chat is a 1:1 private chat. This bot is designed for the
    principal in private — group/channel/supergroup behavior is undefined
    (formatting, FSM state, mentions all break)."""
    return chat is not None and chat.type == "private"


async def _reject(message_or_query, reason: str = "Bu shaxsiy bot."):
    if isinstance(message_or_query, Message):
        await message_or_query.answer(reason)
    elif isinstance(message_or_query, CallbackQuery):
        await message_or_query.answer(reason, show_alert=True)


_FSM_STATE_TTL_SECONDS = 30 * 60  # 30 minutes


@router.message.middleware()
async def auth_message_middleware(handler, event: Message, data):
    if not _is_principal(event.from_user.id if event.from_user else None):
        await _reject(event)
        return
    # Reject group/channel use — even from the principal. Forwarding the bot
    # to a group is the most common accident; this keeps state and formatting
    # behaviour predictable. The principal can DM directly.
    if not _is_private_chat(event.chat):
        await _reject(event, "Bu bot faqat shaxsiy chat'da ishlaydi. Iltimos, men bilan to'g'ridan-to'g'ri yozing.")
        return
    # FSM state TTL: if the user started a flow (say, NewTaskFSM.awaiting_title),
    # walked away for >30 min, and comes back with a random "Salom", we'd
    # store "Salom" as the task title. Detect stale state via timestamp in
    # state.data and auto-clear with a friendly notice.
    state = data.get("state")
    if state is not None:
        try:
            current = await state.get_state()
            if current and current != "default_state":
                sdata = await state.get_data()
                started_at = sdata.get("_fsm_started_at")
                now_ts = datetime.now(database.TZ).timestamp()
                if started_at and (now_ts - float(started_at)) > _FSM_STATE_TTL_SECONDS:
                    await state.clear()
                    await event.answer(
                        "⏱ Avvalgi amaliyot 30 daqiqadan ortiq vaqt o'tgani uchun bekor qilindi. "
                        "Yangi xabar yuboring yoki /start bosing.",
                    )
                    return
                # Refresh the timestamp on each interaction so an active flow
                # never times out mid-conversation.
                await state.update_data(_fsm_started_at=now_ts)
        except Exception:
            logger.debug("FSM TTL check failed (non-fatal)", exc_info=True)
    return await handler(event, data)


@router.callback_query.middleware()
async def auth_callback_middleware(handler, event: CallbackQuery, data):
    if not _is_principal(event.from_user.id if event.from_user else None):
        await _reject(event)
        return
    if event.message and not _is_private_chat(event.message.chat):
        await _reject(event, "Faqat shaxsiy chat'da ishlaydi.")
        return
    return await handler(event, data)


# Inline-query auth happens INSIDE the handler (handle_inline_query) — we return
# empty results to non-principal users without alerting them. Telegram's inline UI
# doesn't tolerate the same kind of rejection messages we use elsewhere.


# ─────────────────────── ACTION EXECUTOR ───────────────────────


def _icloud_push_payload(meeting_id: str, data: dict, end_iso: "str | None") -> dict:
    """Retry-queue payload matching _icloud_retry_sweep's reader (dt_start/dt_end/…)."""
    return {
        "title": data.get("title", "Uchrashuv"),
        "dt_start": data.get("datetime_start"),
        "dt_end": end_iso or data.get("datetime_end") or data.get("datetime_start"),
        "participants": data.get("participants"),
        "location": data.get("location_or_link"),
        "description": data.get("agenda"),
    }


async def _push_meeting_to_icloud(meeting_id: str, data: dict) -> None:
    """Background-task push: doesn't block the bot's reply. On failure (no UID or an
    exception) it ENQUEUES a retry so the scheduler's _icloud_retry_sweep recovers it
    — otherwise a failed push left the meeting permanently 'kutilmoqda' (silent gap)."""
    end_iso = None
    try:
        start_dt = datetime.fromisoformat(data["datetime_start"])
        end_iso = data.get("datetime_end") or (start_dt + timedelta(hours=1)).isoformat()
        end_dt = datetime.fromisoformat(end_iso)
        uid = await asyncio.to_thread(
            calendar_service.push_meeting,
            meeting_id, data.get("title", "Uchrashuv"), start_dt, end_dt,
            data.get("participants"), data.get("location_or_link"),
            data.get("agenda"),
        )
        if uid:
            import aiosqlite
            async with aiosqlite.connect(config.DATABASE_PATH) as db:
                await db.execute("UPDATE meetings SET icloud_uid = ? WHERE id = ?", (uid, meeting_id))
                await db.commit()
            logger.info("iCloud sync complete for %s", meeting_id)
        else:
            logger.warning("iCloud push returned no UID for %s — queuing retry", meeting_id)
            await database.enqueue_icloud_retry(
                "push", meeting_id, _icloud_push_payload(meeting_id, data, end_iso),
                "push returned None")
    except Exception as e:
        logger.exception("Background iCloud push failed for %s — queuing retry", meeting_id)
        try:
            await database.enqueue_icloud_retry(
                "push", meeting_id, _icloud_push_payload(meeting_id, data, end_iso),
                f"{type(e).__name__}: {e}")
        except Exception:
            logger.exception("enqueue_icloud_retry also failed for %s", meeting_id)


_SELF_ASSIGNEE_NAMES = {"men", "o'zim", "ozim", "o'z", "oz", ""}


def _norm_asg_key(s) -> str:
    """Allowlist key for an assignee name: strip + casefold + fold EVERY apostrophe
    variant (ʼ ’ ʻ ` …) to the plain ' — so \"O'zim\"/\"O’zim\"/\"Oʻzim\" (and to_latin's
    U+02BB output for Cyrillic Ўзим/Ғайрат) all match the same key. Mirrors what
    _norm_header does for column names."""
    import translit
    s = str(s or "").strip().casefold()
    for ch in translit._APOS:
        s = s.replace(ch, "'")
    return s


async def _load_contact_map() -> dict:
    """{normalized name key: canonical name} for the assignee allowlist. The
    canonical casing is written back on import so a case-only edit ('Aziz'→'aziz')
    never desyncs the stored assignee from the contact (which would split
    per-person workload/stats into two buckets). Keys are stripped and
    apostrophe-folded (_norm_asg_key) — a trailing space or a typographic
    apostrophe in either side must not defeat the match."""
    return {_norm_asg_key(c.get("name")): (c.get("name") or "").strip()
            for c in await database.list_contacts()}


async def _canon_assignee(raw: str, from_excel: bool, contacts: dict):
    """Canonicalize an assignee cell/value against the contact allowlist.
    Handles the multi-name form: 'Karimov / Aziz' (export display join) and
    'Karimov/Aziz' (stored form) are split into parts; each part is resolved to
    its canonical contact casing. Self-references (men/o'zim, any apostrophe
    variant or script) are dropped — a task the principal does himself is simply
    unassigned. Unknown parts: the trusted Excel path registers them as contacts
    INDIVIDUALLY (never the combined 'A / B' string — that used to create bogus
    contacts); an LLM turn returns None so the caller keeps its old drop behavior.
    Returns the canonical 'A/B' string, or '' when nothing assignable remains."""
    parts = [p.strip() for p in str(raw or "").split("/") if p.strip()]
    out = []
    for p in parts:
        key = _norm_asg_key(p)
        if key in _SELF_ASSIGNEE_NAMES:
            continue
        if key in contacts:
            out.append(contacts[key])
        elif from_excel:
            await _upsert_contacts([p])
            contacts[key] = p
            out.append(p)
        else:
            return None  # unknown executor on an LLM turn — caller drops the field
    return "/".join(out)


async def _upsert_contacts(names: list[str]) -> int:
    """Avtomatik tarzda kontaktlar jadvaliga ismlarni qo'shadi.
    Mavjud bo'lganlar (case-insensitive) o'tkazib yuboriladi. 'men/o'zim' kabi
    self-reference ismlar saqlanmaydi. Yangi yaratilganlar sonini qaytaradi.
    """
    clean = []
    for raw in names:
        name = (raw or "").strip()
        # Apostrophe-folded check: "O’zim"/"Oʻzim"/"Ўзим→Oʻzim" are all self-references.
        if not name or _norm_asg_key(name) in _SELF_ASSIGNEE_NAMES:
            continue
        clean.append(name)
    if not clean:
        return 0
    try:
        existing = await database.list_contacts()
        existing_names = {_norm_asg_key(c["name"]) for c in existing}
    except Exception:
        logger.warning("Auto-contact: failed to load existing contacts")
        return 0
    created = 0
    for name in clean:
        if _norm_asg_key(name) in existing_names:
            continue
        try:
            await database.save_contact({
                "name": name,
                "role": None,
                "formality_level": 3,
            })
            existing_names.add(_norm_asg_key(name))
            created += 1
        except Exception:
            logger.warning("Auto-contact upsert failed for %s", name)
    return created


# Settings reachable by voice/text via the `update_setting` action.
_SETTING_BOOL_KEYS = {"notifications_enabled", "voice_auto_confirm",
                      "confirm_create_actions", "quiet_hours_enabled"}
_SETTING_TIME_KEYS = {"morning_briefing_time", "evening_summary_time",
                      "quiet_hours_start", "quiet_hours_end"}
_SETTING_INT_KEYS = {"meeting_reminder_min", "task_reminder_hours"}
_BRIEFING_TIME_KEYS = {"morning_briefing_time", "evening_summary_time"}


def _coerce_setting_value(key: str, value):
    """Coerce a Claude-provided setting value to the type the DB expects.
    Returns (ok, coerced_value)."""
    if key in _SETTING_BOOL_KEYS:
        if isinstance(value, bool):
            return True, value
        s = str(value).strip().lower()
        if s in ("1", "true", "yes", "on", "yoq", "yoqilgan", "ha"):
            return True, True
        if s in ("0", "false", "no", "off", "ochir", "o'chirilgan", "yo'q"):
            return True, False
        return False, None
    if key in _SETTING_INT_KEYS:
        try:
            return True, int(value)
        except (TypeError, ValueError):
            return False, None
    if key in _SETTING_TIME_KEYS:
        s = str(value).strip()
        return (True, s) if re.match(r"^\d{1,2}:\d{2}$", s) else (False, None)
    return False, None


async def _apply_setting_action(data: dict) -> None:
    """Handle the `update_setting` action — voice/text parity for settings toggles.
    Whitelisted keys only; briefing-time keys reschedule the scheduler live."""
    key = (data.get("key") or "").strip()
    if key not in (_SETTING_BOOL_KEYS | _SETTING_TIME_KEYS | _SETTING_INT_KEYS):
        logger.warning("update_setting: unsupported key %r", key)
        return
    ok, value = _coerce_setting_value(key, data.get("value"))
    if not ok:
        logger.warning("update_setting: bad value for %r: %r", key, data.get("value"))
        return
    await database.set_setting(key, value)
    if key in _BRIEFING_TIME_KEYS:
        await _reschedule_briefings_live()
    logger.info("update_setting via voice/text: %s = %r", key, value)


async def _execute_actions(actions: list[dict]) -> dict[str, list[str]]:
    """Execute Claude-returned actions. Return map of type → list of affected IDs.
    Controlled lists (A): task assignees and categories are added ONLY via Excel
    import (source="excel") or the bot's own UI — never auto-created from an LLM
    turn. Meeting participants are still auto-upserted into contacts.
    """
    created_ids: dict[str, list[str]] = {
        "task": [], "reminder": [], "meeting": [], "contact": [], "correction": [],
        "note": [], "_failed": [], "_refresh": [], "_conflict": [], "_badtime": [],
    }
    _existing_cats: "set | None" = None       # lazy allowlist — categories (manual/Excel only)
    _existing_contacts: "dict | None" = None  # lazy allowlist {casefold: canonical name}
    # Excel hierarchy: "3.1" → subtask of the row numbered "3". Maps each row's №
    # to the task id it created/updated, so a child resolves its parent_id.
    num_to_id: dict[str, str] = {}
    # Order-independent: an exported sheet may be re-sorted so a subtask row comes
    # BEFORE its parent. Seed the map from every UPDATE row (its id is known upfront),
    # so a dotted child resolves its parent regardless of row order (bug: sorting the
    # file detached subtasks). New parents (create) are still added during the loop.
    for _a in actions:
        if _a.get("_num") and _a.get("type") == "update_task" and _a.get("id"):
            num_to_id.setdefault(str(_a["_num"]), _a["id"])

    def _resolve_parent(act: dict, data: dict) -> None:
        """№-driven re-parenting: a dotted № → child of its top-level ancestor; a plain
        № → top-level (parent_id=None). No-op for non-Excel rows and for flat per-assignee
        files (act["_flat"]) where subtasks are numbered plainly and the parent lives in an
        'Asosiy vazifa' column — re-parenting from that would wrongly promote them."""
        num = act.get("_num")
        if not num or act.get("_flat"):
            return
        data["parent_id"] = num_to_id.get(str(num).split(".")[0]) if "." in str(num) else None

    for action in actions:
        atype = action.get("type", "")
        data = action.get("data", {})
        target_id = action.get("id")

        try:
            if atype == "create_task":
                # Controlled lists (A): categories AND assignees are added ONLY via Excel
                # import (source="excel") or the bot's own UI — never auto-created from a
                # voice/text turn. On an LLM create, an unknown category/assignee is
                # DROPPED (task stays uncategorized/unassigned). An Excel import may
                # introduce new ones (creates the category / contact and keeps them).
                _from_excel = (data.get("source") or "") == "excel"
                _cat = (data.get("category") or "").strip()
                if _cat:
                    if _existing_cats is None:
                        _existing_cats = {c.casefold() for c in await database.existing_category_names()}
                    if _cat.casefold() not in _existing_cats:
                        if _from_excel:
                            await database.create_category(_cat)
                            _existing_cats.add(_cat.casefold())
                        else:
                            data["category"] = None
                _asg = (data.get("assignee") or "").strip()
                if _asg:
                    if _existing_contacts is None:
                        _existing_contacts = await _load_contact_map()
                    _canon = await _canon_assignee(_asg, _from_excel, _existing_contacts)
                    # None → unknown on an LLM turn; '' → only self-references → unassigned
                    data["assignee"] = _canon or None
                _resolve_parent(action, data)
                tid = await database.create_task(data)
                created_ids["task"].append(tid)
                if action.get("_num"):
                    num_to_id[action["_num"]] = tid
            elif atype == "create_reminder":
                rid = await database.create_reminder(data)
                created_ids["reminder"].append(rid)
            elif atype == "update_task" and target_id:
                # Same controlled-assignee rule on edits: LLM update to an unknown
                # executor is dropped; an Excel round-trip may introduce one.
                if (data.get("assignee") or "").strip():
                    _asg = data["assignee"].strip()
                    if _existing_contacts is None:
                        _existing_contacts = await _load_contact_map()
                    _canon = await _canon_assignee(
                        _asg, (data.get("source") or "") == "excel", _existing_contacts)
                    if _canon is None:
                        data.pop("assignee", None)   # unknown on an LLM turn — untouched
                    else:
                        data["assignee"] = _canon or None  # '' (self-ref only) → unassign
                _resolve_parent(action, data)
                # A status→"done" flip must go through complete_task so a RECURRING task
                # spawns its next occurrence (update_task alone never does). Apply the
                # other field edits first, then let complete_task own the done transition.
                _mark_done = (data.get("status") == "done")
                if _mark_done:
                    _other = {k: v for k, v in data.items() if k != "status"}
                    if _other:
                        await database.update_task(target_id, _other)
                    _ok = await database.complete_task(target_id)
                else:
                    _ok = await database.update_task(target_id, data)
                # A missing/deleted id → update/complete returns False. Surface it
                # instead of silently reporting success (the LLM may act on a stale
                # OXIRGI KO'RSATILGAN id for a task that was just deleted/completed).
                if _ok:
                    created_ids["task"].append(target_id)
                    if action.get("_num"):
                        num_to_id[action["_num"]] = target_id
                else:
                    created_ids["_failed"].append(atype)
            elif atype == "complete_task" and target_id:
                if await database.complete_task(target_id):
                    created_ids["task"].append(target_id)
                else:
                    created_ids["_failed"].append(atype)
            elif atype == "delete_task" and target_id:
                if not await database.delete_task(target_id):
                    created_ids["_failed"].append(atype)
            elif atype == "schedule_meeting":
                # Vaqtni aniqlab bo'lmasa (bo'sh yoki yaroqsiz datetime_start) —
                # uchrashuvni YARATMAYMIZ. Aks holda kalendar so'rovlari (lexical
                # BETWEEN) topa olmaydigan "ko'rinmas" qator saqlanardi va
                # foydalanuvchiga "qo'shildi" deb yolg'on aytilardi.
                if database.parse_iso_dt(data.get("datetime_start")) is None:
                    created_ids["_badtime"].append(
                        (data.get("title") or "Uchrashuv").strip())
                    logger.info(
                        "schedule_meeting rad etildi — datetime_start yaroqsiz: %r",
                        data.get("datetime_start"))
                    continue
                # To'qnashuvni oldini olish: bir xil vaqtga ustma-ust uchrashuv
                # qo'ymaymiz. Vaqti mavjud uchrashuv bilan kesishsa — yaratmaymiz
                # va foydalanuvchini ogohlantiramiz.
                conflicts = await database.find_meeting_conflicts(
                    data.get("datetime_start"), data.get("datetime_end"))
                if conflicts:
                    created_ids["_conflict"].append(_conflict_summary(data, conflicts))
                    logger.info(
                        "schedule_meeting o'tkazib yuborildi — vaqt to'qnashuvi (%d ta uchrashuv)",
                        len(conflicts))
                else:
                    mid = await database.create_meeting(data)
                    created_ids["meeting"].append(mid)
                    sched = scheduler_module.get_scheduler()
                    if sched and data.get("datetime_start"):
                        sched.schedule_meeting_reminder(mid, data["datetime_start"])
                    # Push to iCloud as a FIRE-AND-FORGET background task — the user gets their
                    # bot reply within ~1 second; iCloud sync happens in parallel (typically 1-3s).
                    if config.ICLOUD_ENABLED and data.get("datetime_start"):
                        _spawn_background(_push_meeting_to_icloud(mid, data), name=f"icloud_push:{mid}")
                    await _upsert_contacts(list(data.get("participants") or []))
            elif atype == "cancel_meeting" and target_id:
                await database.cancel_meeting(target_id)
                sched = scheduler_module.get_scheduler()
                if sched:
                    try:
                        sched.remove_meeting_reminder(target_id)
                    except Exception:
                        logger.exception("Failed to remove meeting reminder for %s", target_id)
                if config.ICLOUD_ENABLED:
                    try:
                        await asyncio.to_thread(calendar_service.delete_meeting, target_id)
                    except Exception:
                        logger.exception("iCloud delete failed (non-fatal)")
            elif atype == "save_contact":
                cid = await database.save_contact(data)
                if cid:
                    created_ids["contact"].append(cid)
            elif atype == "save_correction":
                corr_id = await database.save_correction(data)
                created_ids["correction"].append(corr_id)
            elif atype == "create_note":
                # Quick-capture inbox. Source defaults to "llm" — Claude
                # decided this was a note rather than a task/reminder.
                if data.get("content"):
                    nid = await database.create_note({
                        "content": data.get("content"),
                        "title": data.get("title"),
                        "tags": data.get("tags") or [],
                        "source": data.get("source") or "llm",
                    })
                    if nid:
                        created_ids["note"].append(nid)
            elif atype == "delete_all_tasks":
                n = await database.delete_all_tasks(data.get("status_in"))
                created_ids["_refresh"].append("task")
                logger.info("Bulk delete: %d tasks (status_in=%s)", n, data.get("status_in"))
            elif atype == "delete_all_meetings":
                n = await database.delete_all_meetings()
                created_ids["_refresh"].append("meeting")
                logger.info("Bulk delete: %d meetings", n)
            elif atype == "delete_all_notes":
                n = await database.delete_all_notes()
                created_ids["_refresh"].append("note")
                logger.info("Bulk delete: %d notes", n)
            elif atype == "delete_all_reminders":
                n = await database.delete_all_reminders()
                created_ids["_refresh"].append("reminder")
                logger.info("Bulk delete: %d reminders", n)
            elif atype == "delete_all_contacts":
                n = await database.delete_all_contacts()
                logger.info("Bulk delete: %d contacts", n)
            elif atype == "create_category":
                cat = (data.get("category") or "").strip()
                if cat:
                    await database.create_category(cat, data.get("icon"))
                    logger.info("Created category %r", cat)
            elif atype == "archive_category":
                cat = (data.get("category") or "").strip()
                if cat:
                    await database.archive_category(cat, bool(data.get("archived", True)))
                    created_ids["_refresh"].append("task")
                    logger.info("Archived category %r", cat)
            elif atype == "delete_category":
                cat = (data.get("category") or "").strip()
                if cat:
                    n = await database.clear_category(cat)
                    await database.delete_category_record(cat)  # remove metadata row too
                    created_ids["_refresh"].append("task")
                    logger.info("Cleared+removed category %r from %d tasks", cat, n)
            elif atype == "delete_tasks_by_category":
                cat = (data.get("category") or "").strip()
                if cat:
                    n = await database.delete_tasks_by_category(cat)
                    created_ids["_refresh"].append("task")
                    logger.info("Deleted %d tasks in category %r", n, cat)
            elif atype == "assign_category":
                cat = (data.get("category") or "").strip()
                frm = (data.get("from_category") or "").strip()
                if cat and frm:
                    n = await database.rename_category(frm, cat)
                    created_ids["_refresh"].append("task")
                    logger.info("Reassigned %d tasks: %r → %r", n, frm, cat)
            # ── Parity actions: every button operation also reachable by voice/text ──
            elif atype == "reopen_task" and target_id:
                await database.update_task(target_id, {"status": "todo"}, source="reopen")
                created_ids["task"].append(target_id)
            elif atype == "complete_reminder" and target_id:
                await database.complete_reminder(target_id)
                created_ids["reminder"].append(target_id)
            elif atype == "update_reminder" and target_id:
                # Covers snooze (remind_at), title/note/recurrence edits.
                await database.update_reminder(target_id, data)
                created_ids["reminder"].append(target_id)
            elif atype == "delete_reminder" and target_id:
                await database.cancel_reminder(target_id)
            elif atype == "complete_meeting" and target_id:
                if data.get("undo"):
                    await database.uncomplete_meeting(target_id)
                else:
                    await database.complete_meeting(target_id)
                created_ids["meeting"].append(target_id)
            elif atype == "update_meeting" and target_id:
                # Reschedule (datetime_start/end), duration, title, participants…
                await database.update_meeting(target_id, data)
                created_ids["meeting"].append(target_id)
                if data.get("assignee") or data.get("participants"):
                    await _upsert_contacts(list(data.get("participants") or []))
            elif atype == "note_to_task" and target_id:
                note = await database.get_note(target_id)
                if note:
                    n_title = (note.get("title") or note.get("content", "Qayddan vazifa")).strip()[:200]
                    n_desc = (note.get("content") or "").strip()
                    if n_desc == n_title:
                        n_desc = None
                    n_tags = list(note.get("tags") or []) + [f"note:{target_id}"]
                    ntid = await database.create_task({
                        "title": n_title, "description": n_desc,
                        "priority": data.get("priority") or "P2", "status": "todo",
                        "tags": n_tags, "source": "note",
                        "deadline": data.get("deadline"), "assignee": data.get("assignee"),
                    })
                    created_ids["task"].append(ntid)
                    await database.mark_note_processed(target_id, "task", ntid)
            elif atype == "note_to_reminder" and target_id:
                note = await database.get_note(target_id)
                if note and data.get("remind_at"):
                    nrid = await database.create_reminder({
                        "title": (note.get("title") or note.get("content", "Eslatma")).strip()[:200],
                        "note": note.get("content"), "remind_at": data.get("remind_at"),
                        "source": "note",
                    })
                    created_ids["reminder"].append(nrid)
                    await database.mark_note_processed(target_id, "reminder", nrid)
            elif atype == "update_note" and target_id:
                # status changes: processed (done) / archived / inbox (restore), or edit.
                await database.update_note(target_id, data)
                created_ids["note"].append(target_id)
                created_ids["_refresh"].append("note")
            elif atype == "delete_note" and target_id:
                await database.delete_note(target_id)
                created_ids["_refresh"].append("note")
            elif atype == "update_category":
                cat = (data.get("category") or "").strip()
                if cat:
                    await database.update_category(
                        cat, new_name=(data.get("new_name") or None),
                        icon=(data.get("icon") or None))
                    if "archived" in data:
                        await database.archive_category(cat, bool(data.get("archived")))
                    created_ids["_refresh"].append("task")
            elif atype == "move_category":
                cat = (data.get("category") or "").strip()
                direction = (data.get("direction") or "").strip().lower()
                if cat and direction in ("up", "down"):
                    await database.move_category(cat, direction)
                    created_ids["_refresh"].append("task")
            elif atype == "update_setting":
                await _apply_setting_action(data)
            elif atype == "none":
                pass
            else:
                logger.warning("Unknown action type: %s", atype)
        except Exception:
            logger.exception("Failed to execute action %s", action)
            created_ids["_failed"].append(atype or "unknown")

    return created_ids


# Action type → user-facing noun for the "saqlanmadi" warning.
_ACTION_NOUN_UZ = {
    "create_task": "vazifa", "update_task": "vazifa yangilash",
    "complete_task": "vazifa yakunlash", "delete_task": "vazifa o'chirish",
    "create_reminder": "eslatma",
    "schedule_meeting": "uchrashuv", "cancel_meeting": "uchrashuvni bekor qilish",
    "create_note": "note", "save_contact": "kontakt",
}


def _fmt_meeting_when(start_iso: str | None) -> str:
    """ISO vaqtni 'DD.MM HH:MM' ko'rinishida; o'qilmasa bo'sh satr."""
    dt = database.parse_iso_dt(start_iso) if start_iso else None
    return dt.strftime("%d.%m %H:%M") if dt else ""


def _conflict_summary(data: dict, conflicts: list[dict]) -> str:
    """Yaratilmagan uchrashuv haqida qisqa, foydalanuvchiga ko'rsatiladigan satr."""
    title = (data.get("title") or "Uchrashuv").strip()
    when = _fmt_meeting_when(data.get("datetime_start"))
    clash = conflicts[0]
    clash_title = (clash.get("title") or "Uchrashuv").strip()
    clash_when = _fmt_meeting_when(clash.get("datetime_start"))
    extra = f" (+{len(conflicts) - 1} ta boshqa)" if len(conflicts) > 1 else ""
    head = f"«{title}»" + (f" ({when})" if when else "")
    tail = f"«{clash_title}»" + (f" ({clash_when})" if clash_when else "") + extra
    return f"{head} — {tail} bilan to'qnashadi"


def _conflict_note(ids_by_type: dict[str, list[str]]) -> str:
    """Vaqt to'qnashuvi tufayli qo'yilmagan uchrashuvlar haqida ogohlantirish —
    javobga qo'shiladi, shunda jim qolib ketmaydi."""
    conflicts = ids_by_type.get("_conflict") if ids_by_type else None
    if not conflicts:
        return ""
    lines = "\n".join(f"• {c}" for c in conflicts)
    return (f"\n\n⚠️ **Uchrashuv qo'yilmadi (vaqt band):**\n{lines}\n"
            f"Boshqa vaqt tanlang yoki mavjud uchrashuvni bekor qiling.")


def _badtime_note(ids_by_type: dict[str, list[str]]) -> str:
    """Vaqti aniqlanmagan (bo'sh yoki yaroqsiz sana) uchun YARATILMAGAN
    uchrashuvlar haqida ogohlantirish. Ilgari bunday uchrashuv hech qaysi
    kalendar oynasida ko'rinmaydigan "ko'rinmas" qator bo'lib jim saqlanardi
    va foydalanuvchiga "qo'shildi" deb yolg'on aytilardi — endi ochiq aytamiz."""
    bad = ids_by_type.get("_badtime") if ids_by_type else None
    if not bad:
        return ""
    lines = "\n".join(f"• «{t}»" for t in bad)
    return (f"\n\n⚠️ **Uchrashuv qo'yilmadi (vaqt aniqlanmadi):**\n{lines}\n"
            f"Aniq sana va vaqt bilan qayta yozing (masalan: «ertaga 15:00»).")


def _failed_actions_note(ids_by_type: dict[str, list[str]]) -> str:
    """If any action failed inside _execute_actions, return a short warning to
    append to the reply — so a silent DB error never looks like success."""
    failed = ids_by_type.get("_failed") if ids_by_type else None
    if not failed:
        return ""
    nouns = ", ".join(sorted({_ACTION_NOUN_UZ.get(a, a) for a in failed}))
    return (f"\n\n⚠️ **Saqlanmadi:** {nouns}. Texnik xato yuz berdi — "
            f"qaytadan urinib ko'ring yoki /diagnostics.")


def _humanize_error(exc: BaseException | None) -> str:
    """Turn an exception into a clear, single root-cause message in O'zbek.

    This is a single-user bot (the principal owns it), so surfacing the REAL
    reason beats a generic "Texnik xato": known infra/API errors map to plain
    language, and anything unrecognised shows a short technical detail so the
    actual cause is visible (for reporting / fixing) instead of hidden.
    Classifies by exception class NAME + message text, so no extra imports of
    anthropic/httpx/aiogram exception types are needed."""
    # Plain text only (no markdown): an error notice MUST always deliver, and the
    # raw exception text in the unknown-error branch could contain markdown-breaking
    # characters that would make message.answer(parse_mode="Markdown") throw.
    if exc is None:
        return "⚠️ Noma'lum xato yuz berdi. Qaytadan urinib ko'ring."
    name = type(exc).__name__
    msg = str(exc).strip()
    low = msg.lower()

    # ── Network / connectivity ──
    if (name in ("APIConnectionError", "APITimeoutError", "ConnectError",
                 "ConnectTimeout", "ReadTimeout", "ReadTimeoutError",
                 "TelegramNetworkError", "ClientConnectorError")
            or "timed out" in low or "timeout" in low
            or "network is unreachable" in low or "connection" in low
            or "cannot connect" in low):
        return ("🌐 Sabab: Internet yoki server bilan ulanishda muammo "
                "(tarmoq uzilgan/sekin). Tarmoqni tekshirib, qayta urinib ko'ring.")

    # ── Anthropic: rate / auth / billing ──
    if name == "RateLimitError" or "rate limit" in low or "429" in low:
        return ("⏳ Sabab: Claude'ga juda ko'p so'rov yuborildi. "
                "1-2 daqiqadan keyin qayta urining.")
    if name == "AuthenticationError" or "api key" in low or "authentication" in low:
        return "🔑 Sabab: Claude API kalitida muammo. Sozlamalarni tekshiring."
    if "credit" in low or "billing" in low or "balance" in low or "insufficient" in low:
        return "💳 Sabab: Claude hisobida mablag' tugagan ko'rinadi. To'ldirish kerak."

    # ── Empty / malformed request content ──
    if "non-empty content" in low or ("messages" in low and "content" in low):
        return ("📭 Sabab: So'rov tarkibida bo'sh xabar bor edi. "
                "Bu odatda tuzatiladi — qayta urinib ko'ring.")

    # ── Truncated / unparseable model output ──
    if name == "JSONDecodeError" or ("json" in low and ("delimiter" in low or "expecting" in low)):
        return ("📄 Sabab: Claude javobi to'liq kelmadi (kesildi). "
                "Qisqaroq so'rov bilan qayta urinib ko'ring.")

    # ── Database ──
    # Disk I/O / "database or disk is full" almost always means the server disk (or
    # inodes) filled up — give the principal an actionable hint, not a raw SQLite msg.
    if ("disk i/o" in low or "disk is full" in low or "database or disk" in low
            or "no space" in low or "ioerr" in low):
        return ("🗄 Sabab: Serverда disk to'lgan ko'rinadi (disk I/O error). "
                "Log/backup fayllarni tozalash yoki diskni kengaytirish kerak. "
                "Server: `df -h` va `df -i` bilan tekshiring.")
    if "malformed" in low or "not a database" in low or "corrupt" in low:
        return ("🗄 Sabab: Ma'lumotlar bazasi fayli buzilgan bo'lishi mumkin "
                "(odatda to'satdan to'xtashdan). Serverда: botni to'xtatib "
                "`PRAGMA integrity_check;` bilan tekshiring, kerak bo'lsa backupdan tiklang.")
    if (name in ("OperationalError", "IntegrityError", "ProgrammingError", "DatabaseError")
            or "sqlite" in low or "database" in low):
        return f"🗄 Sabab: Ma'lumotlar bazasi xatosi — {msg[:120]}"

    # ── Telegram rejected the message ──
    if name == "TelegramBadRequest":
        return f"📨 Sabab: Telegram xabarni rad etdi — {msg[:120]}"

    # ── Unknown → show the real technical cause (owner bot) ──
    detail = f"{name}: {msg[:150]}" if msg else name
    return f"⚠️ Kutilmagan xato.\n🔧 Sabab: {detail}\n/diagnostics — batafsil holat."


# ─────────────────────── KEYBOARD BUILDER ───────────────────────


_TEMP_TOKENS = {
    "t-new": ("task", 0),
    "t-latest": ("task", -1),
    "m-new": ("meeting", 0),
    "m-latest": ("meeting", -1),
    "r-new": ("reminder", 0),
    "r-latest": ("reminder", -1),
    "c-new": ("contact", 0),
    "polish": (None, None),
}


def _resolve_callback(callback: str, ids_by_type: dict[str, list[str]]) -> str | None:
    """Replace placeholder tokens in callback data with real IDs from executed actions.

    Returns None if the callback references a placeholder that has no real ID (drops the button).
    """
    for placeholder, (entity_type, idx) in _TEMP_TOKENS.items():
        if placeholder in callback:
            if entity_type is None:
                return callback
            ids = ids_by_type.get(entity_type, [])
            if not ids:
                return None
            real_id = ids[idx] if idx == -1 else ids[0]
            return callback.replace(placeholder, real_id)
    return callback


# Inline ulashish uchun matn keshi — endi DB-backed (share_cache jadvali).
# Avval xotirada edi → bot restart'da token yo'qolib, 'ssilka yo'qolib ketardi'.


def _strip_polish_wrapper(raw: str) -> str:
    """'Tahrirlangan matn:' kartasidan faqat toza matnni ajratadi (sarlavha +
    ─── chiziqlar olib tashlanadi)."""
    out: list[str] = []
    for ln in (raw or "").split("\n"):
        s = ln.strip()
        if not s:
            out.append("")
            continue
        if "Tahrirlangan matn" in s or "Tahrirlangan xat" in s:
            continue
        if set(s) <= set("─—-_=•· "):
            continue
        out.append(ln)
    return "\n".join(out).strip()


def _cache_share_text(text: str) -> str:
    """Matnni inline ulashish uchun DB keshiga yozadi (sync sqlite3 — _build_keyboard
    sync funksiya), token (id) qaytaradi. DB-backed → bot restart'da yo'qolmaydi."""
    import sqlite3
    with sqlite3.connect(config.DATABASE_PATH) as db:
        cur = db.execute(
            "INSERT INTO share_cache (content, created_at) VALUES (?, ?)",
            ((text or "")[:4096], datetime.now(database.TZ).isoformat()))
        token = cur.lastrowid
        db.execute("DELETE FROM share_cache WHERE id < ?", (token - 500,))  # eskilarini tozalash
        db.commit()
    return str(token)


def _build_keyboard(buttons: list, ids_by_type: dict[str, list[str]],
                    share_text: str | None = None) -> InlineKeyboardMarkup | None:
    """Convert Claude's button structure to aiogram InlineKeyboardMarkup.

    Polish (intent B) javobida 'share' tugmasini IKKI xil ulashishga aylantiramiz:
    📋 Nusxa (qo'lda, callback 'share') + 📤 Ulashish (inline, switch_inline_query).
    Ortiqcha 'copy' (no-op) tugmasi tashlanadi — 📋 Nusxa uni qoplaydi."""
    if not buttons:
        return None

    rows: list[list[InlineKeyboardButton]] = []

    if isinstance(buttons, list) and buttons and isinstance(buttons[0], dict):
        buttons = [buttons]

    for row in buttons:
        kb_row: list[InlineKeyboardButton] = []
        for btn in row:
            label = btn.get("label", "")
            callback = btn.get("callback", "")
            if not label or not callback:
                continue
            # Polish ulashish: 'share' → 📋 Nusxa + 📤 Ulashish (inline).
            if callback == "share" and share_text:
                kb_row.append(InlineKeyboardButton(text="📋 Nusxa", callback_data="share"))
                token = _cache_share_text(_strip_polish_wrapper(share_text))
                kb_row.append(InlineKeyboardButton(text="📤 Ulashish",
                                                   switch_inline_query=f"txt:{token}"))
                continue
            if callback == "copy" and share_text:
                continue  # no-op 'copy' tugmasi — 📋 Nusxa o'rnini bosadi
            resolved = _resolve_callback(callback, ids_by_type)
            if resolved is None:
                continue
            if len(resolved) > 64:
                resolved = resolved[:64]
            kb_row.append(InlineKeyboardButton(text=label, callback_data=resolved))
        if kb_row:
            rows.append(kb_row)

    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None


def _append_back_row(kb: InlineKeyboardMarkup | None,
                     callback_data: str = "nav_cockpit") -> InlineKeyboardMarkup:
    rows = [list(row) for row in (kb.inline_keyboard if kb else [])]
    if not rows or all(btn.callback_data != callback_data for row in rows for btn in row):
        rows.append([back_button(callback_data)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ─────────────────────── CORE PROCESSING ───────────────────────


_STREAM_EDIT_MIN_INTERVAL_SEC = 1.2  # Telegram message-edit rate limit headroom
_STREAM_EDIT_MIN_DELTA_CHARS = 24    # don't spam edits for tiny additions


_DESTRUCTIVE_ACTION_TYPES = {"create_task", "schedule_meeting"}

# Field edits overwrite an existing task's value with NO undo, so — like deletes —
# they ALWAYS confirm, independent of the confirm_create_actions setting.
_UPDATE_ACTION_TYPES = {"update_task"}

# Most tasks/meetings creatable from a single CHAT message. A pasted list of ~14 is
# realistic; beyond ~20 the JSON response risks truncation and the confirm card
# grows unwieldy. We keep the first N and tell the principal to resend the rest.
_MAX_CREATE_ACTIONS_PER_MSG = 20

# FILE import (Excel/doc) is parsed deterministically (no LLM → no JSON truncation),
# so it is NOT held to the chat cap. This is a high safety backstop only — it stops a
# pathological/huge file from flooding the DB or hanging the turn, not real use.
_MAX_IMPORT_TASKS = 1000

# Mass-delete actions ("barchasini o'chir"). These ALWAYS require confirmation —
# independent of the confirm_create_actions setting — because they're
# irreversible and a mis-heard voice command could wipe everything.
_BULK_DELETE_ACTION_TYPES = {
    "delete_all_tasks", "delete_all_meetings", "delete_all_notes",
    "delete_all_reminders", "delete_all_contacts",
}
_BULK_DELETE_LABEL = {
    "delete_all_tasks": ("vazifa", "tasks"),
    "delete_all_meetings": ("uchrashuv", "meetings"),
    "delete_all_notes": ("note", "notes"),
    "delete_all_reminders": ("eslatma", "reminders"),
    "delete_all_contacts": ("kontakt", "contacts"),
}

# Single delete/cancel via voice/text. ALWAYS confirm (a mis-heard "X'ni o'chir"
# shouldn't silently delete) — independent of the confirm_create_actions setting.
_SINGLE_DELETE_ACTION_TYPES = {"delete_task", "cancel_meeting", "delete_reminder", "delete_note"}

# Category-scoped destructive actions. ALWAYS confirm — delete_category clears a
# label off many tasks; delete_tasks_by_category wipes them. Previewed with the
# affected count (see _format_create_preview).
_CATEGORY_DELETE_ACTION_TYPES = {"delete_category", "delete_tasks_by_category"}


async def _maybe_refresh_section(
    message: Message,
    state: "FSMContext | None",
    ids_by_type: dict[str, list[str]],
) -> None:
    """When the user is currently inside a section view (`/tasks`, `/meetings`,
    `/reminders`, `/today`) and an item of the matching type was just created,
    send a fresh section render so the new item is visible immediately.

    Without this, the previously-rendered section list (an old Telegram
    message) shows stale data and the user thinks the create silently
    failed — a real reported bug."""
    if state is None or not ids_by_type:
        return
    try:
        current = await state.get_state()
    except Exception:
        return
    if not current:
        return

    # `_refresh` lets non-create actions (e.g. bulk deletes) ask for a section
    # re-render even though they produce no new item IDs.
    refresh = set(ids_by_type.get("_refresh") or [])
    created_tasks = bool(ids_by_type.get("task")) or "task" in refresh
    created_meetings = bool(ids_by_type.get("meeting")) or "meeting" in refresh
    created_reminders = bool(ids_by_type.get("reminder")) or "reminder" in refresh
    created_notes = bool(ids_by_type.get("note")) or "note" in refresh

    # Map current section → render call. Only fire when a matching item was
    # actually created, otherwise we'd spam the user with a redundant list.
    try:
        if current == SectionFSM.in_tasks.state and created_tasks:
            await _render_tasks_for_filter(message, "active")
        elif current == SectionFSM.in_today.state and (created_tasks or created_meetings):
            # /today shows both — re-render via cmd_today.
            await cmd_today(message)
        elif current == SectionFSM.in_meetings.state and created_meetings:
            await _render_meetings_for_filter(message, "week")
        elif current == SectionFSM.in_reminders.state and created_reminders:
            await _render_reminders_for_filter(message, "active")
        elif current == SectionFSM.in_notes.state and created_notes:
            await _render_notes_for_filter(message, "inbox")
    except Exception:
        logger.exception("Section auto-refresh failed (non-fatal)")


# ─────────────────── BO'SH SLOT (kalendar free-time) ───────────────────
# "Seshanba bo'sh slotlarimni ko'rsat" / "bu hafta bo'sh vaqtim" — ish vaqtidan
# (band = uchrashuvlar) bo'sh oraliqlarni hisoblab beradi.
_FREE_WORK_START_H = 9     # ish kuni boshlanishi (09:00)
_FREE_WORK_END_H = 18      # ish kuni tugashi (18:00)
_MIN_FREE_SLOT_MIN = 30    # shundan kichik bo'shliqni ko'rsatmaymiz
_DEFAULT_MEETING_MIN = 60  # datetime_end yo'q bo'lsa — taxminiy davomiylik

_WEEKDAY_NAME_TO_IDX = {
    "dushanba": 0, "seshanba": 1, "chorshanba": 2, "payshanba": 3,
    "juma": 4, "shanba": 5, "yakshanba": 6,
}


def _resolve_target_date(s, now):
    """ISO sana YOKI o'zbekcha hafta-kun / nisbiy so'zni date'ga aylantiradi.
    Parse bo'lmasa None. (LLM sanani noto'g'ri yechsa ham — server fallback.)"""
    if not s:
        return None
    s = str(s).strip().lower()
    try:
        return datetime.fromisoformat(s).date()
    except (ValueError, TypeError):
        pass
    if s in ("bugun", "today"):
        return now.date()
    if s in ("ertaga", "tomorrow"):
        return (now + timedelta(days=1)).date()
    if s == "indin":
        return (now + timedelta(days=2)).date()
    if s in _WEEKDAY_NAME_TO_IDX:
        # bugun shu kun bo'lsa — bugunni nazarda tutadi (kelasi hafta emas)
        delta = (_WEEKDAY_NAME_TO_IDX[s] - now.weekday()) % 7
        return (now + timedelta(days=delta)).date()
    return None


def _compute_free_slots(busy, window_start, window_end, min_minutes=_MIN_FREE_SLOT_MIN):
    """SOF funksiya: band (start, end) oraliqlari va [window_start, window_end] ish
    oynasidan — min_minutes dan katta bo'sh oraliqlarni (start, end) qaytaradi."""
    clipped = []
    for s, e in busy:
        if e <= window_start or s >= window_end:
            continue  # oyna tashqarisidagi uchrashuv — e'tiborsiz
        clipped.append((max(s, window_start), min(e, window_end)))
    clipped.sort(key=lambda x: x[0])
    merged = []  # ustma-ust tushgan band oraliqlarni birlashtirish
    for s, e in clipped:
        if merged and s <= merged[-1][1]:
            if e > merged[-1][1]:
                merged[-1][1] = e
        else:
            merged.append([s, e])
    free = []
    cursor = window_start
    gap = timedelta(minutes=min_minutes)
    for s, e in merged:
        if s - cursor >= gap:
            free.append((cursor, s))
        if e > cursor:
            cursor = e
    if window_end - cursor >= gap:
        free.append((cursor, window_end))
    return free


def _fmt_dur(minutes):
    """45 → '45 daq', 90 → '1s 30daq', 120 → '2 soat'."""
    h, m = divmod(int(minutes), 60)
    if h and m:
        return f"{h}s {m}daq"
    if h:
        return f"{h} soat"
    return f"{m} daq"


async def _free_slots_for_day(target_date):
    """Berilgan sana uchun (free_slots, busy_meetings) — ish vaqti ichida.
    Bugun bo'lsa — o'tib ketgan vaqtni hisobga olmaydi (hozirdan boshlaydi)."""
    # pytz: datetime(..., tzinfo=TZ) LMT (+04:37) beradi — TZ.localize() ishlatamiz
    day_start = database.TZ.localize(datetime.combine(target_date, datetime.min.time()))
    day_end = day_start + timedelta(days=1)
    meetings = await database.list_meetings_in_window(
        day_start.isoformat(), day_end.isoformat())
    window_start = day_start.replace(hour=_FREE_WORK_START_H)
    window_end = day_start.replace(hour=_FREE_WORK_END_H)
    # Bugungi so'rovda — allaqachon o'tgan slotlarni ko'rsatmaymiz
    now = datetime.now(database.TZ)
    if target_date == now.date():
        live = now.replace(second=0, microsecond=0)
        live += timedelta(minutes=(5 - live.minute % 5) % 5)  # keyingi 5 daqiqaga yaxlitlash
        if live > window_start:
            window_start = min(live, window_end)
    busy, busy_meetings = [], []
    for m in meetings:
        try:
            s = datetime.fromisoformat(m["datetime_start"])
        except (ValueError, TypeError, KeyError):
            continue
        end_raw = m.get("datetime_end")
        try:
            e = datetime.fromisoformat(end_raw) if end_raw else s + timedelta(minutes=_DEFAULT_MEETING_MIN)
        except (ValueError, TypeError):
            e = s + timedelta(minutes=_DEFAULT_MEETING_MIN)
        if e <= s:
            e = s + timedelta(minutes=_DEFAULT_MEETING_MIN)
        busy.append((s, e))
        # Ish oynasi bilan kesishadigan uchrashuvlarnigina "Band" ro'yxatida ko'rsatamiz
        if e > window_start and s < window_end:
            busy_meetings.append((s, e, m.get("title") or "Uchrashuv"))
    free = _compute_free_slots(busy, window_start, window_end)
    busy_meetings.sort(key=lambda x: x[0])
    return free, busy_meetings


def _format_free_day(target_date, free, busy_meetings):
    """Bitta kunlik to'liq bo'sh-slot kartasi."""
    weekday = UZ_WEEKDAYS_FULL[target_date.weekday()]
    month = UZ_MONTHS_FULL[target_date.month - 1]
    lines = [f"📅 **{weekday}, {target_date.day}-{month}** — bo'sh vaqt (09:00–18:00):", ""]
    if not free:
        lines.append("🔴 Ish vaqti to'liq band — bo'sh slot yo'q.")
    else:
        total = 0
        for s, e in free:
            mins = int((e - s).total_seconds() // 60)
            total += mins
            lines.append(f"🟢 {s.strftime('%H:%M')}–{e.strftime('%H:%M')}  ({_fmt_dur(mins)})")
        lines.append("")
        lines.append(f"_Jami bo'sh: {_fmt_dur(total)}_")
    if busy_meetings:
        lines.append("")
        lines.append("⚪️ **Band:**")
        for s, e, title in busy_meetings:
            lines.append(f"   {s.strftime('%H:%M')}–{e.strftime('%H:%M')} — {title}")
    return "\n".join(lines)


def _format_free_week_line(target_date, free):
    """Hafta ko'rinishidagi bitta kun — bir qatorli xulosa."""
    weekday = UZ_WEEKDAYS_FULL[target_date.weekday()]
    month = UZ_MONTHS_FULL[target_date.month - 1]
    label = f"**{weekday}, {target_date.day}-{month}**"
    if not free:
        return f"🔴 {label}: to'la band"
    total = sum(int((e - s).total_seconds() // 60) for s, e in free)
    slots = ", ".join(f"{s.strftime('%H:%M')}–{e.strftime('%H:%M')}" for s, e in free)
    return f"🟢 {label}: {_fmt_dur(total)} bo'sh — {slots}"


async def _render_free_slots(message: Message, action: dict) -> None:
    """show_free_slots action — kun yoki hafta bo'yicha bo'sh slotlarni chizadi."""
    data = action.get("data") or {}
    range_kind = (data.get("range") or "day").strip().lower()
    now = datetime.now(database.TZ)
    base = _resolve_target_date(data.get("date"), now)
    if range_kind in ("week", "hafta"):
        start = base or now.date()
        monday = start - timedelta(days=start.weekday())  # shu hafta dushanbasi
        lines = ["📅 **BU HAFTA — bo'sh vaqt** (09:00–18:00, ish kunlari)", ""]
        for i in range(5):  # Dushanba–Juma
            d = monday + timedelta(days=i)
            free, _ = await _free_slots_for_day(d)
            lines.append(_format_free_week_line(d, free))
        await message.answer("\n".join(lines), parse_mode="Markdown")
        return
    target = base or now.date()
    free, busy = await _free_slots_for_day(target)
    await message.answer(_format_free_day(target, free, busy), parse_mode="Markdown")


_SHOW_ACTION_TYPES = {
    "show_tasks", "show_meetings", "show_notes", "show_reminders", "show_contacts",
    "show_free_slots", "show_stats", "run_plan",
}


async def _render_show_action(message: Message, state: "FSMContext | None", action: dict) -> None:
    """Render a full, DB-backed section list for a "ko'rsat/ro'yxat" request.

    Claude only sees today+overdue tasks in its state block, so if it enumerates
    "all tasks" itself the list is INCOMPLETE (reported bug). We instead render
    the real section straight from the DB — every item, with filters/pagination.
    """
    atype = action.get("type")
    filt = (action.get("data") or {}).get("filter") or ""
    if atype == "show_tasks":
        if state is not None:
            await state.set_state(SectionFSM.in_tasks)
        await message.answer("📋 **VAZIFALAR**", parse_mode="Markdown",
                             reply_markup=tasks_section_reply_keyboard())
        await _render_tasks_for_filter(message, filt or "active")
    elif atype == "show_meetings":
        if state is not None:
            await state.set_state(SectionFSM.in_meetings)
        await message.answer("🤝 **UCHRASHUVLAR**", parse_mode="Markdown",
                             reply_markup=meetings_section_reply_keyboard())
        await _render_meetings_for_filter(message, filt or "week")
    elif atype == "show_free_slots":
        await _render_free_slots(message, action)
    elif atype == "show_notes":
        await cmd_notes(message, state)
    elif atype == "show_reminders":
        await cmd_reminders(message, state)
    elif atype == "show_contacts":
        await cmd_team(message, state)
    elif atype == "show_stats":
        d = action.get("data") or {}
        try:
            days = int(d.get("days", 7))
        except (TypeError, ValueError):
            days = 7
        is_report = bool(d.get("report"))
        if days not in ((7, 30) if is_report else (1, 7, 30)):
            days = 7
        label = {1: "Bugun", 7: "Oxirgi 7 kun", 30: "Oxirgi 30 kun"}.get(days, "Oxirgi 7 kun")
        stats = await database.executive_stats(days=days)
        text = _format_executive_report(stats, label) if is_report else _format_stats_dashboard(stats, label)
        await _safe_answer(message, text, parse_mode="Markdown")
    elif atype == "run_plan":
        await _run_planning_session(message, (action.get("data") or {}).get("situation") or "")


# Editable fields shown (Uzbek) in the update_task old→new confirm diff.
_EDIT_FIELD_UZ = {
    "title": "Sarlavha", "description": "Tavsif", "deadline": "Muddat",
    "priority": "Ustuvorlik", "status": "Status", "assignee": "Ijrochi",
    "tags": "Teglar", "category": "Kategoriya",
}


async def _format_create_preview(actions: list[dict], original_input: str | None = None) -> str:
    """Render a confirm-prompt preview for create / edit / delete actions, shown
    before execution. Bulk deletes show the LIVE row count so the user sees
    exactly how much would be wiped.

    When `original_input` is given (the user's typed text or voice transcript), it
    is shown ABOVE the prompt so the principal can verify WHAT WAS UNDERSTOOD
    before confirming — critical for voice, where a mis-transcription would
    otherwise act silently."""
    lines: list[str] = []
    if original_input:
        src = original_input.strip()
        if len(src) > 600:
            src = src[:600] + "…"
        lines.append(f"🎙 **Eshitildi:** {_escape_markdown(src)}")
        lines.append("")
    lines += ["⚠️ **TASDIQLAYSIZMI?**", ""]
    # When many tasks/meetings are created at once (e.g. a pasted list of 14),
    # the full 6-line card per item overflows Telegram's 4096-char message limit
    # and is hard to scan. Switch to one compact line per item past this count.
    _create_types = {"create_task", "schedule_meeting"}
    n_creates = sum(1 for a in actions if a.get("type") in _create_types)
    compact = n_creates > 4
    if compact:
        lines.append(f"📋 **{n_creates} ta yangi yozuv qo'shiladi:**")
        lines.append("")
    create_idx = 0
    for a in actions:
        t = a.get("type")
        d = a.get("data", {}) or {}
        if compact and t == "create_task":
            create_idx += 1
            title = (d.get("title") or "—").strip()
            assignee = (d.get("assignee") or "—").strip()
            dl_label, _ = _format_deadline_short(d.get("deadline"))
            badge = _PRIORITY_BADGE.get(d.get("priority", "P2"), "🔵")
            lines.append(f"{badge} {create_idx}. {title} — 👤 {assignee} · ⏳ {dl_label}")
            continue
        if compact and t == "schedule_meeting":
            create_idx += 1
            title = (d.get("title") or "—").strip()
            start = d.get("datetime_start", "—")
            try:
                dt = datetime.fromisoformat(start).astimezone(database.TZ)
                start_label = dt.strftime("%d-%m %H:%M")
            except (ValueError, TypeError):
                start_label = start
            lines.append(f"🤝 {create_idx}. {title} — 🕐 {start_label}")
            continue
        if t == "update_task":  # an EDIT — show changed fields old→new (no undo → always confirm)
            try:
                task = await database.get_task(a.get("id"))
            except Exception:
                task = None
            cur = task or {}
            title = cur.get("title") or d.get("title") or a.get("id", "—")
            lines.append(f"✏️ **Tahrirlanadi:** {title}")
            for k, v in d.items():
                if k not in _EDIT_FIELD_UZ:
                    continue
                old = cur.get(k)
                if k == "priority":
                    old, v = _PRIORITY_LABEL_UZ.get(old, old), _PRIORITY_LABEL_UZ.get(v, v)
                elif k == "status":
                    old, v = _STATUS_LABEL_UZ.get(old, old), _STATUS_LABEL_UZ.get(v, v)
                elif k == "deadline":
                    old = _format_deadline_short(old)[0] if old else "—"
                    v = _format_deadline_short(v)[0] if v else "—"
                elif k == "tags":
                    old = ", ".join(old) if isinstance(old, list) else old
                    v = ", ".join(v) if isinstance(v, list) else v
                old = "—" if old in (None, "", []) else old
                lines.append(f"   {_EDIT_FIELD_UZ[k]}: {old} → {v}")
            lines.append("   _Eski qiymat saqlanmaydi._")
            lines.append("")
            continue
        if t == "delete_category":
            cat = (d.get("category") or "—").strip()
            try:
                n = await database.count_tasks_in_category(cat)
            except Exception:
                n = 0
            lines.append(f"🗑 **«{cat}» kategoriyasi o'chiriladi**")
            lines.append(f"   _{n} ta vazifaning yorlig'i olinadi — vazifalar saqlanadi ((boshqa)ga o'tadi)._")
            lines.append("")
            continue
        if t == "delete_tasks_by_category":
            cat = (d.get("category") or "—").strip()
            try:
                n = await database.count_tasks_in_category(cat)
            except Exception:
                n = 0
            lines.append(f"🧹 **«{cat}» kategoriyasidagi {n} ta vazifa o'chiriladi**")
            lines.append("   _Qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t == "delete_task":
            try:
                task = await database.get_task(a.get("id"))
            except Exception:
                task = None
            title = (task or {}).get("title", a.get("id", "—"))
            lines.append(f"🗑 **Vazifa o'chiriladi:** {title}")
            lines.append("   _Qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t == "cancel_meeting":
            try:
                m = await database.get_meeting(a.get("id"))
            except Exception:
                m = None
            title = (m or {}).get("title", a.get("id", "—"))
            lines.append(f"🗑 **Uchrashuv bekor qilinadi:** {title}")
            lines.append("   _Qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t == "delete_reminder":
            try:
                r = await database.get_reminder(a.get("id"))
            except Exception:
                r = None
            title = (r or {}).get("title", a.get("id", "—"))
            lines.append(f"🗑 **Eslatma o'chiriladi:** {title}")
            lines.append("   _Qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t == "delete_note":
            try:
                nt = await database.get_note(a.get("id"))
            except Exception:
                nt = None
            title = (nt or {}).get("title") or ((nt or {}).get("content") or a.get("id", "—"))[:50]
            lines.append(f"🗑 **Qayd o'chiriladi:** {title}")
            lines.append("   _Qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t in _BULK_DELETE_ACTION_TYPES:
            noun, table = _BULK_DELETE_LABEL[t]
            scope = ""
            try:
                if t == "delete_all_tasks" and d.get("status_in"):
                    # Count ONLY the filtered rows — must match what the deletion
                    # removes. Was misleading: showed the TOTAL while deleting a subset.
                    n = len(await database.list_tasks(status_in=d["status_in"], limit=100000))
                    scope = f" ({', '.join(d['status_in'])})"
                else:
                    n = await database.count_table(table)
            except Exception:
                n = 0
            lines.append(f"🗑 **Barcha {noun}lar o'chiriladi{scope}** — {n} ta")
            lines.append("   _Bu amalni qaytarib bo'lmaydi._")
            lines.append("")
            continue
        if t == "create_task":
            title = (d.get("title") or "—").strip()
            assignee = (d.get("assignee") or "belgilanmagan").strip()
            deadline_label, _ = _format_deadline_short(d.get("deadline"))
            priority = d.get("priority", "P2")
            pri_label = _PRIORITY_LABEL_UZ.get(priority, priority)
            lines.append("📝 **Yangi vazifa**")
            lines.append(f"   {title}")
            lines.append(f"   👤 Ijrochi: {assignee}")
            lines.append(f"   ⏳ Muddat: {deadline_label}")
            lines.append(f"   🔺 Ustuvorlik: {pri_label}")
            lines.append("")
        elif t == "schedule_meeting":
            title = (d.get("title") or "—").strip()
            participants = ", ".join(d.get("participants", []) or []) or "belgilanmagan"
            location = (d.get("location_or_link") or "belgilanmagan").strip()
            start = d.get("datetime_start", "—")
            try:
                dt = datetime.fromisoformat(start).astimezone(database.TZ)
                start_label = dt.strftime("%d-%m %H:%M")
            except (ValueError, TypeError):
                start_label = start
            lines.append("🤝 **Yangi uchrashuv**")
            lines.append(f"   {title}")
            lines.append(f"   🕐 Vaqt: {start_label}")
            lines.append(f"   👥 Ishtirokchilar: {participants}")
            lines.append(f"   📍 Joy: {location}")
            lines.append("")
    return "\n".join(lines).rstrip()


async def _process_and_reply(message: Message, user_text: str, state: "FSMContext | None" = None) -> None:
    """Send user_text to Claude (streaming), edit a progress message as the
    reply arrives, then attach action buttons once parsing completes.

    If `state` is supplied AND the user has `confirm_create_actions` enabled
    (default), any create_task / schedule_meeting actions in Claude's response
    are deferred: a preview + Tasdiq/Bekor prompt is shown, and execution
    happens only after the user confirms. This protects against mis-transcribed
    voice messages silently creating wrong items.

    Wrapped in a pending_actions row so:
      - a redelivered Telegram update doesn't double-process (dedup on the stable
        (chat_id, message_id) pair — a slow turn that Telegram redelivers used to
        produce a SECOND confirm card),
      - a crash mid-handler is recoverable / observable on next bot start."""

    if not user_text or not user_text.strip():
        await message.answer("Bo'sh xabar. Iltimos, matn yoki ovoz yuboring.")
        return

    pending_id = await database.enqueue_pending_action(
        update_id=None,  # aiogram 3.x doesn't surface update_id; dedup by chat+message
        chat_id=message.chat.id if message.chat else None,
        message_id=message.message_id,
        user_text=user_text,
    )
    if pending_id is None:
        # Redelivered Telegram update for an already-enqueued message — swallow.
        return

    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    progress_msg: Message | None = None
    last_edit_at = 0.0
    last_edit_text = ""
    loop = asyncio.get_event_loop()
    try:
        await database.mark_pending_in_progress(pending_id)
        final_response: dict | None = None

        async for kind, payload in claude_service.process_message_stream(user_text):
            if kind == "partial":
                text = (payload or "").strip()
                if not text:
                    continue
                # Throttle edits: at least N seconds AND N characters of growth
                # before issuing another edit_text call. Telegram's edit rate
                # limit is loose but spammy edits churn the chat for no UX win.
                now = loop.time()
                if (now - last_edit_at) < _STREAM_EDIT_MIN_INTERVAL_SEC:
                    continue
                if abs(len(text) - len(last_edit_text)) < _STREAM_EDIT_MIN_DELTA_CHARS:
                    continue
                if progress_msg is None:
                    progress_msg = await message.answer(text + " ▌")
                else:
                    try:
                        await progress_msg.edit_text(text + " ▌")
                    except TelegramBadRequest:
                        # "message is not modified" or similar — silently skip;
                        # the next edit will be different.
                        pass
                last_edit_at = now
                last_edit_text = text
            elif kind == "complete":
                final_response = payload  # dict envelope
                break

        if final_response is None:
            final_response = claude_service._FALLBACK_RESPONSE

        if final_response.get("needs_clarification"):
            question = (final_response.get("clarification_question")
                        or final_response.get("user_message")
                        or "Aniqlashtiring.")
            if progress_msg is not None:
                try:
                    await progress_msg.edit_text(question, reply_markup=single_back_keyboard())
                except TelegramBadRequest:
                    await message.answer(question, reply_markup=single_back_keyboard())
            else:
                await message.answer(question, reply_markup=single_back_keyboard())
            await database.complete_pending_action(pending_id)
            return

        # ── Export intent ("vazifalarni eksport qil" / "X vazifalarini eksport qil") ──
        # Send the .xlsx file, same as /export — works from voice and text. An
        # optional data.assignee narrows it to one executor.
        _exp = next((a for a in final_response.get("actions", []) if a.get("type") == "export_tasks"), None)
        if _exp is not None:
            await database.complete_pending_action(pending_id)
            if progress_msg is not None:
                try:
                    await progress_msg.delete()
                except TelegramBadRequest:
                    pass
            _exp_data = _exp.get("data") or {}
            _exp_who = (_exp_data.get("assignee") or "").strip() or None
            _exp_status = _EXPORT_STATUS_WORDS.get((_exp_data.get("status") or "").strip().lower()) or None
            _exp_script = "cyr" if (_exp_data.get("script") or "").strip().lower() in _CYR_TOKENS else "lat"
            await _send_tasks_export(message, assignee=_exp_who, status=_exp_status, script=_exp_script)
            return

        # ── "Ko'rsat/ro'yxat" intent — render the REAL section from the DB ──
        # Claude's state block only holds today+overdue, so letting it enumerate
        # "all tasks" yields an incomplete list. Intercept show_* and render the
        # full section instead (reported: "barcha vazifalarni ko'rsat" → faqat bir qismi).
        show_action = next(
            (a for a in final_response.get("actions", []) if a.get("type") in _SHOW_ACTION_TYPES),
            None,
        )
        if show_action is not None:
            await database.complete_pending_action(pending_id)
            if progress_msg is not None:
                try:
                    await progress_msg.delete()
                except TelegramBadRequest:
                    pass
            await _render_show_action(message, state, show_action)
            return

        # ── Tasdiq qatlami ──
        # Bulk deletes ("barchasini o'chir") ALWAYS confirm — irreversible.
        # create_task / schedule_meeting confirm only if the setting is on
        # (default ON; /settings dan o'chirish mumkin).
        actions = final_response.get("actions", [])
        # Cap mass-create requests so a huge pasted list can't truncate the JSON
        # or overflow the confirm card. Keep the first N create actions in order,
        # drop the rest, and tell the principal to resend them.
        _overflow_note = ""
        _n_creates = sum(1 for a in actions if a.get("type") in _DESTRUCTIVE_ACTION_TYPES)
        if _n_creates > _MAX_CREATE_ACTIONS_PER_MSG:
            dropped = _n_creates - _MAX_CREATE_ACTIONS_PER_MSG
            kept, seen = [], 0
            for a in actions:
                if a.get("type") in _DESTRUCTIVE_ACTION_TYPES:
                    seen += 1
                    if seen > _MAX_CREATE_ACTIONS_PER_MSG:
                        continue
                kept.append(a)
            actions = kept
            final_response["actions"] = actions
            _overflow_note = (
                f"\n\n⚠️ Bir xabarda {_MAX_CREATE_ACTIONS_PER_MSG} tagacha yozuv qo'shiladi. "
                f"Qolgan {dropped} tasini alohida xabarda yuboring."
            )
        try:
            _settings = await database.get_settings()
        except Exception:
            _settings = {}
        bulk_deletes = [a for a in actions if a.get("type") in _BULK_DELETE_ACTION_TYPES]
        single_deletes = [a for a in actions if a.get("type") in _SINGLE_DELETE_ACTION_TYPES]
        cat_deletes = [a for a in actions if a.get("type") in _CATEGORY_DELETE_ACTION_TYPES]
        field_updates = [a for a in actions if a.get("type") in _UPDATE_ACTION_TYPES]
        # Deletes AND field-edits ALWAYS confirm: both are irreversible (an edit
        # overwrites the old value with no undo). Creates confirm only if the
        # confirm_create_actions setting is on.
        to_confirm = list(bulk_deletes) + list(single_deletes) + list(cat_deletes) + list(field_updates)
        if _settings.get("confirm_create_actions", True):
            to_confirm += [a for a in actions if a.get("type") in _DESTRUCTIVE_ACTION_TYPES]
        if state is not None and to_confirm:
            # Show the user the original input (transcript/typed text) above the
            # preview so a mis-heard voice command is caught before it executes.
            preview = await _format_create_preview(to_confirm, original_input=user_text) + _overflow_note
            confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="✅ Tasdiqlayman", callback_data="acts_confirm"),
                InlineKeyboardButton(text="✕ Bekor qilish", callback_data="acts_cancel"),
            ]])
            # Remember the prior section state so cb_actions_confirm can
            # restore it and re-render the section after the deferred execute.
            try:
                prior_state = await state.get_state()
            except Exception:
                prior_state = None
            await state.set_state(CreateActionConfirmFSM.awaiting)
            await state.update_data(
                pending_response=final_response,
                _prior_section=prior_state,
            )
            if progress_msg is not None:
                try:
                    await progress_msg.edit_text(
                        preview, parse_mode="Markdown", reply_markup=confirm_kb
                    )
                except TelegramBadRequest:
                    await _safe_answer(message, preview,
                                        reply_markup=confirm_kb, parse_mode="Markdown")
            else:
                await _safe_answer(message, preview,
                                    reply_markup=confirm_kb, parse_mode="Markdown")
            await database.complete_pending_action(pending_id)
            return

        ids_by_type = await _execute_actions(actions)
        keyboard = _build_keyboard(final_response.get("buttons", []), ids_by_type,
                                   share_text=final_response.get("user_message"))
        if keyboard:
            keyboard = _append_back_row(keyboard)

        # Agar to'qnashuv/yaroqsiz-vaqt sabab hech narsa qo'yilmagan bo'lsa, LLM'ning
        # optimistik "qo'shildi" matni ogohlantirish bilan ZID bo'ladi — shu holatda
        # asos matnni tashlab, faqat aniq ogohlantirishni ko'rsatamiz.
        _base = (final_response.get("user_message") or "").strip() or "✅"
        text = "" if (ids_by_type.get("_conflict") or ids_by_type.get("_badtime")) else _base
        text += _conflict_note(ids_by_type)
        text += _badtime_note(ids_by_type)
        text += _failed_actions_note(ids_by_type)
        text += _overflow_note
        text = text.strip() or _base
        if progress_msg is not None:
            # Finalize the same message we've been editing — single chat bubble.
            try:
                await progress_msg.edit_text(text, parse_mode="Markdown", reply_markup=keyboard)
            except TelegramBadRequest:
                # Editing failed (deleted? markdown parse error?) — send fresh.
                await _safe_answer(message, text, reply_markup=keyboard, parse_mode="Markdown")
        else:
            await _safe_answer(message, text, reply_markup=keyboard, parse_mode="Markdown")
        await database.complete_pending_action(pending_id)
        # If the user is currently inside a section view (/tasks, /meetings,
        # /reminders, /today), re-render it with the new item visible.
        await _maybe_refresh_section(message, state, ids_by_type)
    except Exception as e:
        logger.exception("_process_and_reply failed for pending=%s", pending_id)
        await database.fail_pending_action(pending_id, f"{type(e).__name__}: {e}")
        # Tell the user instead of failing silently. Handled here (no re-raise);
        # the global error handler in bot.py is the fallback for everything else.
        try:
            await message.answer(_humanize_error(e))
        except Exception:
            logger.debug("Could not send error notice in _process_and_reply")
    finally:
        typing_task.cancel()


# ─────────────────────── COMMAND HANDLERS ───────────────────────


def _time_greeting() -> str:
    """Time-of-day greeting in the principal's timezone (Asia/Tashkent).
    Evening uses 'Xayrli kech'; late night falls back to neutral 'Salom'
    ('Xayrli tun' reads as a goodbye, not a greeting)."""
    h = datetime.now(database.TZ).hour
    if 5 <= h < 11:
        return "Xayrli tong"
    if 11 <= h < 17:
        return "Xayrli kun"
    if 17 <= h < 22:
        return "Xayrli kech"
    return "Salom"


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext) -> None:
    # /start is the universal "I'm confused, reset" command — clear any FSM
    # state so a user stuck mid-flow can recover without finding a Back button.
    # It also serves as onboarding: a concise tour of what the bot does + examples.
    await state.clear()
    name = (message.from_user.full_name if message.from_user else "").strip() or "Maqsud Rustamov"
    await message.answer(
        f"{_time_greeting()}, {name}! 👋\n\n"
        "🤝 Yordamchi Pro — shaxsiy AI yordamchingiz.\n\n"
        "Vazifa · eslatma · uchrashuv · qayd — yozing yoki ayting,\n"
        "qolganini o'zim qilaman.\n\n"
        "/help · /cockpit",
        parse_mode=None,
        reply_markup=main_reply_keyboard(),
    )


@router.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext) -> None:
    """Universal escape hatch — clear any FSM state and confirm to the user.
    Matches industry-standard /cancel convention."""
    current = await state.get_state()
    await state.clear()
    if current:
        await message.answer(
            "✕ Bekor qilindi. Asosiy menyu.",
            reply_markup=main_reply_keyboard(),
        )
    else:
        await message.answer(
            "Hech qaysi amal aktiv emas. Asosiy menyu.",
            reply_markup=main_reply_keyboard(),
        )


@router.message(Command("new"))
async def cmd_new_command(message: Message, state: FSMContext) -> None:
    """/new slash-command entry point. Delegates to the existing cmd_new()
    submenu so behavior matches both /new typed and the ➕ Yangi button."""
    await cmd_new(message, state)


@router.message(Command("backup"))
async def cmd_backup(message: Message) -> None:
    """Create an on-demand SQLite backup using the .backup API (consistent
    snapshot — works even while the bot is writing). Saves to data/backups/
    with a timestamped name. Sends a status reply with the file size and
    integrity-check result.

    For automated daily backups see DEPLOY.md cron section."""
    import os
    import sqlite3
    from datetime import datetime as _dt

    timestamp = _dt.now(database.TZ).strftime("%Y%m%d-%H%M%S")
    backup_dir = Path(config.DATABASE_PATH).parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"yordamchi-manual-{timestamp}.db"

    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        # sqlite3 .backup runs in a thread; use to_thread to keep the loop free.
        def _do_backup() -> tuple[int, str]:
            with sqlite3.connect(config.DATABASE_PATH) as src, sqlite3.connect(str(backup_path)) as dst:
                src.backup(dst)
            size = os.path.getsize(backup_path)
            # Integrity check on the backup file (not source)
            with sqlite3.connect(str(backup_path)) as check:
                cur = check.execute("PRAGMA integrity_check")
                result = cur.fetchone()[0]
            return size, result

        size, integrity = await asyncio.to_thread(_do_backup)
    except Exception as e:
        logger.exception("Backup failed")
        await message.answer(f"❌ Backup xato: {type(e).__name__}: {e}")
        return
    finally:
        typing_task.cancel()
    _rotate_backups(backup_dir)   # cap local snapshots so the disk can't fill

    kb_size = size / 1024
    ok_mark = "✅" if integrity == "ok" else f"⚠️ {integrity}"
    await _safe_answer(
        message,
        f"💾 **Backup yaratildi**\n\n"
        f"📁 `{backup_path.name}`\n"
        f"📦 Hajmi: {kb_size:,.0f} KB\n"
        f"🔍 Integrity: {ok_mark}\n\n"
        f"_To'liq yo'l: `{backup_path}`_\n"
        f"_Avtomatik kunlik backup GCS'ga ham olinadi (DEPLOY.md)._",
        parse_mode="Markdown",
    )


async def _render_stale_delegations(message: Message) -> None:
    """Ijrochilar panelining sub-ko'rinishi (eski "Delegatsiya trekeri" shu yerga
    ko'chirildi): boshqalarga berilgan aktiv vazifalarni eng uzoq kutilgani
    bo'yicha ko'rsatadi — qotib qolgan topshiriqlarni yuzaga chiqaradi."""
    import aiosqlite
    async with aiosqlite.connect(config.DATABASE_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            """SELECT *, julianday('now') - julianday(created_at) AS age_days
               FROM tasks
               WHERE status IN ('todo', 'in_progress')
                 AND assignee IS NOT NULL
                 AND LOWER(TRIM(assignee)) NOT IN (
                     '', 'men', 'siz', 'belgilanmagan', '—',
                     'oʻzim', 'o''zim', 'o''z', 'ozim'
                 )
               ORDER BY age_days DESC
               LIMIT 20"""
        )
        rows = [dict(r) for r in await cur.fetchall()]

    if not rows:
        # team reply-klaviaturani saqlaymiz (reply_markup bermaymiz)
        await _safe_answer(
            message,
            "⏳  **KUTILAYOTGAN TOPSHIRIQLAR**\n\n_Boshqalarga berilgan aktiv vazifa yo'q._",
            parse_mode="Markdown",
        )
        return

    # Same card concept as the Tasks section (badge + title, blank, aligned
    # labeled lines, dividers). Badge reflects how long it's been pending — a
    # stuck delegation is the thing this view exists to surface.
    DIVIDER = "━" * 20
    shown = rows[:12]
    lines = ["⏳  **KUTILAYOTGAN TOPSHIRIQLAR**", "",
             f"_Boshqalarga berilgan · eng uzoq kutilgani birinchi · {len(rows)} ta_",
             "", DIVIDER, ""]
    for i, t in enumerate(shown, 1):
        age = int(t.get("age_days") or 0)
        badge = "🔴" if age >= 7 else "🟠" if age >= 3 else "🟡" if age >= 1 else "⚪"
        assignee = ((t.get("assignee") or "—").strip() or "—")
        assignee = assignee[0].upper() + assignee[1:]
        deadline = _task_deadline_chip(t)
        age_label = "bugun berilgan" if age == 0 else f"{age} kun oldin"
        lines.append(f"{i}.  {badge}  {(t.get('title') or '—').strip()}")
        lines.append("")
        lines.append(f"      👤  Ijrochi:     {assignee}")
        lines.append(f"      ⏳  Muddat:      {deadline}")
        lines.append(f"      ⏱  Kutilmoqda:  {age_label}")
        if i < len(shown):
            lines.append("")
    if len(rows) > 12:
        lines.extend(["", f"_+{len(rows) - 12} ta yana_"])
    lines.extend(["", DIVIDER, "", "_Raqamni bosing — vazifa kartasi._"])

    # Drill-down: tap a number → open that task's card (taskopen:{id}), + back.
    nums = [InlineKeyboardButton(text=str(i), callback_data=f"taskopen:{t['id']}")
            for i, t in enumerate(shown, 1)]
    kb_rows = [nums[j:j + 5] for j in range(0, len(nums), 5)]
    kb_rows.append([back_button()])
    await _safe_answer(
        message,
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows),
    )


# ─────────────────────── EXPORT / IMPORT (Excel) ───────────────────────

# Shared column schema for export & import. "№" is a row-number column (ignored
# on import — matched by name). Headers are lowercased when read back.
_EXPORT_HEADERS = ["№", "Vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Takroriylik", "Izoh", "Kategoriya"]

# rule code → round-trip-safe Uzbek label (normalize_recurrence_rule accepts these back).
_RECUR_LABEL = {"daily": "har kuni", "weekdays": "ish kunlari", "weekly": "har hafta",
                "monthly": "har oy", "quarterly": "har chorak", "yearly": "har yil"}

# Tokens that request a Cyrillic export (voice/command), e.g. "krillcha yubor".
_CYR_TOKENS = {"krill", "krillcha", "kiril", "kirill", "kirilcha", "cyrillic", "cyr", "кирилл", "кириллча"}


def _export_date(iso):
    """ISO deadline → a naive local datetime for the Excel cell, or None. The cell
    is shown date-only (DD-MM-YYYY number format — the principal asked to drop the
    clock visually), BUT the stored value keeps the time so an UNTOUCHED cell
    round-trips the exact instant back on import (no silent 17:00→00:00 drift).
    tz is dropped because Excel cells are naive; _import_deadline re-localizes."""
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso).astimezone(database.TZ).replace(tzinfo=None)
    except (ValueError, TypeError):
        return None


def _import_deadline(val):
    """Excel/CSV cell → ISO deadline or None. Accepts datetime cells, ISO strings,
    and DD-MM-YYYY[ HH:MM] / DD.MM.YYYY forms; unparseable → None (no deadline).
    Naive datetimes are localized via TZ.localize (pytz) — NOT .replace(tzinfo=),
    which would yield the +04:37 LMT offset instead of +05:00."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return (val if val.tzinfo else database.TZ.localize(val)).isoformat()
    s = str(val).strip()
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s)
        return (dt if dt.tzinfo else database.TZ.localize(dt)).isoformat()
    except ValueError:
        pass
    # Full-date forms (dash / dot / slash, with or without a clock, incl. seconds).
    for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
                "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return database.TZ.localize(datetime.strptime(s, fmt)).isoformat()
        except ValueError:
            continue
    # Year-less forms a user naturally types ("12-10", "12-10 14:30"). A day-month
    # almost always means the NEXT occurrence — if the current year puts it more
    # than a day in the past ("12-01" typed in July), roll to next year instead of
    # silently storing a long-past deadline.
    _now = datetime.now(database.TZ)
    for fmt in ("%d-%m %H:%M", "%d-%m", "%d.%m %H:%M", "%d.%m", "%d/%m %H:%M", "%d/%m"):
        try:
            naive = datetime.strptime(s, fmt).replace(year=_now.year)
        except ValueError:
            continue
        if database.TZ.localize(naive) < _now - timedelta(days=1):
            try:
                naive = naive.replace(year=_now.year + 1)
            except ValueError:
                pass  # Feb-29 → keep the current-year date
        return database.TZ.localize(naive).isoformat()
    return None


def _norm_label(s) -> str:
    """Lowercase + drop apostrophe variants — so a status/priority cell matches
    regardless of script or apostrophe form (Toʻsilgan / To'silgan / Тўсилган)."""
    s = str(s or "").strip().lower()
    for ch in "'ʼ’`ʻ‘":
        s = s.replace(ch, "")
    return s


# Reverse maps built LAZILY from the Uzbek labels in BOTH Latin and Cyrillic, so a
# Cyrillic-exported file ("Бажарилди"/"Шошилинч") round-trips back — otherwise every
# status/priority silently reset to the default on re-import (data corruption).
_PRIO_REVERSE: dict = {}
_STATUS_REVERSE: dict = {}


def _label_reverse(label_map: dict) -> dict:
    import translit
    rev: dict = {}
    for code, label in label_map.items():
        for form in (label, translit.to_cyrillic_pro(label)):
            rev[_norm_label(form)] = code
    return rev


def _import_priority(val) -> str:
    """Cell → P0-P3. Accepts codes (P0..P3) and Uzbek labels in Latin OR Cyrillic."""
    s = _norm_label(val)
    if s in ("p0", "p1", "p2", "p3"):
        return s.upper()
    global _PRIO_REVERSE
    if not _PRIO_REVERSE:
        _PRIO_REVERSE = _label_reverse(_PRIORITY_LABEL_UZ)
    return _PRIO_REVERSE.get(s, "P2")


def _import_status(val) -> str:
    """Cell → status code (Latin OR Cyrillic label); defaults to 'todo'."""
    s = _norm_label(val)
    if s in ("todo", "in_progress", "blocked", "done", "cancelled"):
        return s
    global _STATUS_REVERSE
    if not _STATUS_REVERSE:
        _STATUS_REVERSE = _label_reverse(_STATUS_LABEL_UZ)
    return _STATUS_REVERSE.get(s, "todo")


@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    """Export tasks to an .xlsx. Argument: a STATUS word ('/export bajarilgan',
    'aktiv', "o'tgan", 'muhim', 'bugun'…) → that status; otherwise an executor name
    → that assignee. Voice/text ham ('bajarilgan vazifalarni eksport qil')."""
    parts = (message.text or "").split(maxsplit=1)
    arg = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
    # Strip a 'krillcha'/'kirill' token first → Cyrillic export (else it'd be mis-read
    # as an assignee name and return an empty file).
    script = "lat"
    if arg:
        _kept = [t for t in arg.split() if t.lower() not in _CYR_TOKENS]
        if len(_kept) != len(arg.split()):
            script = "cyr"
        arg = " ".join(_kept).strip() or None
    status = _EXPORT_STATUS_WORDS.get(arg.lower()) if arg else None
    assignee = None if status else arg
    await _send_tasks_export(message, assignee=assignee, status=status, script=script)


@router.callback_query(F.data.startswith("exportby:"))
async def cb_export_by(query: CallbackQuery) -> None:
    """Quick per-assignee export. callback_data carries the assignee INDEX into
    _export_assignee_names() (byte-safe); falls back to a raw name for any legacy
    keyboard still in a chat."""
    raw = query.data.split(":", 1)[1]
    try:
        assignee = (await _export_assignee_names())[int(raw)]
    except ValueError:
        assignee = raw  # legacy raw-name callback
    except IndexError:
        await query.answer("Ijrochi topilmadi", show_alert=True)
        return
    await query.answer(f"📤 {assignee}…")
    await _send_tasks_export(query.message, assignee=assignee)


@router.callback_query(F.data.startswith("exportst:"))
async def cb_export_status(query: CallbackQuery) -> None:
    """Quick per-status export (button under the full export)."""
    status = query.data.split(":", 1)[1]
    await query.answer(f"📤 {_EXPORT_FILTER_LABEL.get(status, status)}…")
    await _send_tasks_export(query.message, status=status)


@router.callback_query(F.data.startswith("exportcyr"))
async def cb_export_cyr(query: CallbackQuery) -> None:
    """Re-export in Uzbek Cyrillic, preserving the current scope:
    'exportcyr' (default active) | 'exportcyr:wi:<idx>' (assignee index) |
    'exportcyr:st:<status>' | legacy 'exportcyr:who:<name>'."""
    parts = (query.data or "").split(":", 2)
    kind = parts[1] if len(parts) > 1 else ""
    val = parts[2] if len(parts) > 2 else ""
    await query.answer("🔤 Krillcha…")
    if kind == "wi" and val:
        try:
            name = (await _export_assignee_names())[int(val)]
            await _send_tasks_export(query.message, assignee=name, script="cyr")
            return
        except (ValueError, IndexError):
            pass
    if kind == "who" and val:          # legacy raw-name callback (back-compat)
        await _send_tasks_export(query.message, assignee=val, script="cyr")
    elif kind == "st" and val:
        await _send_tasks_export(query.message, status=val, script="cyr")
    else:
        await _send_tasks_export(query.message, script="cyr")


_EXPORT_WHO_PER_PAGE = 8


def _export_root_keyboard(has_assignees: bool) -> InlineKeyboardMarkup:
    """Compact narrow-down keyboard attached to a full export: status filters +
    'Hammasi' + a single assignee drill-down (paginated, no silent cap)."""
    rows = [
        [InlineKeyboardButton(text="🔵 Aktiv", callback_data="exportst:active"),
         InlineKeyboardButton(text="✅ Bajarilgan", callback_data="exportst:done")],
        [InlineKeyboardButton(text="⌛ O'tgan", callback_data="exportst:overdue"),
         InlineKeyboardButton(text="⭐ Muhim", callback_data="exportst:important")],
        [InlineKeyboardButton(text="🗓 Shu hafta", callback_data="exportst:week")],
    ]
    bottom = [InlineKeyboardButton(text="📦 Hammasi", callback_data="exportst:all")]
    if has_assignees:
        bottom.append(InlineKeyboardButton(text="👤 Ijrochi bo'yicha →", callback_data="exportwho:0"))
    rows.append(bottom)
    rows.append([InlineKeyboardButton(text="🔤 Krillcha versiya", callback_data="exportcyr")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _export_who_keyboard(names: list, page: int) -> InlineKeyboardMarkup:
    """Paginated assignee picker — shows ALL executors across pages (no 8-cap drop)."""
    total = max(1, (len(names) + _EXPORT_WHO_PER_PAGE - 1) // _EXPORT_WHO_PER_PAGE)
    page = max(0, min(page, total - 1))
    start = page * _EXPORT_WHO_PER_PAGE
    chunk = names[start:start + _EXPORT_WHO_PER_PAGE]
    rows: list = []
    r: list = []
    for off, nm in enumerate(chunk):
        # callback carries the INDEX, not the name — Telegram caps callback_data at
        # 64 BYTES, and a Cyrillic/long patronymic would overflow (crash) or be
        # truncated mid-name (wrong scope). The index resolves back via the list.
        r.append(InlineKeyboardButton(text=f"👤 {nm}"[:64], callback_data=f"exportby:{start + off}"))
        if len(r) == 2:
            rows.append(r)
            r = []
    if r:
        rows.append(r)
    nav: list = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"exportwho:{page - 1}"))
    if total > 1:
        nav.append(InlineKeyboardButton(text=f"{page + 1}/{total}", callback_data="noop"))
    if page < total - 1:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"exportwho:{page + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="exportroot")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _export_assignee_names() -> list:
    """Distinct assignees in the active export scope, INCLUDING subtask assignees
    (the per-assignee export pulls subtasks too, so someone who only owns subtasks
    must still appear in the picker and resolve by index)."""
    tasks = await _fetch_tasks_for_export("active", include_subtasks=True)
    return sorted({(t.get("assignee") or "").strip() for t in tasks if (t.get("assignee") or "").strip()})


@router.callback_query(F.data.startswith("exportwho:"))
async def cb_export_who(query: CallbackQuery) -> None:
    """Drill-down: paginated assignee picker, edits the export keyboard in place."""
    try:
        page = int(query.data.split(":", 1)[1])
    except ValueError:
        page = 0
    names = await _export_assignee_names()
    if not names:
        await query.answer("Ijrochi topilmadi", show_alert=True)
        return
    await query.answer()
    try:
        await query.message.edit_reply_markup(reply_markup=_export_who_keyboard(names, page))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "exportroot")
async def cb_export_root(query: CallbackQuery) -> None:
    """Return from the assignee picker to the root export keyboard."""
    names = await _export_assignee_names()
    await query.answer()
    try:
        await query.message.edit_reply_markup(reply_markup=_export_root_keyboard(bool(names)))
    except TelegramBadRequest:
        pass


# Status/filter labels for the export subtitle + filename + buttons. Keys mirror
# the Tasks-section filters so "export active" == what the "Aktiv" filter shows.
_EXPORT_FILTER_LABEL = {
    "active": "Aktiv", "today": "Bugungi", "week": "Shu haftalik", "overdue": "Muddati o'tgan",
    "important": "Muhim", "urgent": "Shoshilinch", "done": "Bajarilgan",
    "recurring": "Takroriy", "all": "Barchasi",
}


def _export_title(status, assignee) -> str:
    """Dynamic Excel B1 title: department + scope — e.g. 'MARKETING BOSHQARMASI ·
    AKTIV VAZIFALAR', '… · MUDDATI O'TGAN VAZIFALAR', '… · J.KOMILOV VAZIFALARI'."""
    try:
        import protocol_doc
        dept = (protocol_doc.DEFAULTS.get("org_dept") or "").strip()
    except Exception:
        dept = ""
    if assignee:
        scope = f"{assignee.strip()} vazifalari"
    elif status and status != "all":
        scope = f"{_EXPORT_FILTER_LABEL.get(status, status)} vazifalar"
    elif status == "all":
        scope = "Barcha vazifalar"
    else:
        scope = "Aktiv vazifalar"
    return (f"{dept} · {scope}" if dept else scope).upper()
# /export <so'z> va ovoz: O'zbekcha status so'zi → kanonik filtr kaliti.
_EXPORT_STATUS_WORDS = {
    "bajarilgan": "done", "bajarilganlar": "done", "yakunlangan": "done", "done": "done",
    "aktiv": "active", "faol": "active", "active": "active", "ochiq": "active",
    "o'tgan": "overdue", "otgan": "overdue", "muddati o'tgan": "overdue", "overdue": "overdue",
    "muhim": "important", "important": "important",
    "shoshilinch": "urgent", "urgent": "urgent",
    "bugun": "today", "bugungi": "today", "today": "today",
    "hafta": "week", "haftalik": "week", "shu hafta": "week", "shu haftalik": "week", "week": "week",
    "takroriy": "recurring", "recurring": "recurring",
    "barchasi": "all", "hammasi": "all", "all": "all",
}


async def _fetch_tasks_for_export(filt: str, include_subtasks: bool = False) -> list:
    """Status/filter kaliti → vazifalar (eksport uchun, cheklovsiz). Semantikasi
    _render_tasks_for_filter bilan bir xil (izchillik). include_subtasks=True —
    per-ijrochi eksportда sub-vazifalarni ham qamrash uchun."""
    inc = include_subtasks
    if filt == "active":
        return await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=10000, include_subtasks=inc)
    if filt == "today":
        return await database.list_today_tasks()
    if filt == "week":
        # This calendar week (Mon 00:00 → next Mon): active tasks whose deadline falls in range.
        _now = datetime.now(database.TZ)
        _ws = (_now - timedelta(days=_now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        _we = _ws + timedelta(days=7)
        _act = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=10000, include_subtasks=inc)

        def _in_week(t):
            dl = t.get("deadline")
            if not dl:
                return False
            try:
                return _ws <= datetime.fromisoformat(dl).astimezone(database.TZ) < _we
            except (ValueError, TypeError):
                return False
        return [t for t in _act if _in_week(t)]
    if filt == "overdue":
        return await database.list_overdue_tasks()
    if filt == "done":
        return await database.list_tasks(status_in=["done"], limit=10000, include_subtasks=inc)
    if filt in ("important", "urgent"):
        act = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=10000, include_subtasks=inc)
        pset = ("P0", "P1") if filt == "important" else ("P0",)
        return [t for t in act if t.get("priority") in pset]
    if filt == "recurring":
        return await database.list_recurring_tasks(limit=10000)
    return await database.list_tasks(limit=10000, include_subtasks=inc)  # all


async def _send_tasks_export(message: Message, assignee: str | None = None,
                             status: str | None = None, script: str = "lat") -> None:
    """Build and send the tasks workbook (Template style). Shared by /export, the
    export_tasks action (voice/text), per-assignee and per-status buttons.
    `assignee` filters to one executor; `status` filters by Tasks-section filter
    key (active/today/overdue/important/urgent/done/recurring/all).
    `script="cyr"` transliterates all visible text to Uzbek Cyrillic."""
    import io
    from aiogram.types import BufferedInputFile
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        await message.answer("⚠️ Excel kutubxonasi (openpyxl) o'rnatilmagan.")
        return

    # Visible-text formatter. Brand/org names are wrapped in "..." (quote_names)
    # in BOTH scripts for a consistent formal look; cyr additionally transliterates
    # via to_cyrillic_pro (which keeps brands/acronyms/codes — Agrobank, KPI,
    # *9088, humo.uz — in Latin). Identity-ish for non-string cells.
    import translit
    if script == "cyr":
        def _tr(s):
            return translit.to_cyrillic_pro(translit.quote_names(s)) if isinstance(s, str) and s else s
    else:
        def _tr(s):
            return translit.quote_names(s) if isinstance(s, str) and s else s

    _inc_sub = bool(assignee)  # per-assignee export is flat → include subtasks too
    if status == "all":
        tasks = await database.list_tasks(limit=10000, include_subtasks=_inc_sub)  # "hammasi"
    elif status:
        tasks = await _fetch_tasks_for_export(status, include_subtasks=_inc_sub)
    else:
        # DEFAULT: active only (todo/in_progress/blocked) — old done/cancelled tasks
        # don't clutter the working file. Use "/export hammasi" for everything.
        tasks = await _fetch_tasks_for_export("active", include_subtasks=_inc_sub)
    if assignee:
        al = assignee.strip().lower()
        tasks = [t for t in tasks if (t.get("assignee") or "").strip().lower() == al]
    if not tasks:
        _what = []
        if status and status != "all":
            _what.append(f"«{_EXPORT_FILTER_LABEL.get(status, status)}»")
        if assignee:
            _what.append(f"«{assignee}»")
        await message.answer((" · ".join(_what) + " bo'yicha vazifa topilmadi.")
                             if _what else "📭 Eksport qilish uchun vazifa yo'q.")
        return

    # ── Row model: hierarchical (top-level + indented subtasks 1.1) for a full/status
    #    export; flat-with-"Asosiy vazifa" for a per-assignee export (subtasks shown with
    #    their parent's name since the parent itself may belong to someone else). ──
    flat_mode = bool(assignee)
    export_rows: list = []  # (number_str, is_sub, parent_title, task)
    if flat_mode:
        _ptitles: dict = {}
        for _t in tasks:
            _pid = _t.get("parent_id")
            if _pid and _pid not in _ptitles:
                _p = await database.get_task(_pid)
                _ptitles[_pid] = (_p.get("title") or "").strip() if _p else ""
        for _i, _t in enumerate(tasks, 1):
            export_rows.append((str(_i), bool(_t.get("parent_id")),
                                _ptitles.get(_t.get("parent_id"), ""), _t))
    else:
        _i = 0
        for _t in tasks:
            _i += 1
            export_rows.append((str(_i), False, "", _t))
            for _j, _s in enumerate(await database.list_subtasks(_t["id"]), 1):
                export_rows.append((f"{_i}.{_j}", True, "", _s))
    row_tasks = [r[3] for r in export_rows]  # all task dicts (incl. subtasks) for the panel
    # Top-level tasks that actually have subtasks in this export → their title is bold.
    parents_with_kids = {r[3].get("parent_id") for r in export_rows if r[3].get("parent_id")}

    # ── Visual template: a navy title banner + white-on-navy header band, a thin
    #    blue-grey grid, zebra rows, and accent colours (red overdue/P0, grey done).
    #    Visible cols: №..Kategoriya; ID is the last, hidden column (round-trip). ──
    ARIAL = "Arial"
    NAVY = "2E7D32"          # header band / section bands — Agrobank green (not navy)
    ZEBRA = "F4F7FB"         # alternating data rows
    GRIDC = "C8D0DC"         # grid line (light blue-grey)
    thin = Side(style="thin", color=GRIDC)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    NAVY_FILL = PatternFill("solid", fgColor=NAVY)
    ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
    # Per-assignee export is flat → add an "Asosiy vazifa" column so a subtask shows which
    # project it belongs to. The hierarchical export indents subtasks under the parent.
    headers = (["№", "Vazifa", "Asosiy vazifa", "Ijrochi", "Muddat", "Ustuvorlik", "Holat", "Takroriylik", "Izoh", "Kategoriya"]
               if flat_mode else _EXPORT_HEADERS)
    n_visible = len(headers)
    id_col = n_visible + 1
    last_col = get_column_letter(n_visible)
    title_col = headers.index("Vazifa") + 1
    deadline_col = headers.index("Muddat") + 1
    prio_col = headers.index("Ustuvorlik") + 1
    ota_col = (headers.index("Asosiy vazifa") + 1) if flat_mode else None
    left_cols = {title_col} | ({ota_col} if flat_mode else set())  # left-aligned text cols
    # NOTE: every data cell now sets wrap_text=True, so any long value wraps instead
    # of being clipped — no per-column wrap list needed.
    # Column widths (Excel units) — generous so nothing is clipped (matches the
    # principal's print template). Vazifa/Izoh widest; dates/status comfortable.
    widths = ([7, 48, 32, 30, 16, 17, 20, 16, 42, 22] if flat_mode
              else [7, 56, 32, 16, 17, 20, 16, 46, 22])

    def _wrapped_lines(text, col_idx):
        """Estimate display lines for `text` in the column at col_idx (Arial 14),
        honouring explicit newlines — used to size the row so nothing is cramped.
        Datetimes are measured in their DISPLAY form (DD-MM-YYYY, 10 chars) — the
        raw str(datetime) is 19 chars and would inflate every deadline row's height."""
        if isinstance(text, datetime):
            text = text.strftime("%d-%m-%Y")
        cpl = max(6, int(widths[col_idx - 1] * 0.85))
        return sum(max(1, -(-len(seg) // cpl)) for seg in str(text or "").split("\n")) or 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Vazifalar"

    # Full-width navy title banner (A1 … last col).
    ws.merge_cells(f"A1:{last_col}1")
    t1 = ws["A1"]
    t1.value = _tr(_export_title(status, assignee))
    t1.font = Font(name=ARIAL, size=20, bold=True, color="1A1A1A")  # plain bold (no banner fill)
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells(f"A2:{last_col}2")
    t2 = ws["A2"]
    now_s = datetime.now(database.TZ).strftime("%d-%m-%Y %H:%M")
    sub = f"Yaratilgan: {now_s}      Jami: {len(export_rows)} ta"
    if status and status != "all":
        sub += f"      ·  Holat: {_EXPORT_FILTER_LABEL.get(status, status)}"
    if assignee:
        sub += f"      ·  Ijrochi: {assignee}"
    t2.value = _tr(sub)
    t2.font = Font(name=ARIAL, size=11, color="33415C")  # plain grey caption (no fill)
    t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # White-on-navy header band.
    for i in range(1, id_col + 1):
        h = headers[i - 1] if i <= n_visible else "ID"
        c = ws.cell(row=3, column=i, value=_tr(h) if i <= n_visible else "ID")
        c.font = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = grid
    ws.row_dimensions[3].height = 34

    now_dt = datetime.now(database.TZ)
    OVERDUE_FILL = PatternFill("solid", fgColor="FCE4E4")  # light red — muddati o'tgan
    DONE_FILL = PatternFill("solid", fgColor="F0F0F0")     # grey — bajarilgan/bekor
    for rn, (num, is_sub, ptitle, t) in enumerate(export_rows, start=1):
        r = rn + 3
        _st = t.get("status", "todo")
        _is_done = _st in ("done", "cancelled")
        _overdue = False
        if t.get("deadline") and not _is_done:
            try:
                _overdue = datetime.fromisoformat(t["deadline"]).astimezone(database.TZ) < now_dt
            except (ValueError, TypeError):
                _overdue = False
        # Conditional fill wins; otherwise zebra-stripe every other row for readability.
        row_fill = (DONE_FILL if _is_done else OVERDUE_FILL if _overdue
                    else ZEBRA_FILL if rn % 2 == 0 else None)
        base_color = "8C8C8C" if _is_done else "1A1A1A"
        # A top-level task that HAS subtasks → its title is bold (it's a project head).
        _is_parent = (not is_sub) and t.get("id") in parents_with_kids
        title = (t.get("title") or "").strip()
        if is_sub and not flat_mode:
            title = "↳ " + title  # indent subtasks under the parent in hierarchical mode
        # Multi-name cells ("A/B") → " / " so they wrap cleanly into stacked names.
        asg = " / ".join(p.strip() for p in (t.get("assignee") or "").split("/") if p.strip())
        vals = [num, title]
        if flat_mode:
            vals.append(ptitle if is_sub else "")
        vals += [
            asg,
            _export_date(t.get("deadline")),
            _PRIORITY_LABEL_UZ.get(t.get("priority", "P2"), "Rejadagi"),
            _STATUS_LABEL_UZ.get(t.get("status", "todo"), t.get("status", "")),
            _RECUR_LABEL.get(t.get("recurrence_rule") or "", ""),   # Takroriylik (round-trip-safe)
            (t.get("description") or "").strip(),
            (t.get("category") or "").strip(),
        ]
        vals = [_tr(v) if isinstance(v, str) else v for v in vals]  # Cyrillic if script=cyr
        for i, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=val)
            _p0 = (i == prio_col and t.get("priority") == "P0")  # Shoshilinch → red + bold
            _overdue_dl = (i == deadline_col and _overdue)       # muddat o'tgan → red
            c.font = Font(name=ARIAL, size=14, bold=(_p0 or (i == title_col and _is_parent)),
                          color=("C00000" if (_p0 or _overdue_dl) else base_color))
            c.border = grid
            if row_fill:
                c.fill = row_fill
            c.alignment = Alignment(
                horizontal="left" if i in left_cols else "center",
                vertical="center", wrap_text=True, indent=1 if i in left_cols else 0,
            )
            if i == deadline_col and val is not None:
                c.number_format = "DD-MM-YYYY"  # date only — no clock
        idc = ws.cell(row=r, column=id_col, value=t.get("id"))
        idc.font = Font(name=ARIAL, size=11, color="A6A6A6")
        idc.border = grid
        if row_fill:
            idc.fill = row_fill
        # Size the row to the tallest wrapped cell across ALL columns (every cell
        # now wraps), so any long value — title, izoh, ijrochi, kategoriya — breathes
        # instead of being clipped to one line. Arial 14 ≈ 20px/line.
        _lines = max((_wrapped_lines(vals[ci - 1], ci) for ci in range(1, n_visible + 1)), default=1)
        ws.row_dimensions[r].height = max(32, _lines * 20 + 12)

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions[get_column_letter(id_col)].hidden = True  # ID — round-trip, hidden
    ws.freeze_panes = "A4"  # title + subtitle + header stay visible while scrolling
    # Print-ready: landscape, fit all columns to one page width, repeat the header.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:3"
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    # ── "Boshqaruv paneli" — LIVE executive dashboard as the FIRST tab. Every count
    #    is a COUNTIF/SUMPRODUCT formula over the Vazifalar sheet, so it auto-updates
    #    whenever a cell is edited — a real control panel, not a static snapshot.
    #    Counters below only decide WHICH rows to show (skip empty buckets). ──
    from collections import Counter
    _stc = Counter(t.get("status", "todo") for t in row_tasks)
    _prc = Counter(t.get("priority", "P2") for t in row_tasks)
    n_rows = len(export_rows)
    r0, r1 = 4, max(4, n_rows + 3)              # Vazifalar data rows (header on row 3)

    def _C(name):                               # header → column letter (layout-independent)
        return get_column_letter(headers.index(name) + 1)

    def _rng(name):
        c = _C(name)
        return f"Vazifalar!${c}${r0}:${c}${r1}"

    def _countif(name, crit, wild=False):
        c = _tr(str(crit)).replace('"', '""')
        return f'=COUNTIF({_rng(name)},"{("*" + c + "*") if wild else c}")'

    _done_l = _tr(_STATUS_LABEL_UZ["done"])
    _canc_l = _tr(_STATUS_LABEL_UZ["cancelled"])

    dash = wb.create_sheet("Boshqaruv paneli", 0)
    dash.sheet_view.showGridLines = False           # clean dashboard canvas
    # GRID layout: 3 label/value column-pairs side by side (A-B | C-D | E-F) so the
    # sections sit in a balanced dashboard instead of one long cramped strip.
    _dw = {1: 30, 2: 7, 3: 30, 4: 7, 5: 30, 6: 7}   # widths by column index
    for _ci, _w in _dw.items():
        dash.column_dimensions[get_column_letter(_ci)].width = _w
    _drow = Border(bottom=Side(style="thin", color=GRIDC))

    def _bump_height(r, label, col, base=22):
        """Grow the row to fit a (possibly wrapped) label so long category/executor
        names never get clipped into one cramped line. Takes the max across the
        cells already placed on this row (grid rows are shared by 3 panels)."""
        cpl = max(6, int(_dw.get(col, 24) * 0.92))
        lines = max(1, -(-len(_tr(str(label))) // cpl))
        dash.row_dimensions[r].height = max(dash.row_dimensions[r].height or 0, base, lines * 17 + 7)

    def _band(r, c0, c1, label):
        cs, ce = get_column_letter(c0), get_column_letter(c1)
        for cc in range(c0, c1 + 1):
            dash.cell(row=r, column=cc).fill = NAVY_FILL
        dash.merge_cells(f"{cs}{r}:{ce}{r}")
        x = dash.cell(row=r, column=c0, value=_tr(label))
        x.font = Font(name=ARIAL, size=12, bold=True, color="FFFFFF")
        x.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dash.row_dimensions[r].height = 23

    def _kv(r, col, label, value, color="1A1A1A", bold=False):
        a = dash.cell(row=r, column=col, value=_tr(label))
        a.font = Font(name=ARIAL, size=12, bold=bold, color=color)
        # wrap_text → a long name wraps to a 2nd line instead of being clipped.
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        a.border = _drow
        b = dash.cell(row=r, column=col + 1, value=value)
        b.font = Font(name=ARIAL, size=12, bold=True, color=color)
        b.alignment = Alignment(horizontal="center", vertical="center")
        b.border = _drow
        _bump_height(r, label, col)

    def _panel(top, col, title, items):
        """A titled block at (top, col-pair). items: (label, value, color, bold)."""
        _band(top, col, col + 1, title)
        for i, (lab, val, color, bold) in enumerate(items):
            _kv(top + 1 + i, col, lab, val, color, bold)
        return top + 1 + len(items)

    # Title banner + report date across all 6 columns.
    for cc in range(1, 7):
        dash.cell(row=1, column=cc).fill = NAVY_FILL
    dash.merge_cells("A1:F1")
    _t = dash.cell(row=1, column=1, value=_tr(_export_title(status, assignee)))
    _t.font = Font(name=ARIAL, size=16, bold=True, color="FFFFFF")
    _t.alignment = Alignment(horizontal="center", vertical="center")
    dash.row_dimensions[1].height = 34
    _dc = dash.cell(row=2, column=1, value=_tr(f"Hisobot sanasi: {now_s}"))
    _dc.font = Font(name=ARIAL, size=10, color="33415C")
    dash.merge_cells("A2:F2")

    # Asosiy/Sub split: hierarchical № has dots ("3.1") → test for "."; flat (per-
    # assignee) № is dotless ("1","2") so use the "Asosiy vazifa" column (filled only
    # for subtasks) instead, otherwise Sub always reads 0 on a per-person report.
    if flat_mode:
        _sub_f = f'=SUMPRODUCT(--({_rng("Asosiy vazifa")}<>""))'
        _aso_f = f'=COUNTA({_rng("Vazifa")})-{_sub_f[1:]}'
    else:
        _aso_f = f'=SUMPRODUCT(--ISERROR(SEARCH(".",{_rng("№")})))'
        _sub_f = f'=SUMPRODUCT(--ISNUMBER(SEARCH(".",{_rng("№")})))'
    umumiy = [
        ("Jami vazifalar", f"=COUNTA({_rng('Vazifa')})", NAVY, True),
        ("Asosiy vazifa", _aso_f, "1A1A1A", False),
        ("Sub-vazifa", _sub_f, "1A1A1A", False),
        ("Muddati o'tgan",
         f'=SUMPRODUCT(({_rng("Muddat")}<>"")*({_rng("Muddat")}<TODAY())'
         f'*({_rng("Holat")}<>"{_done_l}")*({_rng("Holat")}<>"{_canc_l}"))', "C00000", True),
        ("Bajarilgan", _countif("Holat", _STATUS_LABEL_UZ["done"]), "548235", True),
    ]
    holat = [(_STATUS_LABEL_UZ[c], _countif("Holat", _STATUS_LABEL_UZ[c]), "1A1A1A", False)
             for c in ("todo", "in_progress", "blocked", "done", "cancelled") if _stc.get(c)]
    ustuv = [(_PRIORITY_LABEL_UZ[c], _countif("Ustuvorlik", _PRIORITY_LABEL_UZ[c]), "1A1A1A", False)
             for c in ("P0", "P1", "P2", "P3") if _prc.get(c)]
    _cats = sorted({(t.get("category") or "").strip() for t in row_tasks if (t.get("category") or "").strip()})
    kateg = [(c, _countif("Kategoriya", c), "1A1A1A", False) for c in _cats]
    _names = sorted({_p.strip() for t in row_tasks
                     for _p in (t.get("assignee") or "").split("/") if _p.strip()})

    # Row 1 of blocks: UMUMIY | HOLAT | USTUVORLIK (side by side).
    b1 = _panel(4, 1, "UMUMIY", umumiy)
    b2 = _panel(4, 3, "HOLAT BO'YICHA", holat)
    b3 = _panel(4, 5, "USTUVORLIK BO'YICHA", ustuv)
    row2 = max(b1, b2, b3) + 1

    # Row 2 of blocks: KATEGORIYA (A-B) | IJROCHI YUKLAMASI (C-F, split into 2 columns).
    if kateg:
        _panel(row2, 1, "KATEGORIYA BO'YICHA", kateg)
    if _names:
        _band(row2, 3, 6, "IJROCHI YUKLAMASI")
        half = -(-len(_names) // 2)
        for idx, _nm in enumerate(_names):
            _col = 3 if idx < half else 5
            _kv(row2 + 1 + (idx if idx < half else idx - half), _col,
                _nm, _countif("Ijrochi", _nm, wild=True))

    dash.page_setup.orientation = "landscape"
    dash.page_setup.fitToWidth = 1
    dash.page_setup.fitToHeight = 0
    dash.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    dash.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)

    buf = io.BytesIO()
    wb.save(buf)
    _tag_bits = ([status] if (status and status != "all") else []) + [(assignee or "hammasi")]
    if script == "cyr":
        _tag_bits.append("krill")
    tag = "_".join(_tag_bits).strip().replace(" ", "_").replace("/", "-")
    fname = f"vazifalar_{tag}_{datetime.now(database.TZ).strftime('%Y-%m-%d')}.xlsx"
    cap = f"📤 **{len(export_rows)} ta** eksport qilindi"
    _cap_bits = ([_EXPORT_FILTER_LABEL.get(status, status)] if (status and status != "all") else []) \
        + ([assignee] if assignee else [])
    if status is None:               # default scope is active-only — make it explicit
        _cap_bits.insert(0, "aktiv")
    if script == "cyr":
        _cap_bits.append("krillcha")
    cap += (" — _" + " · ".join(_cap_bits) + "_.") if _cap_bits else "."
    cap += "\n\n_Tahrirlab qaytadan yuborsangiz — import bo'ladi (tasdiqdan keyin)._"
    # Narrow-down keyboard attached to the FILE message itself (one message, no
    # clutter, no separate "Holat yoki ijrochi" prompt). Only after a FULL export
    # (no filter/assignee yet). Assignees live behind a single paginated drill-down
    # so the keyboard stays compact and no executor is silently dropped.
    root_kb = None
    if not assignee and not status:
        has_asg = any((t.get("assignee") or "").strip() for t in row_tasks)
        root_kb = _export_root_keyboard(has_asg)
    elif script != "cyr":
        # Filtered Latin export → offer the SAME scope in Cyrillic. Carry the scope
        # byte-safely: assignee as an INDEX (names overflow the 64-byte callback
        # limit), status as its short ASCII key. If the assignee isn't in the
        # active list (e.g. a done-only filter), skip the button rather than risk
        # a wrong/empty re-export.
        _cyr_cb = None
        if assignee:
            _anames = await _export_assignee_names()
            if assignee in _anames:
                _cyr_cb = f"exportcyr:wi:{_anames.index(assignee)}"
        else:
            _cyr_cb = f"exportcyr:st:{status}"
        if _cyr_cb:
            root_kb = InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🔤 Krillcha versiya", callback_data=_cyr_cb)]])
    await message.answer_document(
        BufferedInputFile(buf.getvalue(), filename=fname), caption=cap,
        parse_mode="Markdown", reply_markup=root_kb,
    )


def _norm_header(c) -> str:
    """Normalize a header cell for matching: back-transliterate Cyrillic → Latin (so a
    Cyrillic export's 'Вазифа'/'Изоҳ'/'Ҳолат' headers match the Latin synonym sets),
    lowercase, drop apostrophe variants and spaces — 'Mas'ul', 'Tugatish sanasi',
    'Изоҳ' all map cleanly. to_latin is a no-op on already-Latin text."""
    import translit
    s = translit.to_latin(str(c or "")).strip().lower()
    for ch in ("'", "ʼ", "’", "`", "ʻ", " "):
        s = s.replace(ch, "")
    return s


def _import_text(val, cyr: bool = False) -> str:
    """Normalize a free-text cell coming back from an export: strip the display
    quotes quote_names added around brand names ('"Agrobank"' → 'Agrobank') and —
    ONLY for a Cyrillic-export file (cyr=True, detected from the header script) —
    back-transliterate the value ('муҳим изоҳ' → 'muhim izoh') so the cyr round-trip
    is deterministic. A LATIN export must NOT transliterate: a stored Cyrillic
    (e.g. Russian) value would otherwise be silently and lossily Latinized on a
    no-op round-trip. Latin text without brands is returned unchanged."""
    import translit
    s = str(val or "").strip()
    if not s:
        return s
    if cyr and any("Ѐ" <= ch <= "ӿ" for ch in s):  # cyr export → restore Latin storage
        s = translit.to_latin(s)
    return translit.unquote_names(s)


# Column synonyms (normalized) — keep the importer forgiving about header names.
_COL_NUMBER = ("№", "no", "tr", "t/r", "nomer", "raqam", "tartib", "tartibraqami", "#")
_COL_TITLE = ("vazifa", "topshiriq", "nomi", "ish", "title", "task")
_COL_ASSIGNEE = ("ijrochi", "masul", "masulxodim", "bajaruvchi", "kim", "kimbajaradi", "assignee")
_COL_DEADLINE = ("muddat", "muddati", "sana", "tugatishsanasi", "tugashsanasi", "deadline")
_COL_PRIORITY = ("ustuvorlik", "muhimlik", "daraja", "darajasi", "priority")
_COL_STATUS = ("holat", "holati", "status")
_COL_DESC = ("izoh", "izohi", "tavsif", "tavsifi", "qoshimcha", "description")
_COL_CATEGORY = ("kategoriya", "category", "turkum", "bolim")
_COL_RECURRENCE = ("takroriylik", "takroriy", "takrorlanish", "recurrence", "rrule")


def _structured_tasks_from_table(table: list) -> list[dict]:
    """Fast path: if the table has a recognizable header row, map columns by name
    → create_task actions. Header names are normalized and common synonyms accepted
    (Mas'ul→ijrochi, Topshiriq→vazifa, …). Returns [] when no header is found."""
    header_idx = next(
        (i for i, row in enumerate(table) if any(_norm_header(c) in _COL_TITLE for c in row)),
        None,
    )
    if header_idx is None:
        return []
    header = [_norm_header(c) for c in table[header_idx]]
    # Header script decides value handling: Cyrillic headers ("Вазифа", "Изоҳ") mean
    # this is OUR cyr export — free-text values are back-transliterated to the Latin
    # storage form. Latin headers → values pass through (stored Cyrillic text must
    # survive a Latin-export round-trip untouched).
    _cyr_file = any("Ѐ" <= ch <= "ӿ" for c in table[header_idx] for ch in str(c or ""))

    def pick(d, names):
        for n in names:
            v = d.get(n)
            if v not in (None, ""):
                return v
        return None

    def has(names):
        """Is a column for this field present in the file's header?"""
        return any(n in header for n in names)

    # A per-assignee (flat) export has an "Asosiy vazifa" column and numbers subtasks
    # plainly — its № carries NO parent/child hierarchy, so re-parenting from it must be
    # suppressed (else every subtask is promoted to top-level on re-import).
    _flat_file = has(("asosiyvazifa", "asosiyvazifasi"))
    out: list[dict] = []
    for r in table[header_idx + 1:]:
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        # Strip the export's subtask indent marker so it doesn't accumulate on
        # re-import ("↳ ↳ Title" → "Title"); hierarchy comes from the № column.
        # _import_text also back-transliterates a Cyrillic export and un-quotes
        # brand names so a round-trip never drifts ('"Agrobank"' → 'Agrobank').
        title = _import_text(pick(d, _COL_TITLE), _cyr_file).lstrip("↳ ").strip()
        if not title:
            continue
        # Only touch a field whose COLUMN exists in the file: a blank cell in a
        # PRESENT column CLEARS that field on update (full round-trip edit), while an
        # ABSENT column is left untouched (a partial file never wipes other fields).
        # Previously blank assignee/category/izoh were skipped → clearing didn't apply.
        data = {"title": title, "source": "excel"}  # source: trusted for new assignee/category (A)
        # priority/status have NO null state — a blank cell must NOT overwrite the
        # stored value with the default (P2/todo). Write only when the cell has a value.
        _pri = pick(d, _COL_PRIORITY)
        if _pri not in (None, ""):
            data["priority"] = _import_priority(_pri)
        _sta = pick(d, _COL_STATUS)
        if _sta not in (None, ""):
            data["status"] = _import_status(_sta)
        _dl_unread = ""
        if has(_COL_DEADLINE):
            _dl_cell = pick(d, _COL_DEADLINE)
            _dl_raw = "" if _dl_cell is None else str(_dl_cell).strip()
            _dl = _import_deadline(_dl_cell)
            if _dl is not None:
                data["deadline"] = _dl            # parsed a real date/time
            elif not _dl_raw:
                data["deadline"] = None           # blank cell → intentional clear
            else:
                # A NON-empty but unrecognized cell must NOT wipe a good deadline —
                # leave the stored value untouched, but remember the raw text so the
                # preview can say "o'zgarmaydi (o'qilmadi)" instead of "Deadline yo'q".
                _dl_unread = _dl_raw
        # Export displays multi-name executors as 'A / B' — normalize back to the
        # stored 'A/B' form so an untouched round-trip is lossless (each part is
        # canonicalized individually in _execute_actions).
        asg = "/".join(p.strip() for p in _import_text(pick(d, _COL_ASSIGNEE), _cyr_file).split("/") if p.strip())
        if has(_COL_ASSIGNEE):
            data["assignee"] = asg          # "" → clears the executor on update
        if has(_COL_DESC):
            desc = _import_text(pick(d, _COL_DESC), _cyr_file)
            # Never let the assignee leak into the description (reported field-mix bug).
            data["description"] = "" if (asg and desc.lower() == asg.lower()) else desc
        if has(_COL_CATEGORY):
            cat = _import_text(pick(d, _COL_CATEGORY), _cyr_file)
            data["category"] = "" if cat.lower() == "(boshqa)" else cat
        if has(_COL_RECURRENCE):
            # Round-trips via normalize_recurrence_rule ("har kuni"→daily); blank → clears.
            data["recurrence_rule"] = database.normalize_recurrence_rule(pick(d, _COL_RECURRENCE))
        act = {"type": "create_task", "data": data,
               # Carry the optional ID (hidden export column) so the caller can turn this
               # into an UPDATE when the task still exists — instead of a duplicate.
               "_id": str(pick(d, ("id",)) or "").strip()}
        if _dl_unread:
            act["_dl_unread"] = _dl_unread  # preview-only flag (never written to DB)
        # Hierarchy from the № column: "3.1" → subtask of "3". _execute_actions
        # resolves _num/parent into a real parent_id at create/update time.
        num = str(pick(d, _COL_NUMBER) or "").strip().rstrip(".")
        if num:
            act["_num"] = num
        if _flat_file:
            act["_flat"] = True   # № is non-hierarchical in a per-assignee export
        out.append(act)
    return out


def _same_deadline_instant(new_iso, stored_iso) -> bool:
    """True when two deadline ISO strings mean the same moment (sub-second
    tolerance): xlsx storage truncates microseconds, so an UNTOUCHED exported cell
    comes back .123456→.123000 — a string compare would see a 'change' and
    database.update_task would reset reminded_at (duplicate reminder)."""
    if not (isinstance(new_iso, str) and isinstance(stored_iso, str)):
        return False
    try:
        return abs(datetime.fromisoformat(new_iso)
                   - datetime.fromisoformat(stored_iso)) < timedelta(seconds=1)
    except (ValueError, TypeError):
        return False


def _count_orphan_subtasks(actions: list) -> int:
    """How many dotted-№ children reference a parent № that is NOT present in this
    batch. _resolve_parent only re-parents from rows in the same file, so an orphan
    silently becomes a top-level task — the caller warns about the count."""
    batch_nums = {(a.get("_num") or "") for a in actions if a.get("_num")}
    return sum(
        1 for a in actions
        if "." in (a.get("_num") or "")
        and a["_num"].split(".")[0] not in batch_nums
    )


def _apply_title_dedup(actions: list, by_title: dict) -> dict:
    """Dedup imported create_tasks by normalized title — two ways:
      • title matches an EXISTING active task → convert to update_task (no duplicate);
      • title repeated WITHIN this batch (same file) → keep the first, DROP the rest.
    The second case is the one that previously slipped through: identical new-title
    rows in one import each became a create → duplicates. Mutates `actions` in place
    (drops intra-batch dupes via slice-assign). Returns {'converted', 'dropped'}."""
    converted = 0
    dropped = 0
    seen_in_batch: set = set()
    out: list = []
    for a in actions:
        if a.get("type") != "create_task":
            out.append(a)
            continue
        key = ((a.get("data") or {}).get("title") or "").strip().lower()
        if key and key in seen_in_batch:
            dropped += 1
            continue  # identical title already handled in this batch — drop the dup
        if key and key in by_title:
            a["type"] = "update_task"
            a["id"] = by_title[key]
            converted += 1
        if key:
            seen_in_batch.add(key)
        out.append(a)
    actions[:] = out
    return {"converted": converted, "dropped": dropped}


def _table_to_text(table: list, max_rows: int = 120, max_chars: int = 6000) -> str:
    """Render a parsed table as plain text for the smart extractor."""
    lines = []
    for row in table[:max_rows]:
        cells = [("" if c is None else str(c)).strip() for c in row]
        if any(cells):
            lines.append(" | ".join(cells))
    return "\n".join(lines)[:max_chars]


async def _smart_tasks_from_table(table: list) -> list[dict]:
    """Flexible import (no fixed format): hand the raw file content to Claude and
    let it find the tasks — whatever the column names, order, or layout. Returns
    create_task actions (deadline as ISO); [] if nothing task-like is found."""
    content = _table_to_text(table)
    if not content.strip():
        return []
    directive = (
        "[INTERNAL] extract_tasks_from_file\n\n"
        "Quyidagi fayl mazmunidan printsipalning VAZIFALARINI ajrat. Fayl istalgan "
        "ko'rinishda bo'lishi mumkin — ustun nomlari boshqacha, tartibsiz yoki erkin matn. "
        "Har bir ANIQ vazifa uchun create_task action chiqar. Maydonlarni ARALASHTIRMA:\n"
        "- title: faqat ish nomi (imperativ). Ichida ijrochi/sana BO'LMASIN.\n"
        "- assignee: ijrochi/mas'ul ismi — bo'lsa; yo'q bo'lsa qo'yma.\n"
        "- deadline: ISO 8601 Asia/Tashkent yoki null.\n"
        "- priority: P0/P1/P2/P3 — matndan tushun, aniqmasa P2.\n"
        "- description: FAQAT manbada mavjud qo'shimcha tushuntirish. Ichiga ijrochi "
        "ismini, sanani yoki ustuvorlikni YOZMA — ular alohida maydonlarda. "
        "Qo'shimcha izoh bo'lmasa — description'ni umuman QO'YMA (bo'sh/null). "
        "HECH NARSA O'YLAB TOPMA (masalan 'Fayldan import qilingan' kabi matn yozma).\n"
        "Sarlavha, 'jami', izoh kabi vazifa BO'LMAGAN qatorlarni tashlab ket. "
        "Hech narsa topilmasa — actions=[].\n\n"
        f"FAYL MAZMUNI:\n{content}"
    )
    try:
        resp = await claude_service.process_message("", internal_directive=directive)
    except Exception:
        logger.exception("smart import extraction failed")
        return []
    _GENERIC_DESC = {
        "fayldan import qilingan vazifa", "import qilingan vazifa", "fayldan import qilingan",
        "import qilingan", "fayldan", "vazifa", "—", "-",
    }
    out: list[dict] = []
    for a in resp.get("actions", []):
        if a.get("type") != "create_task":
            continue
        d = a.get("data") or {}
        if not (d.get("title") or "").strip():
            continue
        # Safety net: strip a description that's just the assignee or an invented filler.
        desc = (d.get("description") or "").strip()
        asg = (d.get("assignee") or "").strip()
        if desc and (desc.lower() in _GENERIC_DESC
                     or (asg and desc.lower() == asg.lower())):
            d.pop("description", None)
        # Normalize priority/status to the allowed set — a hallucinated label
        # ("Critical"/"P5") would otherwise fail create_task's CHECK and silently
        # DROP the row (already-confirmed import). Coerce to the nearest valid code.
        if "priority" in d:
            d["priority"] = _import_priority(d.get("priority"))
        if "status" in d:
            d["status"] = _import_status(d.get("status"))
        out.append(a)
    return out


_IMPORT_PAGE_SIZE = 10


def _import_deadline_label(iso) -> str:
    """Short deadline for the import preview: 'DD-MM, HH:MM' or 'Deadline yo'q'.
    The year is shown whenever it differs from the current one — a year-less cell
    that resolved to another year must be visible in the confirm preview."""
    if not iso:
        return "Deadline yo'q"
    try:
        dt = datetime.fromisoformat(iso).astimezone(database.TZ)
    except (ValueError, TypeError):
        return str(iso)
    fmt = "%d-%m-%Y, %H:%M" if dt.year != datetime.now(database.TZ).year else "%d-%m, %H:%M"
    return dt.strftime(fmt)


def _import_deadline_line(a: dict) -> str:
    """Preview deadline line for one import action — distinguishes the THREE update
    outcomes that used to render identically as 'Deadline yo'q':
    value → set; explicit None → cleared; key absent → untouched (unreadable cell)."""
    d = a.get("data", {}) or {}
    upd = a.get("type") == "update_task"
    if "deadline" in d:
        if d["deadline"] is None:
            return "olib tashlanadi" if upd else "Deadline yo'q"
        return _import_deadline_label(d["deadline"])
    if a.get("_dl_unread"):
        return f"o'zgarmaydi (katak o'qilmadi: {str(a['_dl_unread'])[:24]})"
    return "o'zgarmaydi" if upd else "Deadline yo'q"


def _format_import_page(actions: list[dict], page: int) -> str:
    """One paginated page of the import preview. Each record is a clean 3-line
    block (title / 👤 ijrochi / ⏳ muddat) so long titles never mangle the layout."""
    total = len(actions)
    pages = max(1, (total + _IMPORT_PAGE_SIZE - 1) // _IMPORT_PAGE_SIZE)
    page = max(0, min(page, pages - 1))
    start = page * _IMPORT_PAGE_SIZE
    chunk = actions[start:start + _IMPORT_PAGE_SIZE]
    n_sub = sum(1 for a in actions if "." in (a.get("_num") or ""))
    lines = ["⚠️ **TASDIQLAYSIZMI?**", "", f"📋 {total} ta yozuv topildi"]
    if n_sub:
        lines.append(f"🌳 {total - n_sub} ta asosiy + {n_sub} ta sub-vazifa (№ bo'yicha)")
    if pages > 1:
        lines.append(f"_Ko'rsatilmoqda: {start + 1}–{start + len(chunk)} / {total}_")
    lines.append("")
    for i, a in enumerate(chunk, start=start + 1):
        d = a.get("data", {}) or {}
        title = (d.get("title") or "—").strip()
        assignee = (d.get("assignee") or "belgilanmagan").strip() or "belgilanmagan"
        upd = " ✏️" if a.get("type") == "update_task" else ""
        num = a.get("_num") or ""
        is_sub = "." in num
        prefix = "↳ " if is_sub else ""
        lines.append(f"**{i}. {prefix}{title}**{upd}")
        if is_sub:
            lines.append(f"   🌳 sub-vazifa (ota: №{num.split('.')[0]})")
        lines.append(f"   👤 {assignee}")
        lines.append(f"   ⏳ {_import_deadline_line(a)}")
        lines.append("")
    return "\n".join(lines).rstrip()


def _import_preview_keyboard(page: int, total_pages: int, switch_uid: str = "",
                             allow_mirror: bool = False) -> InlineKeyboardMarkup:
    rows = []
    if total_pages > 1:
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton(text="⬅️ Oldingi", callback_data=f"imppage:{page - 1}"))
        nav.append(InlineKeyboardButton(text=f"{page + 1} / {total_pages}", callback_data="noop"))
        if page < total_pages - 1:
            nav.append(InlineKeyboardButton(text="Keyingi ➡️", callback_data=f"imppage:{page + 1}"))
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="➕ Qo'shish va yangilash", callback_data="acts_confirm")])
    if allow_mirror:  # full-mirror sync: bot tasks become EXACTLY the file (extras archived)
        rows.append([InlineKeyboardButton(text="🔄 To'liq moslashtirish (fayl bo'yicha)",
                                          callback_data="import_mirror")])
    if switch_uid:  # one-tap switch to analysis (when default guessed "import")
        rows.append([InlineKeyboardButton(text="📄 Tahlil qil", callback_data=f"docact:analyze:{switch_uid}")])
    rows.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="acts_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data.startswith("imppage:"))
async def cb_import_page(query: CallbackQuery, state: FSMContext) -> None:
    """Flip import-preview pages without re-parsing — actions live in FSM state."""
    try:
        page = int(query.data.split(":", 1)[1])
    except ValueError:
        return
    data = await state.get_data()
    actions = (data.get("pending_response") or {}).get("actions") or []
    if not actions:
        await query.answer("Vaqt o'tdi — faylni qayta yuboring.", show_alert=True)
        return
    await query.answer()
    switch_uid = data.get("_doc_uid") or ""
    total_pages = max(1, (len(actions) + _IMPORT_PAGE_SIZE - 1) // _IMPORT_PAGE_SIZE)
    try:
        await query.message.edit_text(
            _format_import_page(actions, page), parse_mode="Markdown",
            reply_markup=_import_preview_keyboard(page, total_pages, switch_uid, allow_mirror=True),
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "import_mirror")
async def cb_import_mirror(query: CallbackQuery, state: FSMContext) -> None:
    """🔄 Full-mirror: make the bot's ACTIVE tasks EXACTLY match the imported file.
    ACTIVE tasks the file omitted (their exported ID isn't in the file) are ARCHIVED
    (status → cancelled — recoverable). Requires the hidden ID column; asks a 2nd confirm."""
    data = await state.get_data()
    response = data.get("pending_response")
    if not response or not isinstance(response, dict):
        await query.answer("Tasdiqlash vaqti o'tdi — faylni qayta yuboring.", show_alert=True)
        return
    if not data.get("_import_had_ids"):
        await query.answer(
            "Bu faylda yashirin ID ustuni yo'q — to'liq moslashtirish uchun BOT "
            "eksportini (📤) tahrirlab yuboring. Hozir '➕ Qo'shish/yangilash' ishlaydi.",
            show_alert=True)
        return
    actions = list(response.get("actions") or [])
    referenced = {a["id"] for a in actions if a.get("type") == "update_task" and a.get("id")}
    active = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=10000)
    to_archive = [t for t in active if t["id"] not in referenced]
    n_new = sum(1 for a in actions if a.get("type") == "create_task")
    n_upd = sum(1 for a in actions if a.get("type") == "update_task")
    if not to_archive:
        await query.answer("Fayl bilan bir xil — arxivlanadigan yo'q. '➕' bilan davom eting.",
                           show_alert=True)
        return
    combined = actions + [{"type": "update_task", "id": t["id"], "data": {"status": "cancelled"}}
                          for t in to_archive]
    await state.update_data(pending_response={
        "actions": combined,
        "user_message": (f"📥 Moslashtirildi: {n_new} yangi · {n_upd} yangilangan · "
                         f"{len(to_archive)} arxivlangan."),
        "buttons": []})
    pct = round(100 * len(to_archive) / max(1, len(active)))
    lines = ["🔄 **TO'LIQ MOSLASHTIRISH**", "",
             f"➕ {n_new} yangi · ✏️ {n_upd} yangilanadi · 📦 **{len(to_archive)} arxivlanadi**", ""]
    if pct >= 50:
        lines += [f"⚠️ **DIQQAT:** aktiv vazifalarning **{pct}%** arxivlanadi — fayl chala bo'lmasin!", ""]
    lines.append("_Arxivlanadi (status → Bekor qilingan, qaytarib bo'ladi):_")
    lines += [f"• {(t.get('title') or '—')[:50]}" for t in to_archive[:10]]
    if len(to_archive) > 10:
        lines.append(f"… va yana {len(to_archive) - 10} ta")
    lines += ["", "_Tasdiqlashdan oldin DB zaxirasi olinadi._"]
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Tasdiqlash", callback_data="import_mirror_ok"),
        InlineKeyboardButton(text="❌ Bekor", callback_data="acts_cancel"),
    ]])
    await query.answer()
    await _safe_answer(query.message, "\n".join(lines), parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data == "import_mirror_ok")
async def cb_import_mirror_confirm(query: CallbackQuery, state: FSMContext) -> None:
    """Execute the full-mirror after the 2nd confirm: backup → create/update/archive."""
    data = await state.get_data()
    response = data.get("pending_response")
    await state.clear()
    if not response or not isinstance(response, dict):
        await query.answer("Tasdiqlash vaqti o'tdi.", show_alert=True)
        return
    await query.answer("🔄 Moslashtirilmoqda…")
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    undo_token = None
    try:
        _UNDO_BACKUPS[str(query.message.message_id)] = await _create_db_backup("pre-mirror")
        undo_token = str(query.message.message_id)
    except Exception:
        logger.exception("Pre-mirror backup failed (continuing)")
    try:
        ids_by_type = await _execute_actions(response.get("actions", []))
    except Exception as e:
        logger.exception("mirror _execute_actions failed")
        await query.message.answer(_humanize_error(e))
        return
    msg = response.get("user_message") or "📥 Moslashtirildi."
    # DB xatosi bo'lgan qatorlarni jimgina yutmaymiz — ogohlantirishni qo'shamiz.
    msg += _failed_actions_note(ids_by_type) + _conflict_note(ids_by_type)
    kb = None
    if undo_token:
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="↩️ Qaytarish", callback_data=f"undodelete:{undo_token}")]])
    await query.message.answer(msg, reply_markup=kb)


# Importable document kinds — the task importer can parse these into a table.
_IMPORTABLE_KINDS = {"excel", "csv", "pdf"}
# Caption keywords that mean "load this as a task LIST" (→ importer). Kept
# deliberately narrow; any other caption is treated as an analysis instruction
# (Claude can still extract tasks from prose via proposed actions).
_IMPORT_CAPTION_HINTS = ("import", "yukla", "jadval", "ro'yxatni", "royxatni", "ro'yxat qil")
# Caption keywords that mean "just file this for later" (→ notes inbox).
_NOTE_CAPTION_HINTS = ("qayd", "saqlab qo'y", "saqlab qoy", "inbox", "keyinroq")

# file_unique_id -> {"file_id", "name", "kind", "mime"}. Lets an inline button
# (the caption-less-PDF ask prompt, or a result's switch button) re-fetch a
# just-seen file without stuffing the long file_id into 64-byte callback_data.
# Single-user bot → tiny, short-lived; bounded below.
_DOC_FLIP_CACHE: "dict[str, dict]" = {}
_DOC_FLIP_CACHE_MAX = 20


def _doc_remember(file_unique_id: str, file_id: str, name: str, kind: str, mime: str = "") -> None:
    if not file_unique_id or not file_id:
        return
    if len(_DOC_FLIP_CACHE) >= _DOC_FLIP_CACHE_MAX:
        try:  # drop the oldest entry (dict preserves insertion order)
            _DOC_FLIP_CACHE.pop(next(iter(_DOC_FLIP_CACHE)))
        except StopIteration:
            pass
    _DOC_FLIP_CACHE[file_unique_id] = {"file_id": file_id, "name": name, "kind": kind, "mime": mime}


def _decide_file_route(caption: str, kind: str) -> str:
    """Decide what to do with an incoming file: 'import', 'analyze', or 'note'.
    Pure function (no I/O) so the disambiguation rule stays unit-tested.

    Rule — caption + smart default:
      • caption says note            → note
      • caption says import (and the kind is importable) → import
      • any other caption            → analyze (the caption is the instruction)
      • no caption: excel/csv        → import
      • no caption: everything else  → analyze
    """
    cap = (caption or "").strip().lower()
    if cap:
        if any(h in cap for h in _NOTE_CAPTION_HINTS):
            return "note"
        if any(h in cap for h in _IMPORT_CAPTION_HINTS) and kind in _IMPORTABLE_KINDS:
            return "import"
        return "analyze"
    if kind in ("excel", "csv"):
        return "import"
    return "analyze"


@router.message(F.document | F.photo)
async def handle_incoming_file(message: Message, state: FSMContext) -> None:
    """Single entry point for uploaded/forwarded documents and photos. Routes by
    caption + file type (see _decide_file_route): a task LIST is imported, a
    document/image is analysed by Claude, an explicit 'qayd' caption files it to
    the notes inbox. Registered before the generic 'unsupported attachment'
    handler so it wins for documents and photos."""
    doc = message.document
    photo = message.photo[-1] if message.photo else None
    caption = (message.caption or "").strip()

    if doc is not None:
        file_name = doc.file_name or ""
        mime = doc.mime_type or ""
        kind = document_service.detect_kind(file_name, mime)
        file_id, file_unique_id = doc.file_id, doc.file_unique_id
    elif photo is not None:
        file_name, mime, kind = "photo.jpg", "image/jpeg", "image"
        file_id, file_unique_id = photo.file_id, photo.file_unique_id
    else:
        return  # neither — shouldn't happen given the filter

    # Cache the file so any inline button (ask prompt / switch) can re-fetch it.
    _doc_remember(file_unique_id, file_id, file_name, kind, mime)

    route = _decide_file_route(caption, kind)
    label = file_name or ("rasm" if photo else "fayl")
    is_fwd = bool(getattr(message, "forward_origin", None) or getattr(message, "forward_from", None))

    if route == "note":
        await _capture_file_as_note(message, label=label, caption=caption, is_fwd=is_fwd)
    elif route == "import":
        await _run_task_import(message, state)
    elif kind == "pdf" and not caption:
        # PDF is the one genuinely ambiguous type (contract vs exported task
        # list) and analysis is the expensive path — ask before guessing.
        await _ask_file_intent(message, file_unique_id, label)
    else:
        await _analyze_document(
            message, state, instruction=caption,
            file_id=file_id, file_unique_id=file_unique_id,
            file_name=file_name, mime=mime, kind=kind,
        )


async def _capture_file_as_note(message: Message, *, label: str, caption: str = "",
                                is_fwd: bool = False) -> None:
    """File an attachment to the notes inbox. Stores the caption (+ the file
    name) — not the binary; just the intent to revisit. Reused by the 'Qaydga
    saqla' switch button (docact:note) and the 'qayd' caption route."""
    try:
        nid = await database.create_note({
            "content": caption or f"Yuborilgan fayl: {label}",
            "title": f"📎 {label}",
            "tags": ["fayl"],
            "source": "forward" if is_fwd else "manual",
        })
    except Exception as e:
        await message.answer(_humanize_error(e))
        return
    await message.answer(f"📝 Inbox'ga saqlandi: {label}\n/notes — qaydlar" if nid
                         else "📝 Saqlab bo'lmadi.")


async def _analyze_document(
    message: Message, state: FSMContext, *, instruction: str,
    file_id: str, file_unique_id: str, file_name: str, mime: str, kind: str,
) -> None:
    """Download → extract → Claude multimodal analysis → summary, plus any
    proposed tasks routed through the standard acts_confirm confirm pipeline."""
    try:
        file = await message.bot.get_file(file_id)
        blob = await message.bot.download_file(file.file_path)
        data = blob.read() if hasattr(blob, "read") else blob
    except Exception as e:
        await message.answer(_humanize_error(e))
        return

    try:
        blocks, meta = document_service.build_content_blocks(data, kind, file_name, mime)
    except document_service.DocumentError as e:
        await message.answer(f"📄 {e}")
        return
    except Exception as e:
        logger.exception("Document block build failed")
        await message.answer(_humanize_error(e))
        return

    if instruction.strip():
        directive = (
            f"Foydalanuvchi shu fayl bilan so'radi: «{instruction.strip()}». Shuni bajar. "
            "Aniq muddat yoki majburiyat bo'lsa, create_task/create_reminder action taklif qil."
        )
    else:
        directive = (
            "Bu hujjatni executive uchun tahlil qil: qisqacha xulosa, muhim sanalar/muddatlar, "
            "majburiyatlar, summalar va tomonlar, hamda e'tibor talab qiladigan nuqtalar. Aniq "
            "muddat/majburiyat bo'lsa, create_task yoki create_reminder action sifatida taklif qil."
        )

    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    working = await message.answer("🔍 Hujjatni o'qiyapman…")
    try:
        response = await claude_service.process_document(
            directive, blocks, file_label=file_name or kind)
    except Exception as e:
        logger.exception("process_document failed")
        try:
            await working.edit_text(_humanize_error(e))
        except TelegramBadRequest:
            await message.answer(_humanize_error(e))
        return
    finally:
        typing_task.cancel()

    try:
        await working.delete()
    except TelegramBadRequest:
        pass

    summary = (response.get("user_message") or "").strip() or "📄 Hujjat o'qildi."
    if meta.get("truncated"):
        summary += "\n\n_(Hujjat uzun — boshlang'ich qismi tahlil qilindi.)_"

    # Proposed commitments — already listed in the summary's "➡️ Taklif" lines by
    # Claude; the action bar below acts on them.
    actions = [a for a in response.get("actions", [])
               if a.get("type") in _DESTRUCTIVE_ACTION_TYPES][:_MAX_CREATE_ACTIONS_PER_MSG]

    # Unified action bar — reuses the bot's existing concepts so it never drifts
    # from the rest of the UI: acts_confirm / acts_cancel (create / cancel), the
    # notes inbox (capture), and a revise-by-instruction flow (like polish "✏️ Tahrirla").
    rows = []
    if actions:
        try:
            prior_state = await state.get_state()
        except Exception:
            prior_state = None
        await state.set_state(CreateActionConfirmFSM.awaiting)
        await state.update_data(
            pending_response={"actions": actions,
                              "user_message": "✅ Hujjatdan vazifa(lar) qo'shildi.", "buttons": []},
            _prior_section=prior_state,
        )
        rows.append([
            InlineKeyboardButton(text="✅ Vazifa yarat", callback_data="acts_confirm"),
            InlineKeyboardButton(text="✏️ Tahrirla", callback_data=f"docedit:{file_unique_id}"),
        ])
        rows.append([
            InlineKeyboardButton(text="📝 Qaydga saqla", callback_data=f"docact:note:{file_unique_id}"),
            InlineKeyboardButton(text="✕ Bekor qil", callback_data="acts_cancel"),
        ])
    else:
        # Pure summary — nothing to create; offer revise + capture.
        rows.append([
            InlineKeyboardButton(text="✏️ Tahrirla", callback_data=f"docedit:{file_unique_id}"),
            InlineKeyboardButton(text="📝 Qaydga saqla", callback_data=f"docact:note:{file_unique_id}"),
        ])
    await _safe_answer(message, summary, parse_mode="Markdown",
                       reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


async def _run_task_import(message: Message, state: FSMContext) -> None:
    """Download an uploaded .xlsx/.csv/.pdf and import its tasks. Thin wrapper —
    parsing + preview + confirm lives in _import_tasks_from_file_bytes so the
    PDF→import flip button can reuse it."""
    doc = message.document
    if doc is None:
        await message.answer("📄 Import uchun Excel (.xlsx), CSV yoki PDF fayl yuboring.")
        return
    try:
        file = await message.bot.get_file(doc.file_id)
        blob = await message.bot.download_file(file.file_path)
        data = blob.read() if hasattr(blob, "read") else blob
    except Exception as e:
        await message.answer(_humanize_error(e))
        return
    await _import_tasks_from_file_bytes(message, state, data, doc.file_name or "",
                                        switch_uid=doc.file_unique_id)


def _read_task_sheet(wb) -> list:
    """Pick the worksheet holding the task table. Our export inserts a live stats
    dashboard ('Boshqaruv paneli') at sheet index 0, so wb.active is the DASHBOARD,
    not the tasks — reading .active yields zero task rows and silently routes every
    re-imported export through the LLM fallback (edits lost / duplicates). Prefer the
    named 'Vazifalar' sheet; else the first sheet whose header carries a recognizable
    title column; else fall back to the active sheet (a foreign, non-exported file).
    Each sheet is iterated once (read_only workbooks dislike re-iteration)."""
    def rows(ws):
        return [r for r in ws.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in r)]
    if "Vazifalar" in wb.sheetnames:
        return rows(wb["Vazifalar"])
    # Probe the ACTIVE sheet FIRST — it's the one the user was working on. Otherwise
    # a leading reference/lookup sheet with a generic header ('Nomi') would hijack
    # selection from the real task sheet. Header probe scans the top 20 non-blank
    # rows (some files carry preamble text above the header).
    active_title = wb.active.title if wb.active else None
    order = ([active_title] if active_title in wb.sheetnames else []) + \
            [nm for nm in wb.sheetnames if nm != active_title]
    active_rows = None
    for nm in order:
        tbl = rows(wb[nm])
        if nm == active_title:
            active_rows = tbl
        if any(_norm_header(c) in _COL_TITLE for row in tbl[:20] for c in row):
            return tbl
    return active_rows if active_rows is not None else []


async def _import_tasks_from_file_bytes(
    message: Message, state: FSMContext, data: bytes, name: str, switch_uid: str = ""
) -> None:
    """Parse a task table from raw bytes → preview → confirm → create. Reuses the
    standard acts_confirm pipeline so nothing is created silently. switch_uid (a
    cached file id) adds a one-tap '📄 Tahlil qil' switch under the preview."""
    import io
    name = (name or "").lower()

    # ── Read the file into a raw table (list of row tuples) ──
    try:
        if name.endswith(".csv"):
            import csv
            table = [tuple(r) for r in csv.reader(io.StringIO(data.decode("utf-8-sig", errors="replace")))
                     if any((c or "").strip() for c in r)]
        elif name.endswith(".pdf"):
            try:
                from pypdf import PdfReader
            except ImportError:
                await message.answer(
                    "📄 PDF importi bu yerda hali sozlanmagan. Excel (.xlsx) yoki CSV "
                    "yuboring.\n_(Server: `pip install -r requirements.txt`)_",
                    parse_mode="Markdown",
                )
                return
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
            # PDFs have no columns → each text line becomes a row; the smart
            # extractor (Claude) finds the tasks in whatever the layout is.
            table = [(ln.strip(),) for ln in text.splitlines() if ln.strip()]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            table = _read_task_sheet(wb)
    except Exception as e:
        await message.answer("📄 Faylni o'qib bo'lmadi.\n"
                             f"🔧 {type(e).__name__}: {str(e)[:100]}")
        return
    if not table:
        await message.answer("📄 Fayl bo'sh — vazifa topilmadi.")
        return

    # ── No fixed format required. Stage 1: recognizable columns (fast, free).
    #    Stage 2: if none, Claude reads the raw content and finds tasks in ANY
    #    layout (different headers, free text, mixed). ──
    actions = _structured_tasks_from_table(table)
    # Did the file carry the hidden ID column? Required for full-mirror sync — so we
    # can safely tell which DB tasks the file OMITTED (vs a from-scratch file, where
    # mirror would unsafely archive EVERYTHING).
    _had_ids = any((a.get("_id") or "").strip() for a in actions)
    # Round-trip dedup: a row whose ID (hidden export column) still exists becomes
    # an UPDATE — so editing the exported file and re-sending it changes the task
    # instead of creating a duplicate. Unknown/blank ID → a new task.
    for a in actions:
        rid = a.pop("_id", "")
        if not rid:
            continue
        _stored = await database.get_task(rid)
        if not _stored:
            continue
        a["type"] = "update_task"
        a["id"] = rid
        # Same instant (sub-second tolerance) → drop the deadline key so an
        # untouched cell can't fake a "change" and reset reminded_at.
        if _same_deadline_instant(a.get("data", {}).get("deadline"), _stored.get("deadline")):
            a["data"].pop("deadline", None)
    smart_note = ""
    if not actions:
        thinking = await message.answer("🔍 Fayldan vazifalarni topyapman…")
        actions = await _smart_tasks_from_table(table)
        try:
            await thinking.delete()
        except TelegramBadRequest:
            pass
        # The smart extractor only sees the first 120 rows / 6000 chars of the file.
        # If the input exceeded that, tail rows were never sent to the model — warn so
        # a partial extraction doesn't read as "everything was imported". Size is
        # summed incrementally (no giant throwaway string for a 10k-row file).
        _nonblank = 0
        _chars = 0
        for r in table:
            cells = [str(c).strip() for c in r if c is not None and str(c).strip()]
            if cells:
                _nonblank += 1
                _chars += sum(len(c) for c in cells) + 3 * len(cells)
        if _nonblank > 120 or _chars > 6000:
            smart_note = ("\n\n⚠️ Fayl katta — faqat boshidagi qism o'qildi. "
                          "Hammasi kirgan bo'lsa, qolganini alohida fayl bilan yuboring.")

    if not actions:
        await message.answer(
            "📄 Faylda vazifaga o'xshash ma'lumot topilmadi. Sarlavha, ijrochi yoki "
            "muddat bo'lgan qator/ustunlar bo'lsa — qayta yuboring."
        )
        return

    # ── Content dedup: a task whose title matches an existing ACTIVE one becomes
    #    an UPDATE — identical tasks are NEVER duplicated. Covers files without the
    #    hidden ID column AND the smart-extracted path (ID dedup ran above). ──
    _existing = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=2000)
    _by_title: dict = {}
    for _t in _existing:
        _k = (_t.get("title") or "").strip().lower()
        if _k:
            _by_title.setdefault(_k, _t["id"])
    _dedup = _apply_title_dedup(actions, _by_title)
    # Tag every imported action source="excel" so _execute_actions trusts it to
    # introduce new assignees/categories (LLM turns may only reuse existing ones).
    for _a in actions:
        _a.setdefault("data", {})["source"] = "excel"

    # ── Preview + confirm (reuse the standard acts_confirm pipeline) ──
    # No 20-per-message chat cap here: file imports are parsed deterministically, so
    # all rows are imported. Only a high safety backstop applies.
    overflow = ""
    if len(actions) > _MAX_IMPORT_TASKS:
        dropped = len(actions) - _MAX_IMPORT_TASKS
        actions = actions[:_MAX_IMPORT_TASKS]
        overflow = (f"\n\n⚠️ Juda katta import — {_MAX_IMPORT_TASKS} tasi olindi. "
                    f"Qolgan {dropped} tasini alohida fayl bilan yuboring.")
    overflow += smart_note

    # № hierarchy: a dotted child resolves its parent ONLY from a parent row in this
    # same file (see _resolve_parent). A child whose parent № is missing silently
    # becomes a top-level task — surface that so it isn't mistaken for a lost row.
    _orphans = _count_orphan_subtasks(actions)
    if _orphans:
        overflow += (f"\n\n⚠️ {_orphans} ta sub-vazifaning ota-№ faylda yo'q — "
                     "ular asosiy vazifa sifatida qo'shiladi.")
    _dl_unread_n = sum(1 for a in actions if a.get("_dl_unread"))
    if _dl_unread_n:
        overflow += (f"\n\n⚠️ {_dl_unread_n} ta muddat katagi o'qilmadi — o'sha "
                     "vazifalarda muddat O'ZGARTIRILMAYDI (format: 15-07-2026 yoki 15-07).")

    n_upd = sum(1 for a in actions if a.get("type") == "update_task")
    n_new = len(actions) - n_upd
    done_msg = "📥 Import: " + f"{n_new} yangi" + (f", {n_upd} yangilangan" if n_upd else "") + " vazifa."
    if _dedup.get("dropped"):
        done_msg += f"\n♻️ {_dedup['dropped']} ta takror (bir xil sarlavha) tashlandi."
    await state.set_state(CreateActionConfirmFSM.awaiting)
    await state.update_data(
        pending_response={"actions": actions, "user_message": done_msg, "buttons": []},
        _prior_section=None,
        _doc_uid=switch_uid,
        _import_had_ids=_had_ids,
    )
    # Paginated, scannable preview (one record = 3 lines) instead of a wall of text.
    total_pages = max(1, (len(actions) + _IMPORT_PAGE_SIZE - 1) // _IMPORT_PAGE_SIZE)
    text = _format_import_page(actions, 0) + overflow
    await _safe_answer(message, text, parse_mode="Markdown",
                       reply_markup=_import_preview_keyboard(0, total_pages, switch_uid, allow_mirror=True))


async def _ask_file_intent(message: Message, file_unique_id: str, label: str) -> None:
    """Ask what to do with a genuinely ambiguous file (a caption-less PDF —
    could be a contract to read or an exported task list to import)."""
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="📄 Tahlil", callback_data=f"docact:analyze:{file_unique_id}"),
        InlineKeyboardButton(text="📋 Import", callback_data=f"docact:import:{file_unique_id}"),
        InlineKeyboardButton(text="📝 Qayd", callback_data=f"docact:note:{file_unique_id}"),
    ]])
    await message.answer(f"📎 *{label}* — bu fayl bilan nima qilay?",
                         parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("docact:"))
async def cb_doc_action(query: CallbackQuery, state: FSMContext) -> None:
    """Route a cached file to analyse / import / note — fired by the ambiguous-PDF
    ask prompt or by a result's one-tap switch button."""
    try:
        _, action, uid = query.data.split(":", 2)
    except ValueError:
        await query.answer()
        return
    entry = _DOC_FLIP_CACHE.get(uid)
    if not entry:
        await query.answer("Fayl topilmadi — qayta yuboring.", show_alert=True)
        return
    await query.answer()
    try:  # strip the buttons so the prompt/switch can't be double-tapped
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass

    name = entry.get("name") or "fayl"
    kind = entry.get("kind") or document_service.detect_kind(name, entry.get("mime") or "")

    if action == "note":
        # Choosing "save as note" supersedes any pending task-confirm from the
        # analysis — clear it so a stray follow-up message isn't mis-handled.
        await state.clear()
        await _capture_file_as_note(query.message, label=name)
        return
    if action == "analyze":
        # _analyze_document downloads the file itself (by file_id).
        await _analyze_document(
            query.message, state, instruction="",
            file_id=entry["file_id"], file_unique_id=uid,
            file_name=name, mime=entry.get("mime") or "", kind=kind,
        )
        return
    # import → needs the raw bytes here.
    try:
        file = await query.message.bot.get_file(entry["file_id"])
        blob = await query.message.bot.download_file(file.file_path)
        data = blob.read() if hasattr(blob, "read") else blob
    except Exception as e:
        await query.message.answer(_humanize_error(e))
        return
    await _import_tasks_from_file_bytes(query.message, state, data, name, switch_uid=uid)


@router.callback_query(F.data.startswith("docedit:"))
async def cb_doc_edit(query: CallbackQuery, state: FSMContext) -> None:
    """✏️ Tahrirla — revise the document analysis via a follow-up instruction
    (same concept as the polish '✏️ Tahrirla'). Re-analyses the SAME cached file."""
    uid = query.data.split(":", 1)[1]
    if uid not in _DOC_FLIP_CACHE:
        await query.answer("Fayl keshi eskirdi — qayta yuboring.", show_alert=True)
        return
    await query.answer()
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await state.set_state(DocReviseFSM.awaiting)
    await state.update_data(revise_uid=uid)
    await query.message.answer(
        "✏️ **Qanday o'zgartiray?**\n\nKo'rsatma yuboring (matn yoki ovoz) — masalan: "
        "«summasini ham hisobla», «faqat 1-vazifani yarat», «deadline 10-iyun qil».",
        parse_mode="Markdown",
    )


@router.message(StateFilter(DocReviseFSM.awaiting), F.text | F.voice)
async def handle_doc_revision(message: Message, state: FSMContext) -> None:
    """Revision instruction for a document analysis — re-analyse the same cached
    file with that instruction (text, or transcribed voice)."""
    instr = await _get_text_or_transcribe(message, bot=message.bot)
    if instr is None:
        return
    data = await state.get_data()
    uid = data.get("revise_uid") or ""
    await state.clear()
    entry = _DOC_FLIP_CACHE.get(uid)
    if not entry:
        await message.answer("Hujjat keshi eskirdi — faylni qayta yuboring.")
        return
    kind = entry.get("kind") or document_service.detect_kind(
        entry.get("name") or "", entry.get("mime") or "")
    await _analyze_document(
        message, state, instruction=instr,
        file_id=entry["file_id"], file_unique_id=uid,
        file_name=entry.get("name") or "fayl", mime=entry.get("mime") or "", kind=kind,
    )


@router.message(Command("diagnostics"))
async def cmd_diagnostics(message: Message) -> None:
    """Bot health snapshot — DB size, recent LLM cost, scheduler state, iCloud.
    Useful when something feels broken and you want a single view of internals."""
    import os
    import claude_service
    DIVIDER = "━" * 20
    lines = ["🩺  **DIAGNOSTIKA**", "", DIVIDER, ""]

    # DB size
    try:
        db_bytes = os.path.getsize(config.DATABASE_PATH)
        db_kb = db_bytes / 1024
        lines.append(f"📦 DB: `{config.DATABASE_PATH}` — {db_kb:,.0f} KB")
    except OSError:
        lines.append(f"📦 DB: `{config.DATABASE_PATH}` — ❌ topilmadi")

    # Pending actions
    try:
        stuck = await database.list_stuck_pending_actions(stuck_after_minutes=5)
        lines.append(f"⏳ Stuck pending_actions: {len(stuck)}" if stuck else "⏳ Stuck pending_actions: 0 ✅")
    except Exception:
        lines.append("⏳ Stuck pending_actions: tekshirib bo'lmadi")

    # LLM cost + cache hit rate (last 7 days)
    try:
        breakdown = await database.llm_cost_breakdown(days=7)
        totals = breakdown["totals"]
        lines.append(
            f"💰 7-kunlik: ${totals['cost_usd']:.4f} · "
            f"{totals['calls']} chaqiruv · cache hit {totals['cache_hit_rate'] * 100:.0f}%"
        )
        if breakdown["by_model"]:
            lines.append("   Modellar:")
            for row in breakdown["by_model"][:5]:
                lines.append(f"     • {row['model']}: {row['calls']} (${float(row['cost_usd'] or 0):.4f})")
    except Exception as e:
        lines.append(f"💰 LLM cost: xato ({type(e).__name__})")

    # Circuit breaker
    if claude_service._circuit_is_open():
        remain = claude_service._circuit_open_until - __import__("time").time()
        lines.append(f"⚠️ Claude circuit OPEN ({remain:.0f}s qoldi)")
    else:
        lines.append("✅ Claude circuit closed")

    # Scheduler
    sched = scheduler_module.get_scheduler()
    if sched and sched.scheduler.running:
        jobs = sched.scheduler.get_jobs()
        lines.append(f"⏰ Scheduler: {len(jobs)} aktiv job")
    else:
        lines.append("⏰ Scheduler: ❌ ishlamayapti")

    # iCloud
    if config.ICLOUD_ENABLED:
        lines.append("☁️ iCloud: yoqilgan")
    else:
        lines.append("☁️ iCloud: o'chiq")

    # Redaction
    import redaction
    redaction_label = "o'chiq" if redaction.DISABLED else "yoqilgan"
    lines.append(f"🛡 Redaction: {redaction_label}")

    # Background tasks
    bg = len(_background_tasks)
    lines.append(f"🔧 Background tasks: {bg}")

    # Audit — so'nggi amallar (qayta ishlangan kiritmalar tarixi)
    try:
        recent = await database.list_recent_actions(limit=8)
        if recent:
            lines.extend(["", DIVIDER, "", "📜 **So'nggi amallar:**"])
            for r in recent:
                ts = _short_local_date(r.get("completed_at") or r.get("updated_at"))
                badge = "✅" if r.get("state") == "completed" else "❌"
                txt = (r.get("user_text") or "—").strip().replace("\n", " ")
                if len(txt) > 44:
                    txt = txt[:41] + "…"
                lines.append(f"  {badge} {ts} · {_escape_markdown(txt)}")
    except Exception:
        logger.debug("audit recent-actions section failed", exc_info=True)

    lines.extend(["", DIVIDER])
    await _safe_answer(message, "\n".join(lines), parse_mode="Markdown",
                       reply_markup=single_back_keyboard())


@router.message(Command("help"))
async def cmd_help(message: Message) -> None:
    await message.answer(
        "🤝 **Yordamchi — qo'llanma**\n\n"
        "**1. Boshqaruv**\n"
        "• `/cockpit` — Boshqaruv paneli: umumiy holat, top vazifalar, risklar va tavsiyalar.\n"
        "• `/plan` — murakkab vaziyat uchun aniq ish rejasi.\n\n"
        "**2. Vazifalar**\n"
        "• `/tasks` — aktiv, bugungi, muhim, o'tgan, bajarilgan va takroriy vazifalar.\n\n"
        "**3. Eslatmalar**\n"
        "• `/reminders` — alohida eslatmalar, snooze, takrorlash va bajarildi nazorati.\n\n"
        "**4. Notes (Inbox)**\n"
        "• `/notes` — qayta ishlanmagan qaydlar inbox'i (GTD uslubi).\n"
        "• `/qayd <matn>` — tezkor note qo'shish.\n"
        "• Boshqa chatdan xabarni forward qiling — avto note'ga aylanadi.\n"
        "• Voice: _\"qayd qil: ...\"_ — ovozdan ham mumkin.\n\n"
        "**5. Uchrashuvlar**\n"
        "• `/meetings` — uchrashuvlar, tayyorgarlik brifi va action itemlar.\n\n"
        "**6. Natijalar**\n"
        "• `/stats` — KPI, deadline, delegatsiya, meeting va bot auditi.\n"
        "• Weekly/monthly report statistikadagi tugmalar orqali ochiladi.\n\n"
        "**7. Tizim**\n"
        "• `/settings` — bildirishnomalar, eslatmalar va kalendar holati.\n"
        "• `/help` — ushbu qo'llanma.\n\n"
        "**Avtomatik ishlaydigan funksiyalar**\n"
        "• 08:00 — kunlik brifing\n"
        "• 18:00 — kun yakuni\n"
        "• Uchrashuvdan oldin — prep brief\n"
        "• Uchrashuvdan keyin — action item eslatmasi\n"
        "• Deadline yaqinlashsa — muhim vazifa eslatmasi",
        parse_mode="Markdown",
        reply_markup=main_reply_keyboard(),
    )


UZ_WEEKDAYS_FULL = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
UZ_MONTHS_FULL = ["yanvar", "fevral", "mart", "aprel", "may", "iyun",
                  "iyul", "avgust", "sentyabr", "oktyabr", "noyabr", "dekabr"]


def _fmt_dt_uz(dt, sep: str = " ") -> str:
    """Canonical absolute date+time with an UNAMBIGUOUS Uzbek month — '15-iyun 14:00'
    (numeric '15-03' can be misread as a day in a banking context). Matches the
    meeting/reminder chips so the same date renders the same across sections."""
    return f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]}{sep}{dt.strftime('%H:%M')}"


async def _build_briefing_text() -> str:
    """Deterministic daily operating briefing."""
    now = datetime.now(database.TZ)

    today_tasks = await database.list_today_tasks()
    done_today = await database.list_tasks_done_today()
    today_meetings = await database.list_today_meetings()
    overdue = await database.list_overdue_tasks()

    def _key(t):
        p = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}.get(t.get("priority", "P2"), 2)
        d = t.get("deadline") or "9999"
        return (p, d)

    today_tasks = sorted(today_tasks, key=_key)
    best_task = (overdue[:1] or today_tasks[:1] or [None])[0]
    urgent_count = sum(1 for t in today_tasks if t.get("priority") == "P0")
    missing_assignee = [t for t in today_tasks if not (t.get("assignee") or "").strip()]

    date_label_upper = f"{now.day}-{UZ_MONTHS_FULL[now.month - 1].upper()}"
    DIVIDER = "━" * 20

    # Self-improvement: pending proposal count (Phase 3) — computed once, surfaced in
    # BOTH the quiet-day and the full briefing when > 0.
    try:
        _pending_props = (await database.count_proposals_by_status()).get("new", 0)
    except Exception:
        _pending_props = 0
    _prop_line = (f"💡 **{_pending_props} ta yaxshilanish taklifi** — /improvements"
                  if _pending_props else None)

    if not today_tasks and not today_meetings and not overdue:
        _quiet = [
            f"🗓 **BUGUN · {date_label_upper}**",
            "",
            "Bugun uchun aktiv vazifa yoki uchrashuv yo'q.",
            "",
            "_Kun boshida 1-2 ta muhim vazifani rejalashtiring._",
        ]
        if _prop_line:
            _quiet += ["", _prop_line]
        return "\n".join(_quiet)

    def _muhimlik_emoji(priority: str) -> str:
        # ⚡ only for Shoshilinch (P0); 🔹 for everything else
        return "⚡" if priority == "P0" else "🔹"

    def _task_card(task: dict, prefix: str = "") -> list[str]:
        """One task card: title with priority badge + 3 detail lines.
        prefix: '' for the Eng muhim card; 'N. ' for numbered list items.
        Continuation lines are indented to match the prefix width.
        """
        title = (task.get("title") or "—").strip()
        priority = task.get("priority", "P2")
        badge = _PRIORITY_BADGE.get(priority, "⚪")
        muhimlik = _PRIORITY_LABEL_UZ.get(priority, "Rejadagi")
        assignee = (task.get("assignee") or "belgilanmagan").strip()
        muddat = _muddat_label(task.get("deadline"))
        indent = " " * len(prefix)
        return [
            f"{prefix}{badge} {title}",
            f"{indent}👤 Ijrochi: {assignee}",
            f"{indent}⏳ Muddat: {muddat}",
            f"{indent}{_muhimlik_emoji(priority)} Muhimlik: {muhimlik}",
        ]

    # Inbox count — surfaces "you have N notes waiting to be triaged" so the
    # principal doesn't forget to clear the GTD inbox.
    try:
        inbox_count = await database.count_notes_in_status("inbox")
    except Exception:
        inbox_count = 0

    lines: list[str] = [
        f"🗓 **BUGUN · {date_label_upper}**",
        "",
        "📌 **UMUMIY HOLAT**",
        "",
        f"**{len(today_tasks)}** ta vazifa  ·  **{len(done_today)}** ta yopildi",
        f"**{urgent_count}** ta shoshilinch  ·  **{len(overdue)}** ta muddati o'tgan",
    ]
    if inbox_count > 0:
        lines.append(f"📥 **{inbox_count}** ta note inbox'da kutmoqda")
    lines.append("")

    if best_task:
        lines.extend([DIVIDER, "", "⭐ **ENG MUHIM**", ""])
        lines.extend(_task_card(best_task))
        lines.append("")

    if today_tasks:
        lines.extend([DIVIDER, "", "📌 **VAZIFALAR**", ""])
        for i, task in enumerate(today_tasks[:8], 1):
            lines.extend(_task_card(task, prefix=f"{i}. "))
            lines.append("")
        if len(today_tasks) > 8:
            lines.append(f"_+{len(today_tasks) - 8} ta yana_")
            lines.append("")

    # UCHRASHUVLAR — bot uslubi: 4-qatorli kartochka (Variant A bilan mos)
    if today_meetings:
        lines.extend([DIVIDER, "", "🤝 **UCHRASHUVLAR**", ""])
        for i, meeting in enumerate(today_meetings[:3], 1):
            title = (meeting.get("title") or "—").strip()
            parts = meeting.get("participants") or []
            if not parts:
                plabel = "belgilanmagan"
            elif len(parts) <= 3:
                plabel = ", ".join(parts)
            else:
                plabel = f"{', '.join(parts[:3])} (+{len(parts) - 3} nafar)"
            location = (meeting.get("location_or_link") or "").strip() or "belgilanmagan"
            time_label = _meeting_time_label(meeting.get("datetime_start") or "")
            lines.extend([
                f"{i}.  {title}",
                "",
                f"      ⏰ Vaqt:             {time_label}",
                f"      👥 Ishtirokchilar:   {plabel}",
                f"      📍 Manzil:           {location}",
                "",
            ])
        if len(today_meetings) > 3:
            lines.append(f"_+{len(today_meetings) - 3} ta yana_")
            lines.append("")

    # TAVSIYALAR — 3 buckets: HOZIR / BUGUN / KUN OXIRI
    if best_task and best_task.get("priority") == "P0":
        name = (best_task.get("assignee") or "").strip()
        if name:
            hozir = f"{name} ijrosidagi shoshilinch vazifa bo'yicha status olish."
        else:
            urgent_title = _truncate((best_task.get("title") or "—").strip(), 50)
            hozir = f"Shoshilinch vazifaga ijrochi tayinlash: «{urgent_title}»."
    elif overdue:
        first_title = _truncate((overdue[0].get("title") or "—").strip(), 50)
        hozir = f"Eng kechikkan vazifani («{first_title}») yopish yoki yangi muddat belgilash."
    else:
        hozir = "Eng yuqori ustuvorlikdagi vazifani boshlash."

    if overdue:
        bugun_rec = (f"Muddati o'tgan {len(overdue)} ta vazifani yopish "
                     "yoki yangi aniq muddat belgilash.")
    elif missing_assignee:
        bugun_rec = (f"Ijrochisi belgilanmagan {len(missing_assignee)} ta vazifaga "
                     "mas'ul tayinlash.")
    else:
        bugun_rec = f"Bugungi {len(today_tasks)} ta vazifani ketma-ket yopib borish."

    kun_oxiri = ("Yopilgan, qolgan va kechikayotgan vazifalar bo'yicha "
                 "qisqa qayta hisobot chiqarish.")

    # Self-improvement: pending proposal line (Phase 3) — _prop_line computed above.
    if _prop_line:
        lines += ["", _prop_line]

    lines.extend([
        DIVIDER, "",
        "💡 **TAVSIYALAR**", "",
        "🔴 **HOZIR**",
        hozir, "",
        "⏳ **BUGUN**",
        bugun_rec, "",
        "🧾 **KUN OXIRI**",
        kun_oxiri,
    ])

    return "\n".join(lines).rstrip()


def _parse_dt_safe(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        return dt.astimezone(database.TZ) if dt.tzinfo else database.TZ.localize(dt)
    except (TypeError, ValueError):
        return None





def _today_inline_keyboard(today_tasks: list[dict] | None = None) -> InlineKeyboardMarkup | None:
    """Bugun ekrani uchun INLINE klaviatura — faqat raqamli drill-down.

    Quick actions (➕ Yangi vazifa, 🌙 Kechki yakun, 📌 Vazifalar,
    🤝 Uchrashuvlar) endi pastdagi section reply kbd da bor —
    inline'dan olib tashlandi.
    """
    if not today_tasks:
        return None
    rows: list[list[InlineKeyboardButton]] = []
    nums = [
        InlineKeyboardButton(text=str(i + 1), callback_data=f"taskopen:{t['id']}")
        for i, t in enumerate(today_tasks[:8])
    ]
    for i in range(0, len(nums), 5):
        rows.append(nums[i:i + 5])
    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None



@router.message(Command("today"))
async def cmd_today(message: Message, state: FSMContext | None = None) -> None:
    if state is not None:
        await state.set_state(SectionFSM.in_today)
    # Bo'lim header + section reply kbd
    await message.answer(
        "📅 **BUGUN**", parse_mode="Markdown",
        reply_markup=today_section_reply_keyboard(),
    )
    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        text = await _build_briefing_text()
        today_tasks = await database.list_today_tasks()
    finally:
        typing_task.cancel()

    def _key(t):
        p = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}.get(t.get("priority", "P2"), 2)
        d = t.get("deadline") or "9999"
        return (p, d)

    sorted_today = sorted(today_tasks, key=_key)
    # Inline keyboard: faqat raqamli drill-down + Orqaga
    # (quick actions reply kbd ga ko'chdi)
    keyboard = _today_inline_keyboard(today_tasks=sorted_today)
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=keyboard)


@router.message(Command("briefing"))
async def cmd_briefing(message: Message, state: FSMContext | None = None) -> None:
    await cmd_today(message, state)



_PRIORITY_BADGE = {"P0": "🔴", "P1": "🟠", "P2": "🔵", "P3": "⚪"}


def _task_badge(task: dict) -> str:
    """Single source of truth for a task's leading status dot — used by the list,
    the drill-down card, AND the detail card so the SAME task never shows two
    different dots. Urgency-aware: done > blocked > overdue/P0 > P1 > today > routine."""
    if task.get("status") == "done":
        return "✅"
    if task.get("status") == "blocked":
        return "🚧"  # stuck — needs unblocking
    priority = task.get("priority", "P2")
    deadline = task.get("deadline")
    is_overdue = is_today = False
    if deadline:
        try:
            dt = datetime.fromisoformat(deadline).astimezone(database.TZ)
            now = datetime.now(database.TZ)
            is_overdue = dt < now
            is_today = (not is_overdue) and dt.date() == now.date()
        except (ValueError, TypeError):
            pass
    if is_overdue or priority == "P0":
        return "🔴"
    if priority == "P1":
        return "🟠"
    if is_today:
        return "🟡"
    return "⚪"


# User-facing priority names — replace technical P0/P1/P2/P3 in ALL displays.
_PRIORITY_LABEL_UZ = {
    "P0": "Shoshilinch",
    "P1": "Muhim",
    "P2": "Rejadagi",
    "P3": "Past ustuvorlik",
}


def _muddat_label(iso) -> str:
    """Sodda 'Muddat:' qiymati — list view'lar uchun.
    - overdue → "o'tgan"
    - bugun   → "Bugun HH:MM"
    - ertaga  → "Ertaga HH:MM"
    - boshqa  → "DD-MM, HH:MM"
    - yo'q    → "belgilanmagan"
    """
    if not iso:
        return "belgilanmagan"
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return str(iso)
    now = datetime.now(database.TZ)
    if dt < now:
        return "o'tgan"
    if now.date() == dt.date():
        return f"Bugun {dt.strftime('%H:%M')}"
    if (now + timedelta(days=1)).date() == dt.date():
        return f"Ertaga {dt.strftime('%H:%M')}"
    return dt.strftime("%d-%m, %H:%M")
_STATUS_LABEL_UZ = {"todo": "Aktiv", "in_progress": "Jarayonda",
                    "blocked": "Toʻsilgan", "done": "Bajarildi", "cancelled": "Bekor qilingan"}
_STATUS_EMOJI = {"todo": "⏳", "in_progress": "🔄", "blocked": "⚠️", "done": "✅", "cancelled": "❌"}
_NUMBER_GLYPH = {1: "①", 2: "②", 3: "③", 4: "④", 5: "⑤", 6: "⑥", 7: "⑦", 8: "⑧", 9: "⑨", 10: "⑩"}
_SEP = "━" * 25



async def compute_risk_score() -> dict:
    """Compute a 0-100 risk score from current DB state.

    Components (each contributes 0-25 points, capped):
      • Overdue tasks (count × 5, max 25)
      • Tasks due within 24h × 4, max 20
      • Unassigned + due within 48h × 6, max 18
      • Tasks without deadline × 1.5, max 12
      • Urgent (P0) not done × 5, max 25

    Returns: {score: int, status: str, components: dict}
    """
    # Single SELECT with conditional aggregates — was 6 separate queries
    # opening 6 connections (N+1 anti-pattern). See database.risk_score_counts.
    counts = await database.risk_score_counts()

    components = {
        "overdue": min(25, counts["overdue"] * 5),
        "due_24h": min(20, counts["due_24h"] * 4),
        "unassigned_urgent": min(18, counts["unassigned_due_48h"] * 6),
        "no_deadline": min(12, int(counts["no_deadline"] * 1.5)),
        "urgent_open": min(25, counts["urgent_open"] * 5),
    }
    score = min(100, sum(components.values()))
    if score <= 30:
        status = "Past risk"
        emoji = "🟢"
    elif score <= 60:
        status = "Nazoratda"
        emoji = "🟡"
    elif score <= 80:
        status = "Yuqori risk"
        emoji = "🟠"
    else:
        status = "Kritik risk"
        emoji = "🔴"
    return {
        "score": score,
        "status": status,
        "emoji": emoji,
        "components": components,
        "counts": {
            "overdue": counts["overdue"],
            "due_24h": counts["due_24h"],
            "due_48h": counts["due_48h"],
            "no_deadline": counts["no_deadline"],
            "unassigned": counts["unassigned"],
            "urgent_open": counts["urgent_open"],
        },
    }


def _format_deadline_short(iso) -> tuple[str, bool]:
    """Return (human_label, is_overdue) — bugun/ertaga/o'tgan kabi tabiiy ko'rinish."""
    if not iso:
        return "Deadline yoʻq", False
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return str(iso), False
    now = datetime.now(database.TZ)
    overdue = dt < now
    today = now.date() == dt.date()
    tomorrow = (now + timedelta(days=1)).date() == dt.date()
    time_str = dt.strftime("%H:%M")
    if today:
        return f"Bugun {time_str}" + (" · MUDDATI OʻTDI" if overdue else ""), overdue
    if tomorrow:
        return f"Ertaga {time_str}", overdue
    label = _fmt_dt_uz(dt, sep=", ")
    if overdue:
        label += " · oʻtgan"
    return label, overdue


# ─────────────────────── SETTINGS ───────────────────────


def _format_settings_summary(settings: dict) -> str:
    """Human-readable settings overview shown when entering /settings."""
    voice_status = (
        "AVTO - tasdiqsiz"
        if settings.get("voice_auto_confirm", True)
        else "tasdiq so'raladi"
    )
    create_confirm_status = (
        "tasdiq so'raladi"
        if settings.get("confirm_create_actions", True)
        else "tasdiqsiz yaratiladi"
    )
    return (
        "⚙️ **SOZLAMALAR**\n\n"
        f"🔔 Bildirishnomalar: {'yoqilgan' if settings.get('notifications_enabled', True) else 'oʻchirilgan'}\n"
        f"⏰ Ertalab brifing: `{settings.get('morning_briefing_time', '09:00')}`\n"
        f"🌙 Kechki yakun: `{settings.get('evening_summary_time', '18:00')}`\n"
        f"📞 Uchrashuv eslatmasi: `{settings.get('meeting_reminder_min', 15)} daq oldin`\n"
        f"📌 Vazifa eslatmasi: `{settings.get('task_reminder_hours', 2)} soat oldin`\n"
        f"🎙 Ovoz transkripti: `{voice_status}`\n"
        f"✅ Vazifa/uchrashuv yaratish: `{create_confirm_status}`\n\n"
        "_Pastdagi tugmalardan parametr tanlang._"
    )


@router.message(Command("settings"))
async def cmd_settings(message: Message, state: FSMContext | None = None) -> None:
    if state is not None:
        await state.set_state(SectionFSM.in_settings)
    settings = await database.get_settings()
    text = _format_settings_summary(settings)
    await message.answer(text, parse_mode="Markdown",
                         reply_markup=settings_section_reply_keyboard())


@router.callback_query(F.data == "setting:notifications_toggle")
async def cb_setting_notif(query: CallbackQuery) -> None:
    settings = await database.get_settings()
    new_val = not settings["notifications_enabled"]
    await database.set_setting("notifications_enabled", new_val)
    await query.answer(f"Bildirishnomalar {'yoqildi' if new_val else 'oʻchirildi'} ✅")
    settings["notifications_enabled"] = new_val
    try:
        await query.message.edit_reply_markup(reply_markup=settings_keyboard(settings))
    except Exception:
        pass


@router.callback_query(F.data == "setting:briefing_time")
async def cb_setting_briefing_time(query: CallbackQuery) -> None:
    await query.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t, callback_data=f"brieftime:{t}") for t in ("07:00", "07:30", "08:00")],
        [InlineKeyboardButton(text=t, callback_data=f"brieftime:{t}") for t in ("08:30", "09:00", "10:00")],
        [back_button("nav_settings")],
    ])
    await query.message.answer("Ertalab brifing vaqti:", reply_markup=kb)


@router.callback_query(F.data.startswith("brieftime:"))
async def cb_brief_time(query: CallbackQuery) -> None:
    new_time = query.data.removeprefix("brieftime:")
    await database.set_setting("morning_briefing_time", new_time)
    await _reschedule_briefings_live()
    await query.answer(f"Brifing: {new_time} ✅ (kuchga kirdi)")
    try:
        await query.message.delete()
    except Exception:
        pass
    await cmd_settings(query.message)


@router.callback_query(F.data == "setting:evening_time")
async def cb_setting_evening_time(query: CallbackQuery) -> None:
    await query.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=t, callback_data=f"eveningtime:{t}") for t in ("17:00", "17:30", "18:00")],
        [InlineKeyboardButton(text=t, callback_data=f"eveningtime:{t}") for t in ("18:30", "19:00", "20:00")],
        [back_button("nav_settings")],
    ])
    await query.message.answer("Kechki yakun vaqti:", reply_markup=kb)


@router.callback_query(F.data.startswith("eveningtime:"))
async def cb_evening_time(query: CallbackQuery) -> None:
    new_time = query.data.removeprefix("eveningtime:")
    await database.set_setting("evening_summary_time", new_time)
    await _reschedule_briefings_live()
    await query.answer(f"Kechki yakun: {new_time} ✅ (kuchga kirdi)")
    try:
        await query.message.delete()
    except Exception:
        pass
    await cmd_settings(query.message)


async def _reschedule_briefings_live() -> None:
    """Ask the running scheduler to reload briefing times from settings."""
    sched = scheduler_module.get_scheduler()
    if sched is not None:
        try:
            await sched.apply_briefing_settings()
        except Exception:
            logger.exception("Failed to apply briefing settings live")


@router.callback_query(F.data == "setting:reminders")
async def cb_setting_reminders(query: CallbackQuery) -> None:
    settings = await database.get_settings()
    text = (
        "📲 **Eslatma parametrlari**\n\n"
        f"• Uchrashuv eslatmasi: `{settings['meeting_reminder_min']} daq` oldin\n"
        f"• Vazifa eslatmasi: `{settings['task_reminder_hours']} soat` oldin (Shoshilinch va Muhim uchun)\n\n"
        "_Standart qiymatlarni saqlash tavsiya etiladi._"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="15 daq", callback_data="meetremind:15"),
         InlineKeyboardButton(text="30 daq", callback_data="meetremind:30"),
         InlineKeyboardButton(text="60 daq", callback_data="meetremind:60")],
        [InlineKeyboardButton(text="1 soat", callback_data="taskremind:1"),
         InlineKeyboardButton(text="2 soat", callback_data="taskremind:2"),
         InlineKeyboardButton(text="4 soat", callback_data="taskremind:4")],
        [back_button("nav_settings")],
    ])
    await query.answer()
    await query.message.answer(text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("meetremind:"))
async def cb_meet_remind(query: CallbackQuery) -> None:
    mins = _cb_int(query.data, default=15)
    if mins <= 0 or mins > 24 * 60:
        await query.answer("Noto'g'ri qiymat")
        return
    await database.set_setting("meeting_reminder_min", mins)
    await _apply_reminder_settings_live()
    await query.answer(f"Uchrashuv eslatmasi: {mins} daq oldin ✅")


@router.callback_query(F.data.startswith("taskremind:"))
async def cb_task_remind(query: CallbackQuery) -> None:
    hrs = _cb_int(query.data, default=2)
    if hrs <= 0 or hrs > 168:
        await query.answer("Noto'g'ri qiymat")
        return
    await database.set_setting("task_reminder_hours", hrs)
    await _apply_reminder_settings_live()
    await query.answer(f"Vazifa eslatmasi: {hrs} soat oldin ✅")


async def _apply_reminder_settings_live() -> None:
    """Ask the running scheduler to reload reminder lead times from settings."""
    sched = scheduler_module.get_scheduler()
    if sched is not None:
        try:
            await sched.apply_reminder_settings()
        except Exception:
            logger.exception("Failed to apply reminder settings live")


@router.callback_query(F.data == "setting:calendar")
async def cb_setting_calendar(query: CallbackQuery) -> None:
    await query.answer()
    await cmd_calendar(query.message)


@router.callback_query(F.data == "setting:voice_auto_toggle")
async def cb_setting_voice_auto_toggle(query: CallbackQuery) -> None:
    """Flip the voice_auto_confirm setting between AUTO and confirm-prompt mode."""
    settings = await database.get_settings()
    new_val = not settings.get("voice_auto_confirm", True)
    await database.set_setting("voice_auto_confirm", new_val)
    label = "AVTO — tasdiqsiz" if new_val else "Tasdiq so'raladi"
    await query.answer(f"Ovoz: {label} ✅")
    settings["voice_auto_confirm"] = new_val
    try:
        await query.message.edit_reply_markup(reply_markup=settings_keyboard(settings))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "setting:confirm_create_toggle")
async def cb_setting_confirm_create_toggle(query: CallbackQuery) -> None:
    """Flip the confirm_create_actions setting. When ON (default), the bot
    asks for confirmation before creating any task or meeting."""
    settings = await database.get_settings()
    new_val = not settings.get("confirm_create_actions", True)
    await database.set_setting("confirm_create_actions", new_val)
    label = "yoqildi (xavfsizroq)" if new_val else "o'chirildi (tezroq)"
    await query.answer(f"Yaratish tasdig'i {label} ✅")
    settings["confirm_create_actions"] = new_val
    try:
        await query.message.edit_reply_markup(reply_markup=settings_keyboard(settings))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "setting:quiet_hours")
async def cb_setting_quiet_hours(query: CallbackQuery) -> None:
    """Toggle quiet hours on/off + show the time window picker."""
    settings = await database.get_settings()
    qh_on = settings.get("quiet_hours_enabled", False)
    qh_start = settings.get("quiet_hours_start", "22:00")
    qh_end = settings.get("quiet_hours_end", "07:00")
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=("🔕 Sukunatni O'CHIRISH" if qh_on else "🌙 Sukunatni YOQISH"),
            callback_data="quiet:toggle",
        )],
        [InlineKeyboardButton(text=f"⏰ Boshlanish: {qh_start}", callback_data="quiet:start")],
        [InlineKeyboardButton(text=f"⏰ Tugash: {qh_end}", callback_data="quiet:end")],
        [back_button("nav_settings")],
    ])
    state_label = "yoqilgan" if qh_on else "o'chiq"
    await query.answer()
    await query.message.answer(
        f"🌙 **Sukunat soatlari**\n\n"
        f"Holat: **{state_label}**\n"
        f"Vaqt: `{qh_start}` → `{qh_end}`\n\n"
        f"_Sukunat ichida brifing va eslatmalar yo'naltirilmaydi. "
        f"Faqat /diagnostics, /cancel va sizning xabarlaringizga javob qaytadi._",
        parse_mode="Markdown",
        reply_markup=kb,
    )


@router.callback_query(F.data == "quiet:toggle")
async def cb_quiet_toggle(query: CallbackQuery) -> None:
    settings = await database.get_settings()
    new_val = not settings.get("quiet_hours_enabled", False)
    await database.set_setting("quiet_hours_enabled", new_val)
    label = "yoqildi" if new_val else "o'chirildi"
    await query.answer(f"Sukunat {label} ✅")
    try:
        await query.message.delete()
    except TelegramBadRequest:
        pass
    await cb_setting_quiet_hours(query)


@router.callback_query(F.data.in_({"quiet:start", "quiet:end"}))
async def cb_quiet_time_picker(query: CallbackQuery) -> None:
    which = "start" if query.data == "quiet:start" else "end"
    label = "Boshlanish" if which == "start" else "Tugash"
    presets = ["19:00", "20:00", "21:00", "22:00", "23:00"] if which == "start" else \
              ["06:00", "07:00", "08:00", "09:00", "10:00"]
    rows = [
        [InlineKeyboardButton(text=t, callback_data=f"qtime:{which}:{t}") for t in presets[:3]],
        [InlineKeyboardButton(text=t, callback_data=f"qtime:{which}:{t}") for t in presets[3:]],
        [back_button("nav_settings")],
    ]
    await query.answer()
    await query.message.answer(
        f"⏰ Sukunat {label} vaqti:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data.startswith("qtime:"))
async def cb_quiet_set_time(query: CallbackQuery) -> None:
    parts = query.data.split(":", 3)  # qtime:start:HH:MM
    if len(parts) < 4:
        await query.answer()
        return
    which, hh, mm = parts[1], parts[2], parts[3]
    new_time = f"{hh}:{mm}"
    key = "quiet_hours_start" if which == "start" else "quiet_hours_end"
    await database.set_setting(key, new_time)
    await query.answer(f"Saqlandi: {new_time} ✅")
    try:
        await query.message.delete()
    except TelegramBadRequest:
        pass
    await cb_setting_quiet_hours(query)


# ─────────────────────── YANGI SUBMENU ───────────────────────


async def cmd_new(message: Message, state: FSMContext | None = None) -> None:
    """Yangi menu — section reply kbd ko'rinishida.
    Inline tugmalar olib tashlandi: reply kbd 5 ta turini taqdim etadi.
    """
    if state is not None:
        await state.set_state(SectionFSM.in_new)
    text = (
        "➕ **YANGI**\n\n"
        "Quyidagi turlardan birini tanlang:"
    )
    await message.answer(text, parse_mode="Markdown",
                         reply_markup=new_section_reply_keyboard())


@router.callback_query(F.data.startswith("new:"))
async def cb_new_type(query: CallbackQuery, state: FSMContext) -> None:
    kind = query.data.split(":", 1)[1]
    await query.answer()

    # 📝 Yangi vazifa → offer two paths: 🤖 Tezkor (natural) or 📋 Forma (guided FSM)
    if kind == "task":
        text = (
            "📝 **YANGI VAZIFA**\n" + _SEP + "\n\n"
            "Qaysi usul orqali yaratasiz?\n\n"
            "• 🤖 **Tezkor** — matn yoki ovoz yuboring, men tushunib yarataman.\n"
            "• 📝 **Forma** — bosqichma-bosqich (sarlavha → muddat → ijrochi)."
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="🤖 Tezkor", callback_data="new:task_quick"),
                InlineKeyboardButton(text="📝 Forma", callback_data="new:task_form"),
            ],
            [back_button("nav_new")],
        ])
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)
        return

    if kind == "task_form":
        await _newtask_start(query.message, state)
        return

    if kind == "task_quick":
        prompt = ("📝 **Yangi vazifa — tezkor**\n\nMatn yoki ovoz yuboring. Misol:\n"
                   "_\"Ertaga ertalab Aziz akaga marketing hisobotini yuborish\"_")
        await _safe_answer(query.message, prompt, parse_mode="Markdown",
                            reply_markup=single_back_keyboard("nav_new"))
        return

    if kind == "meeting":
        text = (
            "➕ **YANGI UCHRASHUV**\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            "Uchrashuv haqida ma'lumot yuboring:\n\n"
            "Sarlavha, ishtirokchi, sana/vaqt, manzilni bir matn\n"
            "yoki ovoz xabarda yuborishingiz mumkin.\n\n"
            "_Misollar:_\n"
            "_• \"Ertaga soat 12:00 da Dinislam bilan biznes forum\"_\n"
            "_• \"Juma 15:00 da Olim aka bilan byudjet uchrashuvi\"_\n"
            "_• \"28-may ertalab 10:00 da jamoa stand-up\"_\n\n"
            "Bot ma'lumotni tushunib, iCloud kalendariga ham\n"
            "avtomatik sinxronlaydi."
        )
        await _safe_answer(
            query.message, text, parse_mode="Markdown",
            reply_markup=single_back_keyboard("nav_new", text="✕ Bekor"),
        )
        return

    if kind == "reminder":
        await _newreminder_start(query.message, state)
        return

    prompts = {
        "voice": ("🎙 **Ovozdan vazifa**\n\nMikrofon tugmasini bosib o'zbekcha gapiring. "
                  "Men transkripsiya qilib, vazifani tushunaman."),
        "polish": ("✏️ **Matn tahrirlash**\n\nXabaringizni yuboring va aytib qo'ying — "
                   "kimga, qanday tonda. Misol:\n_\"Aziz akaga rasmiy qil: ertaga hisobot tayyor\"_"),
    }
    await _safe_answer(
        query.message,
        prompts.get(kind, "Matn yoki ovoz yuboring."),
        parse_mode="Markdown",
        reply_markup=single_back_keyboard("nav_new"),
    )


def _format_task_card(t: dict, idx: int = None, show_status: bool = True) -> str:
    """Compact task card for the main task drill-down screen."""
    badge = _task_badge(t)  # unified with the list — same dot for the same task
    priority_name = _PRIORITY_LABEL_UZ.get(t.get("priority", "P2"), "Rejadagi")
    status = t.get("status", "todo")
    status_uz = _STATUS_LABEL_UZ.get(status, status)
    deadline_label, _overdue = _format_deadline_short(t.get("deadline"))
    if not t.get("deadline"):
        deadline_label = "belgilanmagan"
    title = (t.get("title") or "—").strip()
    assignee = (t.get("assignee") or "").strip() or "belgilanmagan"

    num_prefix = f"{idx}. " if idx is not None else ""
    meta = [f"👤 {assignee}", f"⚡ {priority_name}"]
    if show_status and status != "todo":
        meta.append(status_uz)
    out = [
        f"{badge} {num_prefix}**{title}**",
        "",
        " · ".join(meta),
        f"⏰ Muddat: {deadline_label}",
    ]
    return "\n".join(out)


def _format_task_detail_card(t: dict, idx: int = None) -> str:
    """Full task card for ⋯ Batafsil."""
    badge = _task_badge(t)  # unified with the list — same dot for the same task
    priority_name = _PRIORITY_LABEL_UZ.get(t.get("priority", "P2"), "Rejadagi")
    status = t.get("status", "todo")
    status_uz = _STATUS_LABEL_UZ.get(status, status)
    status_emoji = _STATUS_EMOJI.get(status, "•")
    deadline_label, _overdue = _format_deadline_short(t.get("deadline"))
    if not t.get("deadline"):
        deadline_label = "belgilanmagan"
    title = (t.get("title") or "—").strip()
    assignee = (t.get("assignee") or "").strip() or "belgilanmagan"
    description = (t.get("description") or "").strip()
    tags = t.get("tags") or []
    num_prefix = f"{idx}. " if idx is not None else ""

    lines = [
        f"{badge} {num_prefix}**{title}**",
        "",
        f"👤 Ijrochi: {assignee}",
        f"⏰ Muddat: {deadline_label}",
        f"⚡ Ustuvorlik: {priority_name}",
        f"{status_emoji} Holat: {status_uz}",
    ]
    if (t.get("category") or "").strip():
        lines.append(f"📁 Kategoriya: {t.get('category').strip()}")
    if description:
        lines.append(f"📝 Tavsif: {description}")
    if tags:
        lines.append(f"🏷 Teglar: {', '.join(tags)}")
    if t.get("recurrence_rule"):
        lines.append(f"🔁 Takroriy: {_format_recurrence_label(t.get('recurrence_rule'))}")
    return "\n".join(lines)


_MAX_TITLE_COMPACT = 50  # mobil ekranga sig'ish uchun


def _truncate(s: str, n: int = _MAX_TITLE_COMPACT) -> str:
    if not s:
        return "—"
    s = s.strip()
    if len(s) <= n:
        return s
    return s[: n - 1].rstrip() + "…"


def _format_recurrence_label(rule: str | None) -> str:
    labels = {
        "daily": "har kuni",
        "weekdays": "ish kunlari (Dush–Juma)",
        "weekly": "har hafta",
        "monthly": "har oy",
        "quarterly": "har chorak",
        "yearly": "har yil",
    }
    return labels.get(rule or "", rule or "—")



_TASKS_PER_PAGE = 10




def _task_deadline_chip(task: dict) -> str:
    """Compact deadline label for list view."""
    if task.get("status") == "done":
        return "Bajarilgan"
    deadline = task.get("deadline")
    if not deadline:
        return "Muddatsiz"
    try:
        dt = datetime.fromisoformat(deadline).astimezone(database.TZ)
    except (ValueError, TypeError):
        return str(deadline)
    now = datetime.now(database.TZ)
    if dt < now:
        return "O'tgan"
    if dt.date() == now.date():
        return f"Bugun {dt.strftime('%H:%M')}"
    if (now + timedelta(days=1)).date() == dt.date():
        return f"Ertaga {dt.strftime('%H:%M')}"
    return _fmt_dt_uz(dt)


def _format_tasks_compact(
    tasks: list[dict],
    label: str,
    show_status: bool = False,
    stats: dict | None = None,
    page: int = 1,
) -> str:
    """Tasks screen — block-style cards with professional spacing.

    Icon palette (from prior design):
      📋 page · 📌 stats · ⏳ unfinished · ✅ done
      Per-task badge: 🔴 overdue/urgent · 🟠 important · 🟡 today · ⚪ routine · ✅ done
      Detail line icons: 👤 ijrochi · ⏳ muddat · ⚡/⭐/🔹/✅ muhimlik

    Spacing rules: 2 blank lines around dividers, 1 blank between items,
    1 blank between task title and its detail block, 6-space indent for details.
    """
    del show_status
    DIVIDER = "━" * 20

    per_page = _TASKS_PER_PAGE
    total = len(tasks)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_tasks = tasks[start : start + per_page]

    lines: list[str] = [
        "📌  **VAZIFALAR**",
        "",
        f"Ko'rinish · {label}",
        f"Natija · {total} ta",
    ]
    if total_pages > 1:
        lines.append(f"Sahifa · {page} / {total_pages}")
    # Diqqat banner — surface overdue/urgent at the very top so they aren't buried
    # below higher-priority-but-not-urgent items. Uses the global overview counts.
    if total > 0 and stats:
        _alert = []
        if stats.get("overdue"):
            _alert.append(f"{stats['overdue']} muddati o'tgan")
        if stats.get("urgent"):
            _alert.append(f"{stats['urgent']} shoshilinch")
        if _alert:
            lines.append(f"⚠️  **Diqqat:** {' · '.join(_alert)}")

    if total == 0:
        lines.extend([
            "",
            DIVIDER,
            "",
            "Bu bo'limda hozircha vazifa yo'q.",
            "",
            "_Yangi vazifa qo'shish yoki boshqa filter tanlash mumkin._",
        ])
        return "\n".join(lines)

    lines.extend(["", DIVIDER, ""])

    # UMUMIY HOLAT — two stat rows, no blank between them (related data)
    if stats:
        lines.extend([
            "📌  **UMUMIY HOLAT**",
            "",
            f"Jami {stats['total']}   ·   Aktiv {stats['active']}   ·   Bajarilgan {stats['done']}",
            f"Shoshilinch {stats['urgent']}   ·   Muhim {stats['important']}   ·   O'tgan {stats['overdue']}   ·   🚧 To'silgan {stats.get('blocked', 0)}",
            "",
            DIVIDER,
            "",
        ])

    unfinished = [t for t in page_tasks if t.get("status") != "done"]
    done = [t for t in page_tasks if t.get("status") == "done"]

    def _muhimlik_icon(priority: str) -> str:
        return {"P0": "⚡", "P1": "⭐", "P2": "🔹", "P3": "🔹"}.get(priority, "🔹")

    def _task_card_lines(task: dict, num: int) -> list[str]:
        """Render one unfinished task as a labeled card: badge+title, blank, then
        aligned 👤 Ijrochi / ⏳ Muddat / Muhimlik lines, plus 📝 Izoh when present.
        The leading badge signals urgency at a glance (🔴/🟠/🟡/⚪/🚧)."""
        title = (task.get("title") or "—").strip()
        badge = _task_badge(task)
        assignee = ((task.get("assignee") or "Belgilanmagan").strip() or "Belgilanmagan")
        assignee = assignee[0].upper() + assignee[1:]
        deadline = _task_deadline_chip(task)  # smart: "Bugun 17:00" / "08-06 17:00"
        muhimlik_name = _PRIORITY_LABEL_UZ.get(task.get("priority", "P2"), "Rejadagi")
        muhimlik_emoji = _muhimlik_icon(task.get("priority", "P2"))
        # Pad labels to a stable value column. Labels: Ijrochi:(8) Muddat:(7)
        # Muhimlik:(9) Izoh:(5) → value starts 13 chars after the label text.
        card = [
            f"{num}.  {badge}  {title}",
            "",  # blank after title — lets the title stand out as the card header
            f"      👤  Ijrochi:     {assignee}",
            f"      ⏳  Muddat:      {deadline}",
            f"      {muhimlik_emoji}  Muhimlik:    {muhimlik_name}",
        ]
        # Overdue aging — how many days past the deadline. A sharp executive SLA
        # signal: a task that's "🔴 Kechikish: 5 kun" reads as neglected at a glance.
        _dl = _parse_dt_safe(task.get("deadline"))
        if _dl and _dl < datetime.now(database.TZ):
            _days_over = (datetime.now(database.TZ).date() - _dl.date()).days
            if _days_over >= 1:
                card.append(f"      🔴  Kechikish:   {_days_over} kun")
        # Full izoh (description) — aligned label; newlines collapsed so it stays
        # one logical block; Telegram wraps long text on its own.
        description = " ".join((task.get("description") or "").split())
        if description:
            card.append(f"      📝  Izoh:        {description}")
        return card

    abs_idx = start

    if unfinished:
        lines.extend([f"⏳  **BAJARILMAGAN**   ·   {len(unfinished)} ta", ""])
        cards: list[list[str]] = []
        for t in unfinished:
            abs_idx += 1
            cards.append(_task_card_lines(t, abs_idx))
        # Join cards with a single blank line between each.
        joined = []
        for i, card in enumerate(cards):
            if i:
                joined.append("")
            joined.extend(card)
        lines.extend(joined)
        lines.append("")

    if done:
        if unfinished:
            lines.extend([DIVIDER, ""])
        lines.extend([f"✅  **BAJARILGAN**   ·   {len(done)} ta", ""])
        done_cards: list[list[str]] = []
        for t in done:
            abs_idx += 1
            title = (t.get("title") or "—").strip()
            updated_at = t.get("updated_at")
            if updated_at:
                try:
                    dt = datetime.fromisoformat(updated_at).astimezone(database.TZ)
                    date_str = f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]}"
                except (ValueError, TypeError):
                    date_str = "—"
            else:
                date_str = "—"
            assignee = ((t.get("assignee") or "Belgilanmagan").strip() or "Belgilanmagan")
            assignee = assignee[0].upper() + assignee[1:]
            done_cards.append([
                f"{abs_idx}.  ✅  {title}",
                "",
                f"      📅  Yopildi:    {date_str}",
                f"      👤  Ijrochi:    {assignee}",
            ])
        for i, card in enumerate(done_cards):
            if i:
                lines.append("")
            lines.extend(card)
        lines.append("")

    lines.extend([
        DIVIDER,
        "",
        "_Vazifa raqamini bosing — to'liq ma'lumot._",
    ])
    return "\n".join(lines).rstrip()


def tasks_compact_keyboard(
    tasks: list[dict],
    current_filter: str = "active",
    page: int = 1,
) -> InlineKeyboardMarkup:
    """Tasks screen inline keyboard — DRILL-DOWN va sahifalash uchun.

    Filter chiplari va Yangi/Qidirish tugmalari **olib tashlangan** —
    ular allaqachon pastdagi reply kbd (section)'da bor.
    """
    per_page = _TASKS_PER_PAGE
    total = len(tasks)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_tasks = tasks[start : start + per_page]

    rows: list[list[InlineKeyboardButton]] = []

    nums = [
        InlineKeyboardButton(text=str(start + i + 1), callback_data=f"taskopen:{t['id']}")
        for i, t in enumerate(page_tasks)
    ]
    # Remember this numbered view so the LLM can resolve a later "N-vazifani tahrirla"
    # (reference a task by its DISPLAYED number) → the right task id.
    claude_service.set_last_task_view(
        [{"n": start + i + 1, "id": t["id"], "title": (t.get("title") or "—")[:50]}
         for i, t in enumerate(page_tasks)])
    if nums:
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i+5])

    if total_pages > 1:
        pag_row: list[InlineKeyboardButton] = []
        if page > 1:
            pag_row.append(InlineKeyboardButton(
                text="⬅️ Oldingi",
                callback_data=f"taskfilter:{current_filter}:{page-1}",
            ))
        # Non-clickable position indicator between the arrows ("2 / 5"). The "noop"
        # callback has no handler → caught by cb_fallback, which just dismisses it.
        pag_row.append(InlineKeyboardButton(text=f"{page} / {total_pages}", callback_data="noop"))
        if page < total_pages:
            pag_row.append(InlineKeyboardButton(
                text="Keyingi ➡️",
                callback_data=f"taskfilter:{current_filter}:{page+1}",
            ))
        rows.append(pag_row)

    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None


_REMINDERS_PER_PAGE = 10


def _reminder_status_label(status: str | None) -> str:
    return {
        "scheduled": "Rejalangan",
        "sent": "Yuborilgan",
        "done": "Bajarilgan",
        "cancelled": "Bekor qilingan",
    }.get(status or "scheduled", status or "Rejalangan")


def _reminder_status_icon(reminder: dict) -> str:
    status = reminder.get("status", "scheduled")
    if status == "done":
        return "✅"
    if status == "sent":
        return "📤"
    if status == "cancelled":
        return "❌"
    dt = _parse_dt_safe(reminder.get("remind_at"))
    if dt and dt < datetime.now(database.TZ):
        return "🔴"
    if dt and dt.date() == datetime.now(database.TZ).date():
        return "🟡"
    return "⏰"


def _reminder_time_chip(reminder: dict) -> str:
    dt = _parse_dt_safe(reminder.get("remind_at"))
    if not dt:
        return reminder.get("remind_at") or "—"
    now = datetime.now(database.TZ)
    if dt.date() == now.date():
        return f"Bugun {dt.strftime('%H:%M')}"
    if dt.date() == (now + timedelta(days=1)).date():
        return f"Ertaga {dt.strftime('%H:%M')}"
    return f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]} {dt.strftime('%H:%M')}"


def _format_reminder_card(reminder: dict, idx: int | None = None) -> str:
    title = (reminder.get("title") or "—").strip()
    prefix = f"{idx}.  " if idx is not None else ""
    status = _reminder_status_label(reminder.get("status"))
    repeat = _format_recurrence_label(reminder.get("recurrence_rule")) if reminder.get("recurrence_rule") else "bir martalik"
    lines = [
        f"{prefix}{_reminder_status_icon(reminder)} **{title}**",
        "",
        f"⏰ Vaqt: {_reminder_time_chip(reminder)}",
        f"🔁 Takror: {repeat}",
        f"📌 Holat: {status}",
    ]
    note = (reminder.get("note") or "").strip()
    if note:
        lines.append(f"📝 Izoh: {note}")
    return "\n".join(lines)


def _format_reminders_compact(
    reminders: list[dict],
    label: str,
    stats: dict | None = None,
    page: int = 1,
) -> str:
    divider = "━" * 20
    per_page = _REMINDERS_PER_PAGE
    total = len(reminders)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_items = reminders[start:start + per_page]

    lines = [
        "⏰  **ESLATMALAR**",
        "",
        f"Ko'rinish · {label}",
        f"Natija · {total} ta",
    ]
    if total_pages > 1:
        lines.append(f"Sahifa · {page} / {total_pages}")
    if stats:
        lines.extend([
            "",
            divider,
            "",
            "📌  **UMUMIY HOLAT**",
            "",
            f"Rejalangan {stats['scheduled']}   ·   Bugun {stats['today']}   ·   O'tgan {stats['overdue']}",
            f"Yuborilgan {stats['sent']}   ·   Takroriy {stats['recurring']}",
        ])
    if not page_items:
        lines.extend([
            "",
            divider,
            "",
            "Bu bo'limda hozircha eslatma yo'q.",
            "",
            "_Yangi eslatma qo'shish uchun pastdagi tugmadan foydalaning._",
        ])
        return "\n".join(lines)

    lines.extend(["", divider, ""])
    for offset, reminder in enumerate(page_items, start=1):
        if offset > 1:
            lines.append("")
        lines.append(_format_reminder_card(reminder, idx=start + offset))
    lines.extend(["", divider, "", "_Eslatma raqamini bosing — boshqaruv ochiladi._"])
    return "\n".join(lines).rstrip()


def reminders_compact_keyboard(
    reminders: list[dict],
    current_filter: str = "active",
    page: int = 1,
) -> InlineKeyboardMarkup | None:
    per_page = _REMINDERS_PER_PAGE
    total = len(reminders)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_items = reminders[start:start + per_page]

    rows: list[list[InlineKeyboardButton]] = []
    nums = [
        InlineKeyboardButton(text=str(start + i + 1), callback_data=f"remopen:{r['id']}")
        for i, r in enumerate(page_items)
    ]
    if nums:
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
    if total_pages > 1:
        nav: list[InlineKeyboardButton] = []
        if page > 1:
            nav.append(InlineKeyboardButton(text="⬅️ Oldingi", callback_data=f"remfilter:{current_filter}:{page - 1}"))
        nav.append(InlineKeyboardButton(text=f"{page} / {total_pages}", callback_data="noop"))
        if page < total_pages:
            nav.append(InlineKeyboardButton(text="Keyingi ➡️", callback_data=f"remfilter:{current_filter}:{page + 1}"))
        rows.append(nav)
    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None


def reminder_detail_menu(reminder: dict) -> InlineKeyboardMarkup:
    """Eslatma holatiga moslashuvchi tugmalar:
      • done            → ↺ Qayta eslat / o'chir / orqaga
      • takroriy (scheduled) → ⏭ Bu safarni o'tkaz / 🛑 To'xtatish / tahrir
      • bir martalik (scheduled/sent) → ✅ Bajarildi / snooze / tahrir
    Ikonka izchilligi: snooze hammasi ⏰; vaqt tahriri 🕐 (snooze-Ertaga'dan farqli)."""
    rid = reminder["id"]
    status = reminder.get("status")
    is_recurring = bool((reminder.get("recurrence_rule") or "").strip())
    # Konsolidatsiya: bitta '✏️ Tahrirlash' → submenu (Sarlavha/Vaqt/Takror/Izoh).
    edit_row = [
        InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"remeditmenu:{rid}"),
    ]
    tail_row = [
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"remdel:{rid}"),
        InlineKeyboardButton(text="⬅️ Ro'yxatga", callback_data="remfilter:active"),
    ]
    # Done — reopen / EDIT (vaqt tahriri qayta faollashtiradi) / delete / back.
    if status == "done":
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="↺ Qayta eslat", callback_data=f"remsnooze:{rid}:1d")],
            edit_row,
            tail_row,
        ])
    # Recurring (stays 'scheduled' — fires roll forward): skip THIS occurrence
    # vs stop the whole series. NOT a plain "Bajarildi" (that ended the series).
    if is_recurring and status == "scheduled":
        return InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="⏭ Bu safarni o'tkaz", callback_data=f"remskip:{rid}"),
                InlineKeyboardButton(text="🛑 Takrorni to'xtatish", callback_data=f"remstop:{rid}"),
            ],
            edit_row,
            tail_row,
        ])
    # One-time scheduled OR sent (fired): done + snooze (⏰) + edit.
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Bajarildi", callback_data=f"remdone:{rid}")],
        [
            InlineKeyboardButton(text="⏰ 15 daq", callback_data=f"remsnooze:{rid}:15m"),
            InlineKeyboardButton(text="⏰ 1 soat", callback_data=f"remsnooze:{rid}:1h"),
            InlineKeyboardButton(text="⏰ Ertaga", callback_data=f"remsnooze:{rid}:1d"),
        ],
        edit_row,
        tail_row,
    ])


async def _load_reminders_for_filter(filt: str) -> tuple[list[dict], str]:
    if filt == "today":
        return await database.list_today_reminders(limit=200), "Bugungi eslatmalar"
    if filt == "sent":
        return await database.list_reminders(status_in=["sent", "done"], limit=200), "Yuborilgan / bajarilgan"
    if filt == "all":
        return await database.list_reminders(limit=200), "Barchasi"
    if filt == "active":
        # Default ko'rinish — bajarilmaganlar (rejalangan + yuborilgan), done'siz.
        return await database.list_reminders(status_in=["scheduled", "sent"], limit=200), "Aktiv eslatmalar"
    return await database.list_reminders(status_in=["scheduled"], limit=200), "Keyingi eslatmalar"


async def _render_reminders_for_filter(
    message: Message,
    filt: str = "active",
    page: int = 1,
    edit_existing: bool = False,
) -> None:
    reminders, label = await _load_reminders_for_filter(filt)
    stats = await database.reminders_overview()
    text = _format_reminders_compact(reminders, label, stats=stats, page=page)
    kb = reminders_compact_keyboard(reminders, current_filter=filt, page=page)
    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


# ─────────────────────── NOTES (Qaydlar) RENDERING ───────────────────────

def _format_notes_compact(notes: list[dict], label: str,
                            inbox_count: int = 0, page: int = 1) -> str:
    """Sana bo'yicha guruhlangan ro'yxat (Variant B): qaydlar
    BUGUN / KECHA / BU HAFTA / OLDIN bo'lib ajratiladi. Har band oldida manba
    ikonkasi (🎙/✍️/🔁/⚡/🤖); raqamlar saqlanadi — pastdagi 1..N tugmalari
    aynan shu qaydlarga mos keladi."""
    DIVIDER = "━" * 15
    if not notes:
        return "\n".join([
            f"📝 **QAYDLAR · {label.upper()}**", "",
            f"📥 Inbox: **{inbox_count}** ta qayta ishlanmagan", "",
            DIVIDER, "",
            "_Hozircha bu bo'limda qayd yo'q._", "",
            "Tezkor qo'shish: `/qayd <matn>` yoki ovoz orqali _\"qayd qil: …\"_.",
        ]).rstrip()

    per_page = _NOTES_PER_PAGE
    total = len(notes)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_items = notes[start:start + per_page]

    today = datetime.now(database.TZ).date()

    def _bucket(iso) -> int:
        """0=BUGUN, 1=KECHA, 2=BU HAFTA, 3=OLDIN."""
        try:
            d = datetime.fromisoformat(iso).astimezone(database.TZ).date()
        except (TypeError, ValueError):
            return 3
        delta = (today - d).days
        if delta <= 0:
            return 0
        if delta == 1:
            return 1
        if delta <= 6:
            return 2
        return 3

    _BUCKET_LABELS = {0: "BUGUN", 1: "KECHA", 2: "BU HAFTA", 3: "OLDIN"}
    grouped: dict[int, list] = {0: [], 1: [], 2: [], 3: []}
    for i, n in enumerate(page_items, start=start + 1):
        grouped[_bucket(n.get("created_at"))].append((i, n))

    lines = [f"📝 **QAYDLAR · {label.upper()} · {total} ta**", "", DIVIDER]
    for b in (0, 1, 2, 3):
        items = grouped[b]
        if not items:
            continue
        lines.append("")
        lines.append(f"📅 **{_BUCKET_LABELS[b]}**")
        for i, n in items:
            title = (n.get("title") or _derive_note_title_fallback(n.get("content", "")))
            icon = _NOTES_SOURCE_ICON.get(n.get("source", "manual"), "📝")
            preview = (n.get("content") or "").strip().replace("\n", " ")
            if len(preview) > 90:
                preview = preview[:87] + "…"
            lines.append(f"  {icon} {i}. {_escape_markdown(title[:64])}")
            if preview and preview != title:
                lines.append(f"      _{_escape_markdown(preview)}_")
    if total_pages > 1:
        lines += ["", DIVIDER, f"_Sahifa {page}/{total_pages}_"]
    return "\n".join(lines).rstrip()


def _derive_note_title_fallback(content: str) -> str:
    """Mirror of database._derive_title for display use only."""
    first = next((ln.strip() for ln in (content or "").splitlines() if ln.strip()), "")
    if not first:
        return "(bo'sh note)"
    return first if len(first) <= 60 else first[:59] + "…"


def _short_local_date(iso: str | None) -> str:
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso).astimezone(database.TZ)
        return dt.strftime("%d-%m %H:%M")
    except (TypeError, ValueError):
        return "—"


def notes_compact_keyboard(notes: list[dict], current_filter: str = "inbox",
                            page: int = 1) -> InlineKeyboardMarkup | None:
    """Inline numbered keyboard + pagination — matches reminders pattern."""
    per_page = _NOTES_PER_PAGE
    total = len(notes)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_items = notes[start:start + per_page]

    rows: list[list[InlineKeyboardButton]] = []
    nums = [
        InlineKeyboardButton(text=str(start + i + 1), callback_data=f"noteopen:{n['id']}")
        for i, n in enumerate(page_items)
    ]
    if nums:
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
    # KONSEPSIYA: filtrlar (Inbox/Ishlangan/Arxiv) ataylab BU YERDA YO'Q — ular
    # doimiy reply-klaviaturada (notes_section_reply_keyboard) turadi. Inline faqat
    # kontentga xos: drill-down raqamlar + pagination. Shu bilan inline↔reply
    # tugma takrori oldi olinadi (avval bir xil filtrlar ikkala joyda chiqardi).
    if total_pages > 1:
        nav: list[InlineKeyboardButton] = []
        if page > 1:
            nav.append(InlineKeyboardButton(text="⬅️ Oldingi", callback_data=f"notesfilter:{current_filter}:{page - 1}"))
        nav.append(InlineKeyboardButton(text=f"{page} / {total_pages}", callback_data="noop"))
        if page < total_pages:
            nav.append(InlineKeyboardButton(text="Keyingi ➡️", callback_data=f"notesfilter:{current_filter}:{page + 1}"))
        rows.append(nav)
    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None


def note_detail_menu(note: dict) -> InlineKeyboardMarkup:
    """Per-note action menu — Tahlil / Vazifaga / Eslatmaga / Arxiv / O'chir."""
    nid = note["id"]
    status = note.get("status", "inbox")
    rows = []
    # If already processed, expose Reopen instead of conversion actions.
    if status == "processed":
        rows.append([
            InlineKeyboardButton(text="🔁 Inbox'ga qaytar",
                                  callback_data=f"noterestore:{nid}"),
        ])
    elif status == "archived":
        rows.append([
            InlineKeyboardButton(text="🔁 Tiklash", callback_data=f"noterestore:{nid}"),
        ])
    else:
        rows.append([
            InlineKeyboardButton(text="🤖 Tahlil qil", callback_data=f"noteanalyze:{nid}"),
        ])
        rows.append([
            InlineKeyboardButton(text="📝 Vazifaga", callback_data=f"notetotask:{nid}"),
            InlineKeyboardButton(text="⏰ Eslatmaga", callback_data=f"notetorem:{nid}"),
        ])
        rows.append([
            InlineKeyboardButton(text="✂️ Ko'p vazifaga ajrat", callback_data=f"notesplit:{nid}"),
        ])
        rows.append([
            InlineKeyboardButton(text="✅ Ishlandi", callback_data=f"notedone:{nid}"),
            InlineKeyboardButton(text="📦 Arxivga", callback_data=f"notearchive:{nid}"),
            InlineKeyboardButton(text="🗑 O'chir", callback_data=f"notedelete:{nid}"),
        ])
    rows.append([
        InlineKeyboardButton(text="⬅️ Ro'yxatga",
                              callback_data=f"notesfilter:{status if status != 'archived' else 'archived'}"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _load_notes_for_filter(filt: str) -> tuple[list[dict], str]:
    label_map = {
        "inbox": "Inbox",
        "processed": "Qayta ishlangan",
        "archived": "Arxiv",
    }
    if filt not in label_map:
        filt = "inbox"
    notes = await database.list_notes(status=filt, limit=200)
    return notes, label_map[filt]


async def _render_notes_for_filter(message: Message, filt: str = "inbox",
                                     page: int = 1,
                                     edit_existing: bool = False) -> None:
    notes, label = await _load_notes_for_filter(filt)
    inbox_count = await database.count_notes_in_status("inbox")
    text = _format_notes_compact(notes, label, inbox_count=inbox_count, page=page)
    kb = notes_compact_keyboard(notes, current_filter=filt, page=page)
    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


async def _compute_tasks_overview() -> dict:
    """Counts for the UMUMIY HOLAT block — stable across filters."""
    all_tasks = await database.list_tasks(limit=500)
    # "Active" = open work = todo + in_progress + blocked. Blocked is open work
    # that's stuck — it MUST stay visible (previously it only showed in Barchasi).
    active = [t for t in all_tasks if t.get("status") in ("todo", "in_progress", "blocked")]
    done = [t for t in all_tasks if t.get("status") == "done"]
    now = datetime.now(database.TZ)

    def _overdue(task: dict) -> bool:
        d = task.get("deadline")
        if not d:
            return False
        try:
            return datetime.fromisoformat(d).astimezone(database.TZ) < now
        except (ValueError, TypeError):
            return False

    overdue = [t for t in active if _overdue(t)]
    urgent = [t for t in active if t.get("priority") == "P0"]
    important = [t for t in active if t.get("priority") == "P1"]
    blocked = [t for t in active if t.get("status") == "blocked"]
    return {
        "total": len(all_tasks),
        "active": len(active),
        "done": len(done),
        "overdue": len(overdue),
        "urgent": len(urgent),
        "important": len(important),
        "blocked": len(blocked),
    }


async def _render_tasks_for_filter(message: Message, filt: str = "active",
                                    page: int = 1, edit_existing: bool = False) -> None:
    """Render the tasks screen. Pagination = 10 items per page.

    If edit_existing=True (called from a callback), edits the existing message
    instead of sending a new one — keeps the chat tidy.
    """
    global _last_task_filter
    _last_task_filter = filt
    if filt == "active":
        tasks = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=200)
        label = "Aktiv vazifalar"
    elif filt == "today":
        tasks = await database.list_today_tasks()
        label = "Bugungi vazifalar"
    elif filt == "important":
        all_active = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=200)
        tasks = [t for t in all_active if t.get("priority") in ("P0", "P1")]
        label = "Muhim vazifalar"
    elif filt == "urgent":
        all_active = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=200)
        tasks = [t for t in all_active if t.get("priority") == "P0"]
        label = "Shoshilinch"
    elif filt == "overdue":
        tasks = await database.list_overdue_tasks()
        label = "Muddati o'tgan"
    elif filt == "done":
        tasks = await database.list_tasks(status_in=["done"], limit=200)
        label = "Bajarilgan"
    elif filt == "recurring":
        tasks = await database.list_recurring_tasks(limit=100)
        label = "Takroriy vazifalar"
    elif filt == "all":
        tasks = await database.list_tasks(limit=200)
        label = "Barchasi"
    else:
        tasks = await database.list_tasks(status_in=["todo", "in_progress", "blocked"], limit=200)
        label = "Aktiv vazifalar"

    # `_format_tasks_compact` cards'ni "BAJARILMAGAN" → "BAJARILGAN" tartibida chiqaradi.
    # Tugma raqamlari esa `tasks_compact_keyboard`'da page_tasks ketma-ketligida tuziladi.
    # Mos kelishi uchun bu yerda done'larni oxiriga surib qo'yamiz (stable sort —
    # bir status ichida priorityyaki created_at tartibi saqlanadi).
    tasks.sort(key=lambda t: 1 if t.get("status") == "done" else 0)

    stats = await _compute_tasks_overview()
    text = _format_tasks_compact(tasks, label, stats=stats, page=page)
    kb = tasks_compact_keyboard(tasks, current_filter=filt, page=page)

    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass  # message likely too old to edit; fall through to send new
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


@router.message(Command("tasks"))
async def cmd_tasks(message: Message, state: FSMContext | None = None) -> None:
    # Section FSM state — "Bugun"/"Barchasi" kabi noyob bo'lmagan labellar
    # uchun kontekst aniqlash. Reply kbd ni Vazifalar bo'limiga almashtirish.
    if state is not None:
        await state.set_state(SectionFSM.in_tasks)
    await message.answer(
        "📌 **VAZIFALAR**", parse_mode="Markdown",
        reply_markup=tasks_section_reply_keyboard(),
    )
    await _render_tasks_for_filter(message, "active")


_CATEGORY_ICONS = ["📁", "🔴", "🟠", "🟡", "🟢", "🔵", "🟣", "⭐",
                   "💼", "📊", "📢", "🤝", "🛒", "💻", "🎯", "🗓"]


def _categories_keyboard(cats: list[dict]) -> InlineKeyboardMarkup:
    rows, row = [], []
    for c in cats:
        row.append(InlineKeyboardButton(text=f"{c['icon']} {c['name']} ({c['count']})",
                                        callback_data=f"taskcat:{c['name']}"[:64]))
        if len(row) == 2:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([
        InlineKeyboardButton(text="➕ Yangi kategoriya", callback_data="catnew"),
        InlineKeyboardButton(text="📦 Arxiv", callback_data="catarchlist"),
    ])
    rows.append([back_button("nav_cockpit", "⬅️ Orqaga")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _render_categories(message: Message, edit_existing: bool = False) -> None:
    """The 'Kategoriyalar' overview — managed categories (icon + count) with
    add / archive entry points; tap a category for its tasks + management."""
    cats = await database.list_categories()
    DIV = "━" * 20
    if not cats:
        text = ("🗄 **KATEGORIYALAR**\n\n_Hozircha kategoriya yo'q._\n\n"
                "«➕ Yangi kategoriya» bilan qo'shing — yoki vazifa yaratganда avtomatik belgilanadi.")
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Yangi kategoriya", callback_data="catnew")],
            [back_button("nav_cockpit", "⬅️ Orqaga")],
        ])
    else:
        total = sum(c["count"] for c in cats)
        lines = ["🗄 **KATEGORIYALAR**", "", f"_{len(cats)} ta kategoriya · {total} ta aktiv vazifa_", "", DIV, ""]
        for c in cats:
            lines.append(f"{c['icon']} **{c['name']}** — {c['count']} ta")
        lines.extend(["", DIV, "", "_Kategoriyani bosing — vazifalari va boshqaruv._"])
        text = "\n".join(lines)
        kb = _categories_keyboard(cats)
    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


async def _render_archived_categories(message: Message, edit_existing: bool = True) -> None:
    cats = await database.list_categories(include_archived=True)
    DIV = "━" * 20
    if not cats:
        text = "📦 **ARXIV**\n\n_Arxivlangan kategoriya yo'q._"
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="⬅️ Kategoriyalar", callback_data="taskcats")]])
    else:
        lines = ["📦 **ARXIVLANGAN KATEGORIYALAR**", "", DIV, ""]
        for c in cats:
            lines.append(f"{c['icon']} **{c['name']}** — {c['count']} ta")
        lines.extend(["", DIV, "", "_Tiklash uchun bosing._"])
        text = "\n".join(lines)
        rows = [[InlineKeyboardButton(text=f"♻️ {c['icon']} {c['name']}",
                                      callback_data=f"catunarch:{c['name']}"[:64])] for c in cats]
        rows.append([InlineKeyboardButton(text="⬅️ Kategoriyalar", callback_data="taskcats")])
        kb = InlineKeyboardMarkup(inline_keyboard=rows)
    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


def _category_drilldown_kb(cat: str, tasks: list[dict]) -> InlineKeyboardMarkup:
    nums = [InlineKeyboardButton(text=str(i + 1), callback_data=f"taskopen:{t['id']}")
            for i, t in enumerate(tasks[:_TASKS_PER_PAGE])]
    rows = [nums[j:j + 5] for j in range(0, len(nums), 5)]
    if cat == "(boshqa)":  # synthetic uncategorized bucket — no rename/archive/move
        rows.append([
            InlineKeyboardButton(text="➕ Vazifa", callback_data=f"catadd:{cat}"[:64]),
            InlineKeyboardButton(text="🧹 Vazifalar", callback_data=f"cattasksdel:{cat}"[:64]),
        ])
    else:
        rows.append([
            InlineKeyboardButton(text="➕ Vazifa", callback_data=f"catadd:{cat}"[:64]),
            InlineKeyboardButton(text="✏️ Tahrir", callback_data=f"catedit:{cat}"[:64]),
            InlineKeyboardButton(text="📦 Arxiv", callback_data=f"catarch:{cat}"[:64]),
        ])
        rows.append([
            InlineKeyboardButton(text="🗑 Kategoriya", callback_data=f"catdel:{cat}"[:64]),
            InlineKeyboardButton(text="🧹 Vazifalar", callback_data=f"cattasksdel:{cat}"[:64]),
        ])
        rows.append([
            InlineKeyboardButton(text="⬆️ Yuqoriga", callback_data=f"catmove:{cat}:up"[:64]),
            InlineKeyboardButton(text="⬇️ Pastga", callback_data=f"catmove:{cat}:down"[:64]),
        ])
    rows.append([InlineKeyboardButton(text="⬅️ Kategoriyalar", callback_data="taskcats")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _render_category_drilldown(message: Message, cat: str, edit_existing: bool = True) -> None:
    tasks = await database.list_tasks_by_category(cat)
    text = (_format_tasks_compact(tasks, f"Kategoriya: {cat}") if tasks
            else f"📁 **«{cat}»** — bo'sh.")
    kb = _category_drilldown_kb(cat, tasks)
    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


@router.message(Command("categories"))
async def cmd_categories(message: Message) -> None:
    await _render_categories(message)


@router.callback_query(F.data == "taskcats")
async def cb_task_categories(query: CallbackQuery) -> None:
    await query.answer()
    await _render_categories(query.message, edit_existing=True)


@router.callback_query(F.data == "catarchlist")
async def cb_archived_categories(query: CallbackQuery) -> None:
    await query.answer()
    await _render_archived_categories(query.message, edit_existing=True)


@router.callback_query(F.data.startswith("taskcat:"))
async def cb_task_category(query: CallbackQuery) -> None:
    cat = query.data.split(":", 1)[1]
    await query.answer()
    await _render_category_drilldown(query.message, cat, edit_existing=True)


# ── Edit (rename + icon), archive, reorder ──
class CategoryNameFSM(StatesGroup):
    awaiting = State()


def _icon_picker_kb(name: str) -> InlineKeyboardMarkup:
    rows, row = [], []
    for ic in _CATEGORY_ICONS:
        row.append(InlineKeyboardButton(text=ic, callback_data=f"caticon:{name}:{ic}"[:64]))
        if len(row) == 4:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"catedit:{name}"[:64])])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(F.data == "catnew")
async def cb_category_new(query: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(CategoryNameFSM.awaiting)
    await state.update_data(mode="create")
    await query.answer()
    await query.message.answer("➕ Yangi kategoriya nomini yozing (matn yoki ovoz):")


@router.callback_query(F.data.startswith("catedit:"))
async def cb_category_edit_menu(query: CallbackQuery) -> None:
    name = query.data.split(":", 1)[1]
    await query.answer()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Nomini o'zgartirish", callback_data=f"catrename:{name}"[:64])],
        [InlineKeyboardButton(text="🎨 Ikonka", callback_data=f"caticonpick:{name}"[:64])],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"taskcat:{name}"[:64])],
    ])
    await _safe_answer(query.message, f"✏️ **«{name}»** — tahrirlash:", parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("catrename:"))
async def cb_category_rename(query: CallbackQuery, state: FSMContext) -> None:
    name = query.data.split(":", 1)[1]
    await state.set_state(CategoryNameFSM.awaiting)
    await state.update_data(mode="rename", old=name)
    await query.answer()
    await query.message.answer(f"✏️ «{name}» uchun yangi nom yozing:")


@router.message(StateFilter(CategoryNameFSM.awaiting), F.text | F.voice)
async def handle_category_name(message: Message, state: FSMContext, bot: Bot) -> None:
    name = await _get_text_or_transcribe(message, bot=bot)
    if name is None:
        return
    name = name.strip()[:60]
    data = await state.get_data()
    mode = data.get("mode")
    await state.clear()
    if not name:
        await message.answer("Bo'sh nom — bekor qilindi.")
        return
    if mode == "rename":
        old = data.get("old")
        await database.update_category(old, new_name=name)
        await message.answer(f"✅ «{old}» → «{name}»")
    else:
        await database.create_category(name)
        await message.answer(f"✅ «{name}» kategoriyasi yaratildi.")
    await _render_categories(message)


@router.callback_query(F.data.startswith("caticonpick:"))
async def cb_category_icon_pick(query: CallbackQuery) -> None:
    name = query.data.split(":", 1)[1]
    await query.answer()
    await _safe_answer(query.message, f"🎨 «{name}» uchun ikonka tanlang:", reply_markup=_icon_picker_kb(name))


@router.callback_query(F.data.startswith("caticon:"))
async def cb_category_set_icon(query: CallbackQuery) -> None:
    _, name, icon = query.data.split(":", 2)
    await database.update_category(name, icon=icon)
    await query.answer(f"Ikonka → {icon} ✅")
    await _render_categories(query.message, edit_existing=True)


@router.callback_query(F.data.startswith("catarch:"))
async def cb_category_archive(query: CallbackQuery) -> None:
    name = query.data.split(":", 1)[1]
    await database.archive_category(name, True)
    await query.answer(f"📦 «{name}» arxivlandi")
    await _render_categories(query.message, edit_existing=True)


@router.callback_query(F.data.startswith("catunarch:"))
async def cb_category_unarchive(query: CallbackQuery) -> None:
    name = query.data.split(":", 1)[1]
    await database.archive_category(name, False)
    await query.answer(f"♻️ «{name}» tiklandi")
    await _render_archived_categories(query.message, edit_existing=True)


@router.callback_query(F.data.startswith("catmove:"))
async def cb_category_move(query: CallbackQuery) -> None:
    _, name, direction = query.data.split(":", 2)
    ok = await database.move_category(name, direction)
    await query.answer("✅ Ko'chirildi" if ok else "Chekkada")
    await _render_category_drilldown(query.message, name, edit_existing=True)


class AddToCategoryFSM(StatesGroup):
    awaiting = State()


async def _confirm_actions(query: CallbackQuery, state: FSMContext, actions: list[dict], done_msg: str) -> None:
    """Route button-triggered actions through the standard preview + Tasdiq gate."""
    preview = await _format_create_preview(actions)
    await state.set_state(CreateActionConfirmFSM.awaiting)
    await state.update_data(
        pending_response={"actions": actions, "user_message": done_msg, "buttons": []},
        _prior_section=None,
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Tasdiqlayman", callback_data="acts_confirm"),
        InlineKeyboardButton(text="✕ Bekor", callback_data="acts_cancel"),
    ]])
    await _safe_answer(query.message, preview, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("catdel:"))
async def cb_category_clear(query: CallbackQuery, state: FSMContext) -> None:
    """Delete a category label (tasks survive → uncategorized). Confirm-gated."""
    cat = query.data.split(":", 1)[1]
    await query.answer()
    await _confirm_actions(query, state, [{"type": "delete_category", "data": {"category": cat}}],
                           f"✅ «{cat}» kategoriyasi olib tashlandi.")


@router.callback_query(F.data.startswith("cattasksdel:"))
async def cb_category_delete_tasks(query: CallbackQuery, state: FSMContext) -> None:
    """Delete ALL tasks in a category. Confirm-gated."""
    cat = query.data.split(":", 1)[1]
    await query.answer()
    await _confirm_actions(query, state, [{"type": "delete_tasks_by_category", "data": {"category": cat}}],
                           f"✅ «{cat}» kategoriyasidagi vazifalar o'chirildi.")


@router.callback_query(F.data.startswith("catadd:"))
async def cb_category_add_task(query: CallbackQuery, state: FSMContext) -> None:
    """Add a task into a specific category — prompt for the title."""
    cat = query.data.split(":", 1)[1]
    await state.set_state(AddToCategoryFSM.awaiting)
    await state.update_data(category=cat)
    await query.answer()
    await query.message.answer(
        f"➕ **«{cat}»** kategoriyasiga yangi vazifa.\nVazifa nomini yozing (matn yoki ovoz):",
        parse_mode="Markdown",
    )


@router.message(StateFilter(AddToCategoryFSM.awaiting), F.text | F.voice)
async def handle_add_to_category(message: Message, state: FSMContext, bot: Bot) -> None:
    title = await _get_text_or_transcribe(message, bot=bot)
    if title is None:
        return
    title = title.strip()
    data = await state.get_data()
    cat = data.get("category") or None
    await state.clear()
    if not title:
        await message.answer("Bo'sh nom — bekor qilindi.")
        return
    tid = await database.create_task({"title": title, "category": cat, "priority": "P2", "status": "todo"})
    await message.answer(f"✅ «{cat}» kategoriyasiga qo'shildi:\n📌 {title}",
                         reply_markup=_task_card_kb_with_back(await database.get_task(tid)))


@router.message(Command("reminders"))
async def cmd_reminders(message: Message, state: FSMContext | None = None) -> None:
    if state is not None:
        await state.set_state(SectionFSM.in_reminders)
    await message.answer(
        "⏰ **ESLATMALAR**", parse_mode="Markdown",
        reply_markup=reminders_section_reply_keyboard(),
    )
    await _render_reminders_for_filter(message, "active")


@router.callback_query(F.data.startswith("remfilter:"))
async def cb_reminder_filter(query: CallbackQuery) -> None:
    parts = query.data.split(":")
    filt = parts[1] if len(parts) > 1 and parts[1] else "active"
    page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 1
    await query.answer()
    await _render_reminders_for_filter(query.message, filt, page=page, edit_existing=True)


# ─────────────────────── NOTES — COMMANDS + CALLBACKS ───────────────────────


@router.message(Command("notes"))
@router.message(Command("qaydlar"))
async def cmd_notes(message: Message, state: FSMContext | None = None) -> None:
    """Notes bo'limi — Inbox / Ishlangan / Arxiv. Default: Inbox."""
    if state is not None:
        await state.set_state(SectionFSM.in_notes)
    await message.answer(
        "📝 **QAYDLAR**", parse_mode="Markdown",
        reply_markup=notes_section_reply_keyboard(),
    )
    await _render_notes_for_filter(message, "inbox")


@router.callback_query(F.data.startswith("notesfilter:"))
async def cb_notes_filter(query: CallbackQuery) -> None:
    parts = query.data.split(":")
    filt = parts[1] if len(parts) > 1 and parts[1] else "inbox"
    page = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 1
    if filt not in ("inbox", "processed", "archived"):
        filt = "inbox"
    await query.answer()
    await _render_notes_for_filter(query.message, filt, page=page, edit_existing=True)


@router.message(Command("qayd"))
async def cmd_qayd(message: Message, state: FSMContext | None = None) -> None:
    """Tezkor qayd qo'shish. `/qayd <matn>` → darrov inbox'ga saqlaydi.
    `/qayd` (bo'sh) → keyingi xabar qaydga aylanadi (one-shot FSM)."""
    text = (message.text or "").removeprefix("/qayd").removeprefix("/qaydlar").strip()
    if text:
        nid = await database.create_note({
            "content": text, "source": "command",
        })
        await _note_capture_reply(message, nid, "/qayd buyrug'i orqali")
        await _maybe_refresh_section(message, state, {"note": [nid]})
        return
    # Empty body — enter one-shot FSM
    if state is not None:
        await state.set_state(NoteCaptureFSM.awaiting_text)
    await message.answer(
        "📝 **YANGI QAYD**\n\nMatn yoki ovoz yuboring — bot uni Inbox'ga saqlaydi.\n"
        "Bekor qilish: /cancel",
        parse_mode="Markdown",
        reply_markup=single_back_keyboard("nav_notes"),
    )


# ───────────────── SELF-IMPROVEMENT — PROPOSAL GATE (Phase 3) ─────────────────
# /improvements (list + approve/reject/details), /improve <request> (Channel B),
# /autopilot on|off (master kill switch). Principal-gated by the auth middleware.
# Approval ONLY flips status to 'approved' — it starts NO implementation (Phase 4
# does not exist yet). Rejection sets 'rejected' and triggers nothing.

_PROP_FIX_LABELS = {"prompt": "Prompt", "code": "Kod", "config": "Sozlama",
                    "data": "Ma'lumot", "feature": "Yangi imkoniyat"}


def _format_proposal_card(p: dict, full: bool = False) -> str:
    src = "🤖 Avto" if p.get("source") == "auto" else "✍️ So'rov"
    kind = _PROP_FIX_LABELS.get(p.get("fix_kind"), p.get("fix_kind") or "—")
    manual = p.get("status") == "requires_manual"
    lines = ["📋 **" + (p.get("title") or "—") + "**", "",
             f"{src} · {kind}" + ("  ·  ⚠️ qo'lda spec kerak" if manual else "")]
    if p.get("problem"):
        lines += ["", f"❗️ {p['problem']}"]
    if full:
        if p.get("evidence"):
            lines += ["", f"📊 Dalil: {p['evidence']}"]
        if p.get("root_cause"):
            lines += [f"🔍 Sabab: {p['root_cause']}"]
        if p.get("proposed_change"):
            lines += ["", f"🔧 Taklif: {p['proposed_change']}"]
        if p.get("impact_estimate"):
            lines += [f"📈 Ta'sir: {p['impact_estimate']}"]
    return "\n".join(lines)


def _proposal_keyboard(p: dict) -> InlineKeyboardMarkup:
    pid = p["id"]
    rows = []
    if p.get("status") != "requires_manual":   # requires_manual can't be auto-built → no Approve
        rows.append([
            InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=f"impapprove:{pid}"),
            InlineKeyboardButton(text="❌ Rad et", callback_data=f"impreject:{pid}"),
        ])
    else:
        rows.append([InlineKeyboardButton(text="❌ Rad et", callback_data=f"impreject:{pid}")])
    rows.append([InlineKeyboardButton(text="📄 Batafsil", callback_data=f"impdetails:{pid}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(Command("improvements"))
async def cmd_improvements(message: Message) -> None:
    """List pending self-improvement proposals (Channel A + B) for Gate-1 approval."""
    props = await database.list_improvement_proposals(status_in=["new", "requires_manual"], limit=20)
    if not props:
        await message.answer("💡 **Yaxshilanish takliflari**\n\nHozircha yangi taklif yo'q.",
                             parse_mode="Markdown")
        return
    await message.answer(f"💡 **Yaxshilanish takliflari ({len(props)})**", parse_mode="Markdown",
                         reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                             InlineKeyboardButton(text="📜 Tarix / audit", callback_data="silog:refresh")]]))
    for p in props:
        await _safe_answer(message, _format_proposal_card(p), parse_mode="Markdown",
                           reply_markup=_proposal_keyboard(p))


@router.message(Command("improve"))
async def cmd_improve(message: Message) -> None:
    """Channel B: the principal requests a feature/fix. Stores a source=manual
    proposal and shows the Gate-1 scope confirmation. No LLM — the raw request IS
    the scope; the dev_agent refines it in Phase 4."""
    text = (message.text or "").removeprefix("/improve").strip()
    if not text:
        await message.answer(
            "✍️ **Yaxshilanish so'rovi**\n\nFoydalanish: `/improve <nima qo'shay yoki tuzatay>`\n"
            "Masalan: `/improve eslatmalarga snooze tugmasi qo'sh`", parse_mode="Markdown")
        return
    pid = await database.create_improvement_proposal({
        "source": "manual", "title": text[:120], "problem": "Principal so'rovi (Channel B).",
        "proposed_change": text, "fix_kind": "feature", "status": "new",
    })
    p = await database.get_improvement_proposal(pid)
    await _safe_answer(message, "✍️ **So'rov qabul qilindi — tasdiqlaysizmi?**\n\n"
                       + _format_proposal_card(p), parse_mode="Markdown",
                       reply_markup=_proposal_keyboard(p))


@router.message(Command("autopilot"))
async def cmd_autopilot(message: Message) -> None:
    """Master kill switch for the self-improvement loop: `/autopilot on|off`."""
    arg = (message.text or "").removeprefix("/autopilot").strip().lower()
    if arg in ("on", "yoq", "yoqish", "1", "true"):
        await database.set_setting("autopilot_enabled", True)
        await message.answer("🟢 **Autopilot YOQILDI.**\n\nTungi self-diagnoz (02:00) ishlaydi va "
                             "takliflar yaratadi. O'chirish: `/autopilot off`", parse_mode="Markdown")
    elif arg in ("off", "ochir", "o'chir", "0", "false"):
        await database.set_setting("autopilot_enabled", False)
        await message.answer("⚪️ **Autopilot O'CHIRILDI.**\n\nTungi self-diagnoz to'xtaydi.",
                             parse_mode="Markdown")
    else:
        cur = (await database.get_settings()).get("autopilot_enabled", False)
        state_lbl = "🟢 YOQILGAN" if cur else "⚪️ O'CHIRILGAN"
        await message.answer(f"⚙️ **Autopilot holati:** {state_lbl}\n\n"
                             "`/autopilot on` — yoqish · `/autopilot off` — o'chirish",
                             parse_mode="Markdown")


@router.callback_query(F.data.startswith("impapprove:"))
async def cb_improvement_approve(query: CallbackQuery) -> None:
    """Gate 1 ✅ — approve. ONLY flips status to 'approved'; starts NO implementation."""
    pid = query.data.split(":", 1)[1]
    p = await database.get_improvement_proposal(pid)
    if not p:
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    if p.get("status") == "requires_manual":
        await query.answer("Bu qo'lda spec talab qiladi — auto-tasdiqlanmaydi.", show_alert=True)
        return
    await database.update_proposal_status(pid, "approved")
    await query.answer("✅ Tasdiqlandi")
    try:
        await query.message.edit_text(
            "✅ **Tasdiqlandi** (status: approved)\n\n" + _format_proposal_card(p)
            + "\n\n_Tayyor bo'lsangiz **🛠 Implement qil** — izolyatsiyalangan worktree'da "
              "kod yoziladi va testlardan o'tkaziladi. Hech narsa avtomatik deploy bo'lmaydi._",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🛠 Implement qil", callback_data=f"siimpl:{pid}"),
                InlineKeyboardButton(text="❌ Bekor", callback_data=f"sicancel:{pid}"),
            ]]))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("impreject:"))
async def cb_improvement_reject(query: CallbackQuery) -> None:
    """Gate 1 ❌ — reject. Sets status='rejected'; triggers nothing."""
    pid = query.data.split(":", 1)[1]
    if not await database.update_proposal_status(pid, "rejected"):
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    await query.answer("❌ Rad etildi")
    try:
        await query.message.edit_text("❌ **Rad etildi** (status: rejected).", parse_mode="Markdown")
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("impdetails:"))
async def cb_improvement_details(query: CallbackQuery) -> None:
    """📄 Batafsil — show the full proposal card."""
    pid = query.data.split(":", 1)[1]
    p = await database.get_improvement_proposal(pid)
    if not p:
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    await query.answer()
    try:
        await query.message.edit_text(_format_proposal_card(p, full=True), parse_mode="Markdown",
                                      reply_markup=_proposal_keyboard(p))
    except TelegramBadRequest:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Gate 2–3 wiring: approve → implement → deploy → result. HIGHEST-STAKES path.
# Each step is a SEPARATE explicit button press (a human gate); NOTHING is
# automatic. Approving only marks 'approved'; the principal must then tap
# 🛠 Implement (write code in a throwaway worktree, enforce protected paths, run
# the test gate) and, after reviewing the diff, 🚀 Deploy (push + merge + write the
# signal the VM deployer consumes — supervised deploy with auto-rollback). The
# bot never restarts/deploys itself in-process: that is the deployer's job, so it
# survives the bot dying. Every step is audited (database.log_si_audit) and the
# deploy result is reported by _deploy_result_sweep even across the mid-deploy
# restart. Worktree + branch are derived deterministically from the proposal id,
# so no in-memory state is needed between button presses.
# ─────────────────────────────────────────────────────────────────────────────

def _si_data_path(name: str) -> str:
    """A path next to the DB — equals <repo>/data on the VM, matching the deployer's
    SIGNAL_FILE/RESULT_FILE (REPO/data/...)."""
    return os.path.join(os.path.dirname(os.path.abspath(config.DATABASE_PATH)), name)


def _write_deploy_signal(payload: dict) -> None:
    """Write the deploy_request signal ATOMICALLY (temp file + os.replace + fsync).
    Two near-simultaneous writers (a proposal deploy and a manual /deploy) can no
    longer interleave and silently lose a request — the deployer's .path unit sees a
    complete file or nothing, never a half-written one."""
    path = _si_data_path("deploy_request.json")
    tmp = f"{path}.tmp"
    with open(tmp, "w") as f:
        json.dump(payload, f)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)


def _format_diffstat(diff: str) -> str:
    """Compact +N/-N file summary from a unified diff (for the review card)."""
    if not diff or not diff.strip():
        return "—"
    lines = diff.splitlines()
    files = sum(1 for ln in lines if ln.startswith("+++ "))
    adds = sum(1 for ln in lines if ln.startswith("+") and not ln.startswith("+++"))
    dels = sum(1 for ln in lines if ln.startswith("-") and not ln.startswith("---"))
    return f"{max(files, 1)} fayl · +{adds}/-{dels} qator"


def _si_gate2_keyboard(pid: str) -> InlineKeyboardMarkup:
    """Post-implementation review: inspect the diff, then Deploy or cancel."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Deploy qil", callback_data=f"sideploy:{pid}"),
         InlineKeyboardButton(text="👁 Diff", callback_data=f"sidiff:{pid}")],
        [InlineKeyboardButton(text="❌ Bekor qil", callback_data=f"sicancel:{pid}")],
    ])


async def _si_notify(bot: Bot, text: str,
                     reply_markup: "InlineKeyboardMarkup | None" = None) -> None:
    """Proactive Markdown message to the principal from a background SI worker.
    Falls back to plain text on a parse error; never raises."""
    try:
        await bot.send_message(config.PRINCIPAL_USER_ID, text,
                               parse_mode="Markdown", reply_markup=reply_markup)
    except TelegramBadRequest:
        try:
            await bot.send_message(config.PRINCIPAL_USER_ID, text,
                                   parse_mode=None, reply_markup=reply_markup)
        except Exception:
            logger.exception("SI notify failed")
    except Exception:
        logger.exception("SI notify failed")


# In-flight SI worker tasks — tracked so /freeze can HARD-cancel a running
# implement/deploy. The cooperative _si_frozen_abort() re-checks below are the
# graceful stop (bail before the next mutating step); this set is the backstop for
# a task already blocked on slow I/O (SDK call, git push).
_si_inflight_tasks: set = set()


def _si_spawn(coro) -> "asyncio.Task":
    """Spawn an SI worker tracked for /freeze cancellation."""
    task = asyncio.create_task(coro)
    _si_inflight_tasks.add(task)
    task.add_done_callback(_si_inflight_tasks.discard)
    return task


async def _si_frozen_abort(bot: Bot, pid: str, stage: str) -> bool:
    """Kill-switch re-check at an SI await boundary. If frozen → log + notify +
    return True so the worker bails BEFORE the next mutating/irreversible step."""
    if await _si_is_frozen():
        await database.log_si_audit("aborted_frozen", pid, stage)
        await _si_notify(bot, f"🧊 **To'xtatildi** (#{pid}) — '{stage}' bosqichida tizim muzlatilgan.")
        return True
    return False


async def _si_budget_exceeded(notify, op: str) -> bool:
    """Daily SI cost circuit-breaker. Returns True (and notifies via `notify(text)`)
    when today's SI-op count has reached config.SI_DAILY_OP_CAP — so the autonomous
    loop PAUSES instead of burning unbounded LLM spend. Otherwise records this op's
    start (counting it toward the cap) and returns False. Never blocks on a DB error."""
    since = datetime.now(database.TZ).replace(
        hour=0, minute=0, second=0, microsecond=0).isoformat()
    try:
        n = await database.si_daily_op_count(since)
    except Exception:
        return False  # never block the loop on a counting failure
    if n >= config.SI_DAILY_OP_CAP:
        await database.log_si_audit("budget_exceeded", None, f"{n}/{config.SI_DAILY_OP_CAP} ({op})")
        try:
            await notify(f"💸 **SI kunlik limiti** — bugun {config.SI_DAILY_OP_CAP} ta "
                         "o'z-o'zini yaxshilash amali bajarildi. Avtonom halqa to'xtatildi "
                         "(nazoratsiz xarajat oldini olish). Ertaga tiklanadi.")
        except Exception:
            pass
        return True
    await database.log_si_audit(op, None)
    return False


async def _si_run_implementation(bot: Bot, proposal: dict) -> None:
    """Gate-2 background worker: turn an APPROVED proposal into a reviewed diff in an
    isolated worktree. NEVER deploys. dev_agent.prepare enforces protected paths +
    the test gate and audits every step; here we just report the outcome."""
    import dev_agent
    pid = proposal.get("id", "?")
    if await _si_frozen_abort(bot, pid, "implement"):
        return
    if await _si_budget_exceeded(lambda t: _si_notify(bot, t), "implement_started"):
        return
    try:
        res = await dev_agent.prepare(proposal)
    except Exception as e:
        logger.exception("SI implementation crashed")
        await database.log_si_audit("implement_crash", pid, f"{type(e).__name__}: {e}")
        await _si_notify(bot, f"⚠️ **Implementatsiya xatosi** (#{pid})\n\n"
                              f"`{type(e).__name__}`. Worktree tozalandi, hech narsa deploy bo'lmadi.")
        return
    if res.ok:
        await _si_notify(
            bot,
            f"✅ **Kod tayyor — ko'rib chiqing** (#{pid})\n\n"
            f"📋 {proposal.get('title', '—')}\n"
            f"🌿 `{res.branch}`\n"
            f"🧪 Testlar: yashil\n"
            f"📝 {_format_diffstat(res.diff)}\n\n"
            "_`👁 Diff` bilan ko'ring. Tayyor bo'lsangiz **🚀 Deploy qil** — branch "
            "merge qilinadi va VM deployer nazorat ostida yangilaydi. Bekor qilsangiz "
            "worktree tozalanadi, hech narsa o'zgarmaydi._",
            reply_markup=_si_gate2_keyboard(pid))
        return
    # ── failure paths — explain, never deploy ──
    if res.protected_hits:
        await _si_notify(bot, f"⛔️ **Himoyalangan fayl tegildi** (#{pid})\n\n"
                              f"`{', '.join(res.protected_hits)}` — bu fayllar avtomatik "
                              "o'zgartirilmaydi. Taklif **qo'lda spec** uchun belgilandi.")
    elif res.test_summary and not res.tests_passed:
        await _si_notify(bot, f"🧪 **Testlar yiqildi — deploy bloklandi** (#{pid})\n\n"
                              f"```\n{res.test_summary[:1200]}\n```\n_Kod deploy qilinmadi._")
    else:
        await _si_notify(bot, f"ℹ️ **Implementatsiya bajarilmadi** (#{pid})\n\n{res.reason or '—'}")


async def _si_run_deploy(bot: Bot, proposal: dict) -> None:
    """Gate-3 background worker: record a health baseline, push + merge the prepared
    branch, then write the deploy_request signal the VM deployer consumes. It does
    NOT restart/deploy here — the standalone deployer does that (and rolls back on
    failure), so the supervision survives the bot being killed mid-deploy."""
    import dev_agent
    pid = proposal.get("id", "?")
    branch = f"si/{pid}"
    worktree = dev_agent._worktree_path(branch)
    if not os.path.exists(worktree):
        await _si_notify(bot, f"⚠️ **Deploy bekor** (#{pid})\n\nTayyor worktree topilmadi "
                              "(ehtimol bekor qilingan). Qayta **🛠 Implement** qiling.")
        return
    # Phase 6 baseline — capture health BEFORE the new code goes live; suggest-only
    # regression detection (deploy_feedback) compares against this.
    try:
        import feedback
        import metrics
        feedback.record_baseline(pid, await metrics.collect_signals(days=7))
    except Exception:
        logger.exception("baseline capture failed (continuing deploy)")

    title = (proposal.get("title") or "improvement")[:72]
    commit_msg = (f"si({pid}): {title}\n\n"
                  "Supervised self-improvement — approved by the principal (Gate 1-3).")
    if await _si_frozen_abort(bot, pid, "push"):
        return
    ok, msg = await dev_agent.push_branch(branch, worktree, commit_msg)
    if not ok:
        await database.log_si_audit("push_failed", pid, (msg or "")[:500])
        await _si_notify(bot, f"⛔️ **Push muvaffaqiyatsiz** (#{pid})\n\n`{(msg or '')[:300]}`\n"
                              "_Hech narsa deploy bo'lmadi._")
        return
    await database.update_proposal_status(pid, "pr_open")
    if await _si_frozen_abort(bot, pid, "merge"):
        return
    ok, url = await dev_agent.open_and_merge_pr(
        branch, title=f"[SI] {title}",
        body=f"Supervised self-improvement proposal #{pid}.\n\n"
             f"{proposal.get('proposed_change', '')}",
        auto_merge=True)
    if not ok:
        await database.log_si_audit("pr_failed", pid, (url or "")[:500])
        await _si_notify(bot, f"⛔️ **PR/merge muvaffaqiyatsiz** (#{pid})\n\n"
                              f"Branch `{branch}` push qilindi, ammo merge bo'lmadi — qo'lda "
                              f"ko'rib chiqing.\n`{(url or '')[:200]}`")
        return
    await database.update_proposal_status(pid, "merged")
    await database.log_si_audit("merged", pid, url or "")
    if await _si_frozen_abort(bot, pid, "deploy-signal"):
        return  # merged, but principal froze — do NOT trigger the live restart now
    # write the deploy signal the VM deployer's .path unit watches
    try:
        _write_deploy_signal({"target": None, "proposal_id": pid})
        await database.log_si_audit("deploy_signal_written", pid)
    except Exception as e:
        logger.exception("writing deploy signal failed")
        await _si_notify(bot, f"⚠️ **Deploy signali yozilmadi** (#{pid})\n\n"
                              f"`{type(e).__name__}`. Merge bo'ldi, lekin deployer ishga tushmadi.")
        return
    try:                                  # changes are merged → worktree no longer needed
        await dev_agent.cleanup_worktree(worktree)
    except Exception:
        pass
    pr_line = f"\n🔗 PR: {url}" if url else ""
    await _si_notify(bot, f"🚀 **Merge + deploy signali yuborildi** (#{pid}){pr_line}\n\n"
                          "VM deployer nazorat ostida yangilamoqda — buzilsa avtomatik eski "
                          "versiyaga qaytadi. Bot qayta ishga tushishi mumkin; natija tayyor "
                          "bo'lganda xabar beraman.")


# ── /freeze — kill-switch for the autonomous self-improvement loop ──────────────
# Pauses 🛠 Implement, 🚀 Deploy and the nightly diagnosis so a session can change
# code + /deploy without the bot ALSO writing to main (the divergence problem we
# hit repeatedly). Proposals can still be created (they queue); nothing
# executes/deploys while frozen. One source of truth at a time.
async def _si_is_frozen() -> bool:
    try:
        return bool((await database.get_settings()).get("si_frozen", False))
    except Exception:
        return False


_SI_FROZEN_ALERT = "🧊 Tizim muzlatilgan — /unfreeze bilan davom eting."


@router.message(Command("freeze"))
async def cmd_freeze(message: Message) -> None:
    """Kill-switch: pause the autonomous loop AND hard-stop any in-flight worker."""
    await database.set_setting("si_frozen", True)
    cancelled = 0
    for task in list(_si_inflight_tasks):
        if not task.done():
            task.cancel()
            cancelled += 1
    await database.log_si_audit("frozen", None, f"kill-switch ON; cancelled={cancelled}")
    stopped = f"\n\n⏹ {cancelled} ta ishlab turgan jarayon to'xtatildi." if cancelled else ""
    await message.answer(
        "🧊 **Muzlatildi.**\n\n"
        "Avtonom o'z-o'zini yaxshilash to'xtadi: 🛠 Implement, 🚀 Deploy va tungi "
        "diagnoz ishlamaydi. Takliflar yaratilaveradi (navbatда turadi)." + stopped + "\n\n"
        "Davom ettirish: /unfreeze", parse_mode="Markdown")


@router.message(Command("unfreeze"))
async def cmd_unfreeze(message: Message) -> None:
    """Resume the autonomous self-improvement loop."""
    await database.set_setting("si_frozen", False)
    await database.log_si_audit("unfrozen", None, "kill-switch OFF")
    await message.answer(
        "🟢 **Muzlatish olib tashlandi.**\n\n"
        "Avtonom tizim yana faol — Implement / Deploy / tungi diagnoz ishlaydi.",
        parse_mode="Markdown")


@router.callback_query(F.data.startswith("siimpl:"))
async def cb_si_implement(query: CallbackQuery) -> None:
    """Gate-2 🛠 — start implementation of an approved proposal (runs in background)."""
    pid = query.data.split(":", 1)[1]
    p = await database.get_improvement_proposal(pid)
    if not p:
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    if p.get("status") != "approved":
        await query.answer("Faqat tasdiqlangan taklif implement qilinadi.", show_alert=True)
        return
    if await _si_is_frozen():
        await query.answer(_SI_FROZEN_ALERT, show_alert=True)
        return
    await query.answer("🛠 Implementatsiya boshlandi")
    await database.log_si_audit("implement_requested", pid)
    try:
        await query.message.edit_text(
            f"⏳ **Implementatsiya boshlandi** (#{pid})\n\n"
            "Izolyatsiyalangan worktree'da kod yozilmoqda va testlar tekshirilmoqda. "
            "Tayyor bo'lganda diff bilan xabar beraman.", parse_mode="Markdown")
    except TelegramBadRequest:
        pass
    _si_spawn(_si_run_implementation(query.bot, p))


@router.callback_query(F.data.startswith("sidiff:"))
async def cb_si_diff(query: CallbackQuery) -> None:
    """👁 — show the prepared diff (truncated to Telegram message limits)."""
    import dev_agent
    pid = query.data.split(":", 1)[1]
    worktree = dev_agent._worktree_path(f"si/{pid}")
    if not os.path.exists(worktree):
        await query.answer("Worktree topilmadi (tozalangan yoki deploy bo'lgan).", show_alert=True)
        return
    await query.answer()
    diff = await dev_agent.get_diff(worktree)
    if not diff.strip():
        diff = "(diff bo'sh)"
    chunk, more = diff[:3500], ("\n\n… (qisqartirildi)" if len(diff) > 3500 else "")
    await _si_notify(query.bot, f"👁 **Diff — #{pid}**\n\n```diff\n{chunk}\n```{more}")


@router.callback_query(F.data.startswith("sicancel:"))
async def cb_si_cancel(query: CallbackQuery) -> None:
    """❌ — cancel a prepared change: clean up the worktree, mark the proposal rejected.
    Nothing was pushed/deployed, so this fully unwinds the attempt."""
    import dev_agent
    pid = query.data.split(":", 1)[1]
    worktree = dev_agent._worktree_path(f"si/{pid}")
    try:
        if os.path.exists(worktree):
            await dev_agent.cleanup_worktree(worktree)
    except Exception:
        logger.exception("cancel cleanup failed")
    await database.update_proposal_status(pid, "rejected")
    await database.log_si_audit("deploy_cancelled", pid)
    await query.answer("❌ Bekor qilindi")
    try:
        await query.message.edit_text(f"❌ **Bekor qilindi** (#{pid})\n\n"
            "Worktree tozalandi, hech narsa deploy bo'lmadi.", parse_mode="Markdown")
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("sideploy:"))
async def cb_si_deploy(query: CallbackQuery) -> None:
    """Gate-3 🚀 — push + merge + signal the supervised deployer (runs in background).
    Allowed only for a successfully-prepared proposal (status 'in_progress')."""
    pid = query.data.split(":", 1)[1]
    p = await database.get_improvement_proposal(pid)
    if not p:
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    if p.get("status") != "in_progress":
        await query.answer("Avval 🛠 Implement qiling (tayyor diff yo'q).", show_alert=True)
        return
    if await _si_is_frozen():
        await query.answer(_SI_FROZEN_ALERT, show_alert=True)
        return
    await query.answer("🚀 Deploy boshlandi")
    await database.log_si_audit("deploy_requested", pid)
    try:
        await query.message.edit_text(
            f"🚀 **Deploy boshlandi** (#{pid})\n\n"
            "Branch push → PR merge → deploy signali. VM deployer nazorat ostida "
            "yangilaydi (buzilsa avtomatik orqaga qaytaradi). Natijani kutamiz.",
            parse_mode="Markdown")
    except TelegramBadRequest:
        pass
    _si_spawn(_si_run_deploy(query.bot, p))


# ── /silog — self-improvement status + audit trail (read-only, Telegram-native) ──
# Fills the gap: /improvements shows only PENDING proposals; this surfaces in-flight
# AND completed ones (deployed/reverted/rejected) plus each proposal's full audit
# chain — so the principal verifies "did it finish / deploy?" from the phone, no DB
# query. Read-only: it never changes state.
_SI_STATUS_BADGE = {
    "new": "🆕", "approved": "👍", "in_progress": "🛠", "pr_open": "📤",
    "merged": "🔀", "deployed": "✅", "reverted": "↩️", "rejected": "❌",
    "requires_manual": "⚠️", "done": "✔️",
}
_SI_ACTIVE_STATUSES = {"new", "approved", "in_progress", "pr_open", "merged"}


def _format_silog_text(props: list, frozen: bool = False) -> str:
    """The /silog dashboard: in-flight proposals + recently completed ones.
    PLAIN text (sent WITHOUT parse_mode): statuses ('in_progress') and user titles
    ('check_followups …') contain underscores that legacy Markdown mis-parses as
    italics → 'can't find end of entity'. No markup here = no parse errors ever."""
    lines = ["📜 Self-improvement — holat", ""]
    if frozen:
        lines += ["🧊 MUZLATILGAN — Implement/Deploy to'xtatilgan (/unfreeze).", ""]
    if not props:
        lines.append("Hozircha hech qanday taklif yo'q.")
        return "\n".join(lines)
    active = [p for p in props if p.get("status") in _SI_ACTIVE_STATUSES]
    done = [p for p in props if p.get("status") not in _SI_ACTIVE_STATUSES]
    if active:
        lines.append("⏳ Faol:")
        for p in active:
            b = _SI_STATUS_BADGE.get(p.get("status"), "•")
            lines.append(f"  {b} {p['id']} · {p.get('status')} — {(p.get('title') or '—')[:38]}")
        lines.append("")
    if done:
        lines.append("✅ Tugaganlar (oxirgi):")
        for p in done[:6]:
            b = _SI_STATUS_BADGE.get(p.get("status"), "•")
            lines.append(f"  {b} {p['id']} · {p.get('status')} — {(p.get('title') or '—')[:38]}")
    lines += ["", "🔍 tugma — audit zanjirini ko'rish."]
    return "\n".join(lines)


def _silog_keyboard(props: list) -> InlineKeyboardMarkup:
    rows, row = [], []
    for p in props[:6]:                       # drill-in buttons for the 6 most recent
        row.append(InlineKeyboardButton(text=f"🔍 {p['id']}", callback_data=f"siaudit:{p['id']}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="🔄 Yangilash", callback_data="silog:refresh")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(Command("silog"))
async def cmd_silog(message: Message) -> None:
    """Self-improvement status + audit view (read-only)."""
    props = await database.list_improvement_proposals(limit=20)
    await _safe_answer(message, _format_silog_text(props, frozen=await _si_is_frozen()),
                       parse_mode=None, reply_markup=_silog_keyboard(props))


@router.callback_query(F.data == "silog:refresh")
async def cb_silog_refresh(query: CallbackQuery) -> None:
    props = await database.list_improvement_proposals(limit=20)
    await query.answer("🔄 Yangilandi")
    try:
        await query.message.edit_text(_format_silog_text(props, frozen=await _si_is_frozen()),
                                      parse_mode=None, reply_markup=_silog_keyboard(props))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("siaudit:"))
async def cb_si_audit(query: CallbackQuery) -> None:
    """Show one proposal's full audit chain (chronological)."""
    pid = query.data.split(":", 1)[1]
    p = await database.get_improvement_proposal(pid)
    if not p:
        await query.answer("Taklif topilmadi.", show_alert=True)
        return
    entries = await database.list_si_audit(limit=40, proposal_id=pid)
    await query.answer()
    badge = _SI_STATUS_BADGE.get(p.get("status"), "•")
    # PLAIN text — audit action names ('deploy_succeeded') and titles contain
    # underscores that legacy Markdown would mis-parse. No markup = no parse errors.
    lines = [f"🔍 {pid} — audit zanjiri", "",
             f"📋 {(p.get('title') or '—')[:50]}",
             f"holat: {badge} {p.get('status')}", ""]
    if entries:
        for a in reversed(entries):           # list_si_audit is id DESC → reverse to chronological
            ts = (a.get("ts") or "")[11:19]    # HH:MM:SS
            detail = (a.get("detail") or "").strip()
            tail = f" — {detail[:40]}" if detail else ""
            lines.append(f"{ts}  {a.get('action')}{tail}")
    else:
        lines.append("Audit yozuvi yo'q.")
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⬅️ Orqaga", callback_data="silog:refresh")]])
    try:
        await query.message.edit_text("\n".join(lines), parse_mode=None, reply_markup=kb)
    except TelegramBadRequest:
        pass


# ── /deploy — manual supervised deploy of the latest main ──────────────────────
# For changes made OUTSIDE the /improve flow (e.g. a careful session pushes to
# main). Writes the same signal the deployer consumes → it pulls main, restarts,
# health-checks, and auto-rolls-back on failure. Confirmation-gated.
@router.message(Command("deploy"))
async def cmd_deploy(message: Message) -> None:
    """Trigger a supervised deploy of the latest main via the deployer."""
    await _safe_answer(
        message,
        "🚀 **Deploy — eng so'nggi `main`**\n\n"
        "Deployer eng so'nggi kodni tortadi (`git pull`), botni qayta ishga tushiradi "
        "va sog'lig'ini tekshiradi. Buzilsa — **avtomatik orqaga qaytaradi**.\n\n"
        "Davom etamizmi?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="🚀 Ha, deploy", callback_data="mdeploy:yes"),
            InlineKeyboardButton(text="✕ Yo'q", callback_data="mdeploy:no"),
        ]]))


@router.callback_query(F.data == "mdeploy:yes")
async def cb_manual_deploy(query: CallbackQuery) -> None:
    """Write the deploy signal (target=None → git pull main). The VM deployer does the
    supervised restart + rollback; _deploy_result_sweep reports the outcome."""
    if await _si_is_frozen():
        await query.answer(_SI_FROZEN_ALERT, show_alert=True)
        return
    try:
        _write_deploy_signal({"target": None, "proposal_id": "manual"})
        await database.log_si_audit("manual_deploy_requested", "manual")
        ok, detail = True, ""
    except Exception as e:
        logger.exception("manual deploy signal write failed")
        ok, detail = False, type(e).__name__
    await query.answer("🚀 Deploy boshlandi" if ok else "Xato")
    try:
        if ok:
            await query.message.edit_text(
                "🚀 **Deploy signali yuborildi**\n\n"
                "Deployer eng so'nggi `main`ni tortib, botni qayta ishga tushiryapti "
                "(health-check + rollback bilan). Natijani xabar qilaman.",
                parse_mode="Markdown")
        else:
            await query.message.edit_text(f"⚠️ Signal yozilmadi ({detail}).", parse_mode=None)
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "mdeploy:no")
async def cb_manual_deploy_cancel(query: CallbackQuery) -> None:
    await query.answer("Bekor qilindi")
    try:
        await query.message.edit_text("✕ Deploy bekor qilindi.", parse_mode=None)
    except TelegramBadRequest:
        pass


# All reply-keyboard button labels (collected from the *BTN* constants above) —
# used to stop a tapped navigation button from being captured as a note.
_RESERVED_LABELS = {v for k, v in dict(globals()).items()
                    if "BTN" in k and isinstance(v, str) and v}


def _is_note_noise(text: str) -> bool:
    """True if `text` is NOT a real note: empty, a slash command, or a tapped
    reply-keyboard / section button label (e.g. '⬅️ Asosiy menyu', '/team').
    Prevents junk notes from accidental taps/commands."""
    t = (text or "").strip()
    return ((not t) or t.startswith("/")
            or t in _RESERVED_LABELS or t in _SECTION_LABELS)


def _looks_like_bot_output(text: str) -> bool:
    """Heuristic: `text` is THIS bot's own rendered panel forwarded back to it
    (e.g. '📌 VAZIFALAR … Natija · 9 ta … ━━━'), not an external note. Bot panels
    use long ━ dividers together with 'Natija ·' / 'Ko'rinish ·' / 'UMUMIY HOLAT'
    labels that a human-written note would not contain."""
    t = text or ""
    return ("━━━" in t) and any(
        m in t for m in ("Natija ·", "Ko'rinish ·", "Koʻrinish ·", "UMUMIY HOLAT"))


@router.message(StateFilter(NoteCaptureFSM.awaiting_text), F.text | F.voice)
async def handle_note_capture(message: Message, state: FSMContext) -> None:
    """One-shot capture: next text/voice after /qayd or '➕ Yangi qayd' button."""
    content = await _get_text_or_transcribe(message)
    if not content:
        return
    content = content.strip()
    if not content:
        await message.answer("Bo'sh xabar — qayd yaratilmadi.")
        return
    # A tapped nav button or command isn't a note — cancel capture, let it route.
    if _is_note_noise(content):
        await state.clear()
        await message.answer("✕ Note kiritish bekor qilindi.", reply_markup=main_reply_keyboard())
        return
    await state.clear()
    source = "voice" if message.voice else "manual"
    nid = await database.create_note({"content": content, "source": source})
    await _note_capture_reply(message, nid,
                                "ovoz orqali" if source == "voice" else "qo'lda")
    # Restore section if we were in one
    await _maybe_refresh_section(message, state, {"note": [nid]})


async def _capture_new_item(message: Message, state: FSMContext, intent_uz: str, kind: str) -> None:
    """One-shot capture for '➕ Yangi vazifa/uchrashuv': the next text/voice is routed
    to the LLM WITH an explicit create intent, so it can't be reinterpreted as an
    unrelated command. A tapped nav button / command cancels the capture and routes."""
    content = (await _get_text_or_transcribe(message) or "").strip()
    if not content:
        await message.answer(f"Bo'sh xabar — {kind} yaratilmadi.")
        return
    if _is_note_noise(content):  # a tapped button/command isn't the item text
        await state.clear()
        await message.answer("✕ Bekor qilindi.", reply_markup=main_reply_keyboard())
        return
    await state.clear()
    await _process_and_reply(message, f"{intent_uz}: {content}", state=state)


@router.message(StateFilter(NewMeetingTextFSM.awaiting_text), F.text | F.voice)
async def handle_new_meeting_capture(message: Message, state: FSMContext) -> None:
    await _capture_new_item(message, state, "Yangi uchrashuv qo'sh", "uchrashuv")


@router.message(StateFilter(NewTaskTextFSM.awaiting_text), F.text | F.voice)
async def handle_new_task_capture(message: Message, state: FSMContext) -> None:
    await _capture_new_item(message, state, "Yangi vazifa qo'sh", "vazifa")


async def _note_capture_reply(message: Message, note_id: str, source_hint: str) -> None:
    """Compact confirmation right after a note is captured."""
    # 'Inbox'ga' to'liq bo'limga kiradi (nav_notes → reply-filtrlar bilan), shunda
    # ro'yxat doimo reply navigatsiya bilan birga ko'rinadi (notesfilter endi faqat
    # pagination uchun — har doim bo'lim ichida).
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="📥 Inbox'ga", callback_data="nav_notes"),
        InlineKeyboardButton(text="🤖 Hozir tahlil", callback_data=f"noteanalyze:{note_id}"),
    ]])
    await message.answer(
        f"📝 **Qayd saqlandi** · `{note_id}`\n_Manba: {source_hint}_",
        parse_mode="Markdown",
        reply_markup=kb,
    )


@router.callback_query(F.data == "nav_notes")
async def cb_nav_notes(query: CallbackQuery, state: FSMContext) -> None:
    await query.answer()
    await cmd_notes(query.message, state)


def _html_escape(s: str) -> str:
    """Minimal HTML escape for Telegram parse_mode=HTML. The allowed tag set
    is small (b, i, u, s, code, pre, a, blockquote, tg-spoiler), so we only
    need to neutralise &, <, > in user-supplied text."""
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;"))


def _sanitize_forward_html(html: str) -> str:
    """Make a forwarded message's aiogram html_text safe to RE-SEND inside our
    own <blockquote expandable>. Two constructs make Telegram reject the whole
    message (and then we'd fall back to showing raw tags — the reported
    'meaning changed' bug):

      1. <tg-emoji> — a bot can't resend a premium/custom emoji it doesn't own.
         Keep the inner unicode fallback char, drop the wrapper.
      2. Nested <blockquote> — Telegram forbids blockquote-in-blockquote, and we
         already wrap the body in one. Flatten any inner blockquote tags.
    """
    import re as _re
    if not html:
        return html
    html = _re.sub(r'<tg-emoji\b[^>]*>(.*?)</tg-emoji>', r'\1', html, flags=_re.S)
    html = _re.sub(r'</?blockquote\b[^>]*>', '', html)
    return html


def _format_note_detail(note: dict) -> tuple[str, str]:
    """Full-content card shown when a note is opened via noteopen:{id}.
    Returns (html_text, parse_mode). Uses Telegram's native <blockquote>
    so forwarded content renders as a real quote bubble — vertical bar +
    distinct background — and original bold/italic/links survive intact."""
    badge = _NOTES_SOURCE_BADGE.get(note.get("source", "manual"), "📝")
    when = _short_local_date(note.get("created_at"))
    status_uz = {
        "inbox": "📥 Inbox",
        "processed": "⚙️ Ishlangan",
        "archived": "📦 Arxiv",
    }.get(note.get("status"), note.get("status") or "—")

    title = (note.get("title") or "Note").strip()[:80]
    parts: list[str] = [
        f"📝 <b>{_html_escape(title)}</b>",
        "",
        # Bitta ixcham meta qator: manba · sana · holat
        f"{_html_escape(badge)} · 📅 {_html_escape(when)} · {_html_escape(status_uz)}",
    ]
    tags = note.get("tags") or []
    if tags:
        parts.append("🏷 " + " ".join(f"#{_html_escape(str(t))}" for t in tags[:8]))
    if note.get("source_chat"):
        parts.append(f"💬 Chat: <i>{_html_escape(note['source_chat'])}</i>")
    if note.get("source_author"):
        parts.append(f"👤 Muallif: <i>{_html_escape(note['source_author'])}</i>")
    if note.get("converted_to_type") and note.get("converted_to_id"):
        link_label = "Vazifa" if note["converted_to_type"] == "task" else "Eslatma"
        parts.append(
            f"🔗 {link_label}: <code>{_html_escape(note['converted_to_id'])}</code>"
        )
    # GTD: inbox'da qancha kun turgani — eski qaydni qayta ishlashga turtki
    if note.get("status") == "inbox":
        try:
            _cd = datetime.fromisoformat(note.get("created_at")).astimezone(database.TZ)
            _age = (datetime.now(database.TZ) - _cd).days
            if _age >= 1:
                parts.append(f"⏳ <b>{_age} kun</b> inbox'da kutyapti")
        except (ValueError, TypeError):
            pass
    parts.append("")

    # Body: prefer content_html (original Telegram entities preserved) for
    # forwards; fall back to plain content. Render inside a native
    # <blockquote expandable> so long quotes can be collapsed.
    html_body = note.get("content_html")
    plain = (note.get("content") or "").strip()
    if html_body and (note.get("source") == "forward" or html_body.strip()):
        body = _sanitize_forward_html(html_body.strip())
    else:
        body = _html_escape(plain)
    # Hard cap: Telegram message limit is ~4096 chars including markup;
    # blockquote'ning expandable variant uzun matnlarni o'rab qo'yadi.
    if len(body) > 3000:
        body = body[:3000] + "\n\n<i>…(qisqartirildi)</i>"
    parts.append(f"<blockquote expandable>{body}</blockquote>")
    return "\n".join(parts), "HTML"


@router.callback_query(F.data.startswith("noteopen:"))
async def cb_note_open(query: CallbackQuery) -> None:
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer()
    # Mark the opened note as "current" so "shu qaydni ..." resolves to it.
    claude_service.set_last_note_view(
        [{"n": 1, "id": note["id"], "title": (note.get("title") or note.get("content") or "—")[:50]}])
    text, parse_mode = _format_note_detail(note)
    await _safe_answer(
        query.message,
        text,
        parse_mode=parse_mode,
        reply_markup=note_detail_menu(note),
    )


@router.callback_query(F.data.startswith("notearchive:"))
async def cb_note_archive(query: CallbackQuery) -> None:
    nid = query.data.split(":", 1)[1]
    ok = await database.archive_note(nid)
    if not ok:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer("📦 Arxivga ko'chirildi ✅")
    note = await database.get_note(nid)
    if note:
        text, parse_mode = _format_note_detail(note)
        try:
            await query.message.edit_text(
                text, parse_mode=parse_mode,
                reply_markup=note_detail_menu(note),
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("notedone:"))
async def cb_note_done(query: CallbackQuery) -> None:
    """'✅ Ishlandi' — mark a note processed WITHOUT converting (GTD: reviewed,
    no action needed). The note leaves the inbox and moves to '⚙️ Ishlangan'."""
    nid = query.data.split(":", 1)[1]
    ok = await database.mark_note_done(nid)
    if not ok:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer("✅ Ishlandi deb belgilandi")
    note = await database.get_note(nid)
    if note:
        text, parse_mode = _format_note_detail(note)
        try:
            await query.message.edit_text(
                text, parse_mode=parse_mode,
                reply_markup=note_detail_menu(note),
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("noterestore:"))
async def cb_note_restore(query: CallbackQuery) -> None:
    """Restore an archived or processed note back to inbox."""
    nid = query.data.split(":", 1)[1]
    ok = await database.update_note(nid, {
        "status": "inbox",
        "converted_to_id": None,
        "converted_to_type": None,
    })
    if not ok:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer("📥 Inbox'ga qaytarildi ✅")
    note = await database.get_note(nid)
    if note:
        text, parse_mode = _format_note_detail(note)
        try:
            await query.message.edit_text(
                text, parse_mode=parse_mode,
                reply_markup=note_detail_menu(note),
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("notedelete:"))
async def cb_note_delete(query: CallbackQuery) -> None:
    """Two-tap delete: first tap shows confirm, second tap (notedelconfirm:)
    actually deletes. Matches the safety pattern used for task deletes."""
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer()
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🗑 Ha, o'chirilsin", callback_data=f"notedelconfirm:{nid}"),
        InlineKeyboardButton(text="✕ Bekor", callback_data=f"noteopen:{nid}"),
    ]])
    title = (note.get("title") or "note")[:60]
    await query.message.answer(
        f"⚠️ **«{title}»** note o'chirilsinmi?\nBu amalni qaytarib bo'lmaydi.",
        parse_mode="Markdown",
        reply_markup=confirm_kb,
    )


@router.callback_query(F.data.startswith("notedelconfirm:"))
async def cb_note_delete_confirm(query: CallbackQuery) -> None:
    nid = query.data.split(":", 1)[1]
    ok = await database.delete_note(nid)
    if not ok:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer("🗑 O'chirildi ✅")
    try:
        await query.message.edit_text("🗑 Note o'chirildi.")
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("notetotask:"))
async def cb_note_to_task(query: CallbackQuery) -> None:
    """Convert a note straight into a task. Title = note.title, description =
    note.content, deadline=NULL (user can edit afterwards), priority=P2,
    tags include `qayd:{note_id}` for back-reference."""
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    title = (note.get("title") or note.get("content", "Qayddan vazifa")).strip()[:200]
    description = (note.get("content") or "").strip()
    if description == title:
        description = None
    tags = list(note.get("tags") or []) + [f"note:{nid}"]
    tid = await database.create_task({
        "title": title,
        "description": description,
        "priority": "P2",
        "status": "todo",
        "tags": tags,
        "source": "note",
    })
    await database.mark_note_processed(nid, "task", tid)
    await query.answer("📝 Vazifa yaratildi ✅")
    task = await database.get_task(tid)
    if task:
        await _safe_answer(
            query.message,
            "📝 **Note vazifaga aylantirildi**\n\n" + _format_task_card(task),
            parse_mode="Markdown",
            reply_markup=_task_card_kb_with_back(task),
        )


@router.callback_query(F.data.startswith("notetorem:"))
async def cb_note_to_reminder(query: CallbackQuery, state: FSMContext) -> None:
    """Convert a note into a reminder. We don't know the remind_at time, so
    start the NewReminderFSM pre-filled with the note's content as title."""
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    title = (note.get("title") or note.get("content", "Qayddan eslatma")).strip()[:200]
    # Hop into the existing new-reminder flow with title pre-set; the user
    # picks the time, and on submit we run mark_note_processed via a tag.
    await state.set_state(NewReminderFSM.awaiting_time)
    await state.update_data(
        title=title,
        from_note_id=nid,
    )
    await query.answer()
    await query.message.answer(
        f"⏰ **YANGI ESLATMA**\n\nMavzu: «{title[:60]}»\n\n"
        "Eslatma vaqtini tanlang yoki yozing:\n"
        "_Misol: «ertaga 14:00», «1 soatdan keyin»_",
        parse_mode="Markdown",
    )


def _build_note_split_directive(content: str) -> str:
    """Directive: split a note's text into one create_task per actionable item."""
    return (
        "[INTERNAL] split_note_into_tasks\n\n"
        "Quyidagi qayd matnini o'qib, undagi HAR BIR alohida bajariladigan ishni "
        "ALOHIDA `create_task` action qilib chiqar. FAQAT matnda BOR ishlardan "
        "foydalan — yangi ish, mas'ul yoki sana O'YLAB TOPMA. Agar matn aslida "
        "bitta ish bo'lsa — bitta action qaytar. Har task uchun: title (qisqa, "
        "aniq, buyruq ohangida), priority ('P2' default), deadline (matnda aniq "
        "sana bo'lsa ISO8601, aks holda null). user_message — bitta qisqa qator.\n\n"
        f"QAYD MATNI:\n{content}"
    )


@router.callback_query(F.data.startswith("notesplit:"))
async def cb_note_split(query: CallbackQuery, state: FSMContext) -> None:
    """'✂️ Ko'p vazifaga ajrat' — LLM qayd matnidan bir nechta vazifa ajratadi,
    tasdiq so'raydi; tasdiqlangach hammasini yaratadi + qaydni 'ishlangan' qiladi."""
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    content = (note.get("content") or "").strip()
    if not content:
        await query.answer("Qayd bo'sh", show_alert=True)
        return
    await query.answer("✂️ Ajratilmoqda…")
    typing = asyncio.create_task(_keep_typing(query.bot, query.message.chat.id))
    try:
        response = await claude_service.process_message(
            "", internal_directive=_build_note_split_directive(content), complexity="fast")
    finally:
        typing.cancel()
    actions = [a for a in (response.get("actions") or []) if a.get("type") == "create_task"]
    base_tags = list(note.get("tags") or []) + [f"note:{nid}"]
    if not actions:
        await query.message.answer("Vazifa topilmadi — qayd matnida aniq ish yo'q.")
        return
    if len(actions) == 1:
        # Bitta ish — to'g'ridan yaratamiz (oddiy 'Vazifaga' kabi)
        d = actions[0].get("data") or {}
        tid = await database.create_task({
            "title": (d.get("title") or content)[:200], "description": None,
            "priority": d.get("priority") or "P2", "deadline": d.get("deadline"),
            "status": "todo", "tags": base_tags, "source": "note",
        })
        await database.mark_note_processed(nid, "task", tid)
        await query.message.answer("📝 Bitta vazifa yaratildi ✅ _(qayd ishlangan)_",
                                    parse_mode="Markdown")
        return
    # 2+ ish → tasdiq darvozasi (standart create-preview)
    await state.update_data(split_note_id=nid, split_actions=actions)
    preview = await _format_create_preview(actions)
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text=f"✅ {len(actions)} ta vazifa yaratish", callback_data=f"notesplitok:{nid}"),
        InlineKeyboardButton(text="✕ Bekor", callback_data=f"noteopen:{nid}"),
    ]])
    await query.message.answer(preview, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("notesplitok:"))
async def cb_note_split_ok(query: CallbackQuery, state: FSMContext) -> None:
    """Tasdiqlangan ko'p-vazifa: hammasini yaratadi + qaydni 'ishlangan' qiladi."""
    nid = query.data.split(":", 1)[1]
    data = await state.get_data()
    actions = data.get("split_actions") or []
    if data.get("split_note_id") != nid or not actions:
        await query.answer("Muddati o'tdi — qaydni qayta oching.", show_alert=True)
        return
    note = await database.get_note(nid)
    base_tags = list((note or {}).get("tags") or []) + [f"note:{nid}"]
    created = []
    for a in actions:
        d = a.get("data") or {}
        title = (d.get("title") or "").strip()
        if not title:
            continue
        tid = await database.create_task({
            "title": title[:200], "description": None,
            "priority": d.get("priority") or "P2", "deadline": d.get("deadline"),
            "status": "todo", "tags": base_tags, "source": "note",
        })
        created.append(tid)
    if created:
        await database.mark_note_processed(nid, "task", created[0])
    await state.update_data(split_note_id=None, split_actions=None)
    await query.answer(f"✅ {len(created)} ta vazifa yaratildi")
    try:
        await query.message.edit_text(
            f"✅ **{len(created)} ta vazifa yaratildi** _(qayd ishlangan)_\n\n"
            "Vazifalar bo'limida ko'rishingiz mumkin.",
            parse_mode="Markdown",
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("noteanalyze:"))
async def cb_note_analyze(query: CallbackQuery, state: FSMContext) -> None:
    """Send the note content to Claude (fast model) for a 2-line summary +
    one concrete next-action suggestion. The suggested action then flows
    through the existing create-confirm gate."""
    nid = query.data.split(":", 1)[1]
    note = await database.get_note(nid)
    if not note:
        await query.answer("Qayd topilmadi", show_alert=True)
        return
    await query.answer("🤖 Tahlil qilinmoqda...")
    content = (note.get("content") or "").strip()
    if not content:
        await query.message.answer("Bo'sh qayd — tahlil qilish mumkin emas.")
        return
    typing = asyncio.create_task(_keep_typing(query.bot, query.message.chat.id))
    try:
        directive = (
            "[INTERNAL] analyze_note\n\n"
            "Quyidagi qayd matni — uni qisqa tahlil qiling:\n"
            "  1. 1-2 satr xulosa (mavzu nima haqida)\n"
            "  2. 1 ta aniq next-action taklif (vazifa yaratish kerakmi? "
            "     eslatma kerakmi? hech narsa kerakmasmi?)\n"
            "Agar action kerak deb topsangiz — javob ichiga create_task "
            "yoki create_reminder action qo'shing. user_message qisqa "
            "Uzbek xulosa bo'lsin."
            f"\n\nQAYD:\n{content}"
        )
        response = await claude_service.process_message(
            "", internal_directive=directive, complexity="fast",
        )
    finally:
        typing.cancel()
    text = (response.get("user_message") or "").strip() or "_Tahlil bo'sh._"
    actions = response.get("actions", [])
    destructive = [a for a in actions if a.get("type") in _DESTRUCTIVE_ACTION_TYPES]
    # Show the analysis, then ask to CONFIRM before creating anything (no longer
    # auto-creates silently). Reuses the standard acts_confirm pipeline; the
    # _note_id lets it mark this note processed once the action is executed.
    await _safe_answer(query.message, f"🤖 **Tahlil:**\n\n{text}", parse_mode="Markdown")
    if destructive:
        preview = await _format_create_preview(destructive)
        await state.set_state(CreateActionConfirmFSM.awaiting)
        await state.update_data(
            pending_response={"actions": destructive,
                              "user_message": "✅ Tayyor.", "buttons": []},
            _prior_section=None,
            _note_id=nid,
        )
        confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Yarataman", callback_data="acts_confirm"),
            InlineKeyboardButton(text="✕ Yo'q", callback_data="acts_cancel"),
        ]])
        await _safe_answer(query.message, preview, parse_mode="Markdown", reply_markup=confirm_kb)


@router.message(Command("recurring"))
async def cmd_recurring(message: Message) -> None:
    tasks = await database.list_recurring_tasks(limit=30)
    text = _format_tasks_compact(tasks, "Takroriy vazifalar")
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=tasks_compact_keyboard(tasks, "all"))


@router.callback_query(F.data.startswith("taskfilter:"))
async def cb_task_filter(query: CallbackQuery) -> None:
    # Format: taskfilter:<key>  or  taskfilter:<key>:<page>
    parts = query.data.split(":")
    filt = parts[1] if len(parts) > 1 else "active"
    try:
        page = int(parts[2]) if len(parts) > 2 else 1
    except ValueError:
        page = 1
    await query.answer()
    await _render_tasks_for_filter(query.message, filt, page=page, edit_existing=True)



@router.message(StateFilter(TaskSearchFSM.awaiting_query), F.text | F.voice)
async def handle_task_search(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    await state.clear()
    query = message.text.strip()
    if not query:
        await message.answer("Qidiruv so'zi bo'sh bo'lmasin.", reply_markup=single_back_keyboard("taskfilter:active"))
        return
    found = (await database.search_all(query, limit=30)).get("tasks", [])
    text = _format_tasks_compact(found, f"Qidiruv: {query}")
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=tasks_compact_keyboard(found, "search"))


@router.message(StateFilter(ReminderSearchFSM.awaiting_query), F.text | F.voice)
async def handle_reminder_search(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    await state.clear()
    query = message.text.strip()
    if not query:
        await message.answer("Qidiruv so'zi bo'sh bo'lmasin.")
        return
    found = await database.search_reminders(query, limit=50)
    text = _format_reminders_compact(found, f"Qidiruv: {query}", page=1)
    await _safe_answer(
        message,
        text,
        parse_mode="Markdown",
        reply_markup=reminders_compact_keyboard(found, "search"),
    )


@router.callback_query(F.data.startswith("remopen:"))
async def cb_reminder_open(query: CallbackQuery) -> None:
    rid = query.data.split(":", 1)[1]
    reminder = await database.get_reminder(rid)
    if not reminder:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer()
    # Mark the opened reminder as "current" so "shu eslatmani ..." resolves to it.
    claude_service.set_last_reminder_view(
        [{"n": 1, "id": reminder["id"], "title": (reminder.get("title") or "—")[:50]}])
    text = _format_reminder_card(reminder)
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))


def _compute_snooze_time(choice: str) -> str | None:
    now = datetime.now(database.TZ)
    if choice == "15m":
        return (now + timedelta(minutes=15)).replace(second=0, microsecond=0).isoformat()
    if choice == "1h":
        return (now + timedelta(hours=1)).replace(second=0, microsecond=0).isoformat()
    if choice == "1d":
        return (now + timedelta(days=1)).replace(second=0, microsecond=0).isoformat()
    return None


@router.callback_query(F.data.startswith("remdone:"))
async def cb_reminder_done(query: CallbackQuery) -> None:
    rid = query.data.split(":", 1)[1]
    ok = await database.complete_reminder(rid)
    if not ok:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer("Bajarildi ✅")
    reminder = await database.get_reminder(rid)
    if reminder:
        try:
            await query.message.edit_text(
                "✅ **ESLATMA BAJARILDI**\n" + _SEP + "\n\n" + _format_reminder_card(reminder),
                parse_mode="Markdown",
                reply_markup=reminder_detail_menu(reminder),
            )
        except TelegramBadRequest:
            await _safe_answer(query.message, "✅ Eslatma bajarildi.")


@router.callback_query(F.data.startswith("remskip:"))
async def cb_reminder_skip(query: CallbackQuery) -> None:
    """⏭ Bu safarni o'tkaz — takroriy eslatmani KEYINGI takroriga suradi (bu
    martagisi bajarilgan hisoblanadi), seriya saqlanadi. Takror bo'lmasa —
    oddiy 'bajarildi' kabi yopadi."""
    rid = query.data.split(":", 1)[1]
    reminder = await database.get_reminder(rid)
    if not reminder:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    rule = database.normalize_recurrence_rule(reminder.get("recurrence_rule"))
    next_at = database.compute_next_recurrence(reminder.get("remind_at"), rule) if rule else None
    if next_at:
        await database.update_reminder(rid, {"remind_at": next_at, "status": "scheduled", "sent_at": None})
        reminder = await database.get_reminder(rid)
        await query.answer(f"⏭ O'tkazildi · keyingisi: {_reminder_time_chip(reminder)}")
        header = "⏭ **KEYINGI TAKRORGA O'TKAZILDI**"
    else:
        await database.complete_reminder(rid)
        reminder = await database.get_reminder(rid)
        await query.answer("Takror tugadi ✅")
        header = "✅ **BAJARILDI**"
    if reminder:
        try:
            await query.message.edit_text(
                header + "\n" + _SEP + "\n\n" + _format_reminder_card(reminder),
                parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("remstop:"))
async def cb_reminder_stop(query: CallbackQuery) -> None:
    """🛑 Takrorni to'xtatish — butun takroriy seriyani tugatadi (status=done,
    scheduler boshqa ishga tushirmaydi)."""
    rid = query.data.split(":", 1)[1]
    ok = await database.complete_reminder(rid)
    if not ok:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer("🛑 Takror to'xtatildi")
    reminder = await database.get_reminder(rid)
    if reminder:
        try:
            await query.message.edit_text(
                "🛑 **TAKROR TO'XTATILDI**\n" + _SEP + "\n\n" + _format_reminder_card(reminder),
                parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("remsnooze:"))
async def cb_reminder_snooze(query: CallbackQuery) -> None:
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    rid, choice = parts[1], parts[2]
    remind_at = _compute_snooze_time(choice)
    if not remind_at:
        await query.answer("Snooze vaqti noto'g'ri", show_alert=True)
        return
    ok = await database.snooze_reminder(rid, remind_at)
    if not ok:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    reminder = await database.get_reminder(rid)
    label = _reminder_time_chip(reminder or {"remind_at": remind_at})
    await query.answer(f"Keyingi eslatma: {label} ✅")
    if reminder:
        try:
            await query.message.edit_text(
                "⏰ **ESLATMA KO'CHIRILDI**\n" + _SEP + "\n\n" + _format_reminder_card(reminder),
                parse_mode="Markdown",
                reply_markup=reminder_detail_menu(reminder),
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("remdel:"))
async def cb_reminder_delete(query: CallbackQuery) -> None:
    rid = query.data.split(":", 1)[1]
    ok = await database.cancel_reminder(rid)
    if not ok:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer("O'chirildi ✅")
    try:
        await query.message.edit_text("🗑 Eslatma o'chirildi.", reply_markup=single_back_keyboard("remfilter:active", "⬅️ Ro'yxatga"))
    except TelegramBadRequest:
        await _safe_answer(query.message, "🗑 Eslatma o'chirildi.")


# Takror tahriri uchun pick-list variantlari (val, label).
_RECUR_OPTIONS = [
    ("daily", "⏰ Kunlik"),
    ("weekdays", "📅 Ish kunlari (Dush–Juma)"),
    ("weekly", "📆 Haftalik"),
    ("monthly", "🗓 Oylik"),
    ("quarterly", "📊 Choraklik"),
    ("yearly", "🎯 Yillik"),
    ("none", "✖️ Takrorsiz"),
]


@router.callback_query(F.data.startswith("remeditmenu:"))
async def cb_reminder_edit_menu(query: CallbackQuery) -> None:
    """✏️ Tahrirlash — konsolidatsiyalangan tahrir submenyusi."""
    rid = query.data.split(":", 1)[1]
    reminder = await database.get_reminder(rid)
    if not reminder:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer()
    title = (reminder.get("title") or "Eslatma").strip()[:40]
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✏️ Sarlavha", callback_data=f"remedit:{rid}:title"),
            InlineKeyboardButton(text="🕐 Vaqt", callback_data=f"remedit:{rid}:time"),
        ],
        [
            InlineKeyboardButton(text="🔁 Takror", callback_data=f"remrecurmenu:{rid}"),
            InlineKeyboardButton(text="📝 Izoh", callback_data=f"remedit:{rid}:note"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"remopen:{rid}")],
    ])
    text = (f"✏️ **TAHRIRLASH**\n{_SEP}\n\n«{_escape_markdown(title)}»\n\n"
            "Qaysi maydonni o'zgartiramiz?")
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("remrecurmenu:"))
async def cb_reminder_recurrence_menu(query: CallbackQuery) -> None:
    """🔁 Takror — takrorlash qoidasini pick-list bilan o'zgartirish."""
    rid = query.data.split(":", 1)[1]
    reminder = await database.get_reminder(rid)
    if not reminder:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    await query.answer()
    cur = database.normalize_recurrence_rule(reminder.get("recurrence_rule"))
    rows = []
    for val, label in _RECUR_OPTIONS:
        mark = " ✅" if (cur == val or (val == "none" and not cur)) else ""
        rows.append([InlineKeyboardButton(text=label + mark, callback_data=f"remrecur:{rid}:{val}")])
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"remeditmenu:{rid}")])
    cur_label = _format_recurrence_label(cur) if cur else "takrorsiz"
    text = (f"🔁 **TAKROR**\n{_SEP}\n\nHozir: **{cur_label}**\n\n"
            "Yangi takrorlash qoidasini tanlang:")
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                           reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data.startswith("remrecur:"))
async def cb_reminder_set_recurrence(query: CallbackQuery) -> None:
    """Takror qoidasini o'rnatadi (yoki takrorsiz qiladi)."""
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    rid, val = parts[1], parts[2]
    rule = None if val == "none" else database.normalize_recurrence_rule(val)
    ok = await database.update_reminder(rid, {"recurrence_rule": rule})
    if not ok:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    reminder = await database.get_reminder(rid)
    await query.answer(f"🔁 Takror: {_format_recurrence_label(rule) if rule else 'takrorsiz'}")
    try:
        await query.message.edit_text(
            "✅ Saqlandi\n" + _SEP + "\n\n" + _format_reminder_card(reminder),
            parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("remedit:"))
async def cb_reminder_edit(query: CallbackQuery, state: FSMContext) -> None:
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    rid, field = parts[1], parts[2]
    reminder = await database.get_reminder(rid)
    if not reminder:
        await query.answer("Eslatma topilmadi", show_alert=True)
        return
    if field not in {"title", "time", "note"}:
        await query.answer("Maydon noma'lum", show_alert=True)
        return
    await state.set_state(ReminderEditFSM.awaiting_value)
    await state.update_data(reminder_id=rid, field=field)
    await query.answer()
    if field == "title":
        prompt = "✏️ **Eslatma sarlavhasi**\n\nYangi sarlavhani yuboring (matn yoki ovoz)."
    elif field == "note":
        prompt = "📝 **Eslatma izohi**\n\nYangi izoh (tavsif) matnini yuboring. `-` = tozalash."
    else:
        prompt = (
            "🕐 **Eslatma vaqti**\n\n"
            "Yangi vaqtni yuboring: `bugun 17:00`, `ertaga 09:00`, `2 soat`."
        )
    await _safe_answer(
        query.message,
        prompt,
        parse_mode="Markdown",
        reply_markup=single_back_keyboard(f"remopen:{rid}", text="✕ Bekor"),
    )


@router.message(StateFilter(ReminderEditFSM.awaiting_value), F.text | F.voice)
async def handle_reminder_edit_value(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    data = await state.get_data()
    rid = data.get("reminder_id")
    field = data.get("field")
    await state.clear()
    # BUG FIX: ovoz xabarda message.text=None bo'ladi — transkripsiya _msg_text'da.
    # Avval message.text ishlatilgan edi → ovoz orqali tahrir "bo'sh qiymat" deb
    # ishlamasdi. Endi har ikki holatda ham _get_text_or_transcribe natijasidan olamiz.
    raw = (_msg_text or "").strip()
    if not rid or field not in {"title", "time", "note"}:
        await message.answer("Holat yo'qoldi — qayta urinib ko'ring.")
        return
    if not raw:
        await message.answer("Bo'sh qiymat — bekor qilindi.")
        return
    if field == "title":
        ok = await database.update_reminder(rid, {"title": raw[:220]})
    elif field == "note":
        # '-' → izohni tozalash
        ok = await database.update_reminder(rid, {"note": None if raw == "-" else raw[:1000]})
    else:
        parsed, reason = await _parse_deadline_natural(raw)
        if not parsed:
            await _safe_answer(message, _deadline_error_message(reason, kind="time"),
                               parse_mode="Markdown")
            return
        ok = await database.update_reminder(rid, {"remind_at": parsed, "status": "scheduled", "sent_at": None})
    if not ok:
        await message.answer("Eslatma topilmadi.")
        return
    reminder = await database.get_reminder(rid)
    await _safe_answer(
        message,
        "✅ Saqlandi\n\n" + _format_reminder_card(reminder),
        parse_mode="Markdown",
        reply_markup=reminder_detail_menu(reminder),
    )


@router.callback_query(F.data.startswith("taskopen:"))
async def cb_task_open(query: CallbackQuery) -> None:
    """Drill-down: show single task's full card + all action buttons."""
    tid = query.data.split(":", 1)[1]
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()
    # Mark the opened task as "current" so "shu vazifani ..." resolves to it.
    claude_service.set_last_task_view(
        [{"n": 1, "id": task["id"], "title": (task.get("title") or "—")[:50]}])
    text = _format_task_card(task)
    # Subtask summary (top-level tasks only) — real child tasks live under the parent.
    if not task.get("parent_id"):
        subs = await database.list_subtasks(tid)
        if subs:
            done = sum(1 for s in subs if s.get("status") in ("done", "cancelled"))
            text += f"\n\n🌳 **Sub-vazifalar:** {done}/{len(subs)}"
            for s in subs[:8]:
                em = _STATUS_EMOJI.get(s.get("status", "todo"), "•")
                who = (s.get("assignee") or "").strip()
                text += f"\n  {em} {(s.get('title') or '—').strip()[:40]}" + (f" · 👤{who[:12]}" if who else "")
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_task_card_kb_with_back(task))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                           reply_markup=_task_card_kb_with_back(task))


def _format_meeting_card(m: dict, show_date: bool = False) -> str:
    """Single meeting drill-down — Variant A (bot uslubi).

    Tartibi:
      🤝 UCHRASHUV (sarlavha)
      {title}
      ─── divider ───
      📌 MA'LUMOTLAR
      ⏰ Vaqt / 👥 Ishtirokchilar / 📍 Manzil / 📅 Kun tartibi
      📝 Bayonnoma / ☁️ iCloud
      ─── divider ───
      _izoh_
    """
    DIVIDER = "━" * 20

    # Time range label
    try:
        time_start = _meeting_time_label(m.get("datetime_start") or "", with_past_marker=True)
    except (ValueError, TypeError):
        time_start = m.get("datetime_start") or "—"
    end_iso = m.get("datetime_end")
    end_clock = ""
    if end_iso:
        try:
            end_dt = datetime.fromisoformat(end_iso).astimezone(database.TZ)
            end_clock = end_dt.strftime("%H:%M")
        except (ValueError, TypeError):
            pass
    time_label = f"{time_start} – {end_clock}" if end_clock else time_start

    title = (m.get("title") or "—").strip()
    if m.get("completed_at"):
        title = "✅ " + title  # attended/done marker
    parts = m.get("participants") or []
    if not parts:
        plabel = "belgilanmagan"
    elif len(parts) <= 3:
        plabel = ", ".join(parts)
    else:
        plabel = f"{', '.join(parts[:3])} (+{len(parts) - 3} nafar)"
    location = (m.get("location_or_link") or "").strip() or "belgilanmagan"
    agenda = (m.get("agenda") or "").strip() or "belgilanmagan"
    # Truncate long agenda inline; full text visible via Tahrirlash → Kun tartibi.
    if len(agenda) > 80:
        agenda = agenda[:77] + "…"

    # Protocol: stored in follow_up_actions (JSON list of strings, see Phase 7).
    fu = m.get("follow_up_actions") or []
    has_protocol = (isinstance(fu, list) and bool(fu)) or (isinstance(fu, str) and bool(fu.strip()))
    protocol_status = "tayyor" if has_protocol else "yo'q"

    # iCloud sync indicator: only meaningful when iCloud is enabled for this deployment.
    icloud_row: list[str] = []
    if config.ICLOUD_ENABLED:
        icloud_status = "sinxronlangan" if (m.get("icloud_uid") or "").strip() else "kutilmoqda"
        icloud_row = [f"      ☁️ iCloud:           {icloud_status}"]

    lines = [
        "🤝 **UCHRASHUV**",
        "",
        title,
        "",
        DIVIDER,
        "",
        "📌 **MA'LUMOTLAR**",
        "",
        f"      ⏰ Vaqt:             {time_label}",
        f"      👥 Ishtirokchilar:   {plabel}",
        f"      📍 Manzil:           {location}",
        f"      📅 Kun tartibi:      {agenda}",
        f"      📝 Bayonnoma:        {protocol_status}",
        *icloud_row,
        "",
        DIVIDER,
        "",
        "_Quyidagi tugmalardan birini tanlang._",
    ]
    return "\n".join(lines)


_MEETINGS_PER_PAGE = 10


def _meeting_time_label(iso_str: str, *, with_past_marker: bool = True) -> str:
    """Render 'Bugun 14:00' / 'Ertaga 10:00' / '29-may 16:00' / '23-may 11:00 · o'tgan'."""
    try:
        dt = datetime.fromisoformat(iso_str).astimezone(database.TZ)
    except (ValueError, TypeError):
        return iso_str or "—"
    now = datetime.now(database.TZ)
    same_day = dt.date() == now.date()
    tomorrow = dt.date() == (now + timedelta(days=1)).date()
    time_str = dt.strftime("%H:%M")
    if same_day:
        base = f"Bugun {time_str}"
    elif tomorrow:
        base = f"Ertaga {time_str}"
    else:
        base = f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]} {time_str}"
    if with_past_marker and dt < now and not same_day:
        base += " · o'tgan"
    return base


async def _compute_meetings_overview() -> dict:
    """Counts used by 📌 UMUMIY HOLAT in the meetings screen.
    Stats are derived from the full DB, independent of the current filter,
    so the user sees a stable picture regardless of which chip is active.
    """
    now = datetime.now(database.TZ)
    today_start = database.TZ.localize(datetime.combine(now.date(), datetime.min.time()))
    today_end = today_start + timedelta(days=1)
    tomorrow_end = today_end + timedelta(days=1)
    week_end = today_start + timedelta(days=7)
    today = await database.list_meetings_in_window(today_start.isoformat(), today_end.isoformat())
    tomorrow = await database.list_meetings_in_window(today_end.isoformat(), tomorrow_end.isoformat())
    week = await database.list_meetings_in_window(today_start.isoformat(), week_end.isoformat())
    return {
        "today": len(today),
        "tomorrow": len(tomorrow),
        "week": len(week),
        "next_meeting": today[0] if today else (week[0] if week else None),
    }


def _format_meetings_compact(
    meetings: list[dict],
    label: str,
    stats: dict | None = None,
    page: int = 1,
) -> str:
    """Meetings screen — Vazifalar uslubidagi block-stil.

    Ikonkalar: 🤝 sahifa · 📌 stats · ⚡ eng yaqin · 📅 ro'yxat
                ⏰ vaqt · 👥 ishtirokchilar · 📍 manzil
    Bot uslubi: divider atrofida 1 bo'sh qator, item orasida 1 bo'sh qator,
                detal blok 6 ta probel indent, qiymat 12-ustunda.
    """
    DIVIDER = "━" * 20

    per_page = _MEETINGS_PER_PAGE
    total = len(meetings)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    page_meetings = meetings[start : start + per_page]

    lines: list[str] = [
        "🤝 **UCHRASHUVLAR**",
        "",
        f"Ko'rinish · {label}",
        f"Natija · {total} ta",
    ]
    if total_pages > 1:
        lines.append(f"Sahifa · {page} / {total_pages}")

    if total == 0:
        lines.extend([
            "",
            DIVIDER,
            "",
            "Bu bo'limda hozircha uchrashuv yo'q.",
            "",
            "➕ _Yangi uchrashuv yaratish yoki boshqa filter tanlash mumkin._",
        ])
        return "\n".join(lines)

    lines.extend(["", DIVIDER, ""])

    # UMUMIY HOLAT — stats independent of the current filter
    if stats:
        lines.extend([
            "📌 **UMUMIY HOLAT**",
            "",
            f"Jami {total}   ·   Bugun {stats['today']}   ·   Ertaga {stats['tomorrow']}   ·   Hafta {stats['week']}",
            "",
            DIVIDER,
            "",
        ])

    def _participants_label(m: dict) -> str:
        parts = m.get("participants") or []
        if not parts:
            return "belgilanmagan"
        if len(parts) <= 3:
            return ", ".join(parts)
        return f"{', '.join(parts[:3])} (+{len(parts) - 3} nafar)"

    def _meeting_card(m: dict, num: int | None) -> list[str]:
        """One meeting card with 1-line title + 3-line detail block."""
        title = (m.get("title") or "—").strip()
        if m.get("completed_at"):
            title = "✅ " + title  # attended/done marker (seen in O'tgan)
        prefix = f"{num}.  " if num is not None else ""
        return [
            f"{prefix}{title}",
            "",
            f"      ⏰ Vaqt:             {_meeting_time_label(m.get('datetime_start') or '')}",
            f"      👥 Ishtirokchilar:   {_participants_label(m)}",
            f"      📍 Manzil:           {(m.get('location_or_link') or '').strip() or 'belgilanmagan'}",
        ]

    # ENG YAQIN — surface the next upcoming meeting (only on page 1, only if non-past filter)
    next_meeting = stats.get("next_meeting") if stats else None
    if page == 1 and next_meeting:
        lines.append("⚡ **ENG YAQIN**")
        lines.append("")
        lines.extend(_meeting_card_pop(next_meeting))
        lines.extend(["", DIVIDER, ""])

    lines.extend(["📅 **UCHRASHUVLAR**", ""])
    cards = [_meeting_card(m, start + i + 1) for i, m in enumerate(page_meetings)]
    joined = []
    for i, card in enumerate(cards):
        if i:
            joined.append("")
        joined.extend(card)
    lines.extend(joined)
    lines.append("")

    lines.extend([
        DIVIDER,
        "",
        "_Uchrashuv raqamini bosing — bayonnoma va boshqa amallar._",
    ])
    return "\n".join(lines).rstrip()


def _meeting_card_pop(m: dict) -> list[str]:
    """Same shape as a list card but without a leading number (used in ENG YAQIN)."""
    title = (m.get("title") or "—").strip()
    parts = m.get("participants") or []
    if not parts:
        plabel = "belgilanmagan"
    elif len(parts) <= 3:
        plabel = ", ".join(parts)
    else:
        plabel = f"{', '.join(parts[:3])} (+{len(parts) - 3} nafar)"
    return [
        title,
        "",
        f"      ⏰ Vaqt:             {_meeting_time_label(m.get('datetime_start') or '')}",
        f"      👥 Ishtirokchilar:   {plabel}",
        f"      📍 Manzil:           {(m.get('location_or_link') or '').strip() or 'belgilanmagan'}",
    ]


def meetings_filter_keyboard(
    current: str = "week",
    meetings: list[dict] | None = None,
    page: int = 1,
    total_pages: int = 1,
) -> InlineKeyboardMarkup:
    """Meetings inline keyboard — DRILL-DOWN va sahifalash uchun.

    Filter chiplari, Yangi va Qidiruv tugmalari **olib tashlangan** —
    ular pastdagi reply kbd (section)'da bor.
    """
    rows: list[list[InlineKeyboardButton]] = []

    per_page = _MEETINGS_PER_PAGE
    visible = (meetings or [])[(page - 1) * per_page : page * per_page]
    if visible:
        nums = [
            InlineKeyboardButton(text=str((page - 1) * per_page + i + 1),
                                  callback_data=f"meetingopen:{m['id']}")
            for i, m in enumerate(visible)
        ]
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])

    if total_pages > 1:
        nav = []
        if page > 1:
            nav.append(InlineKeyboardButton(
                text="⬅️ Oldingi",
                callback_data=f"meetingfilter:{current}:page:{page - 1}",
            ))
        nav.append(InlineKeyboardButton(text=f"{page} / {total_pages}", callback_data="noop"))
        if page < total_pages:
            nav.append(InlineKeyboardButton(
                text="Keyingi ➡️",
                callback_data=f"meetingfilter:{current}:page:{page + 1}",
            ))
        rows.append(nav)

    return InlineKeyboardMarkup(inline_keyboard=rows) if rows else None


async def _render_meetings_for_filter(message: Message, filt: str = "week",
                                       edit_existing: bool = False,
                                       page: int = 1) -> None:
    global _last_meeting_filter
    _last_meeting_filter = filt
    now = datetime.now(database.TZ)
    today_start = database.TZ.localize(datetime.combine(now.date(), datetime.min.time()))

    if filt == "today":
        meetings = await database.list_today_meetings()
        label = "Bugungi"
    elif filt == "tomorrow":
        start = (today_start + timedelta(days=1)).isoformat()
        end = (today_start + timedelta(days=2)).isoformat()
        meetings = await database.list_meetings_in_window(start, end)
        label = "Ertangi"
    elif filt == "all":
        # Start from the beginning of today (not `now`) so a meeting scheduled
        # earlier today still shows — otherwise it silently drops out the moment
        # its start time passes, which reads as "my meeting disappeared".
        meetings = await database.list_meetings_in_window(
            today_start.isoformat(), (today_start + timedelta(days=30)).isoformat()
        )
        label = "Barchasi"
    elif filt == "past":
        start = (today_start - timedelta(days=7)).isoformat()
        end = now.isoformat()
        past = await database.list_meetings_in_window(start, end)
        # Newest past meetings first for relevance
        meetings = list(reversed(past))
        label = "O'tgan"
    else:  # week (default)
        # Window starts at the beginning of today (not `now`) so meetings earlier
        # today remain visible — the reported "voice shows it, the button doesn't"
        # bug came from `now` excluding an already-started meeting.
        meetings = await database.list_meetings_in_window(
            today_start.isoformat(), (today_start + timedelta(days=7)).isoformat()
        )
        label = "Haftalik"
        filt = "week"

    # Completed (attended) meetings leave the active/upcoming views — they stay
    # visible only under "O'tgan" (marked with ✅).
    if filt != "past":
        meetings = [m for m in meetings if not m.get("completed_at")]

    stats = await _compute_meetings_overview()
    total_pages = max(1, (len(meetings) + _MEETINGS_PER_PAGE - 1) // _MEETINGS_PER_PAGE)
    page = max(1, min(page, total_pages))

    # Remember the meetings shown on this page so a later "N-uchrashuv ..." / "shu
    # uchrashuv" reference resolves to the right id (mirrors the task list view).
    _pm_start = (page - 1) * _MEETINGS_PER_PAGE
    claude_service.set_last_meeting_view(
        [{"n": _pm_start + i + 1, "id": m["id"], "title": (m.get("title") or "—")[:50]}
         for i, m in enumerate(meetings[_pm_start:_pm_start + _MEETINGS_PER_PAGE])])

    text = _format_meetings_compact(meetings, label, stats=stats, page=page)
    kb = meetings_filter_keyboard(filt, meetings, page=page, total_pages=total_pages)

    if edit_existing:
        try:
            await message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


@router.message(Command("meetings"))
async def cmd_meetings(message: Message, state: FSMContext | None = None) -> None:
    if state is not None:
        await state.set_state(SectionFSM.in_meetings)
    await message.answer(
        "🤝 **UCHRASHUVLAR**", parse_mode="Markdown",
        reply_markup=meetings_section_reply_keyboard(),
    )
    await _render_meetings_for_filter(message, filt="week")


def _looks_like_protocol(fu) -> bool:
    """Distinguish a saved bayonnoma (long prose) from the task-id list that the
    post-meeting follow-up flow also stores in `follow_up_actions`."""
    if isinstance(fu, str):
        s = fu.strip()
    elif isinstance(fu, list) and fu:
        s = str(fu[0]).strip()
    else:
        return False
    return len(s) > 40 and (" " in s)


async def _render_protocols(message: Message) -> None:
    """Markaziy 'Bayonnomalar' ro'yxati — barcha bayonnomali uchrashuvlar, OY
    bo'yicha guruhlangan, yangi birinchi. Raqamni bosish → to'liq protokol
    (mavjud viewproto: matn + 📄 Word + 📤 Ulashish)."""
    candidates = await database.list_meetings_with_protocol(limit=100)
    meetings = [m for m in candidates if _looks_like_protocol(m.get("follow_up_actions"))]
    if not meetings:
        await _safe_answer(
            message,
            "📄 **BAYONNOMALAR**\n\n_Hozircha saqlangan bayonnoma yo'q._\n\n"
            "Uchrashuvni oching → 📄 Bayonnoma yarating → ✅ Saqlang.",
            parse_mode="Markdown",
        )
        return
    DIV = "━" * 15
    MAX = 25
    shown = meetings[:MAX]
    lines = [f"📄 **BAYONNOMALAR · {len(meetings)} ta**", "", DIV]
    cur_key = None
    nums: list[tuple[int, str]] = []
    for i, m in enumerate(shown, 1):
        try:
            dt = datetime.fromisoformat(m["datetime_start"]).astimezone(database.TZ)
            mkey = (dt.year, dt.month)
            mlabel = f"{UZ_MONTHS_FULL[dt.month - 1].upper()} {dt.year}"
            day = f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]}"
        except (ValueError, TypeError, KeyError):
            mkey, mlabel, day = (0, 0), "SANASIZ", "—"
        if mkey != cur_key:
            lines.append("")
            lines.append(f"📅 **{mlabel}**")
            cur_key = mkey
        title = (m.get("title") or "Uchrashuv").strip()
        npart = len(m.get("participants") or [])
        part = f" · 👥 {npart}" if npart else ""
        lines.append(f"  {i}. {_escape_markdown(title[:50])} — {day}{part}")
        nums.append((i, m["id"]))
    if len(meetings) > MAX:
        lines.append("")
        lines.append(f"_+{len(meetings) - MAX} ta yana_")
    lines.extend(["", DIV, "Raqamni bosing — to'liq bayonnoma + 📄 Word"])
    btns = [InlineKeyboardButton(text=str(i), callback_data=f"viewproto:{mid}") for i, mid in nums]
    rows = [btns[j:j + 5] for j in range(0, len(btns), 5)]
    rows.append([back_button("meetingfilter:week", "⬅️ Uchrashuvlar")])
    await _safe_answer(message, "\n".join(lines), parse_mode="Markdown",
                       reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.message(Command("bayonnomalar"))
@router.message(Command("protocols"))
async def cmd_protocols(message: Message) -> None:
    """/bayonnomalar — barcha saqlangan bayonnomalar markaziy ro'yxati (oy bo'yicha)."""
    await _render_protocols(message)


@router.callback_query(F.data.startswith("meetingfilter:"))
async def cb_meeting_filter(query: CallbackQuery) -> None:
    parts = query.data.split(":")
    filt = parts[1]
    page = 1
    if len(parts) >= 4 and parts[2] == "page":
        try:
            page = max(1, int(parts[3]))
        except ValueError:
            page = 1
    await query.answer()
    await _render_meetings_for_filter(query.message, filt=filt, edit_existing=True, page=page)


@router.callback_query(F.data.startswith("meetingopen:"))
async def cb_meeting_open(query: CallbackQuery) -> None:
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer()
    # Remember the opened meeting so a follow-up like "uchrashuv sarlavhasini
    # o'zgartir" resolves to THIS meeting (update_meeting{id}) instead of "which?".
    claude_service.set_last_meeting_view(
        [{"n": 1, "id": meeting["id"], "title": (meeting.get("title") or "—")[:50]}])
    text = _format_meeting_card(meeting, show_date=True)
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=meeting_inline_actions(meeting))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=meeting_inline_actions(meeting))


async def _rerender_meeting_card(query: CallbackQuery, mid: str) -> None:
    meeting = await database.get_meeting(mid)
    if not meeting:
        return
    text = _format_meeting_card(meeting, show_date=True)
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=meeting_inline_actions(meeting))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("meeting_done:"))
async def cb_meeting_done(query: CallbackQuery) -> None:
    """✅ Bo'ldi — mark the meeting attended. It leaves Bugun/Haftalik and shows
    with a ✅ under O'tgan."""
    mid = query.data.split(":", 1)[1]
    if not await database.complete_meeting(mid):
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer("✅ Bo'ldi deb belgilandi")
    await _rerender_meeting_card(query, mid)


@router.callback_query(F.data.startswith("meeting_undone:"))
async def cb_meeting_undone(query: CallbackQuery) -> None:
    """↺ Undo the 'done' mark — returns the meeting to the active views."""
    mid = query.data.split(":", 1)[1]
    if not await database.uncomplete_meeting(mid):
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer("↺ Faol ro'yxatga qaytarildi")
    await _rerender_meeting_card(query, mid)


@router.message(StateFilter(MeetingSearchFSM.awaiting_query), F.text | F.voice)
async def handle_meeting_search(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    await state.clear()
    query = message.text.strip()
    if not query:
        await message.answer(
            "Qidiruv so'zi bo'sh bo'lmasin.",
            reply_markup=single_back_keyboard("meetingfilter:week"),
        )
        return
    found = (await database.search_all(query, limit=30)).get("meetings", [])
    text = _format_meetings_compact(found, f"Qidiruv: {query}")
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=meetings_filter_keyboard("search", found))


async def _generate_meeting_prep(meeting: dict) -> str:
    active = await database.list_tasks(status_in=["todo", "in_progress"], limit=20)
    task_lines = "\n".join(
        f"- {t['title']} ({t['priority']}, {t.get('deadline') or 'deadline yoq'})"
        for t in active[:10]
    )
    directive = (
        "[INTERNAL] meeting_prep_brief\n\n"
        "Generate a concise executive prep brief in O'zbek lotin. Include: "
        "1) meeting objective, 2) agenda, 3) decisions needed, 4) questions to ask, "
        "5) related open tasks. Use Markdown, max 18 lines.\n\n"
        f"MEETING:\n{meeting}\n\nACTIVE TASKS:\n{task_lines}"
    )
    response = await claude_service.process_message("", internal_directive=directive)
    text = (response.get("user_message") or "").strip()
    if text:
        return text
    agenda = meeting.get("agenda") or "Agenda kiritilmagan."
    prep = meeting.get("prep_notes") or "Oldingi materiallarni ko'rib chiqing."
    return f"🧠 **Prep brief: {meeting['title']}**\n\n• Agenda: {agenda}\n• Tayyorgarlik: {prep}"


@router.callback_query(F.data.startswith("meeting_prep:"))
async def cb_meeting_prep(query: CallbackQuery) -> None:
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer("Prep tayyorlanmoqda...")
    typing_task = asyncio.create_task(_keep_typing(query.bot, query.message.chat.id))
    try:
        text = await _generate_meeting_prep(meeting)
    finally:
        typing_task.cancel()
    await database.mark_meeting_prep_sent(mid)
    await _safe_answer(
        query.message,
        text,
        parse_mode="Markdown",
        reply_markup=single_back_keyboard("nav_meetings"),
    )


@router.callback_query(F.data.startswith("meeting_followup:"))
async def cb_meeting_followup(query: CallbackQuery, state: FSMContext) -> None:
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await state.set_state(MeetingFollowupFSM.awaiting_notes)
    await state.update_data(meeting_id=mid)
    await query.answer()
    await query.message.answer(
        "📝 **Action items chiqarish**\n\n"
        "Uchrashuvdan keyingi yozuv yoki ovozni yuboring. Men kim-nima-qachon formatida "
        "vazifalarni ajratib, keraklisini task qilib yarataman.",
        parse_mode="Markdown",
        reply_markup=single_back_keyboard("nav_meetings"),
    )


async def _run_meeting_followup_extraction(message: Message, meeting: dict, notes: str) -> None:
    directive = (
        "[INTERNAL] meeting_followup_extract\n\n"
        "Extract concrete action items from the principal's meeting notes. "
        "Create tasks only for items owned by the principal or clearly requiring tracking. "
        "For delegated items, set assignee. Use due dates if mentioned; otherwise deadline=null. "
        "Return JSON envelope with actions=[create_task...] and a short Uzbek confirmation.\n\n"
        f"MEETING:\n{meeting}\n\nNOTES:\n{notes}"
    )
    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        response = await claude_service.process_message(notes, internal_directive=directive)
    finally:
        typing_task.cancel()

    ids_by_type = await _execute_actions(response.get("actions", []))
    created = ids_by_type.get("task", [])
    await database.update_meeting(
        meeting["id"],
        {
            "follow_up_actions": created,
            "followup_sent_at": database.now_iso(),
        },
    )
    text = (response.get("user_message") or "").strip()
    if not text:
        text = f"✅ {len(created)} ta action item yaratildi." if created else "Action item topilmadi."
    text += _failed_actions_note(ids_by_type)
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=tasks_compact_keyboard([]))


@router.message(StateFilter(MeetingFollowupFSM.awaiting_notes), F.text | F.voice)
async def handle_meeting_followup_text(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    data = await state.get_data()
    await state.clear()
    meeting = await database.get_meeting(data.get("meeting_id", ""))
    if not meeting:
        await message.answer("Uchrashuv topilmadi.")
        return
    await _run_meeting_followup_extraction(message, meeting, message.text)


@router.message(StateFilter(MeetingFollowupFSM.awaiting_notes), F.voice)
async def handle_meeting_followup_voice(message: Message, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    await state.clear()
    meeting = await database.get_meeting(data.get("meeting_id", ""))
    if not meeting:
        await message.answer("Uchrashuv topilmadi.")
        return
    if message.voice.file_size and message.voice.file_size > voice_service.MAX_AUDIO_BYTES:
        await message.answer("Ovoz juda katta. Iltimos, qisqaroq yuboring.")
        return
    file = await bot.get_file(message.voice.file_id)
    audio_io = await bot.download_file(file.file_path)
    audio_bytes = audio_io.getvalue() if hasattr(audio_io, "getvalue") else audio_io.read()
    transcript = await voice_service.transcribe(audio_bytes, filename="meeting-followup.ogg", language="uz")
    if not transcript:
        await message.answer("Ovozni o'qiy olmadim. Matn bilan qayta yuboring.")
        return
    _fu_shown = transcript.strip()
    if len(_fu_shown) > 600:
        _fu_shown = _fu_shown[:600] + "…"
    await message.answer(f"_🎙 Tushundim:_ {_escape_markdown(_fu_shown)}", parse_mode="Markdown")
    await _run_meeting_followup_extraction(message, meeting, transcript)


@router.message(Command("calendar"))
async def cmd_calendar(message: Message) -> None:
    if not config.ICLOUD_ENABLED:
        await message.answer(
            "📅 iCloud kalendar **sozlanmagan**.\n\n"
            "Sozlash uchun:\n"
            "1. https://account.apple.com → Sign-In and Security\n"
            "2. App-Specific Passwords → Generate Password\n"
            "3. `.env` faylda `APPLE_ID` va `APPLE_APP_SPECIFIC_PASSWORD` to'ldiring\n"
            "4. Botni qayta ishga tushiring",
            parse_mode="Markdown",
            reply_markup=single_back_keyboard("nav_settings"),
        )
        return
    typing = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        ok, msg = await asyncio.to_thread(calendar_service.test_connection)
    finally:
        typing.cancel()
    emoji = "✅" if ok else "❌"
    await message.answer(
        f"{emoji} **iCloud kalendar**\n\n{msg}",
        parse_mode="Markdown",
        reply_markup=single_back_keyboard("nav_settings"),
    )


def _ascii_bar(value: int, max_value: int, width: int = 10) -> str:
    """Return a `width`-cell unicode bar showing value/max ratio."""
    if max_value <= 0:
        return "░" * width
    filled = round((value / max_value) * width)
    filled = max(0, min(width, filled))
    return "█" * filled + "░" * (width - filled)


def _period_from_text(text: str | None) -> tuple[int, str]:
    raw = (text or "").split(maxsplit=1)
    arg = raw[1].strip().lower() if len(raw) > 1 else "week"
    aliases = {
        "today": (1, "Bugun"),
        "bugun": (1, "Bugun"),
        "day": (1, "Bugun"),
        "week": (7, "Oxirgi 7 kun"),
        "hafta": (7, "Oxirgi 7 kun"),
        "weekly": (7, "Oxirgi 7 kun"),
        "month": (30, "Oxirgi 30 kun"),
        "oy": (30, "Oxirgi 30 kun"),
        "monthly": (30, "Oxirgi 30 kun"),
        "30": (30, "Oxirgi 30 kun"),
        "7": (7, "Oxirgi 7 kun"),
        "1": (1, "Bugun"),
    }
    return aliases.get(arg, aliases["week"])


def _risk_label(score: int) -> tuple[str, str]:
    if score >= 75:
        return "🔴", "Yuqori"
    if score >= 45:
        return "🟡", "O'rta"
    return "🟢", "Nazoratda"


def _risk_bar(score: int, width: int = 10) -> str:
    """Visual risk bar: filled red squares + empty white squares.
    Score is clamped to 0-100 and proportionally mapped to `width` cells.
    """
    score = max(0, min(100, int(score or 0)))
    filled = round(score / 100 * width)
    return "🟥" * filled + "⬜" * (width - filled)


def _stats_period_keyboard(active_days: int) -> InlineKeyboardMarkup:
    items = [(1, "Bugun"), (7, "7 kun"), (30, "30 kun")]
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text=("● " if days == active_days else "") + label,
                callback_data=f"stats_period:{days}",
            )
            for days, label in items
        ],
        [
            InlineKeyboardButton(text="📄 Weekly report", callback_data="report_period:7"),
            InlineKeyboardButton(text="📄 Monthly report", callback_data="report_period:30"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data="stats_back")],
    ])


def _report_keyboard(days: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"stats_period:{days}")],
        [
            InlineKeyboardButton(text="📊 7 kun", callback_data="report_period:7"),
            InlineKeyboardButton(text="📈 30 kun", callback_data="report_period:30"),
        ],
    ])



def _format_uzs_amount(usd: float) -> str:
    # Approximate display rate; keeps stats readable without calling an FX API.
    uzs = round((usd or 0.0) * 12_600 / 1000) * 1000
    return f"{uzs:,.0f}".replace(",", " ")


def _format_stats_dashboard(stats: dict, label: str) -> str:
    tasks = stats["tasks"]
    meetings = stats["meetings"]
    by_p = tasks["by_priority"]
    risk_emoji, risk_label = _risk_label(stats["risk_score"])
    completion = tasks["completion_rate_pct"]
    comp_emoji = "🟢" if completion >= 70 else "🟡" if completion >= 40 else "🔴"
    total_priority = max(1, sum(by_p.get(p, 0) for p in ("P0", "P1", "P2", "P3")))
    priority_labels = {
        "P0": ("🔴", "Shoshilinch"),
        "P1": ("🟠", "Muhim"),
        "P2": ("🔵", "Rejadagi"),
        "P3": ("⚪", "Oddiy"),
    }
    llm_cost = float(stats["llm"]["cost"] or 0.0)

    lines = [
        f"📊 **PROFESSIONAL STATS** · {label}",
        _SEP,
        "",
        "📍 **KPI**",
        f"• Yaratildi: **{tasks['created']} ta**",
        f"• Bajarildi: **{tasks['done']} ta**",
        f"• Aktiv: **{tasks['active']} ta**",
        f"• Bajarilish: {comp_emoji} **{completion}%**",
        f"• Risk: {risk_emoji} **{stats['risk_score']}/100** — {risk_label}",
        "",
        "⏱ **Deadline**",
        f"• Bugun / 24 soat: **{tasks['due_24h']} ta**",
        f"• 48 soat: **{tasks['due_48h']} ta**",
        f"• 7 kun: **{tasks['due_7d']} ta**",
        f"• Deadline yo'q: **{tasks['no_deadline']} ta**",
        "",
        "⚡ **PRIORITET YUKLAMASI**",
    ]
    for p in ("P0", "P1", "P2", "P3"):
        n = by_p.get(p, 0)
        emoji, label_text = priority_labels[p]
        lines.append(f"{emoji} {label_text:<14} {_ascii_bar(n, total_priority, 10)}  **{n} ta**")

    lines.extend([
        "",
        "📋 **Delegatsiya**",
    ])
    if stats["delegation"]:
        for row in stats["delegation"][:4]:
            lines.append(f"• {row['assignee']}: **{row['total']} ta** ochiq · **{row['overdue'] or 0} ta** o'tgan")
    else:
        lines.append("• Ochiq delegatsiya yo'q")

    lines.extend([
        "",
        "🤝 **Meeting**",
        f"• Uchrashuv: **{meetings['count']} ta**",
        f"• Vaqt: **{meetings['hours']} soat**",
        f"• Action items: **{meetings['action_items']} ta**",
        "",
        "🤖 **Bot Audit**",
        f"• Chaqiriqlar: **{stats['llm']['calls']} ta**",
        f"• Bot xarajati: **${llm_cost:.2f} ≈ {_format_uzs_amount(llm_cost)} so'm**",
    ])

    if tasks["risk_tasks"]:
        lines.extend(["", "🚨 **Risklar**"])
        for task in tasks["risk_tasks"][:4]:
            deadline, _ = _format_deadline_short(task.get("deadline"))
            lines.append(f"{_PRIORITY_BADGE.get(task.get('priority'), '🔵')} {_truncate(task['title'], 44)} — {deadline}")

    return "\n".join(lines)


def _format_executive_report(stats: dict, label: str) -> str:
    tasks = stats["tasks"]
    meetings = stats["meetings"]
    risk_score = stats["risk_score"]
    risk_emoji, risk_label = _risk_label(risk_score)
    by_p = tasks["by_priority"]

    # KPI denominator: total work that "touched" this period = done + still-active.
    total_pool = tasks["done"] + tasks["active"]
    yopildi_value = f"**{tasks['done']} / {total_pool}**" if total_pool else f"**{tasks['done']}**"

    recommendations: list[str] = []
    if tasks["overdue"]:
        recommendations.append("O'tgan vazifalarni bugun yopish yoki deadline yangilash.")
    if by_p.get("P0", 0) + by_p.get("P1", 0) >= 5:
        recommendations.append("Shoshilinch va Muhim ro'yxatini qayta saralash, haqiqatan shoshilinch bo'lmaganlarini Rejadagi ga tushirish.")
    if tasks["no_deadline"]:
        recommendations.append("Deadline'siz vazifalarga muddat belgilash.")
    if meetings["count"] and meetings["followup_count"] < meetings["count"]:
        recommendations.append("Uchrashuvlardan keyingi action itemlarni qayd qilish.")
    if not recommendations:
        recommendations.append("Yuklama nazoratda. Top-3 vazifani ketma-ket yopib boring.")

    blocks: list[str] = []

    # Header
    blocks.append(f"📄 **EXECUTIVE REPORT**\n{label}")

    # Risk header + bar
    blocks.append(
        f"{risk_emoji} **Risk: {risk_label.upper()} — {risk_score}/100**\n"
        f"{_risk_bar(risk_score)}"
    )

    # KPI summary
    blocks.append(
        "📌 **Asosiy KPI**\n\n"
        f"✅ Yopildi: {yopildi_value}\n"
        f"📈 Bajarilish: **{tasks['completion_rate_pct']}%**\n"
        f"⏳ Aktiv: **{tasks['active']}**\n"
        f"⚠️ O'tgan: **{tasks['overdue']}**"
    )

    # Immediate attention — overdue tasks
    overdue_tasks = tasks.get("overdue_tasks") or []
    if overdue_tasks:
        section = ["🚨 **Darhol e'tibor**", ""]
        items: list[str] = []
        for i, task in enumerate(overdue_tasks[:3], 1):
            title = _truncate((task.get("title") or "—").strip(), 60)
            deadline_label, _ = _format_deadline_short(task.get("deadline"))
            items.append(f"{i}. {title}\n   Deadline: {deadline_label}")
        section.append("\n\n".join(items))
        blocks.append("\n".join(section))

    # Delegation
    if stats.get("delegation"):
        section = ["📋 **Delegatsiya**", ""]
        for row in stats["delegation"][:4]:
            section.append(f"• {row['assignee']}: **{row['total']}** ochiq · **{row['overdue'] or 0}** o'tgan")
        blocks.append("\n".join(section))

    # Meetings
    if meetings["count"]:
        blocks.append(
            "🤝 **Uchrashuvlar**\n\n"
            f"• Soni: **{meetings['count']}** · Vaqt: **{meetings['hours']} soat**\n"
            f"• Prep: **{meetings['prep_count']}/{meetings['count']}** · "
            f"Follow-up: **{meetings['followup_count']}/{meetings['count']}**\n"
            f"• Action items: **{meetings['action_items']}**"
        )

    # Next steps
    section = ["➡️ **Keyingi qadamlar**", ""]
    items = [f"{i}. {item}" for i, item in enumerate(recommendations[:4], 1)]
    section.append("\n\n".join(items))
    blocks.append("\n".join(section))

    return "\n\n".join(blocks)


class PlanFSM(StatesGroup):
    awaiting_situation = State()


@router.message(Command("plan"))
async def cmd_plan(message: Message, state: FSMContext) -> None:
    """Executive planning mode — user describes situation (or empty=auto from DB), bot returns a structured strategic plan."""
    # If args were passed inline (e.g., `/plan bugun 5 ta vazifa bor...`), use them directly
    args = message.text.split(maxsplit=1)
    if len(args) > 1 and args[1].strip():
        await _run_planning_session(message, args[1].strip())
        return

    await state.set_state(PlanFSM.awaiting_situation)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Hozirgi holatdan reja", callback_data="plan_auto")],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="plan_cancel")],
    ])
    await message.answer(
        "🎯 **Executive Planning**\n\n"
        "📊 **Hozirgi holatdan reja** — hech narsa yozmasdan, bazadagi real "
        "vazifa, uchrashuv va muddatlaringizdan avtomatik reja tuzaman.\n\n"
        "✍️ Yoki vaziyatni o'zingiz yozing (matn/ovoz): qaysi ishlar, qancha vaqt, "
        "uchrashuvlar, kimga bog'liq.\n\n"
        "_Reja: ustuvorliklar, vaqt taqsimoti, yuboriladigan xabarlar, "
        "eskalatsiya, xavflar va tavsiyalar._",
        parse_mode="Markdown",
        reply_markup=kb,
    )


@router.message(StateFilter(PlanFSM.awaiting_situation), F.text | F.voice)
async def handle_plan_situation_text(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    await state.clear()
    # _get_text_or_transcribe patched message.text with the transcript for voice,
    # so _msg_text is the situation for BOTH text and voice. The F.text | F.voice
    # filter above already catches both — no separate F.voice handler is needed.
    await _run_planning_session(message, _msg_text)


@router.callback_query(F.data == "plan_cancel")
async def cb_plan_cancel(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer("Bekor qilindi")
    try:
        await query.message.edit_text(
            "Reja yaratish bekor qilindi.",
            reply_markup=single_back_keyboard(),
        )
    except Exception:
        pass


@router.callback_query(F.data == "plan_auto")
async def cb_plan_auto(query: CallbackQuery, state: FSMContext) -> None:
    """📊 Auto-plan — build the plan straight from the DB state (no manual
    situation). Claude reads the CURRENT PRINCIPAL STATE block (real tasks,
    meetings, overdue, deadlines) and produces the plan from it."""
    await state.clear()
    await query.answer("Bazadan reja tuzaman…")
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await _run_planning_session(query.message, "")


_PLAN_DIRECTIVE = """[INTERNAL] executive_plan

Act as the principal's Chief of Staff for a DELEGATOR. Produce a "TEZKOR NAZORAT"
control board in O'zbek using the EXACT template in 45_planning.md — detailed task
CARDS, color-coded load, firm decisions. A control panel, not prose.

SECTIONS, in this exact order (skip one ONLY if its data is empty):
1) 📌 SARLAVHA — "**<DD-MM> TO'LQINI — TEZKOR NAZORAT**", then "Maqsad: <1 jumla>" and
   "Fokus: **N ta shoshilinch vazifa + N ta qayta taqsimot + N ta mas'ul tayinlash**" (EXACT counts).
2) ⏳ P0 — BUGUN DARHOL — 1-3 must-act-now CARDS (assign a missing owner / pull work off
   the overloaded person / the riskiest item). Omit the section if nothing is urgent today.
3) ⏳ <DD-MM> DEADLINE — <N> TA VAZIFA — every task due that day as a CARD, critical-path order.
   If deadlines do NOT cluster on one day, title it "📋 USTUVOR VAZIFALAR" and order by nearest deadline.
4) 📌 YUK BALANSI — per-owner load, COLOR-CODED: 🔴 overloaded (bottleneck) / 🟡 moderate /
   🟢 free. EXACT counts from the LOAD BY ASSIGNEE block. If NO delegation (all "—"), use 🔑 FAQAT SIZ.
5) ⏳ BUGUNGI HARAKAT REJASI — concrete time blocks ("HH:MM — <harakat>"); compute meeting+deadline conflicts.
6) 📌 QARORLAR — 🔷 firm decisions, including what to defer/drop (the trade-off).
7) 📌 BOSH FORMULA — "Bugun: **...**" / "Ertaga: **...**" / "<DD-MM>: **...**".

CARD format (one blank line between cards):
**<N>. <icon> <sarlavha>**
👤 Ijrochi: <ism / **tayinlanmagan** / qayta taqsimlash kerak>
⏳ Muddat: <DD-MM HH:MM>
⭐ Muhimlik: Yuqori   (yoki "🔷 Muhimlik: Rejadagi/Muhim")
📝 Izoh: <next action; optional "Ichki deadline: DD-MM, HH:MM">
Status icons: 🔴 Yuqori/P0 · 🟠 Muhim/P1 · ⚪ Rejadagi/P2-P3.

DATA SOURCE:
- If the principal's message describes a situation, plan around THAT.
- If EMPTY, build from the CURRENT PRINCIPAL STATE block — REAL active tasks (with their
  `assignee` for YUK BALANSI), today/this-week meetings, overdue and blocked items,
  deadlines. NEVER invent tasks, people, or dates. If state is empty, say so briefly and
  suggest adding tasks — don't fabricate a day.

FORMAT — Telegram-friendly, NO markdown tables: "**bold**" headers, "━━━" dividers, blank line between cards.
Output: user_message = full board (template above); actions=[].
"""


async def _run_planning_session(message: Message, situation: str) -> None:
    # A long Opus plan that lands ~30s later with zero feedback reads as
    # "broken". Stream it so the plan visibly builds, exactly like the main
    # chat path. The directive routes to Opus via its "executive_plan" keyword.
    directive = _PLAN_DIRECTIVE
    if situation and situation.strip():
        # The principal typed/dictated a situation — feed it in. (Previously this
        # was silently dropped: an internal_directive overrides user_text, so the
        # free-text `/plan <vaziyat>` was ignored and only the DB was planned.)
        directive = _PLAN_DIRECTIVE + "\n\n## PRINTSIPAL YOZGAN VAZIYAT\n" + situation.strip()

    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    progress_msg: Message | None = None
    last_edit_at = 0.0
    last_edit_text = ""
    loop = asyncio.get_event_loop()
    response: dict | None = None
    try:
        async for kind, payload in claude_service.process_message_stream(
            # Sonnet (was Opus) — principal's explicit cost choice. The executive_plan
            # directive would otherwise route to Opus via _COMPLEX_DIRECTIVE_KEYWORDS;
            # complexity="default" overrides that. Revert: drop this arg to restore Opus.
            "", internal_directive=directive, complexity="default",
        ):
            if kind == "partial":
                text = (payload or "").strip()
                if not text:
                    continue
                now = loop.time()
                if (now - last_edit_at) < _STREAM_EDIT_MIN_INTERVAL_SEC:
                    continue
                if abs(len(text) - len(last_edit_text)) < _STREAM_EDIT_MIN_DELTA_CHARS:
                    continue
                # Partials go out WITHOUT parse_mode — mid-stream markdown is
                # often unbalanced (open ** with no close) and would 400.
                if progress_msg is None:
                    progress_msg = await message.answer(text + " ▌")
                else:
                    try:
                        await progress_msg.edit_text(text + " ▌")
                    except TelegramBadRequest:
                        pass
                last_edit_at = now
                last_edit_text = text
            elif kind == "complete":
                response = payload
                break
    finally:
        typing_task.cancel()

    if response is None:
        response = claude_service._FALLBACK_RESPONSE
    plan_text = response.get("user_message", "")
    if not plan_text:
        if progress_msg is not None:
            try:
                await progress_msg.delete()
            except TelegramBadRequest:
                pass
        await message.answer("Reja yaratib boʻlmadi. Qaytadan urinib koʻring.")
        return

    plan_id = await database.save_plan(situation, plan_text)

    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Qabul qilaman", callback_data=f"plan_accept:{plan_id}"),
        InlineKeyboardButton(text="📌 Vazifalar yaratish", callback_data=f"plan_tasks:{plan_id}"),
    ], [
        back_button(),
    ]])
    # Finalize: drop the cursor, render markdown, attach buttons. A board longer
    # than one Telegram message can't be edited into the streamed bubble
    # (edit_text 400s past ~4096 chars), so in that case — or if the markdown
    # edit fails — delete the stale bubble and send the board split at section
    # boundaries via _safe_answer, instead of orphaning a truncated stream bubble.
    edited = False
    if progress_msg is not None and len(plan_text) <= _TG_SOFT_LIMIT:
        try:
            await progress_msg.edit_text(plan_text, parse_mode="Markdown", reply_markup=kb)
            edited = True
        except TelegramBadRequest:
            edited = False
    if not edited:
        if progress_msg is not None:
            try:
                await progress_msg.delete()
            except TelegramBadRequest:
                pass
        await _safe_answer(message, plan_text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("plan_accept:"))
async def cb_plan_accept(query: CallbackQuery) -> None:
    plan_id = query.data.split(":", 1)[1]
    await database.mark_plan_accepted(plan_id)
    await query.answer("Reja qabul qilindi ✅")
    try:
        await query.message.edit_reply_markup(reply_markup=single_back_keyboard())
    except Exception:
        pass


@router.callback_query(F.data.startswith("plan_tasks:"))
async def cb_plan_tasks(query: CallbackQuery, state: FSMContext) -> None:
    """Extract the principal's own tasks from a saved plan, then show a PREVIEW
    and require confirmation before creating them (reuses the standard
    acts_confirm pipeline — so nothing is created silently)."""
    plan_id = query.data.split(":", 1)[1]
    await query.answer("Vazifalarni ajrataman…")

    plans = await database.list_recent_plans(limit=50)
    plan = next((p for p in plans if p["id"] == plan_id), None)
    if not plan:
        await query.answer("Reja topilmadi", show_alert=True)
        return

    extract_directive = (
        "[INTERNAL] extract_tasks_from_plan\n\n"
        "Extract EVERY actionable task the principal must do himself (NOT delegated "
        "ones — skip items marked 'Topshiring' / Mas'ul ≠ Siz).\n\n"
        # 6000 (was 3000) so tasks deep in a long plan aren't dropped.
        f"PLAN:\n{plan['output_text'][:6000]}\n\n"
        "Output JSON envelope with actions=[create_task...]. Each task: title (imperative), "
        "priority (P0/P1/P2/P3 — map from Status), deadline (ISO 8601 Asia/Tashkent or null). "
        "user_message: one short Uzbek line."
    )
    response = await claude_service.process_message("", internal_directive=extract_directive)
    actions = [a for a in response.get("actions", []) if a.get("type") == "create_task"]
    if not actions:
        await query.message.answer("Rejadan bajariladigan (o'zingizning) vazifa topilmadi.")
        return

    # Confirm before creating — route through the existing acts_confirm flow.
    preview = await _format_create_preview(actions)
    await state.set_state(CreateActionConfirmFSM.awaiting)
    await state.update_data(
        pending_response={"actions": actions,
                          "user_message": f"📌 Rejadan {len(actions)} ta vazifa yaratildi.",
                          "buttons": []},
        _prior_section=None,
    )
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Tasdiqlayman", callback_data="acts_confirm"),
        InlineKeyboardButton(text="✕ Bekor qilish", callback_data="acts_cancel"),
    ]])
    await _safe_answer(query.message, preview, parse_mode="Markdown", reply_markup=confirm_kb)



@router.message(Command("stats"))
async def cmd_stats(message: Message, state: FSMContext | None = None) -> None:
    """Professional stats dashboard with KPI, risks, delegation, meetings, and audit.
    Section reply kbd bilan ishlaydi — davrlar pastdagi reply tugmalardan tanlanadi.
    """
    if state is not None:
        await state.set_state(SectionFSM.in_stats)
    days, label = _period_from_text(message.text)
    await message.answer(
        "📊 **STATISTIKA**", parse_mode="Markdown",
        reply_markup=stats_section_reply_keyboard(),
    )
    stats = await database.executive_stats(days=days)
    await _safe_answer(
        message,
        _format_stats_dashboard(stats, label),
        parse_mode="Markdown",
    )


@router.callback_query(F.data.startswith("stats_period:"))
async def cb_stats_period(query: CallbackQuery) -> None:
    days = _cb_int(query.data, default=7)
    if days not in (1, 7, 30):
        days = 7
    label = {1: "Bugun", 7: "Oxirgi 7 kun", 30: "Oxirgi 30 kun"}.get(days, "Oxirgi 7 kun")
    await query.answer()
    stats = await database.executive_stats(days=days)
    text = _format_stats_dashboard(stats, label)
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=_stats_period_keyboard(days))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=_stats_period_keyboard(days))


@router.message(Command("report"))
async def cmd_report(message: Message) -> None:
    days, label = _period_from_text(message.text.replace("/report", "/stats", 1) if message.text else None)
    stats = await database.executive_stats(days=days)
    await _safe_answer(
        message,
        _format_executive_report(stats, label),
        parse_mode="Markdown",
        reply_markup=_report_keyboard(days),
    )


@router.callback_query(F.data.startswith("report_period:"))
async def cb_report_period(query: CallbackQuery) -> None:
    days = _cb_int(query.data, default=7)
    if days not in (7, 30):
        days = 7
    label = {7: "Oxirgi 7 kun", 30: "Oxirgi 30 kun"}.get(days, "Oxirgi 7 kun")
    await query.answer()
    stats = await database.executive_stats(days=days)
    text = _format_executive_report(stats, label)
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=_report_keyboard(days))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=_report_keyboard(days))


@router.callback_query(F.data == "stats_back")
async def cb_stats_back(query: CallbackQuery, state: FSMContext) -> None:
    await query.answer()
    await state.clear()
    try:
        await query.message.delete()
    except Exception:
        pass
    await _restore_main_keyboard(query.message)


@router.message(Command("team"))
async def cmd_team(message: Message, state: FSMContext | None = None) -> None:
    """👥 Ijrochilar paneli — section reply kbd bilan."""
    if state is not None:
        await state.set_state(SectionFSM.in_team)
    await message.answer(
        "👥 **IJROCHILAR**", parse_mode="Markdown",
        reply_markup=team_section_reply_keyboard(),
    )
    await _render_team_panel(message)


@router.message(Command("risks"))
async def cmd_risks(message: Message, state: FSMContext | None = None) -> None:
    """🚨 Risklar paneli — section reply kbd bilan."""
    if state is not None:
        await state.set_state(SectionFSM.in_risks)
    await message.answer(
        "🚨 **RISKLAR**", parse_mode="Markdown",
        reply_markup=risks_section_reply_keyboard(),
    )
    await _render_risks_panel(message)


@router.message(Command("search"))
async def cmd_search(message: Message, state: FSMContext | None = None) -> None:
    """🔍 Qidiruv — section reply kbd bilan. Scope tanlanmasa, default `all`."""
    if state is not None:
        await state.set_state(SectionFSM.in_search)
        await state.update_data(scope="all")
    text = (
        "🔍 **QIDIRUV**\n\n"
        "Scope tanlang yoki to'g'ridan-to'g'ri so'z yuboring.\n\n"
        "_Default: hammasi (vazifa + uchrashuv + kontakt)._"
    )
    await message.answer(text, parse_mode="Markdown",
                         reply_markup=search_section_reply_keyboard())


# ─────────────────────── IJROCHILAR PANELI (Batch 4) ───────────────────────

import base64 as _b64_team  # local alias to avoid top-level dependency


def _encode_assignee(name: str) -> str:
    """Encode an assignee name for safe inclusion in callback_data (≤64 bytes total)."""
    return _b64_team.urlsafe_b64encode(name.encode("utf-8")).decode("ascii").rstrip("=")


def _decode_assignee(token: str) -> str:
    pad = "=" * (-len(token) % 4)
    return _b64_team.urlsafe_b64decode((token + pad).encode("ascii")).decode("utf-8")


def _load_band(active: int) -> tuple[str, str]:
    """Categorise an assignee's active-task load. Returns (label, emoji)."""
    if active >= _ASSIGNEE_OVERLOAD:
        return "yuqori", "🔴"
    if active >= 3:
        return "o'rta", "🟠"
    if active >= 1:
        return "past", "🟢"
    return "bo'sh", "⚪"


def _team_recommendation(loads: dict, top_loaded: dict | None,
                          unassigned_n: int, total_overdue: int) -> str:
    """Heuristic recommendation for the team panel."""
    if not loads or (len(loads) == 1 and "belgilanmagan" in loads):
        return ("Hozircha ijrochilar belgilanmagan. Yangi vazifa yaratishda 👤 Ijrochi tugmasidan "
                "foydalanib mas'ul tayinlang.")
    if unassigned_n >= 5:
        return (f"**{unassigned_n} ta vazifa ijrochisiz**. Avval ularga mas'ul belgilang — "
                "javobgarliksiz vazifa amalga oshmaydi.")
    if top_loaded and top_loaded["active"] >= _ASSIGNEE_OVERLOAD:
        # Find someone less loaded
        light = [
            d for n, d in loads.items()
            if n not in ("belgilanmagan", top_loaded["name"]) and d["active"] < 3
        ]
        if light:
            target = min(light, key=lambda d: d["active"])
            return (f"**{top_loaded['name']}** ortiqcha yuklangan ({top_loaded['active']} ta). "
                    f"Ba'zi vazifalarni **{target['name']}** ga ko'chirib ko'ring "
                    f"(hozir {target['active']} ta).")
        return (f"**{top_loaded['name']}** ortiqcha yuklangan ({top_loaded['active']} ta). "
                "Yangi vazifa berishni vaqtincha to'xtating yoki muddatlarni cho'zing.")
    if total_overdue >= 3:
        return (f"Jamoada **{total_overdue} ta muddati o'tgan** vazifa bor. Har bir ijrochi bilan "
                "alohida ishlab, muddatlarni qayta belgilang.")
    return ("Jamoa yuklamasi muvozanatda. Yangi vazifa berishda eng kam yuklangan ijrochini tanlang.")


def _format_assignees_overview(loads: dict, unassigned_count: int) -> tuple[str, list[dict]]:
    """Ijrochilar paneli — Vazifalar/Uchrashuvlar uslubidagi block-stil.

    Ikonkalar (oldingi paletadan):
      👥 sahifa · 📌 stats · ⚠️ diqqat · 💡 tavsiya
    Detal qator (har ijrochi uchun) ikonkalari:
      📊 Aktiv · ⚡ Yuklama · ⚡/⭐/⏰ holat · 📅 keyingi muddat
    Spacing: divider atrofida 1 bo'sh qator, item orasida 1 bo'sh qator,
    detal blok 6 ta probel indent.
    """
    DIVIDER = "━" * 20

    real = [d for name, d in loads.items() if name != "belgilanmagan"]
    real.sort(key=lambda d: (-d["active"], -d["urgent"], -d["overdue"]))

    busy = [d for d in real if d["active"] > 0]
    overloaded = [d for d in real if d["active"] >= _ASSIGNEE_OVERLOAD]
    total_overdue = sum(d["overdue"] for d in real)

    lines = [
        "👥 **IJROCHILAR**",
        "",
        f"Jami {len(real)} nafar  ·  Yuklangan {len(busy)}  ·  Ortiqcha {len(overloaded)}",
    ]

    if not real:
        lines.extend([
            "",
            DIVIDER,
            "",
            "Bu bo'limda hozircha ijrochilar yo'q.",
            "",
            "_Yangi vazifa yaratishda ijrochi tayinlang._",
        ])
        return "\n".join(lines), []

    if unassigned_count > 0:
        lines.append(f"Ijrochisiz vazifa  ·  {unassigned_count} ta")

    lines.extend(["", DIVIDER, "", "📌 **RO'YXAT**", ""])

    def _next_deadline_label(next_dl: str | None) -> str:
        if not next_dl:
            return "yo'q"
        try:
            dl_dt = datetime.fromisoformat(next_dl)
        except (ValueError, TypeError):
            return "—"
        now = datetime.now(database.TZ)
        if dl_dt.date() == now.date():
            return f"Bugun {dl_dt.strftime('%H:%M')}"
        if dl_dt.date() == (now + timedelta(days=1)).date():
            return f"Ertaga {dl_dt.strftime('%H:%M')}"
        return dl_dt.strftime("%d-%m, %H:%M")

    items: list[str] = []
    for i, d in enumerate(real[:10], start=1):
        band, _badge = _load_band(d["active"])
        dl_label = _next_deadline_label(d.get("next_deadline"))
        block = [
            f"{i}.  {d['name']}",
            "",
            f"      📊 Aktiv:           {d['active']} ta",
            f"      ⚡ Yuklama:         {band}",
        ]
        # Selective signal lines — only render if non-zero
        if d.get("urgent"):
            block.append(f"      ⚡ Shoshilinch:     {d['urgent']} ta")
        if d.get("important"):
            block.append(f"      ⭐ Muhim:           {d['important']} ta")
        if d.get("overdue"):
            block.append(f"      ⏰ O'tgan:          {d['overdue']} ta")
        block.append(f"      📅 Keyingi muddat:  {dl_label}")
        items.append("\n".join(block))

    lines.append("\n\n".join(items))
    lines.append("")

    if len(real) > 10:
        lines.append(f"_+{len(real) - 10} ta yana_")
        lines.append("")

    # ⚠️ DIQQAT — surface specific anomalies (kept as bullets for quick scanning)
    flags = []
    for d in real[:10]:
        if d["active"] >= _ASSIGNEE_OVERLOAD:
            flags.append(f"  • **{d['name']}** — yuklamasi yuqori ({d['active']} ta)")
        if d["overdue"] >= 2:
            flags.append(f"  • **{d['name']}** — {d['overdue']} ta muddati o'tgan")
    if unassigned_count >= 3:
        flags.append(f"  • Belgilanmagan — {unassigned_count} ta vazifa")
    if flags:
        lines.extend([DIVIDER, "", "⚠️ **DIQQAT**", ""])
        lines.extend(flags[:5])
        lines.append("")

    top_loaded = real[0] if real else None
    rec = _team_recommendation(loads, top_loaded, unassigned_count, total_overdue)
    lines.extend([DIVIDER, "", "💡 **TAVSIYA**", "", rec])

    return "\n".join(lines), real[:10]


async def _render_team_panel(message: Message) -> None:
    """Ijrochilar paneli — overview + drill-down + reassign shortcuts."""
    loads = await database.assignee_load_map()
    unassigned = await database.list_unassigned_tasks(limit=500)

    text, ordered = _format_assignees_overview(loads, len(unassigned))

    # Build keyboard: number drill-downs for top assignees, then actions, then back.
    rows: list[list[InlineKeyboardButton]] = []
    if ordered:
        nums = [
            InlineKeyboardButton(text=str(i + 1),
                                  callback_data=f"team:open:{_encode_assignee(d['name'])}")
            for i, d in enumerate(ordered)
        ]
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])

    # Manual executor add (controlled list — A): new assignees come only from here or
    # Excel import, never auto from a voice/text turn.
    rows.append([InlineKeyboardButton(text="➕ Ijrochi qo'shish", callback_data="contactadd")])
    await _safe_answer(message, text, parse_mode="Markdown",
                       reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


class ContactAddFSM(StatesGroup):
    awaiting = State()


@router.callback_query(F.data == "contactadd")
async def cb_contact_add(query: CallbackQuery, state: FSMContext) -> None:
    """Manually add an executor (contact) — the controlled-list add path (A)."""
    await state.set_state(ContactAddFSM.awaiting)
    await query.answer()
    await query.message.answer(
        "➕ Ijrochi ism(lar)ini yozing (matn/ovoz). Bir nechta — har qatorga bittadan.")


@router.message(StateFilter(ContactAddFSM.awaiting), F.text | F.voice)
async def handle_contact_add(message: Message, state: FSMContext, bot: Bot) -> None:
    text = await _get_text_or_transcribe(message, bot=bot)
    if text is None:
        return
    await state.clear()
    added = 0
    for line in (text or "").splitlines():
        nm = line.strip(" -•\t·")
        if nm:
            await database.save_contact({"name": nm[:80]})
            added += 1
    if not added:
        await message.answer("Bo'sh — bekor qilindi.")
        return
    await message.answer(f"✅ {added} ta ijrochi qo'shildi — endi ularga vazifa tayinlasa bo'ladi.")
    await _render_team_panel(message)


def _format_assignee_profile(profile: dict) -> str:
    """Render a single assignee's profile card."""
    DIVIDER = "━" * 20
    name = profile["name"]
    band, _badge = _load_band(profile["active"])

    lines = [
        "👤 **IJROCHI**",
        "",
        name,
        "",
        DIVIDER,
        "",
        "📌 **STATISTIKA**",
        "",
        f"      📊 Aktiv:           {profile['active']} ta",
        f"      ✅ Bajarilgan:      {profile['completed']} ta",
        f"      📈 Bajarilish:      {profile['completion_rate']}%",
        f"      ⏰ O'tgan:          {profile['overdue']} ta",
        f"      ⚡ Shoshilinch:     {profile['urgent']} ta",
        f"      ⭐ Muhim:           {profile['important']} ta",
    ]
    if profile.get("avg_closing_hours"):
        lines.append(f"      ⌛ O'rtacha:        {profile['avg_closing_hours']} soat")
    lines.append("")

    tasks = profile.get("tasks", [])
    if tasks:
        lines.extend([DIVIDER, "", f"📋 **AKTIV VAZIFALAR**   ·   {len(tasks)} ta", ""])
        items: list[str] = []
        for i, t in enumerate(tasks, start=1):
            title = (t.get("title") or "—").strip()
            muhimlik = _PRIORITY_LABEL_UZ.get(t.get("priority", "P2"), "Rejadagi")
            block = [
                f"{i}.  {title}",
                "",
                f"      ⏳ Muddat:          {_muddat_label(t.get('deadline'))}",
                f"      🔹 Muhimlik:        {muhimlik}",
            ]
            items.append("\n".join(block))
        lines.append("\n\n".join(items))
        lines.append("")
    else:
        lines.extend([DIVIDER, "", "_Hozircha aktiv vazifa yo'q._", ""])

    # Yuklama signali
    lines.extend([DIVIDER, "", "⚡ **YUKLAMA**", ""])
    if band == "yuqori":
        lines.append(f"Yuqori — {profile['active']}+ aktiv vazifa.")
        lines.append("_Yangi vazifa berishni cheklang yoki muddatlarni cho'zing._")
    elif band == "o'rta":
        lines.append(f"O'rta — {profile['active']} ta aktiv.")
        lines.append("_Yuklamani kuzatib turing, muvozanatda._")
    elif band == "past":
        lines.append(f"Past — {profile['active']} ta aktiv.")
        lines.append("_Qo'shimcha vazifa qabul qila oladi._")
    else:
        lines.append("Bo'sh — aktiv vazifa yo'q.")
        lines.append("_Yangi vazifalar uchun ochiq._")

    return "\n".join(lines)


@router.callback_query(F.data.startswith("team:open:"))
async def cb_team_open(query: CallbackQuery) -> None:
    """Drill-down: show one assignee's full profile."""
    token = query.data.split(":", 2)[2]
    try:
        name = _decode_assignee(token)
    except Exception:
        await query.answer("Ijrochini tahlil qilib bo'lmadi", show_alert=True)
        return
    profile = await database.assignee_profile(name)
    if not profile:
        await query.answer("Ijrochi topilmadi", show_alert=True)
        return
    await query.answer()

    # Build per-task drill-down buttons (numbered) + reassign + back
    rows: list[list[InlineKeyboardButton]] = []
    tasks = profile.get("tasks", [])
    if tasks:
        nums = [
            InlineKeyboardButton(text=str(i + 1), callback_data=f"taskopen:{t['id']}")
            for i, t in enumerate(tasks)
        ]
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
    rows.append([
        InlineKeyboardButton(text="🔄 Qayta taqsimlash",
                              callback_data=f"team:reassign_one:{token}"),
        InlineKeyboardButton(text="⬅️ Jamoa", callback_data="team:refresh"),
    ])

    text = _format_assignee_profile(profile)
    try:
        await query.message.edit_text(
            text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=InlineKeyboardMarkup(inline_keyboard=rows))


@router.callback_query(F.data == "team:refresh")
async def cb_team_refresh(query: CallbackQuery) -> None:
    await query.answer("Yangilanmoqda...")
    try:
        await query.message.delete()
    except Exception:
        pass
    await _render_team_panel(query.message)





@router.callback_query(F.data.startswith("team:reassign_one:"))
async def cb_team_reassign_one(query: CallbackQuery) -> None:
    """Show one assignee's tasks for picking which to reassign."""
    token = query.data.split(":", 2)[2]
    try:
        name = _decode_assignee(token)
    except Exception:
        await query.answer("Ijrochini tahlil qilib bo'lmadi", show_alert=True)
        return
    profile = await database.assignee_profile(name)
    if not profile or not profile.get("tasks"):
        await query.answer("Ijrochida aktiv vazifa yo'q", show_alert=True)
        return
    await query.answer()
    tasks = profile["tasks"][:10]
    lines = [
        f"🔄 **QAYTA TAQSIMLASH** — {name}",
        _SEP,
        "",
        "Qaysi vazifani boshqa ijrochiga ko'chirmoqchisiz?",
        "",
    ]
    nums: list[InlineKeyboardButton] = []
    for i, t in enumerate(tasks, start=1):
        badge = _PRIORITY_BADGE.get(t.get("priority"), "🔵")
        title = (t.get("title") or "—").strip()
        dl_label, _ovd = _format_deadline_short(t.get("deadline"))
        lines.append(f"**{i}. {badge} {title}**")
        lines.append(f"   ⏰ {dl_label}")
        lines.append("")
        nums.append(InlineKeyboardButton(text=str(i), callback_data=f"set_assignee:{t['id']}"))
    grid = [nums[i:i + 5] for i in range(0, len(nums), 5)]
    grid.append([InlineKeyboardButton(text="⬅️ Profil", callback_data=f"team:open:{token}")])
    await _safe_answer(
        query.message, "\n".join(lines).rstrip(),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=grid),
    )


# ─────────────────────── RISKLAR PANELI (Batch 3) ───────────────────────

# Threshold: an assignee with this many active tasks is considered "overloaded".
_ASSIGNEE_OVERLOAD = 5


async def _classify_risks() -> dict:
    """Group active tasks into yuqori / o'rta / past risk buckets per the spec.

    Each task lands in only ONE bucket — the highest-risk one that matches.
    Returns: {high: [(task, reason)], medium: [...], low: [...]}.
    """
    now = datetime.now(database.TZ)
    horizon_24 = (now + timedelta(hours=24)).isoformat()
    horizon_48 = (now + timedelta(hours=48)).isoformat()
    horizon_3d = (now + timedelta(days=3)).isoformat()
    now_iso_str = now.isoformat()

    active = await database.list_tasks(status_in=["todo", "in_progress"], limit=500)
    loads = await database.assignee_load_map()
    overloaded = {
        name for name, d in loads.items()
        if name != "belgilanmagan" and d["active"] >= _ASSIGNEE_OVERLOAD
    }

    high: list[tuple[dict, str]] = []
    medium: list[tuple[dict, str]] = []
    low: list[tuple[dict, str]] = []

    for t in active:
        deadline = t.get("deadline")
        priority = t.get("priority", "P2")
        assignee = (t.get("assignee") or "").strip()
        is_unassigned = (not assignee) or assignee.lower() == "belgilanmagan"
        status = t.get("status", "todo")

        # HIGH (first match wins)
        if deadline and deadline < now_iso_str:
            high.append((t, "Deadline o'tdi"))
            continue
        if priority == "P0":
            high.append((t, "Shoshilinch va hali bajarilmagan"))
            continue
        if deadline and deadline <= horizon_24:
            high.append((t, "Deadline 24 soat ichida"))
            continue
        if is_unassigned and deadline and deadline <= horizon_48:
            high.append((t, "Ijrochi yo'q + muddat 48 soatda"))
            continue

        # MEDIUM
        if deadline and deadline <= horizon_48:
            medium.append((t, "Deadline 48 soat ichida"))
            continue
        if priority == "P1" and status == "todo":
            medium.append((t, "Muhim, jarayonda emas"))
            continue
        if (not is_unassigned) and assignee in overloaded:
            medium.append((t, f"Ijrochi yuklamasi yuqori ({assignee})"))
            continue

        # LOW: deadline + assignee + in progress + > 3 days away
        if (deadline and not is_unassigned and status == "in_progress"
                and deadline > horizon_3d):
            low.append((t, "Rejada, jarayonda"))
            continue

    # Order each bucket: P0 first, then nearest deadline first (None last).
    pri_rank = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}

    def _sort_key(item):
        t, _ = item
        return (
            pri_rank.get(t.get("priority", "P2"), 2),
            t.get("deadline") or "9999",
        )

    high.sort(key=_sort_key)
    medium.sort(key=_sort_key)
    low.sort(key=_sort_key)

    return {"high": high, "medium": medium, "low": low}


def _risks_recommendation(high: list, medium: list, low: list, risk: dict) -> str:
    """Heuristic system recommendation for the risks panel."""
    overdue_n = sum(1 for _, r in high if "Deadline o'tdi" in r)
    unassigned_n = sum(1 for _, r in high if "Ijrochi yo'q" in r)

    if not high and not medium:
        return ("Hozir kritik risk yo'q. Past riskdagi vazifalarni rejada davom ettiring "
                "va keyingi vazifalarga muddat belgilashga vaqt ajrating.")
    if overdue_n >= 1:
        return (f"**{overdue_n} ta vazifa muddati o'tgan**. Avval shularni yoping yoki "
                "realistik yangi muddat tayinlang — eski deadline'lar yangi vazifalarni ham buzadi.")
    if unassigned_n >= 1:
        return (f"**{unassigned_n} ta yuqori riskli vazifa ijrochisiz**. Mas'ul belgilamasdan "
                "ular kechikadi. Avval [👤 Ijrochi tayinlash] tugmasidan foydalaning.")
    if len(high) >= 5:
        return (f"**{len(high)} ta yuqori risk** to'planib qoldi. Bir kunda hammasini yopish "
                "qiyin — eng yaqin deadline'dan boshlang yoki ba'zilarini delegatsiya qiling.")
    if risk["score"] >= 70:
        return ("Risk darajasi yuqori. Bugun faqat yuqori risk vazifalariga e'tibor qarating, "
                "yangi vazifalar yaratishni vaqtincha to'xtating.")
    if len(medium) >= 5:
        return (f"**{len(medium)} ta o'rta risk** kuzatuvda. Yuqori bo'lib ketmasligi uchun "
                "shu hafta ichida boshlang yoki muddatlarni qayta ko'rib chiqing.")
    return ("Yuqori risklarni bugun yoki ertaga yoping. O'rta risklarga 24-48 soat ichida "
            "vaqt ajrating — proaktiv yondashish kechikishni oldini oladi.")


async def _render_risks_panel(message: Message) -> None:
    """Risklar paneli — comprehensive risk classification + action shortcuts."""
    buckets = await _classify_risks()
    high = buckets["high"]
    medium = buckets["medium"]
    low = buckets["low"]
    risk = await compute_risk_score()
    now = datetime.now(database.TZ)

    lines = [
        f"🚨 **RISKLAR PANELI** · {now.strftime('%d-%m %H:%M')}",
        _SEP,
        "",
        f"Risk score: {risk['emoji']} **{risk['score']}/100** — {risk['status']}",
        f"Jami risklar: 🔴 {len(high)} · 🟠 {len(medium)} · ⚪ {len(low)}",
        "",
    ]

    def fmt_block(tasks: list, badge: str, label: str, limit: int) -> list[str]:
        out = [f"{badge} **{label.upper()}** ({len(tasks)} ta)", ""]
        if not tasks:
            out.append("_Hozircha bunday vazifa yo'q._")
            out.append("")
            return out
        for i, (t, reason) in enumerate(tasks[:limit], start=1):
            badge_p = _PRIORITY_BADGE.get(t.get("priority", "P2"), "🔵")
            title = (t.get("title") or "—").strip()
            deadline_label, _ovd = _format_deadline_short(t.get("deadline"))
            assignee = (t.get("assignee") or "").strip() or "belgilanmagan"
            out.append(f"**{i}. {badge_p} {title}**")
            out.append(f"• Sabab: _{reason}_")
            out.append(f"• Muddat: {deadline_label}")
            out.append(f"• Ijrochi: {assignee}")
            out.append("")
        if len(tasks) > limit:
            out.append(f"_+{len(tasks) - limit} ta yana ko'rsatilmadi_")
            out.append("")
        return out

    lines.extend(fmt_block(high, "🔴", "Yuqori risk", limit=5))
    lines.extend(fmt_block(medium, "🟠", "O'rta risk", limit=5))
    lines.extend(fmt_block(low, "⚪", "Past risk", limit=3))

    lines.extend([
        "🤖 **TIZIM TAVSIYASI**",
        _risks_recommendation(high, medium, low, risk),
    ])

    # Build keyboard: numbered drill-down for HIGH tasks (open task card),
    # then global action shortcuts, then back.
    rows: list[list[InlineKeyboardButton]] = []
    if high:
        nums = [
            InlineKeyboardButton(text=str(i + 1), callback_data=f"taskopen:{t['id']}")
            for i, (t, _) in enumerate(high[:5])
        ]
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])

    # Object-specific tezkor amallar (Ijrochi tayinlash, Eslatma, Muddat) —
    # inline'da qoladi (kontekstli). 🔄 Yangilash va ⬅️ Cockpit reply kbd'da bor.
    rows.extend([
        [
            InlineKeyboardButton(text="👤 Ijrochi tayinlash", callback_data="risks_assign"),
            InlineKeyboardButton(text="⏰ Eslatma qo'yish", callback_data="risks_remind"),
        ],
        [
            InlineKeyboardButton(text="📅 Muddatni o'zgartirish", callback_data="risks_deadline"),
        ],
    ])

    await _safe_answer(
        message,
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows) if rows else None,
    )


def _risks_sublist(title_line: str, tasks: list[dict],
                   callback_prefix: str, empty_msg: str) -> tuple[str, InlineKeyboardMarkup]:
    """Build a numbered sublist for one of the risk-action shortcuts.

    callback_prefix: prefix for each numbered button's callback_data, e.g.
                     "set_assignee", "taskopen", "editfield". The task id is appended.
                     For "editfield", we append ":deadline" too.
    """
    if not tasks:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Risklar", callback_data="risks_refresh")],
        ])
        return empty_msg, kb

    lines = [title_line, _SEP, ""]
    nums: list[InlineKeyboardButton] = []
    for i, t in enumerate(tasks[:10], start=1):
        badge = _task_badge(t)  # unified with the list — same dot for the same task
        title = (t.get("title") or "—").strip()
        deadline_label, _ovd = _format_deadline_short(t.get("deadline"))
        assignee = (t.get("assignee") or "").strip() or "belgilanmagan"
        lines.append(f"**{i}. {badge} {title}**")
        lines.append(f"• Muddat: {deadline_label}")
        lines.append(f"• Ijrochi: {assignee}")
        lines.append("")

        if callback_prefix == "editfield":
            cb = f"editfield:{t['id']}:deadline"
        else:
            cb = f"{callback_prefix}:{t['id']}"
        nums.append(InlineKeyboardButton(text=str(i), callback_data=cb))

    grid = [nums[i:i + 5] for i in range(0, len(nums), 5)]
    grid.append([InlineKeyboardButton(text="⬅️ Risklar", callback_data="risks_refresh")])
    return "\n".join(lines).rstrip(), InlineKeyboardMarkup(inline_keyboard=grid)


@router.callback_query(F.data == "risks_refresh")
async def cb_risks_refresh(query: CallbackQuery) -> None:
    await query.answer("Yangilanmoqda...")
    try:
        await query.message.delete()
    except Exception:
        pass
    await _render_risks_panel(query.message)



@router.callback_query(F.data == "risks_assign")
async def cb_risks_assign(query: CallbackQuery) -> None:
    """Shortcut: list unassigned risky tasks → tap a number → set_assignee flow."""
    await query.answer()
    candidates = await database.list_unassigned_tasks(limit=10)
    text, kb = _risks_sublist(
        title_line="👤 **Ijrochi tayinlash kerak**",
        tasks=candidates,
        callback_prefix="set_assignee",
        empty_msg="✅ Ijrochisiz vazifa qolmagan.",
    )
    await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data == "risks_remind")
async def cb_risks_remind(query: CallbackQuery) -> None:
    """Shortcut: list high-risk tasks → tap a number → open task card (reminder there)."""
    await query.answer()
    buckets = await _classify_risks()
    candidates = [t for t, _ in buckets["high"][:10]]
    text, kb = _risks_sublist(
        title_line="⏰ **Eslatma qo'yish uchun vazifani tanlang**",
        tasks=candidates,
        callback_prefix="taskopen",
        empty_msg="✅ Hozir yuqori risk-vazifa yo'q.",
    )
    await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data == "risks_deadline")
async def cb_risks_deadline(query: CallbackQuery) -> None:
    """Shortcut: list overdue + 24h tasks → tap a number → opens deadline picker."""
    await query.answer()
    buckets = await _classify_risks()
    candidates = [t for t, _ in buckets["high"][:10]]
    text, kb = _risks_sublist(
        title_line="📅 **Muddat o'zgartirish kerak**",
        tasks=candidates,
        callback_prefix="editfield",  # → editfield:<id>:deadline
        empty_msg="✅ Muddat o'zgartirish talab qilinadigan vazifa yo'q.",
    )
    await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


# ─────────────────────── /search (Batch 5) ───────────────────────


async def _render_search_prompt(message: Message, state: FSMContext = None) -> None:
    """Global search across tasks + meetings + contacts."""
    if state is not None:
        await state.set_state(GlobalSearchFSM.awaiting_query)
    text = (
        "🔍 **QIDIRUV**\n" + _SEP + "\n\n"
        "Sarlavha, tavsif, ijrochi, ishtirokchi yoki teg bo'yicha so'z yuboring.\n\n"
        "_Misol: «marketing», «Aziz aka», «shoshilinch»._"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="search_cancel")],
    ])
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data == "search_cancel")
async def cb_search_cancel(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer("Bekor qilindi")
    try:
        await query.message.delete()
    except Exception:
        pass


@router.message(StateFilter(GlobalSearchFSM.awaiting_query), F.text | F.voice)
async def handle_global_search(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    await state.clear()
    q = (message.text or "").strip()
    if not q or q.startswith("/"):
        await _safe_answer(
            message,
            "Qidiruv so'zi bo'sh bo'lmasin. `/search` yana yuboring.",
            parse_mode="Markdown", reply_markup=main_reply_keyboard(),
        )
        return

    results = await database.search_all(q, limit=20)
    tasks = results.get("tasks", [])
    reminders = results.get("reminders", [])
    meetings = results.get("meetings", [])

    lines = [
        f"🔍 **QIDIRUV NATIJASI** · «{q}»",
        _SEP,
        "",
        f"Topildi: **{len(tasks)} vazifa** · **{len(reminders)} eslatma** · **{len(meetings)} uchrashuv**",
        "",
    ]

    if not tasks and not reminders and not meetings:
        lines.extend([
            "_Hech narsa topilmadi._",
            "",
            "Boshqa so'z yoki ism bilan urinib ko'ring.",
        ])
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 Qayta qidirish", callback_data="search_again")],
            [InlineKeyboardButton(text="⬅️ Boshqaruv paneli", callback_data="search_back")],
        ])
        await _safe_answer(message, "\n".join(lines), parse_mode="Markdown", reply_markup=kb)
        return

    rows: list[list[InlineKeyboardButton]] = []

    if tasks:
        lines.append("📌 **VAZIFALAR**")
        lines.append("")
        nums: list[InlineKeyboardButton] = []
        for i, ttask in enumerate(tasks[:8], start=1):
            badge = _PRIORITY_BADGE.get(ttask.get("priority", "P2"), "🔵")
            title = (ttask.get("title") or "—").strip()
            dl_label, _ovd = _format_deadline_short(ttask.get("deadline"))
            assignee = (ttask.get("assignee") or "").strip() or "belgilanmagan"
            status_uz = _STATUS_LABEL_UZ.get(ttask.get("status", "todo"), ttask.get("status", "todo"))
            lines.append(f"**{i}. {badge} {title}**")
            lines.append(f"   ⏰ {dl_label} · 👤 {assignee} · 📊 {status_uz}")
            lines.append("")
            nums.append(InlineKeyboardButton(text=str(i), callback_data=f"taskopen:{ttask['id']}"))
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
        if len(tasks) > 8:
            lines.append(f"_+{len(tasks) - 8} ta vazifa yana ko'rsatilmadi_")
            lines.append("")

    if reminders:
        lines.append("⏰ **ESLATMALAR**")
        lines.append("")
        nums: list[InlineKeyboardButton] = []
        for i, reminder in enumerate(reminders[:6], start=1):
            title = (reminder.get("title") or "—").strip()
            lines.append(f"**R{i}. {_reminder_status_icon(reminder)} {title}**")
            lines.append(f"   ⏰ {_reminder_time_chip(reminder)} · 📌 {_reminder_status_label(reminder.get('status'))}")
            lines.append("")
            nums.append(InlineKeyboardButton(text=f"R{i}", callback_data=f"remopen:{reminder['id']}"))
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
        if len(reminders) > 6:
            lines.append(f"_+{len(reminders) - 6} ta eslatma yana ko'rsatilmadi_")
            lines.append("")

    if meetings:
        lines.append("🤝 **UCHRASHUVLAR**")
        lines.append("")
        nums: list[InlineKeyboardButton] = []
        for i, m in enumerate(meetings[:5], start=1):
            try:
                dt = datetime.fromisoformat(m["datetime_start"]).astimezone(database.TZ)
                when_label = dt.strftime("%d-%m %H:%M")
            except (ValueError, TypeError):
                when_label = m.get("datetime_start") or "—"
            title = (m.get("title") or "—").strip()
            participants = ", ".join(m.get("participants", [])) or "—"
            lines.append(f"**M{i}. {title}**")
            lines.append(f"   🕐 {when_label} · 👥 {participants}")
            lines.append("")
            nums.append(InlineKeyboardButton(
                text=f"M{i}", callback_data=f"meetingopen:{m['id']}"))
        for i in range(0, len(nums), 5):
            rows.append(nums[i:i + 5])
        if len(meetings) > 5:
            lines.append(f"_+{len(meetings) - 5} ta uchrashuv yana ko'rsatilmadi_")
            lines.append("")

    rows.append([
        InlineKeyboardButton(text="🔄 Qayta qidirish", callback_data="search_again"),
        InlineKeyboardButton(text="⬅️ Boshqaruv paneli", callback_data="search_back"),
    ])

    await _safe_answer(
        message, "\n".join(lines).rstrip(),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


@router.callback_query(F.data == "search_again")
async def cb_search_again(query: CallbackQuery, state: FSMContext) -> None:
    await query.answer()
    try:
        await query.message.delete()
    except Exception:
        pass
    await _render_search_prompt(query.message, state)


@router.callback_query(F.data == "search_back")
async def cb_search_back(query: CallbackQuery) -> None:
    await query.answer()
    try:
        await query.message.delete()
    except Exception:
        pass
    await cmd_cockpit(query.message)


# ─────────────────────── YANGI VAZIFA — guided FSM (Batch 5) ───────────────────────


_NEWTASK_PRIORITY_MAP = {"P0": "Shoshilinch", "P1": "Muhim", "P2": "Rejadagi", "P3": "Oddiy"}

# Voice/text → priority code. Longest phrase wins (so "juda muhim" → P0, not P1).
_PRIORITY_WORDS = {
    "shoshilinch": "P0", "juda muhim": "P0", "zudlik": "P0", "zarur": "P0",
    "favqulodda": "P0", "tezkor": "P0",
    "muhim": "P1", "kerakli": "P1",
    "rejadagi": "P2", "rejali": "P2", "oddiy": "P2", "o'rtacha": "P2",
    "past ustuvorlik": "P3", "muhim emas": "P3", "shoshilmaydi": "P3", "past": "P3",
}


def _parse_priority_word(text: str) -> str | None:
    """Map an Uzbek priority word/phrase (or raw 'P0'..'P3') to a priority code.
    Returns None when nothing matches so the caller can re-prompt."""
    s = (text or "").strip().lower()
    if s in {"p0", "p1", "p2", "p3"}:
        return s.upper()
    for phrase in sorted(_PRIORITY_WORDS, key=len, reverse=True):
        if phrase in s:
            return _PRIORITY_WORDS[phrase]
    return None


def _newtask_summary(data: dict) -> str:
    """Render the running summary shown above each form step."""
    title = data.get("title") or "_(kiritilmagan)_"
    pri = data.get("priority")
    pri_label = (f"{_PRIORITY_BADGE.get(pri, '🔵')} {_PRIORITY_LABEL_UZ.get(pri, '—')}"
                 if pri else "_(tanlanmagan)_")
    deadline = data.get("deadline")
    if deadline:
        dl_label, _ovd = _format_deadline_short(deadline)
    else:
        dl_label = "_(o'tkazib yuborilgan)_"
    assignee = data.get("assignee") or "_(o'zim)_"
    return (
        f"📝 Sarlavha: {title}\n"
        f"⚡ Ustuvorlik: {pri_label}\n"
        f"📅 Muddat: {dl_label}\n"
        f"👤 Ijrochi: {assignee}"
    )


def _newtask_priority_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔴 Shoshilinch", callback_data="newtask:pri:P0"),
            InlineKeyboardButton(text="🟠 Muhim",      callback_data="newtask:pri:P1"),
        ],
        [
            InlineKeyboardButton(text="🔵 Rejadagi",   callback_data="newtask:pri:P2"),
            InlineKeyboardButton(text="⚪ Oddiy",      callback_data="newtask:pri:P3"),
        ],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
    ])


def _newtask_deadline_kb() -> InlineKeyboardMarkup:
    """Deadline step 1 — pick a DAY. The time is chosen on the next screen."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Bugun", callback_data="newtask:dl:today"),
            InlineKeyboardButton(text="📅 Ertaga", callback_data="newtask:dl:tomorrow"),
            InlineKeyboardButton(text="📅 Indin", callback_data="newtask:dl:indin"),
        ],
        [
            InlineKeyboardButton(text="📅 +3 kun", callback_data="newtask:dl:plus3"),
            InlineKeyboardButton(text="📅 Hafta oxiri", callback_data="newtask:dl:weekend"),
        ],
        [
            InlineKeyboardButton(text="✏️ Qo'lda", callback_data="newtask:dl:manual"),
            InlineKeyboardButton(text="⏭ O'tkazib yuborish", callback_data="newtask:dl:skip"),
        ],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
    ])


_NEWTASK_TIME_SLOTS = ["09:00", "12:00", "14:00", "17:00", "18:00"]


def _newtask_time_kb() -> InlineKeyboardMarkup:
    """Deadline step 2 — pick a TIME for the already-chosen day. Time callbacks
    use 'HHMM' (no colon) so the ':' delimiter split stays unambiguous."""
    slots = [
        InlineKeyboardButton(text=t, callback_data=f"newtask:tm:{t.replace(':', '')}")
        for t in _NEWTASK_TIME_SLOTS
    ]
    return InlineKeyboardMarkup(inline_keyboard=[
        slots[0:3],
        slots[3:5] + [InlineKeyboardButton(text="⌨️ Boshqa", callback_data="newtask:tm:custom")],
        [
            InlineKeyboardButton(text="⬅️ Orqaga", callback_data="newtask:dl:back"),
            InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel"),
        ],
    ])


def _newtask_assignee_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="👤 Men", callback_data="newtask:as:me"),
            InlineKeyboardButton(text="⏭ O'tkazib yuborish", callback_data="newtask:as:skip"),
        ],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
    ])


def _newtask_confirm_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Tasdiqlash", callback_data="newtask:confirm"),
            InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel"),
        ],
    ])


async def _newtask_start(message: Message, state: FSMContext) -> None:
    """Entry to the guided new-task form."""
    await state.set_state(NewTaskFSM.awaiting_title)
    await state.update_data(
        title=None, priority=None, deadline=None, assignee=None,
    )
    text = (
        "➕ **YANGI VAZIFA — Forma**\n" + _SEP + "\n\n"
        "1️⃣ **Sarlavha** kiriting:\n"
        "_Masalan: «Marketing strategiyasini Aziz akaga yuborish»_"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
    ])
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


@router.message(StateFilter(NewTaskFSM.awaiting_title), F.text | F.voice)
async def newtask_title(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    title = (message.text or "").strip()
    if not title or title.startswith("/"):
        await message.answer("Sarlavha bo'sh bo'lmasin. Iltimos, qaytadan yuboring.")
        return
    if len(title) > 200:
        title = title[:200]
    await state.update_data(title=title)
    await state.set_state(NewTaskFSM.awaiting_priority)
    data = await state.get_data()
    await _safe_answer(
        message,
        f"{_newtask_summary(data)}\n" + _SEP + "\n\n"
        "2️⃣ **Ustuvorlik** tanlang:",
        parse_mode="Markdown",
        reply_markup=_newtask_priority_kb(),
    )


@router.callback_query(F.data.startswith("newtask:pri:"))
async def newtask_priority(query: CallbackQuery, state: FSMContext) -> None:
    pri = query.data.split(":", 2)[2]
    if pri not in {"P0", "P1", "P2", "P3"}:
        await query.answer("Noto'g'ri ustuvorlik", show_alert=True)
        return
    await state.update_data(priority=pri)
    await state.set_state(NewTaskFSM.awaiting_deadline)
    data = await state.get_data()
    await query.answer()
    text = (
        f"{_newtask_summary(data)}\n" + _SEP + "\n\n"
        "3️⃣ **Muddat** tanlang yoki o'tkazib yuboring:"
    )
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_newtask_deadline_kb())
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=_newtask_deadline_kb())


@router.message(StateFilter(NewTaskFSM.awaiting_priority), F.text | F.voice)
async def newtask_priority_typed(message: Message, state: FSMContext) -> None:
    """Voice/text parity for the priority step — saying «shoshilinch» (or typing
    it) lands the same P0-P3 outcome as tapping the inline button. Mirrors
    newtask_priority's state transition exactly."""
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return
    pri = _parse_priority_word(message.text or "")
    if not pri:
        await _safe_answer(
            message,
            "⚡ Ustuvorlikni ayting yoki tugmadan tanlang: "
            "«shoshilinch», «muhim», «rejadagi» yoki «oddiy».",
            parse_mode="Markdown", reply_markup=_newtask_priority_kb(),
        )
        return
    await state.update_data(priority=pri)
    await state.set_state(NewTaskFSM.awaiting_deadline)
    data = await state.get_data()
    await _safe_answer(
        message,
        f"{_newtask_summary(data)}\n" + _SEP + "\n\n"
        "3️⃣ **Muddat** tanlang yoki o'tkazib yuboring:",
        parse_mode="Markdown", reply_markup=_newtask_deadline_kb(),
    )


def _newtask_compute_day(key: str) -> tuple[str | None, str | None]:
    """Convert a day-preset key into (iso_date 'YYYY-MM-DD', human_label).
    Returns (None, None) for an unknown key."""
    now = datetime.now(database.TZ)
    if key == "today":
        target, label = now, "Bugun"
    elif key == "tomorrow":
        target, label = now + timedelta(days=1), "Ertaga"
    elif key == "indin":
        target, label = now + timedelta(days=2), "Indin"
    elif key == "plus3":
        target, label = now + timedelta(days=3), "+3 kun"
    elif key == "weekend":
        days_until_sat = (5 - now.weekday()) % 7 or 7
        target, label = now + timedelta(days=days_until_sat), "Shanba"
    else:
        return None, None
    return target.strftime("%Y-%m-%d"), label


def _combine_day_time(day_iso: str | None, hh: int, mm: int) -> str | None:
    """Combine a 'YYYY-MM-DD' day with HH:MM into a TZ-aware ISO timestamp.
    Returns None on malformed input or an impossible time (e.g. hour 25)."""
    if not day_iso:
        return None
    try:
        y, mo, d = map(int, day_iso.split("-"))
        return database.TZ.localize(datetime(y, mo, d, hh, mm)).isoformat()
    except (ValueError, TypeError):
        return None


async def _newtask_show_assignee(message: Message, data: dict, *, edit: bool) -> None:
    """Render the assignee step. edit=True edits the bot's message (callback
    path); edit=False sends a fresh message (user-typed path)."""
    text = (
        f"{_newtask_summary(data)}\n" + _SEP + "\n\n"
        "4️⃣ **Ijrochi** (ixtiyoriy):\n\n"
        "Pastdagi tugmalardan biri yoki ism yuboring "
        "(masalan: «Komilov Javohir»)."
    )
    if edit:
        try:
            await message.edit_text(text, parse_mode="Markdown",
                                    reply_markup=_newtask_assignee_kb())
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown",
                       reply_markup=_newtask_assignee_kb())


@router.callback_query(F.data.startswith("newtask:dl:"))
async def newtask_deadline(query: CallbackQuery, state: FSMContext) -> None:
    preset = query.data.split(":", 2)[2]
    if preset == "manual":
        await state.set_state(NewTaskFSM.awaiting_deadline_manual)
        await query.answer()
        text = (
            "📅 **Muddatni qo'lda kiriting**\n\n"
            "Formatlar:\n"
            "• `2026-05-25 14:30`\n"
            "• `25-05 14:30` (joriy yil)\n"
            "• `ertaga 09:00` / `bugun 17:00`\n"
            "• `juma 12:00` (yaqin payshanba/juma...)\n"
            "• `2 soat` / `15 daqiqa`"
        )
        try:
            await query.message.edit_text(
                text, parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="⬅️ Orqaga", callback_data="newtask:dl:back")],
                    [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
                ]),
            )
        except TelegramBadRequest:
            await _safe_answer(query.message, text, parse_mode="Markdown")
        return

    if preset == "back":
        # Return to step 1 — the day picker.
        await state.set_state(NewTaskFSM.awaiting_deadline)
        data = await state.get_data()
        await query.answer()
        try:
            await query.message.edit_text(
                f"{_newtask_summary(data)}\n" + _SEP + "\n\n3️⃣ **Muddat** — kunni tanlang:",
                parse_mode="Markdown", reply_markup=_newtask_deadline_kb(),
            )
        except TelegramBadRequest:
            pass
        return

    if preset == "skip":
        await state.update_data(deadline=None)
        await state.set_state(NewTaskFSM.awaiting_assignee)
        await query.answer()
        await _newtask_show_assignee(query.message, await state.get_data(), edit=True)
        return

    # A day was chosen → remember it and show step 2 (the time picker).
    day_iso, label = _newtask_compute_day(preset)
    if not day_iso:
        await query.answer("Noto'g'ri kun preseti", show_alert=True)
        return
    await state.update_data(_dl_day_iso=day_iso, _dl_day_label=label)
    await state.set_state(NewTaskFSM.awaiting_deadline_time)
    await query.answer()
    text = f"🕐 **{label}** — soatni tanlang yoki `HH:MM` yozing:"
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_newtask_time_kb())
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=_newtask_time_kb())


@router.callback_query(F.data.startswith("newtask:tm:"))
async def newtask_deadline_time(query: CallbackQuery, state: FSMContext) -> None:
    """Step 2: a time slot (or 'custom') was tapped. Combine with the day chosen
    in step 1 and advance to the assignee step."""
    choice = query.data.split(":", 2)[2]
    data = await state.get_data()

    if choice == "custom":
        await query.answer()
        label = data.get("_dl_day_label", "")
        text = f"⌨️ **{label}** uchun soatni yozing — masalan `15:30`:"
        try:
            await query.message.edit_text(text, parse_mode="Markdown",
                                           reply_markup=_newtask_time_kb())
        except TelegramBadRequest:
            await _safe_answer(query.message, text, parse_mode="Markdown")
        return

    try:  # choice is 'HHMM'
        hh, mm = int(choice[:2]), int(choice[2:])
    except (ValueError, IndexError):
        await query.answer("Noto'g'ri vaqt", show_alert=True)
        return
    deadline = _combine_day_time(data.get("_dl_day_iso"), hh, mm)
    if not deadline:
        await query.answer("Vaqtni hisoblab bo'lmadi", show_alert=True)
        return
    await state.update_data(deadline=deadline)
    await state.set_state(NewTaskFSM.awaiting_assignee)
    await query.answer()
    await _newtask_show_assignee(query.message, await state.get_data(), edit=True)


@router.message(StateFilter(NewTaskFSM.awaiting_deadline_time), F.text | F.voice)
async def newtask_deadline_time_text(message: Message, state: FSMContext) -> None:
    """User typed instead of tapping a time slot. Accept a bare 'HH:MM' (combined
    with the day chosen in step 1) or a full natural-language date as fallback."""
    txt = await _get_text_or_transcribe(message, bot=message.bot)
    if txt is None:
        return
    raw = (message.text or "").strip()
    data = await state.get_data()

    # 1) Bare time: "15:30", "15.30", "15 30", or "9" → combine with chosen day.
    import re as _re
    deadline = None
    m = _re.match(r"^(\d{1,2})(?:[:.\s](\d{2}))?$", raw)
    if m:
        hh, mm = int(m.group(1)), int(m.group(2) or 0)
        deadline = _combine_day_time(data.get("_dl_day_iso"), hh, mm)
    # 2) Otherwise try a full natural-language date/time.
    if not deadline:
        parsed, _reason = await _parse_deadline_natural(raw)
        deadline = parsed
    if not deadline:
        await _safe_answer(
            message,
            "🕐 Soatni `HH:MM` ko'rinishida yuboring (masalan `15:30`), "
            "yoki tugmalardan tanlang.",
            parse_mode="Markdown", reply_markup=_newtask_time_kb(),
        )
        return
    await state.update_data(deadline=deadline)
    await state.set_state(NewTaskFSM.awaiting_assignee)
    await _newtask_show_assignee(message, await state.get_data(), edit=False)


@router.message(StateFilter(NewTaskFSM.awaiting_deadline), F.text | F.voice)
async def newtask_deadline_typed(message: Message, state: FSMContext) -> None:
    """At the day picker the user can also just TYPE a full date/time instead of
    tapping a day — parse it directly and jump to the assignee step."""
    txt = await _get_text_or_transcribe(message, bot=message.bot)
    if txt is None:
        return
    parsed, reason = await _parse_deadline_natural((message.text or "").strip())
    if not parsed:
        await _safe_answer(
            message,
            _deadline_error_message(reason, kind="deadline") + "\n\nYoki pastdagi tugmalardan tanlang.",
            parse_mode="Markdown", reply_markup=_newtask_deadline_kb(),
        )
        return
    await state.update_data(deadline=parsed)
    await state.set_state(NewTaskFSM.awaiting_assignee)
    await _newtask_show_assignee(message, await state.get_data(), edit=False)


@router.message(StateFilter(NewTaskFSM.awaiting_deadline_manual), F.text | F.voice)
async def newtask_deadline_manual(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    deadline_iso, reason = await _parse_deadline_natural((message.text or "").strip())
    if not deadline_iso:
        await _safe_answer(
            message,
            _deadline_error_message(reason, kind="deadline") + "\n\nYoki pastdagi presetlardan tanlang.",
            parse_mode="Markdown",
            reply_markup=_newtask_deadline_kb(),
        )
        await state.set_state(NewTaskFSM.awaiting_deadline)
        return
    await state.update_data(deadline=deadline_iso)
    await state.set_state(NewTaskFSM.awaiting_assignee)
    data = await state.get_data()
    await _safe_answer(
        message,
        f"{_newtask_summary(data)}\n" + _SEP + "\n\n"
        "4️⃣ **Ijrochi** (ixtiyoriy):\n\n"
        "Pastdagi tugmalardan biri yoki ism yuboring.",
        parse_mode="Markdown",
        reply_markup=_newtask_assignee_kb(),
    )


@router.callback_query(F.data.startswith("newtask:as:"))
async def newtask_assignee_btn(query: CallbackQuery, state: FSMContext) -> None:
    choice = query.data.split(":", 2)[2]
    if choice == "me":
        await state.update_data(assignee="Men")
    elif choice == "skip":
        await state.update_data(assignee=None)
    await _newtask_show_confirm(query.message, state, edit=True)
    await query.answer()


@router.message(StateFilter(NewTaskFSM.awaiting_assignee), F.text | F.voice)
async def newtask_assignee_text(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    name = (message.text or "").strip()
    if not name or name.startswith("/"):
        await message.answer("Ism bo'sh bo'lmasin. Qaytadan yuboring yoki tugma bosing.",
                              reply_markup=_newtask_assignee_kb())
        return
    if len(name) > 80:
        name = name[:80]
    await state.update_data(assignee=name)
    await _newtask_show_confirm(message, state, edit=False)


async def _newtask_show_confirm(message: Message, state: FSMContext, edit: bool) -> None:
    await state.set_state(NewTaskFSM.awaiting_confirm)
    data = await state.get_data()
    text = (
        "5️⃣ **TASDIQLASH**\n" + _SEP + "\n\n"
        f"{_newtask_summary(data)}\n\n"
        "_Pastdagi tugma orqali yarating yoki bekor qiling._"
    )
    if edit:
        try:
            await message.edit_text(text, parse_mode="Markdown",
                                     reply_markup=_newtask_confirm_kb())
            return
        except TelegramBadRequest:
            pass
    await _safe_answer(message, text, parse_mode="Markdown",
                        reply_markup=_newtask_confirm_kb())


@router.callback_query(F.data == "newtask:confirm")
async def newtask_confirm(query: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    title = data.get("title")
    if not title:
        await query.answer("Sarlavha kiritilmagan", show_alert=True)
        return
    payload = {
        "title": title,
        "priority": data.get("priority") or "P2",
        "status": "todo",
    }
    if data.get("deadline"):
        payload["deadline"] = data["deadline"]
    if data.get("assignee"):
        payload["assignee"] = data["assignee"]

    # Past-deadline guard: warn once and require an explicit "create anyway" tap.
    # Overdue items are legitimate (there's an "O'tgan" view), so we warn, not block.
    if payload.get("deadline") and database.is_past_deadline(payload["deadline"]) \
            and not data.get("_confirm_past"):
        await state.update_data(_confirm_past=True)
        await query.answer()
        dl_label, _ = _format_deadline_short(payload["deadline"])
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Baribir yaratish", callback_data="newtask:confirm")],
            [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newtask:cancel")],
        ])
        try:
            await query.message.edit_text(
                f"⚠️ **Muddat o'tib ketgan** ({dl_label})\n\n"
                "Bu vazifaning muddati o'tmishda. Baribir yaratilsinmi?",
                parse_mode="Markdown", reply_markup=kb)
        except TelegramBadRequest:
            await _safe_answer(
                query.message,
                f"⚠️ Muddat o'tib ketgan ({dl_label}). Baribir yaratilsinmi?",
                reply_markup=kb)
        return

    tid = await database.create_task(payload)
    await state.clear()
    await query.answer("✅ Vazifa yaratildi")

    task = await database.get_task(tid)
    if not task:
        await _safe_answer(query.message, "✅ Vazifa yaratildi, lekin ko'rsatib bo'lmadi.",
                            reply_markup=main_reply_keyboard())
        return
    text = "✅ **VAZIFA YARATILDI**\n" + _SEP + "\n\n" + _format_task_card(task)
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_task_card_kb_with_back(task))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=_task_card_kb_with_back(task))


@router.callback_query(F.data == "newtask:cancel")
async def newtask_cancel(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer("Bekor qilindi")
    # edit_text faqat inline kbd qabul qiladi — ReplyKeyboardMarkup'ni alohida
    # `_restore_main_keyboard` orqali jo'natamiz.
    try:
        await query.message.edit_text("↩ Yangi vazifa yaratish bekor qilindi.")
    except TelegramBadRequest:
        pass
    await _restore_main_keyboard(query.message)


# ─────────────────────── YANGI ESLATMA — guided FSM ───────────────────────

_NEWREMINDER_REPEAT_LABELS = {
    "once": "bir martalik",
    "daily": "har kuni",
    "weekly": "har hafta",
    "monthly": "har oy",
}


def _newreminder_time_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⏰ 15 daq", callback_data="newrem:time:15m"),
            InlineKeyboardButton(text="🕐 1 soat", callback_data="newrem:time:1h"),
        ],
        [
            InlineKeyboardButton(text="📅 Bugun 17:00", callback_data="newrem:time:today17"),
            InlineKeyboardButton(text="📅 Ertaga 09:00", callback_data="newrem:time:tomorrow9"),
        ],
        [
            InlineKeyboardButton(text="✏️ Qo'lda kiritish", callback_data="newrem:time:manual"),
            InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newrem:cancel"),
        ],
    ])


def _newreminder_repeat_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Bir martalik", callback_data="newrem:repeat:once"),
            InlineKeyboardButton(text="Har kuni", callback_data="newrem:repeat:daily"),
        ],
        [
            InlineKeyboardButton(text="Har hafta", callback_data="newrem:repeat:weekly"),
            InlineKeyboardButton(text="Har oy", callback_data="newrem:repeat:monthly"),
        ],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newrem:cancel")],
    ])


def _newreminder_confirm_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Saqlash", callback_data="newrem:confirm"),
            InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newrem:cancel"),
        ],
    ])


def _newreminder_summary(data: dict) -> str:
    title = data.get("title") or "_(kiritilmagan)_"
    remind_at = data.get("remind_at")
    time_label = _reminder_time_chip({"remind_at": remind_at}) if remind_at else "_(tanlanmagan)_"
    repeat = _NEWREMINDER_REPEAT_LABELS.get(data.get("recurrence_rule") or "once", "bir martalik")
    return "\n".join([
        f"📝 Matn: {title}",
        f"⏰ Vaqt: {time_label}",
        f"🔁 Takror: {repeat}",
    ])


def _newreminder_compute_time(preset: str) -> str | None:
    now = datetime.now(database.TZ)
    if preset == "15m":
        return (now + timedelta(minutes=15)).replace(second=0, microsecond=0).isoformat()
    if preset == "1h":
        return (now + timedelta(hours=1)).replace(second=0, microsecond=0).isoformat()
    if preset == "today17":
        target = now.replace(hour=17, minute=0, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        return target.isoformat()
    if preset == "tomorrow9":
        return (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0).isoformat()
    return None


async def _newreminder_start(message: Message, state: FSMContext) -> None:
    await state.set_state(NewReminderFSM.awaiting_title)
    await state.update_data(title=None, remind_at=None, recurrence_rule=None)
    await _safe_answer(
        message,
        "⏰ **YANGI ESLATMA**\n" + _SEP + "\n\n"
        "1️⃣ Nimani eslatay?\n"
        "_Masalan: «Aziz akaga qo'ng'iroq qilish»_",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newrem:cancel")],
        ]),
    )


@router.message(StateFilter(NewReminderFSM.awaiting_title), F.text | F.voice)
async def newreminder_title(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    title = (message.text or "").strip()
    if not title or title.startswith("/"):
        await message.answer("Eslatma matni bo'sh bo'lmasin. Qaytadan yuboring.")
        return
    if len(title) > 220:
        title = title[:220]
    await state.update_data(title=title)
    await state.set_state(NewReminderFSM.awaiting_time)
    data = await state.get_data()
    await _safe_answer(
        message,
        f"{_newreminder_summary(data)}\n" + _SEP + "\n\n"
        "2️⃣ Qachon eslatay?",
        parse_mode="Markdown",
        reply_markup=_newreminder_time_kb(),
    )


@router.callback_query(F.data.startswith("newrem:time:"))
async def newreminder_time(query: CallbackQuery, state: FSMContext) -> None:
    preset = query.data.split(":", 2)[2]
    if preset == "manual":
        await state.set_state(NewReminderFSM.awaiting_time_manual)
        await query.answer()
        text = (
            "📆 **Eslatma vaqtini kiriting**\n\n"
            "Formatlar:\n"
            "• `2026-05-25 14:30`\n"
            "• `25-05 14:30`\n"
            "• `bugun 17:00`, `ertaga 09:00`\n"
            "• `15 daqiqa`, `2 soat`"
        )
        try:
            await query.message.edit_text(
                text, parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="⬅️ Orqaga", callback_data="newrem:time:back")],
                    [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="newrem:cancel")],
                ]),
            )
        except TelegramBadRequest:
            await _safe_answer(query.message, text, parse_mode="Markdown")
        return
    if preset == "back":
        await state.set_state(NewReminderFSM.awaiting_time)
        data = await state.get_data()
        await query.answer()
        try:
            await query.message.edit_text(
                f"{_newreminder_summary(data)}\n" + _SEP + "\n\n2️⃣ Qachon eslatay?",
                parse_mode="Markdown",
                reply_markup=_newreminder_time_kb(),
            )
        except TelegramBadRequest:
            pass
        return
    remind_at = _newreminder_compute_time(preset)
    if not remind_at:
        await query.answer("Vaqt preseti noto'g'ri", show_alert=True)
        return
    await state.update_data(remind_at=remind_at)
    await state.set_state(NewReminderFSM.awaiting_repeat)
    data = await state.get_data()
    await query.answer()
    try:
        await query.message.edit_text(
            f"{_newreminder_summary(data)}\n" + _SEP + "\n\n3️⃣ Takrorlansinmi?",
            parse_mode="Markdown",
            reply_markup=_newreminder_repeat_kb(),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, f"{_newreminder_summary(data)}\n\n3️⃣ Takrorlansinmi?",
                           parse_mode="Markdown", reply_markup=_newreminder_repeat_kb())


@router.message(StateFilter(NewReminderFSM.awaiting_time_manual), F.text | F.voice)
async def newreminder_time_manual(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    remind_at, reason = await _parse_deadline_natural((message.text or "").strip())
    if not remind_at:
        await _safe_answer(
            message,
            _deadline_error_message(reason, kind="time"),
            parse_mode="Markdown",
            reply_markup=_newreminder_time_kb(),
        )
        await state.set_state(NewReminderFSM.awaiting_time)
        return
    await state.update_data(remind_at=remind_at)
    await state.set_state(NewReminderFSM.awaiting_repeat)
    data = await state.get_data()
    await _safe_answer(
        message,
        f"{_newreminder_summary(data)}\n" + _SEP + "\n\n3️⃣ Takrorlansinmi?",
        parse_mode="Markdown",
        reply_markup=_newreminder_repeat_kb(),
    )


@router.callback_query(F.data.startswith("newrem:repeat:"))
async def newreminder_repeat(query: CallbackQuery, state: FSMContext) -> None:
    repeat = query.data.split(":", 2)[2]
    if repeat not in _NEWREMINDER_REPEAT_LABELS:
        await query.answer("Takrorlash turi noto'g'ri", show_alert=True)
        return
    await state.update_data(recurrence_rule=None if repeat == "once" else repeat)
    await state.set_state(NewReminderFSM.awaiting_confirm)
    data = await state.get_data()
    await query.answer()
    try:
        await query.message.edit_text(
            "4️⃣ **TASDIQLASH**\n" + _SEP + "\n\n"
            f"{_newreminder_summary(data)}",
            parse_mode="Markdown",
            reply_markup=_newreminder_confirm_kb(),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, _newreminder_summary(data), parse_mode="Markdown",
                           reply_markup=_newreminder_confirm_kb())


@router.callback_query(F.data == "newrem:confirm")
async def newreminder_confirm(query: CallbackQuery, state: FSMContext) -> None:
    data = await state.get_data()
    if not data.get("title") or not data.get("remind_at"):
        await query.answer("Ma'lumot to'liq emas", show_alert=True)
        return
    rid = await database.create_reminder({
        "title": data["title"],
        "remind_at": data["remind_at"],
        "recurrence_rule": data.get("recurrence_rule"),
        "source": "manual_form",
    })
    # If this reminder was started from a note (📝 Qaydlar → Eslatmaga), mark the
    # note processed so it leaves the inbox. Without this the note lingered as
    # unprocessed even after a successful conversion.
    from_note = data.get("from_note_id")
    if from_note:
        await database.mark_note_processed(from_note, "reminder", rid)
    await state.clear()
    reminder = await database.get_reminder(rid)
    await query.answer("⏰ Eslatma saqlandi")
    text = "✅ **ESLATMA SAQLANDI**\n" + _SEP + "\n\n" + _format_reminder_card(reminder)
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=reminder_detail_menu(reminder))


@router.callback_query(F.data == "newrem:cancel")
async def newreminder_cancel(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer("Bekor qilindi")
    try:
        await query.message.edit_text("↩ Yangi eslatma yaratish bekor qilindi.")
    except TelegramBadRequest:
        await _safe_answer(query.message, "↩ Bekor qilindi.")


_PRINCIPAL_NAME_SHORT = "Maqsud aka"

_UZ_WEEKDAYS = [
    "Dushanba", "Seshanba", "Chorshanba", "Payshanba",
    "Juma", "Shanba", "Yakshanba",
]


def _greeting_for_hour(hour: int) -> str:
    """Kontekstga qarab dinamik salomlashish matni."""
    if 6 <= hour < 11:
        return f"Xayrli tong, {_PRINCIPAL_NAME_SHORT}. Ish kuni boshlanmoqda."
    if 11 <= hour < 16:
        return "Xayrli kun. Kunning yarmi yopildi."
    if 16 <= hour < 20:
        return "Xayrli kech. Kun yakunlanmoqda."
    return "Kech vaqt. Ertangi rejani ko'rib chiqing."


def _date_label(now: datetime) -> str:
    """`Dushanba, 25-may · 14:23`"""
    weekday = _UZ_WEEKDAYS[now.weekday()]
    month = UZ_MONTHS_FULL[now.month - 1]
    return f"{weekday}, {now.day}-{month} · {now.strftime('%H:%M')}"


def _pick_top_task(active: list[dict], overdue: list[dict], unassigned: list[dict],
                    today: list[dict]) -> dict | None:
    """ENG MUHIM tanlash mantig'i (spec'dagi tartibda)."""
    # 1) P0 ochiq + o'tgan
    p0_overdue = [t for t in overdue if t.get("priority") == "P0"]
    if p0_overdue:
        return p0_overdue[0]
    # 2) P0 ochiq
    p0_open = sorted(
        [t for t in active if t.get("priority") == "P0"],
        key=lambda t: t.get("deadline") or "9999",
    )
    if p0_open:
        return p0_open[0]
    # 3) P1 ochiq + bugun deadline
    today_ids = {t["id"] for t in today}
    p1_today = [t for t in active if t.get("priority") == "P1" and t["id"] in today_ids]
    if p1_today:
        return sorted(p1_today, key=lambda t: t.get("deadline") or "9999")[0]
    # 4) Ijrochisiz P0/P1
    p01_unassigned = [t for t in unassigned if t.get("priority") in ("P0", "P1")]
    if p01_unassigned:
        return p01_unassigned[0]
    # 5) yo'q
    return None


def _top_task_status(task: dict, now: datetime) -> str:
    """Top task uchun `Shoshilinch / O'tgan / Bugun` label."""
    parts: list[str] = []
    priority = task.get("priority", "P2")
    if priority == "P0":
        parts.append("Shoshilinch")
    elif priority == "P1":
        parts.append("Muhim")
    deadline = task.get("deadline")
    if deadline:
        try:
            dt = datetime.fromisoformat(deadline).astimezone(database.TZ)
            if dt < now:
                hours = int((now - dt).total_seconds() / 3600)
                parts.append(f"muddati o'tgan ({hours} soat)")
            elif dt.date() == now.date():
                parts.append("bugun")
        except (ValueError, TypeError):
            pass
    return ", ".join(parts) or "—"


def _format_top_task_deadline(deadline_iso: str | None, now: datetime) -> str:
    """ENG MUHIM bloki uchun deadline label: aniq vaqt + farq."""
    if not deadline_iso:
        return "belgilanmagan"
    try:
        dt = datetime.fromisoformat(deadline_iso).astimezone(database.TZ)
    except (ValueError, TypeError):
        return str(deadline_iso)
    if dt < now:
        delta = now - dt
        if delta.total_seconds() < 24 * 3600:
            hours = int(delta.total_seconds() / 3600)
            label = f"{hours} soat kechikkan"
        else:
            days = delta.days
            label = f"{days} kun kechikkan"
        return f"{dt.strftime('%d-%m %H:%M')} · {label}"
    if dt.date() == now.date():
        return f"Bugun {dt.strftime('%H:%M')}"
    if (now + timedelta(days=1)).date() == dt.date():
        return f"Ertaga {dt.strftime('%H:%M')}"
    return dt.strftime("%d-%m %H:%M")


async def _cockpit_anomalies(now: datetime, loads: dict) -> list[str]:
    """Anomaliya signali — odatdagidan og'ish bo'lgan pattern'larni topish.

    Topiladigan signallar:
    1. Aktiv vazifasi bor ijrochi 3+ kun hech narsa yopmagan
    2. So'nggi haftada Shoshilinch vazifalar avvalgi haftadan 2x+ ko'p
    3. Ijrochisiz vazifalar chegaradan oshib ketdi
    """
    import aiosqlite
    anomalies: list[str] = []
    cutoff_3d = (now - timedelta(days=3)).isoformat()
    cutoff_7d = (now - timedelta(days=7)).isoformat()
    cutoff_14d = (now - timedelta(days=14)).isoformat()

    async with aiosqlite.connect(config.DATABASE_PATH) as db:
        db.row_factory = aiosqlite.Row

        # 1. Ijrochilar bo'sh turishi (active va done = 0)
        for name, info in loads.items():
            if name == "belgilanmagan" or info.get("active", 0) < 2:
                continue
            cur = await db.execute(
                "SELECT COUNT(*) AS n FROM tasks WHERE assignee = ? AND status = 'done' AND updated_at >= ?",
                (name, cutoff_3d),
            )
            closed = (await cur.fetchone())["n"]
            if closed == 0:
                anomalies.append(
                    f"**{name}** so'nggi 3 kunda 0 ta vazifa yopdi ({info['active']} ta aktiv)"
                )

        # 2. P0 vazifalar spike
        cur = await db.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE priority = 'P0' AND created_at >= ?",
            (cutoff_7d,),
        )
        p0_recent = (await cur.fetchone())["n"]
        cur = await db.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE priority = 'P0' AND created_at >= ? AND created_at < ?",
            (cutoff_14d, cutoff_7d),
        )
        p0_prior = (await cur.fetchone())["n"]
        if p0_recent >= 3 and p0_recent >= 2 * max(1, p0_prior):
            ratio = round(p0_recent / max(1, p0_prior), 1)
            anomalies.append(
                f"So'nggi haftada **{p0_recent} ta yangi Shoshilinch** vazifa — avvalgi haftadan {ratio}x ko'p"
            )

        # 3. Eskalatsiya: 3+ overdue va 3+ P0 — eskirayotgan vaziyat
        cur = await db.execute(
            "SELECT COUNT(*) AS n FROM tasks WHERE status IN ('todo','in_progress') "
            "AND priority = 'P0' AND deadline IS NOT NULL AND deadline < ?",
            (now.isoformat(),),
        )
        p0_overdue = (await cur.fetchone())["n"]
        if p0_overdue >= 2:
            anomalies.append(
                f"**{p0_overdue} ta Shoshilinch vazifa muddati o'tgan** — kritik eskalatsiya"
            )

    return anomalies[:4]  # max 4 signal


def _cockpit_time_budget(today_meetings: list, today_tasks: list, now: datetime) -> dict:
    """Bugungi vaqt taqsimoti — uchrashuvlar, deadline'lar, zich oraliqlar."""
    # Uchrashuv soatlari
    meeting_hours = 0.0
    for m in today_meetings:
        try:
            start = datetime.fromisoformat(m["datetime_start"])
            end_iso = m.get("datetime_end")
            if end_iso:
                end = datetime.fromisoformat(end_iso)
            else:
                end = start + timedelta(hours=1)  # default 1 soat
            meeting_hours += max(0, (end - start).total_seconds() / 3600)
        except (ValueError, TypeError):
            meeting_hours += 1.0

    # Deadline'larni soat bo'yicha tasniflash (bugun yopilishi kerak bo'lganlar)
    deadline_buckets: dict[int, int] = {}
    for t in today_tasks:
        deadline = t.get("deadline")
        if not deadline:
            continue
        try:
            dt = datetime.fromisoformat(deadline).astimezone(database.TZ)
            if dt.date() == now.date():
                deadline_buckets[dt.hour] = deadline_buckets.get(dt.hour, 0) + 1
        except (ValueError, TypeError):
            pass

    # Eng zich oraliq (3 soatlik window)
    peak_window = None
    peak_count = 0
    if deadline_buckets:
        sorted_hours = sorted(deadline_buckets.keys())
        for h in sorted_hours:
            window_count = sum(deadline_buckets.get(h + i, 0) for i in range(3))
            if window_count > peak_count:
                peak_count = window_count
                peak_window = (h, h + 2)

    return {
        "meeting_hours": round(meeting_hours, 1),
        "meeting_count": len(today_meetings),
        "deadline_total": sum(deadline_buckets.values()),
        "peak_window": peak_window,
        "peak_count": peak_count,
    }


def _cockpit_recommendation(signals: dict) -> str:
    """Spec asosida tavsiya — prioritet bo'yicha qoidalar.
    Risk skor olib tashlanganligi sababli endi to'g'ridan-to'g'ri counts asosida ishlaydi.
    """
    overdue_n = signals["counts"]["overdue"]
    unassigned_n = signals["counts"]["unassigned"]
    p0_open = signals["by_priority"].get("P0", 0)
    no_dl_n = signals["counts"]["no_deadline"]
    top_loaded = signals.get("top_loaded")
    next_title = signals.get("next_task_title")
    overload_threshold = 4

    # Kritik holat: ko'p o'tgan + ko'p shoshilinch — vaziyat boshqaruvdan chiqqan
    if overdue_n >= 5 and p0_open >= 2:
        return ("Vaziyat kritik. Avval **muddati o'tgan** vazifalarni yoping yoki "
                "yangi deadline qo'ying. Keyin Shoshilinch ro'yxatini ko'ring.")
    if overdue_n >= 3:
        first_title = signals.get("oldest_overdue_title") or "vazifa"
        return (f"**{overdue_n} ta vazifa muddati o'tgan**. Eng eskisini "
                f"(«{first_title}») bugun yoping yoki delegatsiya qiling.")
    if p0_open >= 3:
        return (f"Bir vaqtning o'zida **{p0_open} ta Shoshilinch** vazifa ochiq. "
                "Haqiqatdan hammasi shoshilinchmi? Ba'zilarini Muhim ga ko'chiring.")
    if unassigned_n >= 5:
        return (f"**{unassigned_n} ta vazifaga ijrochi tayinlanmagan**. Avval "
                "mas'ul belgilang — javobgarliksiz vazifa kechikadi.")
    if no_dl_n >= 5:
        return (f"**{no_dl_n} ta vazifada deadline yo'q**. Muddatsiz vazifalar "
                "amalga oshmaydi. Har biriga sana belgilang.")
    if top_loaded and top_loaded.get("active", 0) >= overload_threshold:
        return (f"**{top_loaded['name']}** ortiqcha yuklangan ({top_loaded['active']} ta). "
                "Yangi vazifa berishni vaqtincha to'xtating yoki muddatlarni cho'zing.")
    if next_title:
        return f"Vaziyat nazoratda. Bugun **{_truncate(next_title, 50)}** bilan boshlang."
    return "Kun rejasini ko'rib chiqing va prioritetlarni belgilang."


async def _cockpit_compute_signals(now: datetime) -> dict:
    """Barcha DB chaqiruvlarni parallel bajarib, signallarni jamlaydi."""
    # Parallel
    (active, overdue, today, today_meetings, unassigned, no_deadline,
     risk, loads, upcoming) = await asyncio.gather(
        database.list_tasks(status_in=["todo", "in_progress"], limit=500),
        database.list_overdue_tasks(),
        database.list_today_tasks(),
        database.list_today_meetings(),
        database.list_unassigned_tasks(limit=200),
        database.list_tasks_without_deadline(limit=200),
        compute_risk_score(),
        database.assignee_load_map(),
        database.list_upcoming_meetings(within_minutes=1440 * 2),  # 2 kun
    )
    # Anomaliya va vaqt bo'limi — loads va today ma'lumotlariga bog'liq, ketma-ket
    anomalies = await _cockpit_anomalies(now, loads)
    time_budget = _cockpit_time_budget(today_meetings, today, now)
    by_p = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for t in active:
        p = t.get("priority", "P2")
        if p in by_p:
            by_p[p] += 1

    # 24h/48h deadline
    due_24 = [t for t in active if (t.get("deadline") or "") and
              t["deadline"] <= (now + timedelta(hours=24)).isoformat() and
              t["deadline"] >= now.isoformat()]
    due_48 = [t for t in active if (t.get("deadline") or "") and
              t["deadline"] <= (now + timedelta(hours=48)).isoformat() and
              t["deadline"] >= now.isoformat()]

    # Team load — eng yuklangan va eng bo'sh
    real_assignees = [d for n, d in loads.items() if n != "belgilanmagan" and d["active"] > 0]
    top_loaded = max(real_assignees, key=lambda d: d["active"], default=None)
    lightest = min(real_assignees, key=lambda d: d["active"], default=None) if real_assignees else None

    top_task = _pick_top_task(active, overdue, unassigned, today)

    # Next task (top_task'dan boshqa — recommendation uchun)
    next_task = None
    for t in sorted(active, key=lambda t: (
        {"P0": 0, "P1": 1, "P2": 2, "P3": 3}.get(t.get("priority", "P2"), 2),
        t.get("deadline") or "9999",
    )):
        if top_task and t["id"] == top_task["id"]:
            continue
        next_task = t
        break

    # Eng yaqin uchrashuv (bugun yoki ertaga)
    next_meeting = None
    for m in upcoming:
        try:
            dt = datetime.fromisoformat(m["datetime_start"]).astimezone(database.TZ)
            if dt >= now and (dt - now).total_seconds() <= 48 * 3600:
                next_meeting = m
                break
        except (ValueError, TypeError):
            continue

    return {
        "now": now,
        "risk": risk,
        "by_priority": by_p,
        "counts": {
            "active": len(active),
            "today": len(today),
            "overdue": len(overdue),
            "unassigned": len(unassigned),
            "no_deadline": len(no_deadline),
            "due_24h": len(due_24),
            "due_48h": len(due_48),
        },
        "top_task": top_task,
        "top_loaded": top_loaded,
        "lightest": lightest,
        "next_meeting": next_meeting,
        "next_task_title": (next_task.get("title") if next_task else None) or
                            (top_task.get("title") if top_task else None),
        "oldest_overdue_title": (overdue[0].get("title") if overdue else None),
        "anomalies": anomalies,
        "time_budget": time_budget,
        "overload_threshold": _ASSIGNEE_OVERLOAD,
    }


def _build_cockpit_panel(signals: dict) -> str:
    """Cockpit matnini blok-stilda yaratish."""
    DIVIDER = "━" * 20
    now = signals["now"]

    # ── Block 1: Sarlavha ──
    blocks: list[list[str]] = [[
        "🎛 **BOSHQARUV PANELI**",
        "",
        _greeting_for_hour(now.hour),
        _date_label(now),
    ]]

    # Empty state — 0 aktiv vazifa + 0 uchrashuv
    if signals["counts"]["active"] == 0 and not signals["next_meeting"]:
        blocks.append([
            "✨ **Vaziyat tinch**",
            "",
            "Hozir aktiv vazifa va uchrashuv yo'q.",
        ])
        blocks.append([
            "💡 **TAVSIYA**",
            "",
            "Bugun keyingi haftani rejalashtirish uchun yaxshi vaqt. Yangi "
            "maqsadlar belgilang yoki uzoq muddatli loyihaga vaqt ajrating.",
        ])
        return _join_blocks(blocks, DIVIDER)

    # ── Block 2: Bugungi holat (Risk skor olib tashlandi — foydalanuvchi so'rovi) ──
    c = signals["counts"]
    by_p = signals["by_priority"]
    overdue_prefix = "🚨 " if c["overdue"] > 0 else ""
    unassigned_prefix = "⚠️ " if c["unassigned"] > 0 else ""
    blocks.append([
        "📌 **BUGUNGI HOLAT**",
        "",
        f"Aktiv vazifalar: **{c['active']}** ta (🔴 {by_p.get('P0', 0)} shoshilinch · 🟠 {by_p.get('P1', 0)} muhim)",
        f"Bugun yopilishi kerak: **{c['today']}** ta",
        f"{overdue_prefix}Muddati o'tgan: **{c['overdue']}** ta",
        f"{unassigned_prefix}Ijrochisiz: **{c['unassigned']}** ta",
        "",
        f"24 soat ichida deadline: {c['due_24h']} ta",
        f"48 soat ichida deadline: {c['due_48h']} ta",
    ])

    # ── Block 4: Eng muhim (top_task) ──
    top = signals.get("top_task")
    if top:
        title = (top.get("title") or "—").strip()
        assignee = (top.get("assignee") or "").strip() or "belgilanmagan"
        deadline_label = _format_top_task_deadline(top.get("deadline"), now)
        status_label = _top_task_status(top, now)
        blocks.append([
            "⚡ **ENG MUHIM**",
            "",
            title,
            "",
            f"      ⏰ Muddat:        {deadline_label}",
            f"      👤 Ijrochi:        {assignee}",
            f"      ⚡ Holat:          {status_label}",
        ])

    # ── Block 5: Jamoa yuklamasi ──
    top_loaded = signals.get("top_loaded")
    lightest = signals.get("lightest")
    overload = signals["overload_threshold"]
    if top_loaded or lightest or c["unassigned"]:
        team_lines = ["👥 **JAMOA YUKLAMASI**", ""]
        if top_loaded:
            overload_mark = " 🔴 ortiqcha" if top_loaded["active"] >= overload else ""
            team_lines.append(f"Eng yuklangan: **{top_loaded['name']}** — {top_loaded['active']} ta aktiv{overload_mark}")
        if lightest and (not top_loaded or lightest["name"] != top_loaded["name"]):
            team_lines.append(f"Eng bo'sh: **{lightest['name']}** — {lightest['active']} ta aktiv")
        if c["unassigned"]:
            team_lines.append(f"Belgilanmagan: **{c['unassigned']}** ta vazifa")
        blocks.append(team_lines)

    # ── Block 6: Keyingi uchrashuv ──
    nm = signals.get("next_meeting")
    if nm:
        try:
            dt = datetime.fromisoformat(nm["datetime_start"]).astimezone(database.TZ)
            if dt.date() == now.date():
                time_label = f"Bugun {dt.strftime('%H:%M')}"
            elif (now + timedelta(days=1)).date() == dt.date():
                time_label = f"Ertaga {dt.strftime('%H:%M')}"
            else:
                time_label = dt.strftime("%d-%m %H:%M")
        except (ValueError, TypeError):
            time_label = "—"
        title = (nm.get("title") or "—").strip()
        parts = nm.get("participants") or []
        plabel = ", ".join(parts[:3]) if parts else "belgilanmagan"
        if len(parts) > 3:
            plabel += f" (+{len(parts) - 3} nafar)"
        blocks.append([
            "📆 **KEYINGI UCHRASHUV**",
            "",
            f"{time_label} — {title}",
            f"👥 {plabel}",
        ])

    # ── Block 7: AI Tavsiya ──
    blocks.append([
        "💡 **TIZIM TAVSIYASI**",
        "",
        _cockpit_recommendation(signals),
    ])

    # ── Block 8: 🚨 ANOMALIYA SIGNALI ──
    anomalies = signals.get("anomalies") or []
    if anomalies:
        anomaly_lines = ["🚨 **ANOMALIYA**", ""]
        for a in anomalies:
            anomaly_lines.append(f"• {a}")
        blocks.append(anomaly_lines)

    # ── Block 9: ⏳ VAQT BO'LIMI ──
    tb = signals.get("time_budget") or {}
    if tb.get("meeting_count", 0) > 0 or tb.get("deadline_total", 0) > 0:
        tb_lines = ["⏳ **BUGUNGI VAQT**", ""]
        if tb.get("meeting_count", 0) > 0:
            tb_lines.append(
                f"      🤝 Uchrashuv:         {tb['meeting_hours']} soat ({tb['meeting_count']} ta)"
            )
        if tb.get("deadline_total", 0) > 0:
            tb_lines.append(
                f"      📌 Vazifa deadline:   {tb['deadline_total']} ta"
            )
        if tb.get("peak_window") and tb.get("peak_count", 0) >= 2:
            start_h, end_h = tb["peak_window"]
            tb_lines.append("")
            tb_lines.append(
                f"      🚨 Zich oraliq:       "
                f"{start_h:02d}:00–{end_h+1:02d}:00 ({tb['peak_count']} ta deadline)"
            )
        blocks.append(tb_lines)

    return _join_blocks(blocks, DIVIDER)


def _join_blocks(blocks: list[list[str]], divider: str) -> str:
    """Block ro'yxatini divider + bo'sh qator bilan birlashtirish."""
    result: list[str] = []
    for i, block in enumerate(blocks):
        if i > 0:
            result.extend(["", divider, ""])
        result.extend(block)
    return "\n".join(result)


@router.message(Command("cockpit"))
async def cmd_cockpit(message: Message, state: FSMContext | None = None) -> None:
    """🎛 Boshqaruv Paneli — professional operational dashboard.

    Spec-based 8-block layout: sarlavha + risk + bugungi holat + eng muhim
    + jamoa + keyingi uchrashuv + AI tavsiya + 24h delta.
    """
    now = datetime.now(database.TZ)
    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        signals = await _cockpit_compute_signals(now)
        text = _build_cockpit_panel(signals)
        # B5: qayta ishlanmagan qaydlar sonini yuzaga chiqaramiz (GTD turtki)
        inbox_n = await database.count_notes_in_status("inbox")
    finally:
        typing_task.cancel()

    # Inline kbd with drill-down rows so users can jump straight from the
    # cockpit into the underlying panels (team, risks, stats, qaydlar) without
    # leaving and finding the section buttons in the main keyboard. (Delegatsiya
    # trekeri endi Ijrochilar paneli ichida — "⏳ Kutilayotganlar" tugmasi.)
    notes_label = f"📝 Qaydlar ({inbox_n})" if inbox_n else "📝 Qaydlar"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="👥 Ijrochilar", callback_data="cockpit_team"),
            InlineKeyboardButton(text="🚨 Risklar", callback_data="cockpit_risks"),
        ],
        [
            InlineKeyboardButton(text="📊 Statistika", callback_data="cockpit_stats"),
            InlineKeyboardButton(text=notes_label, callback_data="nav_notes"),
        ],
        [InlineKeyboardButton(text="🔄 Yangilash", callback_data="cockpit_refresh")],
    ])

    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)




@router.callback_query(F.data == "cockpit_refresh")
async def cb_cockpit_refresh(query: CallbackQuery) -> None:
    await query.answer("Yangilanmoqda...")
    await cmd_cockpit(query.message)


@router.callback_query(F.data == "cockpit_team")
async def cb_cockpit_team(query: CallbackQuery) -> None:
    await query.answer()
    await cmd_team(query.message)


@router.callback_query(F.data == "cockpit_risks")
async def cb_cockpit_risks(query: CallbackQuery) -> None:
    await query.answer()
    await cmd_risks(query.message)






@router.callback_query(F.data == "cockpit_stats")
async def cb_cockpit_stats(query: CallbackQuery) -> None:
    await query.answer()
    days, label = 7, "Oxirgi 7 kun"
    stats = await database.executive_stats(days=days)
    await _safe_answer(
        query.message,
        _format_stats_dashboard(stats, label),
        parse_mode="Markdown",
        reply_markup=_stats_period_keyboard(days),
    )


# ─────────────────────── MESSAGE HANDLERS ───────────────────────


# ─────────────────────── VOICE CONFIRMATION ───────────────────────


def _escape_markdown(text: str) -> str:
    """Telegram Markdown V1 special chars: *, _, `, [, ]."""
    if not text:
        return ""
    for c in ("_", "*", "`", "[", "]"):
        text = text.replace(c, "\\" + c)
    return text


_VOICE_CONFIRM_KEYWORDS = {
    "confirm": {"tasdiqlayman", "tasdiq", "ha", "to'g'ri", "togri", "to'gri",
                "yes", "ok", "okay", "qabul", "davom"},
    "cancel":  {"yo'q", "yoq", "bekor", "no", "qayt", "to'xtat", "toxtat"},
    "edit":    {"tahrir", "tahrirla", "boshqacha", "noto'g'ri", "notogri",
                "qayta", "o'zgartir", "ozgartir"},
}


def _classify_voice_response(text: str) -> str | None:
    """Quick keyword classifier for confirm/cancel/edit. Returns None if ambiguous."""
    t = (text or "").lower().strip()
    if not t:
        return None
    # Direct single-word match
    words = set(t.replace(",", " ").replace(".", " ").split())
    for key, vocab in _VOICE_CONFIRM_KEYWORDS.items():
        if words & vocab:
            return key
    # Substring match (handles "tasdiqlayman.", "bekor qilaman" etc.)
    for key, vocab in _VOICE_CONFIRM_KEYWORDS.items():
        if any(v in t for v in vocab):
            return key
    return None


async def _send_voice_confirm_prompt(message: Message, state: FSMContext, transcript: str) -> None:
    """Show the confirm/edit/cancel prompt and store the transcript in FSM state."""
    DIVIDER = "━" * 20
    text = (
        "🎙 **OVOZ XABARINGIZ**\n\n"
        f"{DIVIDER}\n\n"
        f"\"{_escape_markdown(transcript)}\"\n\n"
        f"{DIVIDER}\n\n"
        "Tushundimi to'g'ri?"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Tasdiqlayman", callback_data="voice_ok"),
            InlineKeyboardButton(text="✏️ Tahrirlayman", callback_data="voice_edit"),
        ],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="voice_cancel")],
    ])
    await state.set_state(VoiceConfirmFSM.awaiting_action)
    await state.update_data(transcript=transcript)
    await _safe_answer(message, text, parse_mode="Markdown", reply_markup=kb)


async def _download_voice(message: Message, bot: Bot) -> bytes | None:
    """Download a voice file from Telegram. Returns None if size exceeds limit."""
    if message.voice.file_size and message.voice.file_size > voice_service.MAX_AUDIO_BYTES:
        await message.answer(
            f"Ovoz xabari juda katta ({message.voice.file_size // 1024} KB). "
            f"Iltimos, {voice_service.MAX_AUDIO_BYTES // (1024 * 1024)} MB dan kichikroq yuboring."
        )
        return None
    file = await bot.get_file(message.voice.file_id)
    audio_io = await bot.download_file(file.file_path)
    if hasattr(audio_io, "getvalue"):
        return audio_io.getvalue()
    if hasattr(audio_io, "read"):
        return audio_io.read()
    return bytes(audio_io)


def _forward_signal(message: Message) -> tuple[str | None, str | None]:
    """Extract (source_chat_label, author_label) from a Telegram forward.
    Returns (None, None) if the message isn't a forward.

    aiogram 3 unified forward metadata into `message.forward_origin`
    (MessageOrigin variants). Older clients/updates still populate the
    legacy `forward_from` / `forward_from_chat` attributes — check both."""
    chat_label: str | None = None
    author_label: str | None = None

    origin = getattr(message, "forward_origin", None)
    if origin is not None:
        # Try the union variants we care about.
        chat = getattr(origin, "chat", None)
        if chat is not None:
            chat_label = getattr(chat, "title", None) or getattr(chat, "username", None)
        sender_user = getattr(origin, "sender_user", None)
        if sender_user is not None:
            author_label = " ".join(filter(None, [
                getattr(sender_user, "first_name", None),
                getattr(sender_user, "last_name", None),
            ])).strip() or None
        sender_name = getattr(origin, "sender_user_name", None)
        if sender_name and not author_label:
            author_label = sender_name

    # Legacy fields (still set by some clients)
    legacy_chat = getattr(message, "forward_from_chat", None)
    if legacy_chat and not chat_label:
        chat_label = getattr(legacy_chat, "title", None) or getattr(legacy_chat, "username", None)
    legacy_user = getattr(message, "forward_from", None)
    if legacy_user and not author_label:
        author_label = " ".join(filter(None, [
            getattr(legacy_user, "first_name", None),
            getattr(legacy_user, "last_name", None),
        ])).strip() or None
    legacy_name = getattr(message, "forward_sender_name", None)
    if legacy_name and not author_label:
        author_label = legacy_name

    if chat_label or author_label or origin or legacy_chat or legacy_user or legacy_name:
        return chat_label, author_label
    return None, None


def _forward_is_bot_echo(message: Message) -> bool:
    """True if the forward originates from a bot (often THIS bot) — i.e. the user
    forwarded bot output back to it. That's not external info worth capturing as
    a note (reported: bot's own '📌 VAZIFALAR' panel saved as a forward note)."""
    bot_id = getattr(getattr(message, "bot", None), "id", None)
    origin = getattr(message, "forward_origin", None)
    for u in (getattr(origin, "sender_user", None), getattr(message, "forward_from", None)):
        if u is None:
            continue
        if getattr(u, "is_bot", False):
            return True
        if bot_id and getattr(u, "id", None) == bot_id:
            return True
    return False


@router.message(
    # A forward is an explicit "save this" gesture — capture it as a note in the
    # default chat AND while browsing any section. Previously this was
    # default_state only, so forwarding *inside a section* fell through to the
    # section's text handler -> _process_and_reply -> Claude, which reinterpreted
    # the forwarded message instead of saving it (reported: "meaning changed").
    # Active input FSMs (NewTask, reminders, meeting edit/protocol, …) are
    # intentionally NOT listed, so a forward there still serves as that flow's input.
    StateFilter(
        default_state,
        SectionFSM.in_tasks, SectionFSM.in_reminders, SectionFSM.in_meetings,
        SectionFSM.in_stats, SectionFSM.in_team, SectionFSM.in_risks,
        SectionFSM.in_today, SectionFSM.in_new, SectionFSM.in_search,
        SectionFSM.in_settings, SectionFSM.in_notes,
    ),
    F.forward_origin | F.forward_from | F.forward_from_chat | F.forward_sender_name,
)
async def handle_forwarded_message(message: Message, state: FSMContext) -> None:
    """Auto-capture: any forwarded message in default_state becomes a note.

    Empty forwards (stickers without caption, etc.) are rejected with a
    short explanation. Otherwise we save the text/caption with provenance
    metadata and confirm with two quick-actions (Inbox / Tahlil)."""
    chat_label, author_label = _forward_signal(message)
    content = (message.text or message.caption or "").strip()
    if not content:
        await message.answer(
            "📎 Bo'sh forward — qayd yaratilmadi. Matn yoki izoh bo'lgan xabarni forward qiling."
        )
        return
    # Bot'ning o'z chiqishi / tugma / buyruq forward qilinsa — qayd yaratmaymiz
    # (reported: '📌 VAZIFALAR' paneli, '⬅️ Asosiy menyu' tugmasi qayd bo'lib qolgan).
    if _forward_is_bot_echo(message) or _looks_like_bot_output(content) or _is_note_noise(content):
        await message.answer("↩️ Bu botning o'z xabari yoki tugma matni — qayd yaratilmadi.")
        return
    # html_text preserves Telegram entities (bold, italic, code, links) as HTML
    # tags. Used by _format_note_detail to render the forwarded text inside a
    # native <blockquote> with original formatting intact.
    html_content: str | None = None
    try:
        # aiogram 3 provides html_text for text and caption_html_text for media.
        html_content = (getattr(message, "html_text", None)
                         or getattr(message, "caption_html_text", None))
    except Exception:
        html_content = None
    nid = await database.create_note({
        "content": content,
        "content_html": html_content,
        "source": "forward",
        "source_chat": chat_label,
        "source_author": author_label,
        "source_message_id": message.message_id,
    })
    source_hint = "forward"
    if chat_label:
        source_hint = f"forward · {chat_label}"
    elif author_label:
        source_hint = f"forward · {author_label}"
    await _note_capture_reply(message, nid, source_hint)
    await _maybe_refresh_section(message, state, {"note": [nid]})


@router.message(StateFilter(default_state), F.voice)
async def handle_voice(message: Message, bot: Bot, state: FSMContext) -> None:
    """Free-form ovoz handler: transkripsiya → auto-process (default) yoki
    confirm prompt (agar foydalanuvchi /settings da yoqib qo'ygan bo'lsa).

    Faqat FSM state'siz holatda ishga tushadi. Boshqa FSM state aktiv bo'lsa
    (MeetingRescheduleFSM, MeetingEditFSM, MeetingProtocolFSM va boshqalar),
    o'sha state'ning maxsus handler'i ovozni qabul qiladi.
    """
    audio_bytes = await _download_voice(message, bot)
    if audio_bytes is None:
        return

    await message.bot.send_chat_action(message.chat.id, "typing")
    transcript = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
    if not transcript:
        await message.answer("Ovozni o'qiy olmadim. Iltimos, qaytadan urinib ko'ring yoki matn yozing.")
        return

    # Variant A+B: respect the voice_auto_confirm setting.
    settings = await database.get_settings()
    if settings.get("voice_auto_confirm", True):
        # Default — show the transcript first (so the user sees what was heard)
        # then immediately dispatch to Claude without an extra tap. Any
        # destructive action (create_task / schedule_meeting) still gets a
        # confirm prompt downstream via confirm_create_actions.
        # Show the FULL transcript (600-char safety cap) so the user sees exactly
        # what was heard — a non-destructive action runs immediately afterward.
        _shown = transcript.strip()
        if len(_shown) > 600:
            _shown = _shown[:600] + "…"
        await message.answer(
            f"_🎙 Tushundim:_ {_escape_markdown(_shown)}",
            parse_mode="Markdown",
        )
        await _process_and_reply(message, transcript, state=state)
    else:
        # Legacy flow — explicit confirm/edit/cancel buttons before processing.
        await _send_voice_confirm_prompt(message, state, transcript)


@router.callback_query(F.data == "voice_ok")
async def cb_voice_ok(query: CallbackQuery, state: FSMContext) -> None:
    """Confirmed transcript → process with Claude."""
    data = await state.get_data()
    transcript = (data.get("transcript") or "").strip()
    await state.clear()
    if not transcript:
        await query.answer("Matn topilmadi", show_alert=True)
        return
    await query.answer("✅ Tasdiqlandi")
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await _process_and_reply(query.message, transcript, state=state)


@router.callback_query(F.data == "voice_cancel")
async def cb_voice_cancel(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer("✕ Bekor qilindi")
    try:
        await query.message.edit_text(
            "✕ Bekor qilindi. Yangi xabar yuboring.",
            reply_markup=None,
        )
    except TelegramBadRequest:
        pass


# Data safety: snapshots taken right before a bulk delete, keyed by the
# confirm message_id so the "↩️ Qaytarish" button can restore them.
_UNDO_BACKUPS: dict[str, str] = {}


_BACKUP_KEEP = 20   # cap local snapshots so undo backups can't fill the disk over time


def _rotate_backups(backup_dir: Path, keep: int = _BACKUP_KEEP) -> None:
    """Keep only the newest `keep` snapshot files — auto undo-backups accumulate on
    every bulk/category delete and, unrotated, silently fill a small VPS disk (a
    'disk I/O error' cause). Never raises (best-effort cleanup)."""
    try:
        files = sorted(backup_dir.glob("yordamchi-*.db"),
                       key=lambda p: p.stat().st_mtime, reverse=True)
        for old in files[keep:]:
            try:
                old.unlink()
            except OSError:
                pass
    except Exception:
        logger.debug("Backup rotation skipped", exc_info=True)


async def _create_db_backup(tag: str) -> str:
    """Consistent SQLite snapshot via the .backup API (safe while writing).
    Returns the backup file path under data/backups/."""
    import sqlite3
    ts = datetime.now(database.TZ).strftime("%Y%m%d-%H%M%S")
    backup_dir = Path(config.DATABASE_PATH).parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    path = str(backup_dir / f"yordamchi-{tag}-{ts}.db")

    def _do() -> None:
        with sqlite3.connect(config.DATABASE_PATH) as src, sqlite3.connect(path) as dst:
            src.backup(dst)

    await asyncio.to_thread(_do)
    _rotate_backups(backup_dir)
    return path


@router.callback_query(F.data.startswith("undodelete:"))
async def cb_undo_delete(query: CallbackQuery) -> None:
    """'↩️ Qaytarish' — restore the pre-delete snapshot over the live DB. Safe
    immediately after a bulk delete (no intervening writes); uses the .backup
    API in reverse so WAL is handled correctly."""
    token = query.data.split(":", 1)[1]
    path = _UNDO_BACKUPS.get(token)
    if not path or not Path(path).exists():
        await query.answer("Qaytarish muddati o'tdi — backup topilmadi.", show_alert=True)
        return
    await query.answer("↩️ Tiklanmoqda…")
    import sqlite3

    def _restore() -> None:
        with sqlite3.connect(path) as src, sqlite3.connect(config.DATABASE_PATH) as dst:
            src.backup(dst)

    try:
        await asyncio.to_thread(_restore)
    except Exception as e:
        await query.message.answer(_humanize_error(e))
        return
    _UNDO_BACKUPS.pop(token, None)
    try:
        await query.message.edit_text(
            "↩️ **Tiklandi** — o'chirilgan ma'lumotlar qaytarildi.",
            parse_mode="Markdown")
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "acts_confirm")
async def cb_actions_confirm(query: CallbackQuery, state: FSMContext) -> None:
    """User tapped ✅ on a create-action preview — execute the deferred Claude
    response now, send the original user_message reply, AND restore the prior
    section state so the section auto-refresh sees the new item."""
    data = await state.get_data()
    response = data.get("pending_response")
    prior_state = data.get("_prior_section")
    note_id = data.get("_note_id")  # set when confirming a note-analyze action
    await state.clear()
    if not response or not isinstance(response, dict):
        await query.answer("Tasdiqlash vaqti o'tdi — yangi so'rov yuboring.",
                            show_alert=True)
        return
    await query.answer("✅ Tasdiqlandi")
    # Strip the confirm keyboard so it can't be re-tapped while we execute.
    try:
        await query.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    actions = response.get("actions", [])
    # Data safety: snapshot the DB before any irreversible bulk delete so we can
    # offer one-tap "↩️ Qaytarish". Failure to back up never blocks the delete.
    undo_token = None
    if any(a.get("type") in (_BULK_DELETE_ACTION_TYPES | _CATEGORY_DELETE_ACTION_TYPES)
           for a in actions):
        try:
            _UNDO_BACKUPS[str(query.message.message_id)] = await _create_db_backup("pre-delete")
            undo_token = str(query.message.message_id)
        except Exception:
            logger.exception("Pre-delete backup failed (continuing without undo)")
    try:
        ids_by_type = await _execute_actions(actions)
    except Exception as e:
        logger.exception("Deferred _execute_actions failed after confirm")
        await query.message.answer(_humanize_error(e))
        return
    # If this came from a note-analyze confirm, mark the source note processed.
    if note_id:
        for key in ("task", "reminder"):
            if ids_by_type.get(key):
                try:
                    await database.mark_note_processed(note_id, key, ids_by_type[key][0])
                except Exception:
                    logger.debug("mark_note_processed failed for %s", note_id)
                break
    keyboard = _build_keyboard(response.get("buttons", []), ids_by_type,
                               share_text=response.get("user_message"))
    if keyboard:
        keyboard = _append_back_row(keyboard)
    # Surface one-tap undo right after a bulk delete (snapshot taken above).
    if undo_token:
        undo_row = [InlineKeyboardButton(text="↩️ Qaytarish",
                                         callback_data=f"undodelete:{undo_token}")]
        if keyboard:
            keyboard.inline_keyboard.insert(0, undo_row)
        else:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[undo_row])
    # Uchrashuv tasdiq yo'li orqali o'tadi (schedule_meeting confirm-gated), shu
    # sabab to'qnashuv/yaroqsiz-vaqt ogohlantirishlari SHU YERDA ham qo'shilishi
    # shart. To'qnashuv/yaroqsiz-vaqt bo'lsa LLM'ning "✅ Yaratildi" matni ZID
    # bo'lardi — uni tashlab, faqat ogohlantirishni ko'rsatamiz.
    _base = (response.get("user_message") or "").strip() or "✅ Yaratildi"
    text = "" if (ids_by_type.get("_conflict") or ids_by_type.get("_badtime")) else _base
    text += _conflict_note(ids_by_type)
    text += _badtime_note(ids_by_type)
    text += _failed_actions_note(ids_by_type)
    text = text.strip() or _base
    await _safe_answer(query.message, text,
                        parse_mode="Markdown", reply_markup=keyboard)
    # Restore prior section state and auto-refresh the section list so the
    # new item is visible (otherwise the user sees a stale list higher in
    # the chat and thinks the create failed).
    if prior_state and isinstance(prior_state, str):
        try:
            await state.set_state(prior_state)
        except Exception:
            logger.debug("Could not restore prior state %r", prior_state)
    await _maybe_refresh_section(query.message, state, ids_by_type)


@router.callback_query(F.data == "acts_cancel")
async def cb_actions_cancel(query: CallbackQuery, state: FSMContext) -> None:
    """User tapped ✕ on a create-action preview — drop the deferred response."""
    await state.clear()
    await query.answer("✕ Bekor qilindi")
    try:
        await query.message.edit_text(
            "✕ **Yaratish bekor qilindi.**\n\nYangi xabar yuboring.",
            parse_mode="Markdown",
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data == "voice_edit")
async def cb_voice_edit(query: CallbackQuery, state: FSMContext) -> None:
    """Switch to revision state — next message (text or voice) replaces the transcript."""
    await state.set_state(VoiceConfirmFSM.awaiting_revision)
    await query.answer()
    text = (
        "✏️ **TAHRIRLASH**\n\n"
        "Yangi matn yoki ovoz yuboring — transkripsiyani almashtiraman."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="voice_cancel")],
    ])
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


@router.message(StateFilter(VoiceConfirmFSM.awaiting_revision), F.text | F.voice)
async def handle_voice_revision(message: Message, bot: Bot, state: FSMContext) -> None:
    """User sent a correction during voice edit flow — re-transcribe (if voice) and show confirm again."""
    if message.voice:
        audio_bytes = await _download_voice(message, bot)
        if audio_bytes is None:
            return
        await message.bot.send_chat_action(message.chat.id, "typing")
        new_transcript = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not new_transcript:
            await message.answer("Ovozni o'qiy olmadim. Matn yozing.")
            return
    else:
        new_transcript = (message.text or "").strip()
        if not new_transcript:
            await message.answer("Bo'sh xabar. Yangi matn yuboring.")
            return
    await _send_voice_confirm_prompt(message, state, new_transcript)


# ─────────────── POLISH OUTPUT ACTIONS (📋 Nusxa / 📤 Yuborish / ✎ Yana tahrir) ───────────────
# These callbacks ('copy', 'share', 'edit:polish') are emitted by Claude in the
# POLISH response's buttons array — they had no handlers, so the buttons did
# nothing (copy/share) or hit the task editor with id='polish' → "Vazifa topilmadi".

def _extract_polished_body(message: Message) -> str:
    """Pull just the polished letter out of a 'Tahrirlangan matn:' card —
    header line and ─── rules stripped — so it can be re-sent clean."""
    raw = (getattr(message, "text", None) or getattr(message, "caption", None) or "").strip()
    return _strip_polish_wrapper(raw)


@router.callback_query(F.data == "copy")
async def cb_polish_copy(query: CallbackQuery) -> None:
    """📋 Nusxa olish — re-send the polished text alone so it's easy to
    long-press → copy (a bot can't write to the clipboard directly)."""
    body = _extract_polished_body(query.message)
    if not body:
        await query.answer("Nusxa olinadigan matn topilmadi", show_alert=True)
        return
    await query.answer("📋 Toza matn pastda — bosib turib nusxa oling")
    await query.message.answer(body)


@router.callback_query(F.data == "share")
async def cb_polish_share(query: CallbackQuery) -> None:
    """📤 Boshqaga yuborish — re-send the polished text as a standalone message
    the user can forward to the recipient."""
    body = _extract_polished_body(query.message)
    if not body:
        await query.answer("Yuboriladigan matn topilmadi", show_alert=True)
        return
    await query.answer()
    await query.message.answer(body)
    await query.message.answer(
        "📤 _Yuqoridagi xabarni kerakli odamga forward qiling._",
        parse_mode="Markdown",
    )


@router.callback_query(F.data == "edit:polish")
async def cb_polish_edit(query: CallbackQuery, state: FSMContext) -> None:
    """✎ Yana tahrir — refine the polished text. Registered BEFORE the generic
    edit:<task-id> handler so it isn't mis-routed to the task editor."""
    body = _extract_polished_body(query.message)
    await state.set_state(PolishRevisionFSM.awaiting)
    await state.update_data(polish_original=body)
    await query.answer()
    await query.message.answer(
        "✏️ **Qanday o'zgartiray?**\n\nKo'rsatma yoki yangi matn yuboring "
        "(masalan: «qisqartir», «rasmiyroq qil», «iliqroq ohang»).",
        parse_mode="Markdown",
    )


@router.message(StateFilter(PolishRevisionFSM.awaiting), F.text | F.voice)
async def handle_polish_revision(message: Message, state: FSMContext) -> None:
    """User sent a revision instruction for the polished text — re-polish it
    (original + instruction) through the normal pipeline, which returns a fresh
    polished card with the same buttons."""
    instr = await _get_text_or_transcribe(message, bot=message.bot)
    if instr is None:
        return
    data = await state.get_data()
    original = data.get("polish_original", "")
    await state.clear()
    combined = (
        "Quyidagi rasmiy matnni ko'rsatma bo'yicha qayta tahrirla (polish). "
        f"Ko'rsatma: {(message.text or '').strip()}\n\nAsl matn:\n{original}"
    )
    await _process_and_reply(message, combined, state=state)


@router.message(StateFilter(VoiceConfirmFSM.awaiting_action), F.text | F.voice)
async def handle_voice_action(message: Message, bot: Bot, state: FSMContext) -> None:
    """User responded with text/voice instead of pressing a button — classify intent."""
    if message.voice:
        audio_bytes = await _download_voice(message, bot)
        if audio_bytes is None:
            return
        response_text = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not response_text:
            await message.answer("Ovozni o'qiy olmadim. Tugmalarni bosing yoki matn yuboring.")
            return
    else:
        response_text = (message.text or "").strip()

    intent = _classify_voice_response(response_text)
    data = await state.get_data()
    transcript = (data.get("transcript") or "").strip()

    if intent == "confirm":
        await state.clear()
        if transcript:
            await message.answer("✅ Tasdiqlandi")
            await _process_and_reply(message, transcript, state=state)
        return
    if intent == "cancel":
        await state.clear()
        await message.answer("✕ Bekor qilindi. Yangi xabar yuboring.")
        return
    if intent == "edit":
        await state.set_state(VoiceConfirmFSM.awaiting_revision)
        await message.answer(
            "✏️ Yangi matn yoki ovoz yuboring.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✕ Bekor qilish", callback_data="voice_cancel")],
            ]),
        )
        return

    # Ambiguous — treat as a brand-new transcript and re-ask.
    await _send_voice_confirm_prompt(message, state, response_text)


_REPLY_BUTTON_LABELS: set[str] = {
    BTN_COCKPIT, BTN_TODAY, BTN_TASKS, BTN_REMINDERS, BTN_TEAM, BTN_RISKS,
    BTN_NEW, BTN_STATS, BTN_SEARCH, BTN_MEETINGS, BTN_SETTINGS,
}



async def _restore_main_keyboard(message: Message) -> None:
    """Pastdagi reply kbd ni Asosiy menyuga qaytarish."""
    await message.answer(
        "🎛 **Asosiy menyu**\n\nKerakli bo'limni tanlang yoki matn/ovoz yuboring.",
        parse_mode="Markdown",
        reply_markup=main_reply_keyboard(),
    )


@router.message(F.text.func(lambda t: t and t.strip() in _REPLY_BUTTON_LABELS | set(_LEGACY_BTN_TASKS)))
async def handle_main_reply_button(message: Message, state: FSMContext) -> None:
    """Asosiy menyu reply tugmalari — har qanday holatda ishlasin
    (section'dan chiqib boshqa bo'limga o'tish uchun)."""
    _msg_text = message.text  # matn-only handler; section handlerlari bilan bir xil nom
    label = _msg_text.strip()
    current = await state.get_state()
    if current is not None:
        await state.clear()
    if label == BTN_COCKPIT:
        await _restore_main_keyboard(message); await cmd_cockpit(message); return
    if label == BTN_TODAY:
        await cmd_today(message, state); return
    if label == BTN_TASKS or label in _LEGACY_BTN_TASKS:
        await cmd_tasks(message, state); return
    if label == BTN_REMINDERS:
        await cmd_reminders(message, state); return
    if label == BTN_TEAM:
        await cmd_team(message, state); return
    if label == BTN_RISKS:
        await cmd_risks(message, state); return
    if label == BTN_NEW:
        await cmd_new(message, state); return
    if label == BTN_STATS:
        await cmd_stats(message, state); return
    if label == BTN_SEARCH:
        await cmd_search(message, state); return
    if label == BTN_MEETINGS:
        await cmd_meetings(message, state); return
    if label == BTN_SETTINGS:
        await cmd_settings(message, state); return


@router.message(F.text == BTN_BACK_MAIN)
async def handle_back_to_main(message: Message, state: FSMContext) -> None:
    """⬅️ Asosiy menyu — har qanday section'dan asosiy menyuga qaytadi."""
    await state.clear()
    await _restore_main_keyboard(message)


@router.message(StateFilter(SectionFSM.in_tasks), F.text | F.voice)
async def handle_tasks_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Vazifalar bo'limidagi reply tugmalari (state == in_tasks)."""
    label = _msg_text.strip()
    if label in _TASKS_SECTION_FILTERS:
        await _render_tasks_for_filter(message, _TASKS_SECTION_FILTERS[label])
        return
    if label == TBTN_TASKS_CATEGORIES:
        await _render_categories(message)
        return
    if label == TBTN_TASKS_NEW:
        await state.set_state(NewTaskTextFSM.awaiting_text)  # capture next msg → create-task
        await _safe_answer(
            message,
            "➕ **YANGI VAZIFA**\n\nMatn yoki ovoz yuboring. Misol:\n"
            "_\"Ertaga ertalab Aziz akaga marketing hisobotini yuborish\"_\n"
            "Bekor qilish: /cancel",
            parse_mode="Markdown",
        )
        return
    if label == TBTN_TASKS_SEARCH:
        # Vazifalar bo'limidan kelganda task-only qidiruv (TaskSearchFSM)
        # ishga tushiriladi — global Qidiruv bo'limiga sakramaymiz.
        await state.set_state(TaskSearchFSM.awaiting_query)
        await _safe_answer(
            message,
            "🔍 **VAZIFA QIDIRISH**\n\n"
            "Sarlavha, tavsif, teg yoki ijrochi bo'yicha so'z yuboring.",
            parse_mode="Markdown",
        )
        return
    # Boshqa matn — Claude'ga yuborish (section state'da ham erkin xabar mumkin)
    await _process_and_reply(message, _msg_text, state=state)


@router.message(F.text.func(lambda t: bool(t) and t.strip() in _REMINDERS_SECTION_FILTERS))
async def handle_reminder_filter_anystate(message: Message, state: FSMContext) -> None:
    """Eslatma filtr tugmasi (⏰ Bugun / 📤 Yuborilgan / 📋 Barchasi) HAR QANDAY
    holatda FAQAT eslatmalarni ko'rsatadi. FSM state yo'qolsa (TTL/navigatsiya)
    ham Claude'ga tushib 'Bugun → bugungi vazifalar' bo'lib ketmaydi (reported)."""
    await state.set_state(SectionFSM.in_reminders)
    await _render_reminders_for_filter(message, _REMINDERS_SECTION_FILTERS[message.text.strip()])


@router.message(StateFilter(SectionFSM.in_reminders), F.text | F.voice)
async def handle_reminders_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Eslatmalar bo'limidagi reply tugmalari."""
    label = _msg_text.strip()
    if label in _REMINDERS_SECTION_FILTERS:
        await _render_reminders_for_filter(message, _REMINDERS_SECTION_FILTERS[label])
        return
    if label == RBTN_REMINDERS_NEW:
        await _newreminder_start(message, state)
        return
    if label == RBTN_REMINDERS_SEARCH:
        await state.set_state(ReminderSearchFSM.awaiting_query)
        await _safe_answer(
            message,
            "🔍 **ESLATMA QIDIRISH**\n\nQidiruv so'zini yuboring.",
            parse_mode="Markdown",
        )
        return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_meetings), F.text | F.voice)
async def handle_meetings_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Uchrashuvlar bo'limidagi reply tugmalari (state == in_meetings)."""
    label = _msg_text.strip()
    if label in _MEETINGS_SECTION_FILTERS:
        await _render_meetings_for_filter(message, _MEETINGS_SECTION_FILTERS[label])
        return
    if label == MBTN_MEETINGS_NEW:
        await state.set_state(NewMeetingTextFSM.awaiting_text)  # capture next msg → create-meeting
        await _safe_answer(
            message,
            "➕ **YANGI UCHRASHUV**\n\nMatn yoki ovoz yuboring. Misol:\n"
            "_\"Ertaga soat 12:00 da Dinislam bilan biznes forum\"_\n"
            "Bekor qilish: /cancel",
            parse_mode="Markdown",
        )
        return
    if label == MBTN_MEETINGS_SEARCH:
        await cmd_search(message); return
    if label == MBTN_MEETINGS_PROTOCOLS:
        await _render_protocols(message); return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_stats), F.text | F.voice)
async def handle_stats_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Statistika bo'limidagi reply tugmalari (state == in_stats)."""
    label = _msg_text.strip()
    if label in _STATS_SECTION_PERIODS:
        days = _STATS_SECTION_PERIODS[label]
        stats = await database.executive_stats(days=days)
        period_label = {1: "Bugun", 7: "7 kun", 30: "30 kun"}.get(days, f"{days} kun")
        await _safe_answer(
            message, _format_stats_dashboard(stats, period_label),
            parse_mode="Markdown",
        )
        return
    if label == SBTN_STATS_REPORT_WEEK:
        stats = await database.executive_stats(days=7)
        await _safe_answer(message, _format_executive_report(stats, "Oxirgi 7 kun"),
                            parse_mode="Markdown", reply_markup=_report_keyboard(7))
        return
    if label == SBTN_STATS_REPORT_MONTH:
        stats = await database.executive_stats(days=30)
        await _safe_answer(message, _format_executive_report(stats, "Oxirgi 30 kun"),
                            parse_mode="Markdown", reply_markup=_report_keyboard(30))
        return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_team), F.text | F.voice)
async def handle_team_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Ijrochilar bo'limidagi reply tugmalari (state == in_team)."""
    label = _msg_text.strip()
    if label == YBTN_TEAM_REFRESH:
        await _render_team_panel(message); return
    if label == YBTN_TEAM_STALE:
        await _render_stale_delegations(message); return
    if label == YBTN_TEAM_UNASSIGNED:
        tasks = await database.list_unassigned_tasks(limit=50)
        text = _format_tasks_compact(tasks, "Ijrochisiz vazifalar")
        await _safe_answer(message, text, parse_mode="Markdown",
                            reply_markup=tasks_compact_keyboard(tasks, "all"))
        return
    if label == YBTN_TEAM_REASSIGN:
        await _safe_answer(
            message,
            "🔄 **QAYTA TAQSIMLASH**\n\n"
            "Vazifa raqamini va yangi ijrochini yuboring.\n"
            "_Misol: \"3-vazifani Sanjarga ber\"_",
            parse_mode="Markdown",
        )
        return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_risks), F.text | F.voice)
async def handle_risks_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Risklar bo'limidagi reply tugmalari (state == in_risks)."""
    label = _msg_text.strip()
    if label == RBTN_RISKS_REFRESH:
        await _render_risks_panel(message); return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_today), F.text | F.voice)
async def handle_today_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Bugun bo'limidagi reply tugmalari (state == in_today)."""
    label = _msg_text.strip()
    if label == DBTN_TODAY_EVENING:
        # Re-use cb_today_evening logic without callback object
        typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
        try:
            response = await claude_service.process_message(
                "", internal_directive="[INTERNAL] generate_evening_summary"
            )
        finally:
            typing_task.cancel()
        text = (response.get("user_message") or "").strip()
        if text:
            await _safe_answer(message, text, parse_mode="Markdown")
        else:
            await message.answer("Hozircha yakun chiqarib bo'lmadi.")
        return
    if label == DBTN_TODAY_ALL_TASKS:
        await cmd_tasks(message, state); return
    if label == DBTN_TODAY_NEW_TASK:
        await _safe_answer(
            message,
            "📝 **YANGI VAZIFA**\n\nMatn yoki ovoz yuboring. Misol:\n"
            "_\"Ertaga ertalab Aziz akaga marketing hisobotini yuborish\"_",
            parse_mode="Markdown",
        )
        return
    if label == DBTN_TODAY_MEETINGS:
        await cmd_meetings(message, state); return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_new), F.text | F.voice)
async def handle_new_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Yangi bo'limidagi reply tugmalari (state == in_new)."""
    label = _msg_text.strip()
    prompts = {
        NBTN_NEW_TASK: ("📝 **YANGI VAZIFA**\n\nMatn yoki ovoz yuboring. Misol:\n"
                        "_\"Ertaga ertalab Aziz akaga marketing hisoboti\"_"),
        NBTN_NEW_MEETING: ("🤝 **YANGI UCHRASHUV**\n\nMatn yoki ovoz yuboring. Misol:\n"
                           "_\"Juma 15:00 da Olim aka bilan byudjet uchrashuvi\"_"),
        NBTN_NEW_VOICE: ("🎙 **OVOZLI VAZIFA**\n\nMikrofon tugmasini bosib o'zbekcha gapiring. "
                         "Men transkripsiya qilib, vazifani tushunaman."),
        NBTN_NEW_POLISH: ("✏️ **MATN TAHRIRLASH**\n\nXabaringizni yuboring va aytib qo'ying — "
                          "kimga, qanday tonda. Misol:\n_\"Aziz akaga rasmiy qil: ertaga hisobot tayyor\"_"),
    }
    if label == NBTN_NEW_REMINDER:
        await _newreminder_start(message, state)
        return
    if label == NBTN_NEW_NOTE:
        # Enter the one-shot capture FSM — next text/voice becomes a note.
        await state.set_state(NoteCaptureFSM.awaiting_text)
        await _safe_answer(
            message,
            "📥 **YANGI QAYD**\n\nMatn yoki ovoz yuboring — bot uni Inbox'ga saqlaydi.\n"
            "Bekor qilish: /cancel",
            parse_mode="Markdown",
        )
        return
    if label == NBTN_NEW_MEETING:
        await state.set_state(NewMeetingTextFSM.awaiting_text)  # capture next msg → create-meeting
        await _safe_answer(message, prompts[label] + "\nBekor qilish: /cancel", parse_mode="Markdown")
        return
    if label == NBTN_NEW_TASK:
        await state.set_state(NewTaskTextFSM.awaiting_text)     # capture next msg → create-task
        await _safe_answer(message, prompts[label] + "\nBekor qilish: /cancel", parse_mode="Markdown")
        return
    if label in prompts:
        await _safe_answer(message, prompts[label], parse_mode="Markdown")
        return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_search), F.text | F.voice)
async def handle_search_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Qidiruv bo'limidagi reply tugmalari (state == in_search).
    Scope tugmasi tanlanmasa, matn to'g'ridan-to'g'ri qidiruv so'zi sifatida ishlatiladi.
    """
    label = _msg_text.strip()
    scope_map = {
        QBTN_SEARCH_TASKS:    "tasks",
        QBTN_SEARCH_MEETINGS: "meetings",
        QBTN_SEARCH_CONTACTS: "contacts",
        QBTN_SEARCH_ALL:      "all",
    }
    if label in scope_map:
        await state.update_data(scope=scope_map[label])
        scope_label = {
            "tasks": "📌 Faqat vazifalar",
            "meetings": "🤝 Faqat uchrashuvlar",
            "contacts": "👥 Faqat kontaktlar",
            "all": "🗂 Hammasi",
        }[scope_map[label]]
        await message.answer(
            f"✅ Scope: **{scope_label}**\n\nQidiruv so'zini yuboring.",
            parse_mode="Markdown",
        )
        return
    # Matn — qidirish bajariladi
    data = await state.get_data()
    scope = data.get("scope", "all")
    results = await database.search_all(label, limit=30)
    parts = []
    if scope in ("tasks", "all") and results.get("tasks"):
        parts.append(_format_tasks_compact(results["tasks"][:10], f"Qidiruv: {label}"))
    if scope in ("meetings", "all") and results.get("meetings"):
        parts.append(_format_meetings_compact(results["meetings"][:10], f"Qidiruv: {label}"))
    if scope == "all" and results.get("reminders"):
        parts.append(_format_reminders_compact(results["reminders"][:10], f"Qidiruv: {label}"))
    if scope in ("contacts", "all") and results.get("contacts"):
        cl = results["contacts"][:10]
        lines = [f"👥 **KONTAKTLAR** · {len(cl)} ta", ""]
        for c in cl:
            lines.append(f"• {c['name']} ({c.get('role') or '—'})")
        parts.append("\n".join(lines))
    if not parts:
        await message.answer(f"Hech narsa topilmadi: «{label}»")
        return
    await _safe_answer(message, "\n\n━━━━━━━━━━━━━━━━━━━━\n\n".join(parts),
                       parse_mode="Markdown")


@router.message(StateFilter(SectionFSM.in_notes), F.text | F.voice)
async def handle_notes_section_button(message: Message, state: FSMContext) -> None:
    """Reply-keyboard tugmalari Qaydlar bo'limida. Inbox/Ishlangan/Arxiv —
    filter; '➕ Yangi qayd' → one-shot capture FSM; '🔍 Qidirish' → search
    flow; matn fall-through Claude'ga, voice ham."""
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return
    label = _msg_text.strip()
    if label in _NOTES_SECTION_FILTERS:
        await _render_notes_for_filter(message, _NOTES_SECTION_FILTERS[label])
        return
    if label == NBTN_NOTES_NEW:
        await state.set_state(NoteCaptureFSM.awaiting_text)
        await _safe_answer(
            message,
            "📥 **YANGI QAYD**\n\nMatn yoki ovoz yuboring — bot uni Inbox'ga saqlaydi.\n"
            "Bekor qilish: /cancel",
            parse_mode="Markdown",
        )
        return
    if label == NBTN_NOTES_SEARCH:
        # Reuse the global search FSM but pre-flag a notes-only view via
        # a hint in state.data; the existing flow handles the rest.
        await state.set_state(GlobalSearchFSM.awaiting_query)
        await state.update_data(_search_scope="notes")
        await _safe_answer(
            message,
            "🔍 **QAYD QIDIRISH**\n\nKalit so'z yoki ibora yuboring.",
            parse_mode="Markdown",
        )
        return
    # Fall-through: free text/voice → Claude (might create a note via LLM).
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(SectionFSM.in_settings), F.text | F.voice)
async def handle_settings_section_button(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """Sozlamalar bo'limidagi reply tugmalari (state == in_settings)."""
    label = _msg_text.strip()
    if label == GBTN_SETTINGS_NOTIFY:
        settings = await database.get_settings()
        new_val = not settings["notifications_enabled"]
        await database.set_setting("notifications_enabled", new_val)
        await message.answer(
            f"🔔 Bildirishnomalar: **{'yoqildi ✅' if new_val else 'oʻchirildi ❌'}**",
            parse_mode="Markdown",
        )
        return
    if label == GBTN_SETTINGS_BRIEFING:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=t, callback_data=f"brieftime:{t}") for t in ("07:00", "07:30", "08:00")],
            [InlineKeyboardButton(text=t, callback_data=f"brieftime:{t}") for t in ("08:30", "09:00", "10:00")],
        ])
        await message.answer("⏰ **Ertalab brifing vaqti:**", parse_mode="Markdown", reply_markup=kb)
        return
    if label == GBTN_SETTINGS_EVENING:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=t, callback_data=f"eveningtime:{t}") for t in ("17:00", "17:30", "18:00")],
            [InlineKeyboardButton(text=t, callback_data=f"eveningtime:{t}") for t in ("18:30", "19:00", "20:00")],
        ])
        await message.answer("🌙 **Kechki yakun vaqti:**", parse_mode="Markdown", reply_markup=kb)
        return
    if label == GBTN_SETTINGS_REMINDER:
        settings = await database.get_settings()
        text = (
            "📲 **Eslatma parametrlari**\n\n"
            f"• Uchrashuv: `{settings['meeting_reminder_min']} daq` oldin\n"
            f"• Vazifa: `{settings['task_reminder_hours']} soat` oldin (Shoshilinch va Muhim)"
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="15 daq", callback_data="meetremind:15"),
             InlineKeyboardButton(text="30 daq", callback_data="meetremind:30"),
             InlineKeyboardButton(text="60 daq", callback_data="meetremind:60")],
            [InlineKeyboardButton(text="1 soat", callback_data="taskremind:1"),
             InlineKeyboardButton(text="2 soat", callback_data="taskremind:2"),
             InlineKeyboardButton(text="4 soat", callback_data="taskremind:4")],
        ])
        await message.answer(text, parse_mode="Markdown", reply_markup=kb)
        return
    if label == GBTN_SETTINGS_VOICE:
        settings = await database.get_settings()
        new_val = not settings.get("voice_auto_confirm", True)
        await database.set_setting("voice_auto_confirm", new_val)
        if new_val:
            text = (
                "🎙 **Ovoz tasdig'i:** AVTO - tasdiqsiz ishlaydi ✅\n\n"
                "Ovoz transkripti darhol qayta ishlanadi."
            )
        else:
            text = (
                "🎙 **Ovoz tasdig'i:** tasdiq so'raladi ✅\n\n"
                "Har bir ovoz xabarida transkript ko'rsatiladi; "
                "tasdiqlasangizgina davom etadi."
            )
        await message.answer(text, parse_mode="Markdown")
        return
    if label == GBTN_SETTINGS_CREATE_CONFIRM:
        settings = await database.get_settings()
        new_val = not settings.get("confirm_create_actions", True)
        await database.set_setting("confirm_create_actions", new_val)
        if new_val:
            text = (
                "✅ **Yaratish tasdig'i:** yoqildi ✅\n\n"
                "Vazifa yoki uchrashuv yaratishdan oldin tasdiq so'raladi."
            )
        else:
            text = (
                "✅ **Yaratish tasdig'i:** o'chirildi ✅\n\n"
                "Vazifa va uchrashuvlar tasdiqsiz yaratiladi."
            )
        await message.answer(text, parse_mode="Markdown")
        return
    if label == GBTN_SETTINGS_CALENDAR:
        status = "yoqilgan" if config.ICLOUD_ENABLED else "oʻchirilgan"
        cal_name = config.ICLOUD_CALENDAR_NAME or "(default)"
        text = (
            "📅 **Kalendar holati**\n\n"
            f"iCloud sync: **{status}**\n"
            f"Kalendar: `{cal_name}`\n"
            f"Sync interval: `{config.ICLOUD_SYNC_INTERVAL_MIN} daq`"
        )
        await message.answer(text, parse_mode="Markdown")
        return
    await _process_and_reply(message, _msg_text, state=state)


@router.message(StateFilter(default_state), F.text)
async def handle_text(message: Message, state: FSMContext) -> None:
    """Free-form text (faqat FSM state'siz holatda). Aktiv state'da
    o'sha state'ning maxsus handler'i tomonidan qabul qilinadi."""
    if message.text.startswith("/"):
        return
    _msg_text = message.text
    await _process_and_reply(message, _msg_text, state=state)


# ─────────────────────── CALLBACK HANDLERS ───────────────────────


async def _delete_if_possible(message: Message | None) -> None:
    if not message:
        return
    try:
        await message.delete()
    except TelegramBadRequest:
        # Message too old / already deleted / not modifiable — benign.
        # Narrow except prevents masking unexpected errors (network, programming).
        pass


@router.callback_query(F.data == "nav_cockpit")
async def cb_nav_cockpit(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer()
    await _delete_if_possible(query.message)
    await cmd_cockpit(query.message)


@router.callback_query(F.data == "nav_new")
async def cb_nav_new(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer()
    await _delete_if_possible(query.message)
    await cmd_new(query.message)


@router.callback_query(F.data == "nav_settings")
async def cb_nav_settings(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer()
    await _delete_if_possible(query.message)
    await cmd_settings(query.message)




@router.callback_query(F.data == "nav_meetings")
async def cb_nav_meetings(query: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await query.answer()
    await _render_meetings_for_filter(query.message, "week", edit_existing=True)



@router.callback_query(F.data.startswith("confirm:"))
async def cb_confirm(query: CallbackQuery) -> None:
    await query.answer("Tasdiqlandi ✅")
    try:
        await query.message.edit_reply_markup(reply_markup=single_back_keyboard())
    except Exception:
        pass


@router.callback_query(F.data.startswith("task_del:"))
async def cb_task_del_confirm(query: CallbackQuery) -> None:
    """Two-step task delete: first show confirmation prompt."""
    tid = query.data.split(":", 1)[1]
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()
    title = (task.get("title") or "—").strip()
    text = (
        "🗑 **VAZIFANI O'CHIRISH**\n\n"
        f"{title}\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Haqiqatdan ham o'chirmoqchimisiz?"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Ha, o'chir", callback_data=f"task_del_do:{tid}"),
            InlineKeyboardButton(text="⬅️ Yo'q", callback_data=f"taskopen:{tid}"),
        ],
    ])
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


# Single-task delete undo: full deleted-task dict + timestamp, keyed by an opaque
# token. Lighter than _UNDO_BACKUPS (no full-DB copy) — restored via restore_task.
_UNDO_TASKS: dict[str, tuple[dict, datetime]] = {}
_UNDO_TASK_TTL = timedelta(minutes=5)


def _undo_task_gc() -> None:
    """Drop undo entries older than the TTL so the dict can't grow unbounded."""
    cutoff = datetime.now(database.TZ) - _UNDO_TASK_TTL
    for k in [k for k, (_, ts) in _UNDO_TASKS.items() if ts < cutoff]:
        _UNDO_TASKS.pop(k, None)


@router.callback_query(F.data.startswith("task_del_do:"))
async def cb_task_del_do(query: CallbackQuery) -> None:
    """Execute task deletion after confirmation, offering a one-tap undo."""
    tid = query.data.split(":", 1)[1]
    # Snapshot the WHOLE subtree + linked reminders before delete — else the undo
    # would silently restore only the parent and drop its subtasks/reminders.
    snapshot = await database.snapshot_task_tree(tid)
    await database.delete_task(tid)
    await query.answer("✅ Vazifa o'chirildi")
    kb = single_back_keyboard("taskfilter:active")
    if snapshot and snapshot.get("tasks"):
        _undo_task_gc()
        token = database.new_id("undo-")
        _UNDO_TASKS[token] = (snapshot, datetime.now(database.TZ))
        undo_row = [InlineKeyboardButton(text="↩️ Qaytarish",
                                         callback_data=f"undotask:{token}")]
        kb = InlineKeyboardMarkup(inline_keyboard=[undo_row] + list(kb.inline_keyboard))
    try:
        await query.message.edit_text("🗑 Vazifa o'chirildi.", reply_markup=kb)
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("undotask:"))
async def cb_undo_task_delete(query: CallbackQuery) -> None:
    """'↩️ Qaytarish' after a single-task delete — re-insert the captured row with
    its original id. Mirrors the bulk-delete undo, but per-task (no DB snapshot)."""
    token = query.data.split(":", 1)[1]
    entry = _UNDO_TASKS.pop(token, None)
    if not entry:
        await query.answer("Qaytarish muddati o'tdi.", show_alert=True)
        return
    snapshot, _ts = entry
    try:
        ok = await database.restore_task_tree(snapshot) > 0
    except Exception as e:
        await query.message.answer(_humanize_error(e))
        return
    if not ok:
        await query.answer("Allaqachon tiklangan.", show_alert=True)
        return
    await query.answer("↩️ Tiklandi ✅")
    try:
        await query.message.edit_text(
            "↩️ Vazifa tiklandi.",
            reply_markup=single_back_keyboard("taskfilter:active"),
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("cancel:"))
async def cb_cancel(query: CallbackQuery) -> None:
    parts = query.data.split(":", 1)
    target_id = parts[1] if len(parts) > 1 else ""
    if target_id.startswith("t-"):
        await database.delete_task(target_id)
        await query.answer("Vazifa bekor qilindi ✕")
        back_target = "taskfilter:active"
    elif target_id.startswith("m-"):
        await database.cancel_meeting(target_id)
        sched = scheduler_module.get_scheduler()
        if sched:
            try:
                sched.remove_meeting_reminder(target_id)
            except Exception:
                logger.exception("Failed to remove meeting reminder for %s", target_id)
        if config.ICLOUD_ENABLED:
            try:
                await asyncio.to_thread(calendar_service.delete_meeting, target_id)
            except Exception:
                pass
        await query.answer("Uchrashuv bekor qilindi ✕")
        back_target = "meetingfilter:week"
    else:
        await query.answer("Bekor qilindi")
        back_target = "nav_cockpit"
    try:
        await query.message.edit_reply_markup(reply_markup=single_back_keyboard(back_target))
    except Exception:
        pass


@router.callback_query(F.data.startswith("complete:"))
async def cb_complete(query: CallbackQuery) -> None:
    parts = query.data.split(":", 1)
    target_id = parts[1] if len(parts) > 1 else ""
    if not target_id:
        await query.answer()
        return
    await database.complete_task(target_id)
    await query.answer("Bajarildi ✅")
    # Cardni vizual yangilash — foydalanuvchi yangi statusni darrov ko'rsin.
    # Plus add an "Undo" button so a fat-finger tap can be reversed within
    # the next minute without hunting through the task detail menu.
    task = await database.get_task(target_id)
    if task:
        undo_row = InlineKeyboardButton(text="↶ Bekor qilish", callback_data=f"unclomp:{target_id}")
        kb = _task_card_kb_with_back(task)
        # Prepend the undo row so it's the most prominent action.
        kb = InlineKeyboardMarkup(inline_keyboard=[[undo_row]] + list(kb.inline_keyboard))
        try:
            await query.message.edit_text(
                _format_task_card(task), parse_mode="Markdown",
                reply_markup=kb,
            )
            return
        except TelegramBadRequest:
            pass
    # Fallback — agar matn yangilanmasa, kamida tugmalarni yangilab qo'yamiz.
    try:
        await query.message.edit_reply_markup(
            reply_markup=single_back_keyboard("taskfilter:active"))
    except Exception:
        pass


@router.callback_query(F.data.startswith("unclomp:"))
async def cb_uncomplete(query: CallbackQuery) -> None:
    """Undo a task completion. Flips status back to 'todo' and re-renders
    the card so the user can immediately recover from a fat-finger tap."""
    parts = query.data.split(":", 1)
    target_id = parts[1] if len(parts) > 1 else ""
    if not target_id:
        await query.answer()
        return
    ok = await database.update_task(target_id, {"status": "todo"}, source="undo_complete")
    if not ok:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer("↶ Yana aktiv ✅")
    task = await database.get_task(target_id)
    if task:
        try:
            await query.message.edit_text(
                _format_task_card(task), parse_mode="Markdown",
                reply_markup=_task_card_kb_with_back(task),
            )
        except TelegramBadRequest:
            pass


def _snooze_deadline(task: dict, when: str) -> datetime:
    """Compute a new deadline for a one-tap reschedule. The clock time is preserved
    from the task's existing deadline (else defaults to 18:00 — end of workday).
    Base is ALWAYS relative to now (snooze = push it forward from today)."""
    now = datetime.now(database.TZ)
    cur = _parse_dt_safe(task.get("deadline"))
    hh, mm = (cur.hour, cur.minute) if cur else (18, 0)
    if when == "week":
        base = now + timedelta(days=7)
    elif when == "monday":
        # the coming Monday; if today is already Monday, jump a full week ahead
        base = now + timedelta(days=((7 - now.weekday()) % 7) or 7)
    else:  # "tomorrow" (default)
        base = now + timedelta(days=1)
    return base.replace(hour=hh, minute=mm, second=0, microsecond=0)


@router.callback_query(F.data.startswith("snooze:"))
async def cb_task_snooze(query: CallbackQuery) -> None:
    """One-tap reschedule from the task card — moves the deadline to tomorrow /
    +1 week / next Monday (keeping the clock time), then re-renders the card."""
    try:
        _, tid, when = query.data.split(":", 2)
    except ValueError:
        await query.answer()
        return
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    new_dt = _snooze_deadline(task, when)
    ok = await database.update_task(tid, {"deadline": new_dt.isoformat()}, source="snooze")
    if not ok:
        await query.answer("Saqlanmadi", show_alert=True)
        return
    task = await database.get_task(tid)
    await query.answer(f"📅 Muddat: {_task_deadline_chip(task)}")
    try:
        await query.message.edit_text(
            _format_task_card(task), parse_mode="Markdown",
            reply_markup=_task_card_kb_with_back(task),
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("task_detail:"))
async def cb_task_detail(query: CallbackQuery) -> None:
    """⋯ Batafsil — show the full action menu for a task."""
    tid = query.data.split(":", 1)[1]
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()
    text = _format_task_detail_card(task)
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=task_detail_menu(task))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                           reply_markup=task_detail_menu(task))


@router.callback_query(F.data.startswith("set_assignee:"))
async def cb_set_assignee(query: CallbackQuery, state: FSMContext) -> None:
    """👤 Ijrochi — prompt user for assignee name (free text or 'Men' for self)."""
    tid = query.data.split(":", 1)[1]
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await state.set_state(AssigneeFSM.awaiting_name)
    await state.update_data(task_id=tid)
    await query.answer()
    current = task.get("assignee") or "belgilanmagan"
    cancel_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🧹 Tozalash", callback_data=f"clear_assignee:{tid}")],
        [InlineKeyboardButton(text="✕ Bekor qilish", callback_data=f"taskopen:{tid}")],
    ])
    await query.message.answer(
        f"👤 **Ijrochi tayinlash**\n\nHozir: {current}\n\n"
        f"Yangi ism yuboring (masalan: «Komilov Javohir» yoki «Men»):",
        parse_mode="Markdown",
        reply_markup=cancel_kb,
    )


@router.callback_query(F.data.startswith("clear_assignee:"))
async def cb_clear_assignee(query: CallbackQuery, state: FSMContext) -> None:
    tid = query.data.split(":", 1)[1]
    await state.clear()
    await database.update_task(tid, {"assignee": None}, source="edit")
    await query.answer("Ijrochi tozalandi ✅")
    task = await database.get_task(tid)
    if task:
        try:
            await query.message.edit_text(
                _format_task_card(task), parse_mode="Markdown",
                reply_markup=_task_card_kb_with_back(task),
            )
        except TelegramBadRequest:
            pass


@router.message(StateFilter(AssigneeFSM.awaiting_name), F.text | F.voice)
async def handle_assignee_input(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    data = await state.get_data()
    tid = data.get("task_id")
    if not tid:
        await state.clear()
        return
    name = message.text.strip()[:80]
    if not name:
        await message.answer("Boʻsh nom qabul qilinmadi. Qaytadan yuboring yoki bekor qiling.")
        return
    await state.clear()
    await database.update_task(tid, {"assignee": name}, source="edit")
    task = await database.get_task(tid)
    if task:
        await _safe_answer(
            message, f"✅ Ijrochi: **{name}**\n\n" + _format_task_card(task),
            parse_mode="Markdown",
            reply_markup=_task_card_kb_with_back(task),
        )


def _reschedule_presets_keyboard(mid: str) -> InlineKeyboardMarkup:
    """6 ta tezkor preset + qo'lda kiritish + bekor."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⏰ Bugun 16:00", callback_data=f"resched_preset:{mid}:today_16"),
            InlineKeyboardButton(text="⏰ Bugun 18:00", callback_data=f"resched_preset:{mid}:today_18"),
        ],
        [
            InlineKeyboardButton(text="📅 Ertaga 09:00", callback_data=f"resched_preset:{mid}:tomorrow_9"),
            InlineKeyboardButton(text="📅 Ertaga 14:00", callback_data=f"resched_preset:{mid}:tomorrow_14"),
        ],
        [
            InlineKeyboardButton(text="📅 +3 kun, 10:00", callback_data=f"resched_preset:{mid}:plus3"),
            InlineKeyboardButton(text="📅 Keyingi hafta", callback_data=f"resched_preset:{mid}:next_week"),
        ],
        [InlineKeyboardButton(text="✏️ Qo'lda kiritish", callback_data=f"resched_manual:{mid}")],
        [InlineKeyboardButton(text="⬅️ Bekor", callback_data=f"meetingopen:{mid}")],
    ])


def _resched_preset_to_datetime(
    key: str,
    now: datetime,
    meeting_start: datetime | None = None,
) -> datetime | None:
    """Map preset key to a concrete future datetime (TZ-aware).

    Today-presets roll to tomorrow if the chosen clock time has already passed.
    `next_week` preserves the original meeting's clock time (or 10:00 if missing).
    `meeting_start` — current datetime_start of the meeting (used for next_week).
    """
    today = now.replace(second=0, microsecond=0)

    def _today_or_tomorrow_at(hour: int) -> datetime:
        candidate = today.replace(hour=hour, minute=0)
        if candidate <= now:
            candidate += timedelta(days=1)
        return candidate

    if key == "today_16":
        return _today_or_tomorrow_at(16)
    if key == "today_18":
        return _today_or_tomorrow_at(18)
    if key == "tomorrow_9":
        return (today + timedelta(days=1)).replace(hour=9, minute=0)
    if key == "tomorrow_14":
        return (today + timedelta(days=1)).replace(hour=14, minute=0)
    if key == "plus3":
        # +3 kun, asl uchrashuv vaqtini saqlash (yo'q bo'lsa 10:00)
        if meeting_start:
            base = meeting_start.astimezone(database.TZ)
            return (today + timedelta(days=3)).replace(hour=base.hour, minute=base.minute)
        return (today + timedelta(days=3)).replace(hour=10, minute=0)
    if key == "next_week":
        # +7 kun — asl uchrashuv vaqtini saqlash
        if meeting_start:
            base = meeting_start.astimezone(database.TZ)
            return (today + timedelta(days=7)).replace(hour=base.hour, minute=base.minute)
        return (today + timedelta(days=7)).replace(hour=10, minute=0)
    return None


async def _apply_reschedule(mid: str, new_start: datetime) -> dict | None:
    """Persist new start (and shift end by same delta), trigger iCloud re-sync,
    and re-register the meeting reminder job at the new time.
    Returns the updated meeting dict or None if not found.
    """
    meeting = await database.get_meeting(mid)
    if not meeting:
        return None
    try:
        old_start = datetime.fromisoformat(meeting["datetime_start"])
    except (ValueError, TypeError):
        return None
    if new_start.tzinfo is None:
        new_start = database.TZ.localize(new_start)
    delta = new_start - old_start.astimezone(database.TZ)
    new_end_iso = None
    if meeting.get("datetime_end"):
        try:
            old_end = datetime.fromisoformat(meeting["datetime_end"])
            new_end_iso = (old_end + delta).isoformat()
        except (ValueError, TypeError):
            pass
    import aiosqlite
    async with aiosqlite.connect(config.DATABASE_PATH) as db:
        if new_end_iso:
            await db.execute(
                "UPDATE meetings SET datetime_start=?, datetime_end=?, "
                "reminded_at=NULL, prep_sent_at=NULL, followup_sent_at=NULL "
                "WHERE id=?",
                (new_start.isoformat(), new_end_iso, mid),
            )
        else:
            await db.execute(
                "UPDATE meetings SET datetime_start=?, "
                "reminded_at=NULL, prep_sent_at=NULL, followup_sent_at=NULL "
                "WHERE id=?",
                (new_start.isoformat(), mid),
            )
        await db.commit()

    # Re-register the reminder at the new time. APScheduler replaces
    # the existing job with the same id, so the old reminder is cancelled.
    sched = scheduler_module.get_scheduler()
    if sched:
        try:
            sched.schedule_meeting_reminder(mid, new_start.isoformat())
        except Exception:
            logger.exception("Failed to reschedule reminder for %s", mid)

    # iCloud re-sync (background — delete old event, push new one)
    if config.ICLOUD_ENABLED:
        _spawn_background(_resync_meeting_to_icloud(mid), name=f"icloud_resync:{mid}")

    return await database.get_meeting(mid)


def _resched_interval(meeting: dict, new_start: datetime) -> tuple[str, str | None]:
    """(start_iso, end_iso) for a duration-preserving reschedule — for the conflict
    check. Mirrors _apply_reschedule's delta logic so the check matches the result."""
    if new_start.tzinfo is None:
        new_start = database.TZ.localize(new_start)
    end_iso = None
    try:
        old_start = datetime.fromisoformat(meeting["datetime_start"]).astimezone(database.TZ)
        if meeting.get("datetime_end"):
            old_end = datetime.fromisoformat(meeting["datetime_end"])
            end_iso = (old_end + (new_start - old_start)).isoformat()
    except (ValueError, TypeError, KeyError):
        pass
    return new_start.isoformat(), end_iso


def _conflict_warn_text(conflicts: list[dict], verb: str = "ko'chirilsinmi") -> str:
    """Band-vaqt ogohlantirish matni — reschedule/duration band slotга tushganda."""
    lines = "\n".join(
        f"• {_meeting_time_label(c.get('datetime_start') or '', with_past_marker=False)} — "
        f"{(c.get('title') or '—').strip()}" for c in conflicts[:5])
    return f"⚠️ **Vaqt band — to'qnashuv:**\n{lines}\n\nBaribir {verb}?"


def _resched_force_kb(mid: str, new_start: datetime) -> InlineKeyboardMarkup:
    """Override keyboard: ko'chirishni tasdiqlash (band bo'lsa ham) yoki bekor."""
    compact = new_start.astimezone(database.TZ).strftime("%Y%m%d%H%M")
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⚠️ Baribir ko'chir", callback_data=f"rsforce:{mid}:{compact}"),
        InlineKeyboardButton(text="✕ Bekor", callback_data=f"meetingopen:{mid}"),
    ]])


@router.callback_query(F.data.startswith("reschedule:"))
async def cb_reschedule(query: CallbackQuery) -> None:
    """Show the time-preset picker for a meeting."""
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer()
    current = _meeting_time_label(meeting.get("datetime_start") or "", with_past_marker=False)
    DIVIDER = "━" * 20
    text = "\n".join([
        "🔄 **YANGI VAQT TANLANG**",
        "",
        f"Hozirgi vaqt: {current}",
        "",
        DIVIDER,
        "",
        "Tezkor variantlar:",
    ])
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_reschedule_presets_keyboard(mid))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=_reschedule_presets_keyboard(mid))


@router.callback_query(F.data.startswith("resched_preset:"))
async def cb_resched_preset(query: CallbackQuery) -> None:
    """Apply a preset reschedule."""
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    mid, key = parts[1], parts[2]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    now = datetime.now(database.TZ)
    meeting_start: datetime | None = None
    try:
        meeting_start = datetime.fromisoformat(meeting["datetime_start"]).astimezone(database.TZ)
    except (ValueError, TypeError, KeyError):
        meeting_start = None
    new_start = _resched_preset_to_datetime(key, now, meeting_start)
    if not new_start:
        await query.answer("Preset noma'lum", show_alert=True)
        return
    _si, _ei = _resched_interval(meeting, new_start)
    _conf = await database.find_meeting_conflicts(_si, _ei, exclude_id=mid)
    if _conf:
        await query.answer("⚠️ Vaqt band")
        await _safe_answer(query.message, _conflict_warn_text(_conf), parse_mode="Markdown",
                           reply_markup=_resched_force_kb(mid, new_start))
        return
    updated = await _apply_reschedule(mid, new_start)
    if not updated:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    label = _meeting_time_label(updated.get("datetime_start") or "", with_past_marker=False)
    await query.answer(f"✅ Vaqt {label} ga ko'chirildi")
    try:
        await query.message.edit_text(
            _format_meeting_card(updated, show_date=True),
            parse_mode="Markdown",
            reply_markup=meeting_inline_actions(updated),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, _format_meeting_card(updated, show_date=True),
                            parse_mode="Markdown", reply_markup=meeting_inline_actions(updated))


@router.callback_query(F.data.startswith("rsforce:"))
async def cb_resched_force(query: CallbackQuery) -> None:
    """Apply a reschedule the principal CONFIRMED despite a time conflict
    (rsforce:{mid}:{YYYYMMDDHHMM} — compact start, colon-free for the callback)."""
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer()
        return
    mid, compact = parts[1], parts[2]
    try:
        new_start = database.TZ.localize(datetime.strptime(compact, "%Y%m%d%H%M"))
    except (ValueError, TypeError):
        await query.answer("Vaqt xato", show_alert=True)
        return
    updated = await _apply_reschedule(mid, new_start)
    if not updated:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    label = _meeting_time_label(updated.get("datetime_start") or "", with_past_marker=False)
    await query.answer(f"⚠️ {label} ga ko'chirildi (band edi)")
    try:
        await query.message.edit_text(
            _format_meeting_card(updated, show_date=True),
            parse_mode="Markdown", reply_markup=meeting_inline_actions(updated))
    except TelegramBadRequest:
        await _safe_answer(query.message, _format_meeting_card(updated, show_date=True),
                            parse_mode="Markdown", reply_markup=meeting_inline_actions(updated))


class MeetingRescheduleFSM(StatesGroup):
    awaiting_datetime = State()


@router.callback_query(F.data.startswith("resched_manual:"))
async def cb_resched_manual(query: CallbackQuery, state: FSMContext) -> None:
    """Ask the user to type or speak the new date/time. Claude parses on receipt."""
    mid = query.data.split(":", 1)[1]
    await state.set_state(MeetingRescheduleFSM.awaiting_datetime)
    await state.update_data(meeting_id=mid)
    await query.answer()
    text = (
        "✏️ **YANGI VAQT**\n\n"
        "Yangi sana va vaqtni yozib yuboring yoki ovoz xabar yuboring.\n\n"
        "_Misol: \"28-may 11:00\" yoki \"ertaga ertalab 09:00\"_"
    )
    await query.message.edit_text(text, parse_mode="Markdown",
                                   reply_markup=single_back_keyboard(f"meetingopen:{mid}",
                                                                     text="✕ Bekor qilish"))


@router.message(StateFilter(MeetingRescheduleFSM.awaiting_datetime), F.text | F.voice)
async def handle_resched_manual(message: Message, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    mid = data.get("meeting_id")
    await state.clear()
    if not mid:
        await message.answer("Uchrashuv topilmadi.")
        return
    # Resolve user text — accept voice too.
    if message.voice:
        await message.bot.send_chat_action(message.chat.id, "typing")
        file = await bot.get_file(message.voice.file_id)
        audio_io = await bot.download_file(file.file_path)
        audio_bytes = audio_io.getvalue() if hasattr(audio_io, "getvalue") else audio_io.read()
        user_text = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not user_text:
            await message.answer("Ovozni o'qiy olmadim. Matn yuboring.")
            return
    else:
        user_text = (message.text or "").strip()
    if not user_text:
        await message.answer("Bo'sh xabar — bekor qilindi.")
        return
    # Ask Claude to parse the date/time. Use an internal directive that returns
    # a normalized ISO datetime in user_message, or "INVALID" if it cannot parse.
    directive = (
        "[INTERNAL] parse_reschedule_datetime\n\n"
        f"User input: {user_text!r}\n"
        f"Current time: {datetime.now(database.TZ).isoformat()}\n"
        "Return ONLY the ISO 8601 datetime (Asia/Tashkent, +05:00) in user_message, "
        "for example '2026-05-28T11:00:00+05:00'. "
        "If you cannot determine a clear future datetime, return user_message='INVALID'."
    )
    response = await claude_service.process_message("", internal_directive=directive, complexity="fast")
    raw = (response.get("user_message") or "").strip()
    new_start: datetime | None = None
    if raw and raw.upper() != "INVALID":
        try:
            new_start = datetime.fromisoformat(raw)
            if new_start.tzinfo is None:
                new_start = database.TZ.localize(new_start)
        except (ValueError, TypeError):
            new_start = None
    if not new_start:
        await message.answer(
            "Vaqtni tushuna olmadim. Misol: \"28-may 11:00\" yoki \"ertaga 09:00\".",
            reply_markup=single_back_keyboard(f"meetingopen:{mid}", text="✕ Bekor"),
        )
        return
    _m_for_conf = await database.get_meeting(mid)
    if _m_for_conf:
        _si, _ei = _resched_interval(_m_for_conf, new_start)
        _conf = await database.find_meeting_conflicts(_si, _ei, exclude_id=mid)
        if _conf:
            await _safe_answer(message, _conflict_warn_text(_conf), parse_mode="Markdown",
                               reply_markup=_resched_force_kb(mid, new_start))
            return
    updated = await _apply_reschedule(mid, new_start)
    if not updated:
        await message.answer("Uchrashuv topilmadi.")
        return
    label = _meeting_time_label(updated.get("datetime_start") or "", with_past_marker=False)
    await _safe_answer(
        message,
        f"✅ Vaqt **{label}** ga ko'chirildi.\n\n" + _format_meeting_card(updated, show_date=True),
        parse_mode="Markdown",
        reply_markup=meeting_inline_actions(updated),
    )


async def _resync_meeting_to_icloud(meeting_id: str) -> None:
    """Delete and re-push a meeting after reschedule."""
    try:
        await asyncio.to_thread(calendar_service.delete_meeting, meeting_id)
        meeting = await database.get_meeting(meeting_id)
        if not meeting:
            return
        start_dt = datetime.fromisoformat(meeting["datetime_start"])
        end_iso = meeting.get("datetime_end") or (start_dt + timedelta(hours=1)).isoformat()
        end_dt = datetime.fromisoformat(end_iso)
        uid = await asyncio.to_thread(
            calendar_service.push_meeting,
            meeting_id, meeting["title"], start_dt, end_dt,
            meeting.get("participants"), meeting.get("location_or_link"), meeting.get("agenda"),
        )
        if uid:
            import aiosqlite
            async with aiosqlite.connect(config.DATABASE_PATH) as db:
                await db.execute("UPDATE meetings SET icloud_uid=? WHERE id=?", (uid, meeting_id))
                await db.commit()
    except Exception:
        logger.exception("Meeting re-sync to iCloud failed")


_EDIT_FIELD_LABELS = {
    "title":        ("📝 Sarlavha",      "Sarlavhani yuboring."),
    "participants": ("👥 Ishtirokchilar", "Ishtirokchilarni vergul bilan ajratib yuboring. Misol: Dinislam, Sanjar"),
    "location":     ("📍 Manzil",        "Manzil yoki havolani yuboring."),
    "agenda":       ("📅 Kun tartibi",   "Kun tartibini yuboring."),
}


def _meeting_edit_menu_keyboard(mid: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=label, callback_data=f"medit:{mid}:{field}")
         for field, (label, _) in [(k, v) for k, v in _EDIT_FIELD_LABELS.items() if k == "title"]],
        [InlineKeyboardButton(text=_EDIT_FIELD_LABELS["participants"][0],
                              callback_data=f"medit:{mid}:participants")],
        [InlineKeyboardButton(text=_EDIT_FIELD_LABELS["location"][0],
                              callback_data=f"medit:{mid}:location")],
        [InlineKeyboardButton(text=_EDIT_FIELD_LABELS["agenda"][0],
                              callback_data=f"medit:{mid}:agenda")],
        [InlineKeyboardButton(text="⏱ Davomiylik",
                              callback_data=f"medit:{mid}:duration")],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"meetingopen:{mid}")],
    ])


def _duration_picker_keyboard(mid: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="30 daqiqa", callback_data=f"mdur:{mid}:30"),
            InlineKeyboardButton(text="1 soat", callback_data=f"mdur:{mid}:60"),
        ],
        [
            InlineKeyboardButton(text="1.5 soat", callback_data=f"mdur:{mid}:90"),
            InlineKeyboardButton(text="2 soat", callback_data=f"mdur:{mid}:120"),
        ],
        [
            InlineKeyboardButton(text="3 soat", callback_data=f"mdur:{mid}:180"),
            InlineKeyboardButton(text="4 soat", callback_data=f"mdur:{mid}:240"),
        ],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"meeting_edit:{mid}")],
    ])


@router.callback_query(F.data.startswith("meeting_edit:"))
async def cb_meeting_edit(query: CallbackQuery) -> None:
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer()
    title = (meeting.get("title") or "—").strip()
    text = (
        "✏️ **TAHRIRLASH**\n\n"
        f"Uchrashuv: {title}\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Qaysi maydonni o'zgartirmoqchisiz?"
    )
    try:
        await query.message.edit_text(text, parse_mode="Markdown",
                                       reply_markup=_meeting_edit_menu_keyboard(mid))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=_meeting_edit_menu_keyboard(mid))


@router.callback_query(F.data.startswith("medit:"))
async def cb_meeting_edit_field(query: CallbackQuery, state: FSMContext) -> None:
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    mid, field = parts[1], parts[2]
    if field == "duration":
        await query.answer()
        text = (
            "⏱ **DAVOMIYLIK**\n\n"
            "Uchrashuv qancha davom etadi?"
        )
        try:
            await query.message.edit_text(text, parse_mode="Markdown",
                                           reply_markup=_duration_picker_keyboard(mid))
        except TelegramBadRequest:
            await _safe_answer(query.message, text, parse_mode="Markdown",
                                reply_markup=_duration_picker_keyboard(mid))
        return
    if field not in _EDIT_FIELD_LABELS:
        await query.answer("Maydon noma'lum", show_alert=True)
        return
    label, prompt = _EDIT_FIELD_LABELS[field]
    await state.set_state(MeetingEditFSM.awaiting_value)
    await state.update_data(meeting_id=mid, field=field)
    await query.answer()
    text = f"{label}\n\n{prompt}"
    try:
        await query.message.edit_text(
            text, parse_mode="Markdown",
            reply_markup=single_back_keyboard(f"meeting_edit:{mid}", text="✕ Bekor"),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=single_back_keyboard(f"meeting_edit:{mid}", text="✕ Bekor"))


@router.message(StateFilter(MeetingEditFSM.awaiting_value), F.text | F.voice)
async def handle_meeting_edit_value(message: Message, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    mid = data.get("meeting_id")
    field = data.get("field")
    await state.clear()
    if not mid or field not in _EDIT_FIELD_LABELS:
        await message.answer("Holat yo'qoldi — qayta urinib ko'ring.")
        return
    if message.voice:
        await message.bot.send_chat_action(message.chat.id, "typing")
        file = await bot.get_file(message.voice.file_id)
        audio_io = await bot.download_file(file.file_path)
        audio_bytes = audio_io.getvalue() if hasattr(audio_io, "getvalue") else audio_io.read()
        new_value = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not new_value:
            await message.answer("Ovozni o'qiy olmadim. Matn yuboring.")
            return
    else:
        new_value = (message.text or "").strip()
    if not new_value:
        await message.answer("Bo'sh qiymat — bekor qilindi.")
        return

    update_data: dict = {}
    if field == "participants":
        parts = [p.strip() for p in new_value.replace(";", ",").split(",") if p.strip()]
        update_data["participants"] = parts
    elif field == "title":
        update_data["title"] = new_value
    elif field == "location":
        update_data["location_or_link"] = new_value
    elif field == "agenda":
        update_data["agenda"] = new_value

    await database.update_meeting(mid, update_data)
    if config.ICLOUD_ENABLED:
        _spawn_background(_resync_meeting_to_icloud(mid), name=f"icloud_resync:{mid}")

    updated = await database.get_meeting(mid)
    if not updated:
        await message.answer("Uchrashuv topilmadi.")
        return
    await _safe_answer(
        message,
        "✅ Yangilandi.\n\n" + _format_meeting_card(updated, show_date=True),
        parse_mode="Markdown",
        reply_markup=meeting_inline_actions(updated),
    )


@router.callback_query(F.data.startswith("mdur:"))
async def cb_meeting_duration(query: CallbackQuery) -> None:
    """Set duration: keeps start as-is, computes new end = start + N minutes."""
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer("Xato format", show_alert=True)
        return
    mid, minutes_str = parts[1], parts[2]
    try:
        minutes = int(minutes_str)
    except ValueError:
        await query.answer("Davomiylik xato", show_alert=True)
        return
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    try:
        start = datetime.fromisoformat(meeting["datetime_start"])
    except (ValueError, TypeError):
        await query.answer("Vaqt formatida xato", show_alert=True)
        return
    new_end = start + timedelta(minutes=minutes)
    # Davomiylikni uzaytirsa keyingi uchrashuvga kirib ketishi mumkin — tekshiramiz.
    _conf = await database.find_meeting_conflicts(
        meeting["datetime_start"], new_end.isoformat(), exclude_id=mid)
    if _conf:
        await query.answer("⚠️ Vaqt band")
        _kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="⚠️ Baribir saqla", callback_data=f"mdurf:{mid}:{minutes}"),
            InlineKeyboardButton(text="✕ Bekor", callback_data=f"meetingopen:{mid}"),
        ]])
        await _safe_answer(query.message, _conflict_warn_text(_conf, verb="o'zgartirilsinmi"),
                           parse_mode="Markdown", reply_markup=_kb)
        return
    await query.answer(f"✅ Davomiylik {minutes} daq ga o'zgartirildi")
    await _apply_meeting_duration(query, mid, minutes)


async def _apply_meeting_duration(query: CallbackQuery, mid: str, minutes: int) -> None:
    """Persist new end = start + minutes; iCloud re-sync; re-render card.
    Shared by cb_meeting_duration (after the conflict check) and the force path."""
    meeting = await database.get_meeting(mid)
    if not meeting:
        return
    try:
        start = datetime.fromisoformat(meeting["datetime_start"])
    except (ValueError, TypeError):
        return
    new_end = start + timedelta(minutes=minutes)
    await database.update_meeting(mid, {"datetime_end": new_end.isoformat()})
    if config.ICLOUD_ENABLED:
        _spawn_background(_resync_meeting_to_icloud(mid), name=f"icloud_resync:{mid}")
    updated = await database.get_meeting(mid)
    if updated:
        try:
            await query.message.edit_text(_format_meeting_card(updated, show_date=True),
                                           parse_mode="Markdown",
                                           reply_markup=meeting_inline_actions(updated))
        except TelegramBadRequest:
            await _safe_answer(query.message, _format_meeting_card(updated, show_date=True),
                                parse_mode="Markdown", reply_markup=meeting_inline_actions(updated))


@router.callback_query(F.data.startswith("mdurf:"))
async def cb_meeting_duration_force(query: CallbackQuery) -> None:
    """Apply a duration change the principal CONFIRMED despite a time conflict."""
    parts = query.data.split(":")
    if len(parts) < 3:
        await query.answer()
        return
    mid = parts[1]
    try:
        minutes = int(parts[2])
    except ValueError:
        await query.answer("Davomiylik xato", show_alert=True)
        return
    await query.answer(f"⚠️ {minutes} daq (band edi)")
    await _apply_meeting_duration(query, mid, minutes)


def _build_protocol_directive(meeting: dict, user_notes: str) -> str:
    """Build the [INTERNAL] generate_meeting_protocol directive for Claude.
    Asks Claude to return a JSON envelope where:
      - user_message: full formatted protocol text (with markdown sections)
      - actions: list of create_task actions for the topshiriqlar
    """
    try:
        dt = datetime.fromisoformat(meeting["datetime_start"]).astimezone(database.TZ)
        date_str = f"{dt.year}-yil {dt.day}-{UZ_MONTHS_FULL[dt.month - 1]}"
        time_str = dt.strftime("%H:%M")
    except (ValueError, TypeError):
        date_str = meeting.get("datetime_start") or "—"
        time_str = ""
    end_time_str = ""
    if meeting.get("datetime_end"):
        try:
            end_dt = datetime.fromisoformat(meeting["datetime_end"]).astimezone(database.TZ)
            end_time_str = end_dt.strftime("%H:%M")
        except (ValueError, TypeError):
            pass
    participants = ", ".join(meeting.get("participants") or []) or "belgilanmagan"
    location = (meeting.get("location_or_link") or "").strip() or "belgilanmagan"
    agenda = (meeting.get("agenda") or "").strip() or "belgilanmagan"

    return (
        "[INTERNAL] generate_meeting_protocol\n\n"
        "Sen rasmiy bayonnoma MUHARRIRISAN, muallif EMAS. Foydalanuvchi bergan "
        "qaydlarni rasmiy, professional o'zbek uslubiga keltirasan — XOLOS. Yangi "
        "muhokama tafsiloti, qaror, topshiriq, ishtirokchi yoki lavozim O'YLAB TOPMA. "
        "Faqat berilgan qaydlar + uchrashuv metadata'sidan foydalan.\n\n"
        "UCHRASHUV MA'LUMOTLARI:\n"
        f"  Mavzu: {meeting.get('title') or '—'}\n"
        f"  Sana: {date_str}\n"
        f"  Vaqt: soat {time_str}" + (f" dan {end_time_str} gacha" if end_time_str else "") + "\n"
        f"  Joy: {location}\n"
        f"  Ishtirokchilar: {participants}\n"
        f"  Kun tartibi: {agenda}\n\n"
        f"FOYDALANUVCHI QAYDLARI (yagona mazmun manbai):\n{user_notes}\n\n"
        "1-QADAM — USLUBNI TANLA (uchrashuv mavzusi va qaydlar ohangiga qarab):\n"
        "  A) RASMIY — kengash, hay'at, vazirlik/rasmiy muzokara, qaror qabul "
        "qiluvchi yig'ilish. To'liq formal: har masala bo'yicha 'ESHITILDI:' va "
        "'QAROR QILINDI:' (qarorlar raqamli — 1.1, 1.2); oxirida 'Rais: … · Kotib: …'.\n"
        "  B) ISHCHI (STANDART) — jamoa yig'ilishi, loyiha/operatsion muhokama. "
        "Har mavzu BITTA blok: 🔹 Eshitildi (qisqa) → ✅ Qaror → 📌 Topshiriq.\n"
        "  C) QISQA — 1:1, tezkor sync, qisqa muloqot. Muhokamasiz: faqat "
        "✅ **QARORLAR** va 📌 **TOPSHIRIQLAR** ro'yxati.\n"
        "  Aniq belgilanmasa — B ni tanla.\n\n"
        "2-QADAM — BAYONNOMANI YOZ (markdown). Doimiy sarlavha (ma'lumot bo'lmagan "
        "qatorni TASHLA, placeholder qo'yma):\n"
        "📝 **UCHRASHUV BAYONNOMASI**\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📅 **Sana va vaqt** — metadata'dan\n"
        "📍 **Joy** — bo'lsa (yo'q bo'lsa qatorni tashla)\n"
        "👥 **Ishtirokchilar** — faqat berilgan ismlar (lavozim yo'q bo'lsa — faqat ism)\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "📋 **KUN TARTIBI** — mavzular ro'yxati\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "ASOSIY QISM — tanlangan uslub (A/B/C) tuzilmasida.\n\n"
        "QAT'IY QOIDALAR (eng muhim):\n"
        "- HECH NARSA QO'SHMA — qaydда yo'q bo'lsa, bayonnomada ham bo'lmaydi.\n"
        "- Eshitildi/Muhokama — faqat aytilgani, qisqa; tasvirlanmagan bo'lsa tashla. "
        "Tafsilot O'YLAB TOPMA.\n"
        "- Noma'lum maydon (joy/lavozim/mas'ul/muddat) — tashlab ket, "
        "'[aniqlashtirish kerak]' kabi placeholder YOZMA.\n"
        "- Uslub doimo RASMIY: faol nisbat, hissiy so'zsiz, aniq sanalar. Qarorlar "
        "buyruq fe'li bilan ('...sin'). Mazmun — foydalanuvchiniki; uslub — rasmiy.\n\n"
        "TOPSHIRIQLAR uchun `actions` ga create_task joyla (title, assignee, "
        "deadline ISO 8601 yoki null, priority='P1'). Qaydда aniq topshiriq yo'q "
        "bo'lsa actions: []. Qaror yoki uchrashuvни topshiriq qilib QO'SHMA."
    )


_SCRIPT_LABEL = {"lat": "Lotin", "kir": "Kiril"}


def _proto_export_row(mid: str, script: str, ctx: str) -> list:
    """[📄 Word] [📄 PDF] [🔤 <boshqa yozuv>] — eksport + Lotin/Kiril almashtirgich.
    script: 'lat'/'kir' (joriy holat); ctx: 'res' (yangi natija kbd) / 'view' (saqlangan)."""
    other = "kir" if script == "lat" else "lat"
    return [
        InlineKeyboardButton(text="📄 Word", callback_data=f"proto_export:{mid}:word:{script}"),
        InlineKeyboardButton(text="📄 PDF", callback_data=f"proto_export:{mid}:pdf:{script}"),
        InlineKeyboardButton(text=f"🔤 {_SCRIPT_LABEL[other]}", callback_data=f"proto_script:{mid}:{ctx}:{other}"),
    ]


def _viewproto_kb(mid: str, script: str = "lat") -> InlineKeyboardMarkup:
    """Saqlangan bayonnoma ko'rinishidagi klaviatura: Word/PDF + yozuv + nusxa/ulashish."""
    return InlineKeyboardMarkup(inline_keyboard=[
        _proto_export_row(mid, script, "view"),
        [InlineKeyboardButton(text="📋 Nusxa", callback_data=f"proto_share:{mid}"),
         InlineKeyboardButton(text="📤 Ulashish", switch_inline_query=f"proto:{mid}")],
        [back_button(f"meetingopen:{mid}", "⬅️ Orqaga")],
    ])


def _protocol_result_kb(mid: str, n_pending: int, saved: bool = False,
                        tasks_done: bool = False, script: str = "lat") -> InlineKeyboardMarkup:
    """Buttons under a generated protocol. Save + create-tasks are DECOUPLED.
    Eksport qatori: Word/PDF + Lotin/Kiril almashtirgich (Agrobank shabloni)."""
    rows: list = []
    first = []
    if not saved:
        first.append(InlineKeyboardButton(text="✅ Bayonnomani saqlash", callback_data=f"proto_ok:{mid}"))
    if n_pending and not tasks_done:
        first.append(InlineKeyboardButton(text=f"📌 Vazifalarni qo'shish ({n_pending})",
                                          callback_data=f"proto_tasks:{mid}"))
    if first:
        rows.append(first)
    rows.append([
        InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"proto_edit:{mid}"),
        InlineKeyboardButton(text="📋 Nusxa", callback_data=f"proto_share:{mid}"),
        InlineKeyboardButton(text="📤 Ulashish", switch_inline_query=f"proto:{mid}"),
    ])
    rows.append(_proto_export_row(mid, script, "res"))
    rows.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"meetingopen:{mid}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# Emoji/markdown tozalagich — `_proto_tasks_from_text` topshiriqlarni ajratishda
# ishlatadi. (Eski rang-kodli Word builder protocol_doc.py — Agrobank shabloni —
# bilan almashtirildi; eski rang/keyword konstantalari olib tashlandi.)
_PROTO_EMOJI_RE = re.compile("[\U0001F300-\U0001FAFF☀-➿←-⇿⬀-⯿️⃣]")


def _proto_clean(line: str) -> str:
    """Emoji va markdown belgilarini (**, ━, ─) tozalaydi."""
    return (_PROTO_EMOJI_RE.sub("", line).replace("**", "")
            .replace("━", "").replace("─", "").strip())


def _proto_format_deadline(iso) -> str:
    """Jadval 'Muddat' ustuni — sana yoki 'Aniqlashtirilsin' (null bo'lsa)."""
    if not iso:
        return "Aniqlashtirilsin"
    try:
        dt = datetime.fromisoformat(iso).astimezone(database.TZ)
        return f"{dt.day}-{UZ_MONTHS_FULL[dt.month - 1]}"
    except (ValueError, TypeError):
        return str(iso)


def _proto_tasks_from_actions(actions) -> list:
    """create_task action'laridan jadval uchun topshiriqlarni ajratadi."""
    out = []
    for a in (actions or []):
        if a.get("type") == "create_task":
            d = a.get("data") or {}
            out.append({"assignee": (d.get("assignee") or "").strip(),
                        "title": (d.get("title") or "").strip(),
                        "deadline": d.get("deadline")})
    return out


def _proto_tasks_from_text(protocol_text: str) -> list:
    """Fallback: TOPSHIRIQ qatorlaridan jadval topshiriqlarini ajratadi
    (saqlangan bayonnomani eksport qilganda actions bo'lmasa)."""
    out = []
    for raw in protocol_text.split("\n"):
        c = _proto_clean(raw)
        if c.upper().startswith("TOPSHIRIQ"):
            rest = c[len("TOPSHIRIQ"):].lstrip(" :—-").strip()
            if " — " in rest:
                assignee, _, task = rest.partition(" — ")
                out.append({"assignee": assignee.strip(), "title": task.strip(), "deadline": None})
            elif rest:
                out.append({"assignee": "", "title": rest, "deadline": None})
    return out


@router.callback_query(F.data.startswith("protocol:"))
async def cb_meeting_protocol_start(query: CallbackQuery, state: FSMContext) -> None:
    """Open the protocol-creation prompt for a meeting."""
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await state.set_state(MeetingProtocolFSM.awaiting_notes)
    await state.update_data(meeting_id=mid)
    await query.answer()
    title = (meeting.get("title") or "—").strip()
    when = _meeting_time_label(meeting.get("datetime_start") or "", with_past_marker=False)
    text = (
        "📝 **BAYONNOMA**\n\n"
        f"{title}\n"
        f"{when}\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Uchrashuvda nimalar muhokama qilindi?\n\n"
        "Qisqacha yozing yoki ovozli xabar yuboring.\n"
        "Bot sizning matningiz asosida rasmiy bayonnoma tayyorlaydi.\n\n"
        "_Eslatib o'ting:_\n"
        "_• Muhokama mavzulari_\n"
        "_• Qabul qilingan qarorlar_\n"
        "_• Topshiriqlar (kim, nima, qachon)_"
    )
    try:
        await query.message.edit_text(
            text, parse_mode="Markdown",
            reply_markup=single_back_keyboard(f"meetingopen:{mid}", text="✕ Bekor"),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=single_back_keyboard(f"meetingopen:{mid}", text="✕ Bekor"))


@router.message(StateFilter(MeetingProtocolFSM.awaiting_notes), F.text | F.voice)
async def handle_protocol_notes(message: Message, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    mid = data.get("meeting_id")
    if not mid:
        await state.clear()
        await message.answer("Holat yo'qoldi — qayta urinib ko'ring.")
        return
    if message.voice:
        await message.bot.send_chat_action(message.chat.id, "typing")
        file = await bot.get_file(message.voice.file_id)
        audio_io = await bot.download_file(file.file_path)
        audio_bytes = audio_io.getvalue() if hasattr(audio_io, "getvalue") else audio_io.read()
        user_notes = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not user_notes:
            await message.answer("Ovozni o'qiy olmadim. Matn yuboring.")
            return
    else:
        user_notes = (message.text or "").strip()
    if not user_notes:
        await message.answer("Bo'sh xabar — bekor qilindi.")
        await state.clear()
        return

    meeting = await database.get_meeting(mid)
    if not meeting:
        await state.clear()
        await message.answer("Uchrashuv topilmadi.")
        return

    await message.bot.send_chat_action(message.chat.id, "typing")
    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        directive = _build_protocol_directive(meeting, user_notes)
        response = await claude_service.process_message("", internal_directive=directive)
    finally:
        typing_task.cancel()

    protocol_text = (response.get("user_message") or "").strip()
    pending_actions = response.get("actions") or []
    if not protocol_text:
        await message.answer("Bayonnomani tuzib bo'lmadi — qayta urinib ko'ring.")
        await state.clear()
        return

    # Cache the result in FSM data so confirm/edit can find it.
    await state.update_data(
        protocol_text=protocol_text,
        pending_actions=pending_actions,
        user_notes=user_notes,
    )

    pending_count = sum(1 for a in pending_actions if a.get("type") == "create_task")
    await state.update_data(proto_saved=False, proto_tasks_done=False, proto_pending_count=pending_count)
    await _safe_answer(message, protocol_text, parse_mode="Markdown",
                       reply_markup=_protocol_result_kb(mid, pending_count, saved=False, tasks_done=False))


@router.callback_query(F.data.startswith("proto_ok:"))
async def cb_protocol_confirm(query: CallbackQuery, state: FSMContext) -> None:
    """Save the protocol text ONLY. Task creation is a SEPARATE, optional step
    (📌 Vazifalarni qo'shish) — saving never auto-creates tasks."""
    mid = query.data.split(":", 1)[1]
    data = await state.get_data()
    protocol_text = data.get("protocol_text", "")
    if not protocol_text:
        await query.answer("Bayonnoma topilmadi — qayta yarating.", show_alert=True)
        return
    await database.update_meeting(mid, {
        "follow_up_actions": [protocol_text],
        "followup_sent_at": datetime.now(database.TZ).isoformat(),
    })
    await state.update_data(proto_saved=True)
    await query.answer("✅ Bayonnoma saqlandi")
    try:
        await query.message.edit_reply_markup(reply_markup=_protocol_result_kb(
            mid, data.get("proto_pending_count", 0),
            saved=True, tasks_done=data.get("proto_tasks_done", False)))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("proto_tasks:"))
async def cb_protocol_tasks(query: CallbackQuery, state: FSMContext) -> None:
    """Create tasks from the protocol's topshiriqlar — separate, optional step."""
    mid = query.data.split(":", 1)[1]
    data = await state.get_data()
    pending = data.get("pending_actions") or []
    if not pending:
        await query.answer("Vazifa topilmadi — bayonnomani qayta yarating.", show_alert=True)
        return
    created = await _execute_actions(pending)
    n = len(created.get("task", []))
    await state.update_data(proto_tasks_done=True)
    _fail = "  ⚠️ ba'zilari saqlanmadi" if created.get("_failed") else ""
    await query.answer((f"✅ {n} ta vazifa qo'shildi" if n else "Vazifa qo'shilmadi") + _fail)
    try:
        await query.message.edit_reply_markup(reply_markup=_protocol_result_kb(
            mid, data.get("proto_pending_count", n),
            saved=data.get("proto_saved", False), tasks_done=True))
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("proto_export:"))
async def cb_protocol_export(query: CallbackQuery, state: FSMContext) -> None:
    """Export the bayonnoma as Word/PDF in the chosen script (Agrobank template).
    Callback: proto_export:{mid}:{fmt}:{script}  (fmt=word|pdf, script=lat|kir).
    Backward-compat: proto_export:{mid} → word/lat."""
    parts = query.data.split(":")
    mid = parts[1] if len(parts) > 1 else ""
    fmt = parts[2] if len(parts) > 2 else "word"
    script = "cyrillic" if (len(parts) > 3 and parts[3] == "kir") else "latin"
    data = await state.get_data()
    text = (data.get("protocol_text") or "").strip()
    m = await database.get_meeting(mid)
    if m and not text:
        fu = m.get("follow_up_actions") or []
        text = (fu[0] if (isinstance(fu, list) and fu) else str(fu or "")).strip()
    if not m or not text:
        await query.answer("Bayonnoma topilmadi — avval saqlang.", show_alert=True)
        return
    await query.answer("📄 Tayyorlayapman…")
    from aiogram.types import BufferedInputFile
    try:
        import protocol_doc
        proto_tasks = _proto_tasks_from_actions(data.get("pending_actions")) or _proto_tasks_from_text(text)
        settings = await database.get_settings()
        fields = protocol_doc.build_fields(m, text, proto_tasks, settings)
        stamp = datetime.now(database.TZ).strftime("%Y-%m-%d")
        if fmt == "pdf":
            blob = protocol_doc.build_pdf(fields, script)
            fname, caption = f"bayonnoma_{stamp}.pdf", "📄 Bayonnoma (PDF)"
        else:
            blob = protocol_doc.build_docx(fields, script)
            fname, caption = f"bayonnoma_{stamp}.docx", "📄 Bayonnoma (Word)"
        if script == "cyrillic":
            caption += " · Kiril"
        await query.message.answer_document(BufferedInputFile(blob, filename=fname), caption=caption)
    except RuntimeError as e:
        # reportlab o'rnatilmagan (PDF) — aniq xabar; Word ishlayveradi.
        await query.message.answer(f"⚠️ {e}")
    except Exception as e:
        await query.message.answer(_humanize_error(e))


@router.callback_query(F.data.startswith("proto_script:"))
async def cb_protocol_script(query: CallbackQuery, state: FSMContext) -> None:
    """Lotin/Kiril almashtirgich — bayonnoma klaviaturasini boshqa yozuvga qayta chizadi
    (matn o'zgarmaydi; faqat Word/PDF eksport yozuvi + tugma yorlig'i)."""
    parts = query.data.split(":")
    if len(parts) < 4:
        await query.answer()
        return
    mid, ctx, script = parts[1], parts[2], parts[3]
    if ctx == "view":
        kb = _viewproto_kb(mid, script)
    else:
        data = await state.get_data()
        kb = _protocol_result_kb(mid, data.get("proto_pending_count", 0),
                                 saved=data.get("proto_saved", False),
                                 tasks_done=data.get("proto_tasks_done", False),
                                 script=script)
    try:
        await query.message.edit_reply_markup(reply_markup=kb)
        await query.answer(f"Yozuv: {_SCRIPT_LABEL.get(script, script)}")
    except TelegramBadRequest:
        await query.answer()


@router.callback_query(F.data.startswith("viewproto:"))
async def cb_view_protocol(query: CallbackQuery) -> None:
    """Show the saved meeting protocol (bayonnoma) text — fixes 'generated but
    can't find it later'. Stored in the meeting's follow_up_actions column."""
    mid = query.data.split(":", 1)[1]
    m = await database.get_meeting(mid)
    if not m:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    fu = m.get("follow_up_actions") or []
    if isinstance(fu, list) and fu:
        text = fu[0] if (len(fu) == 1 and isinstance(fu[0], str)) else "\n".join(f"• {x}" for x in fu)
    else:
        text = str(fu or "").strip()
    if not text.strip():
        await query.answer("Bu uchrashuvda bayonnoma yo'q", show_alert=True)
        return
    await query.answer()
    await _safe_answer(
        query.message,
        f"📄 **BAYONNOMA — {m.get('title', '')}**\n\n{text}",
        parse_mode="Markdown",
        reply_markup=_viewproto_kb(mid, "lat"),
    )


@router.callback_query(F.data.startswith("proto_edit:"))
async def cb_protocol_edit(query: CallbackQuery, state: FSMContext) -> None:
    """Ask the user for additional corrections; re-run Claude with combined notes."""
    mid = query.data.split(":", 1)[1]
    data = await state.get_data()
    prev_notes = data.get("user_notes", "")
    await state.set_state(MeetingProtocolFSM.awaiting_revision)
    await state.update_data(meeting_id=mid, prev_notes=prev_notes)
    await query.answer()
    text = (
        "✏️ **BAYONNOMANI TUZATISH**\n\n"
        "Qaysi joyini o'zgartirish kerak? Qo'shimcha ma'lumot yoki tuzatishni\n"
        "yozib yuboring (matn yoki ovoz).\n\n"
        "_Misol: \"Dinislamning lavozimi — hamkorlik direktori\"_"
    )
    try:
        await query.message.edit_text(
            text, parse_mode="Markdown",
            reply_markup=single_back_keyboard(f"meetingopen:{mid}", text="✕ Bekor"),
        )
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown",
                            reply_markup=single_back_keyboard(f"meetingopen:{mid}", text="✕ Bekor"))


@router.message(StateFilter(MeetingProtocolFSM.awaiting_revision), F.text | F.voice)
async def handle_protocol_revision(message: Message, state: FSMContext, bot: Bot) -> None:
    data = await state.get_data()
    mid = data.get("meeting_id")
    prev_notes = data.get("prev_notes", "")
    if not mid:
        await state.clear()
        await message.answer("Holat yo'qoldi — qayta urinib ko'ring.")
        return
    if message.voice:
        file = await bot.get_file(message.voice.file_id)
        audio_io = await bot.download_file(file.file_path)
        audio_bytes = audio_io.getvalue() if hasattr(audio_io, "getvalue") else audio_io.read()
        correction = await voice_service.transcribe(audio_bytes, filename="voice.ogg", language="uz")
        if not correction:
            await message.answer("Ovozni o'qiy olmadim. Matn yuboring.")
            return
    else:
        correction = (message.text or "").strip()
    if not correction:
        await message.answer("Bo'sh xabar — bekor qilindi.")
        await state.clear()
        return

    meeting = await database.get_meeting(mid)
    if not meeting:
        await state.clear()
        await message.answer("Uchrashuv topilmadi.")
        return

    combined = f"{prev_notes}\n\nTUZATISH: {correction}"
    await message.bot.send_chat_action(message.chat.id, "typing")
    typing_task = asyncio.create_task(_keep_typing(message.bot, message.chat.id))
    try:
        directive = _build_protocol_directive(meeting, combined)
        response = await claude_service.process_message("", internal_directive=directive)
    finally:
        typing_task.cancel()

    protocol_text = (response.get("user_message") or "").strip()
    pending_actions = response.get("actions") or []
    if not protocol_text:
        await message.answer("Tuzata olmadim — qayta urinib ko'ring.")
        return

    await state.set_state(MeetingProtocolFSM.awaiting_notes)
    await state.update_data(
        protocol_text=protocol_text,
        pending_actions=pending_actions,
        user_notes=combined,
    )

    pending_count = sum(1 for a in pending_actions if a.get("type") == "create_task")
    confirm_label = f"✅ Tasdiqlash ({pending_count} ta vazifa)" if pending_count else "✅ Tasdiqlash"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text=confirm_label, callback_data=f"proto_ok:{mid}"),
            InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"proto_edit:{mid}"),
        ],
        [InlineKeyboardButton(text="📋 Nusxa", callback_data=f"proto_share:{mid}"),
        InlineKeyboardButton(text="📤 Ulashish", switch_inline_query=f"proto:{mid}")],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"meetingopen:{mid}")],
    ])
    await _safe_answer(message, protocol_text, parse_mode="Markdown", reply_markup=kb)


# NOTE: inline 'proto:' / 'txt:' share AND the general search are handled by the
# SINGLE handle_inline_query() below. There must be exactly ONE @router.inline_query()
# handler: previously a second catch-all handler offered the raw "proto:<id>" string
# as a result, so sharing leaked "@bot proto:<id>" text into the chat instead of the
# protocol (reported bug).


@router.callback_query(F.data.startswith("proto_share:"))
async def cb_protocol_share(query: CallbackQuery, state: FSMContext) -> None:
    """Re-send the protocol as a clean, standalone forwardable message. Resolves
    the text from FSM state (freshly generated) OR the saved meeting record —
    the latter is needed when sharing from the central 'Bayonnomalar' list,
    where the text isn't in state (reported: 'ulashish ishlamadi')."""
    mid = query.data.split(":", 1)[1]
    data = await state.get_data()
    protocol_text = (data.get("protocol_text") or "").strip()
    if not protocol_text:
        m = await database.get_meeting(mid)
        if m:
            fu = m.get("follow_up_actions") or []
            protocol_text = (fu[0] if (isinstance(fu, list) and fu) else str(fu or "")).strip()
    if not protocol_text:
        await query.answer("Bayonnoma topilmadi — avval saqlang.", show_alert=True)
        return
    await query.answer("📤 Quyidagi xabarni uzun bosib → Forward qiling.")
    # Standalone copyable/forwardable message (inline buttons don't forward, so a
    # clean text-only copy is the shareable artifact).
    await _safe_answer(query.message, protocol_text, parse_mode="Markdown")


@router.callback_query(F.data.startswith("meeting_cancel:"))
async def cb_meeting_cancel_confirm(query: CallbackQuery) -> None:
    """Two-step cancel: first shows the confirmation prompt."""
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await query.answer()
    title = (meeting.get("title") or "—").strip()
    when = _meeting_time_label(meeting.get("datetime_start") or "", with_past_marker=False)
    text = (
        "✕ **UCHRASHUVNI BEKOR QILISH**\n\n"
        f"{title}\n"
        f"{when}\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Haqiqatdan ham bekor qilmoqchimisiz?\n\n"
        "_iCloud kalendaridan ham o'chiriladi._"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Ha, bekor", callback_data=f"mcanc_do:{mid}"),
            InlineKeyboardButton(text="⬅️ Yo'q", callback_data=f"meetingopen:{mid}"),
        ],
    ])
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("mcanc_do:"))
async def cb_meeting_cancel_do(query: CallbackQuery) -> None:
    """Execute the cancel + iCloud delete."""
    mid = query.data.split(":", 1)[1]
    meeting = await database.get_meeting(mid)
    if not meeting:
        await query.answer("Uchrashuv topilmadi", show_alert=True)
        return
    await database.cancel_meeting(mid)
    # Scheduler'dan 15-daqiqali eslatma jobini olib tashlash — bekor qilingan
    # uchrashuv uchun eslatma chiqmasin.
    sched = scheduler_module.get_scheduler()
    if sched:
        try:
            sched.remove_meeting_reminder(mid)
        except Exception:
            logger.exception("Failed to remove meeting reminder for %s", mid)
    if config.ICLOUD_ENABLED:
        try:
            await asyncio.to_thread(calendar_service.delete_meeting, mid)
        except Exception:
            logger.exception("iCloud meeting delete failed (will retry)")
    await query.answer("✅ Uchrashuv bekor qilindi")
    try:
        await query.message.edit_text(
            "✕ Uchrashuv bekor qilindi.\n\nKalendar bilan sinxronlanmoqda...",
            parse_mode="Markdown",
            reply_markup=single_back_keyboard("meetingfilter:week"),
        )
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("reopen:"))
async def cb_reopen(query: CallbackQuery) -> None:
    parts = query.data.split(":", 1)
    target_id = parts[1] if len(parts) > 1 else ""
    if target_id:
        await database.update_task(target_id, {"status": "todo"})
        await query.answer("Qaytarildi ↺")
        task = await database.get_task(target_id)
        if task:
            try:
                await query.message.edit_text(_format_task_card(task), parse_mode="Markdown",
                                              reply_markup=_task_card_kb_with_back(task))
            except Exception:
                pass
    else:
        await query.answer()



@router.callback_query(F.data.startswith("edit:"))
async def cb_edit(query: CallbackQuery) -> None:
    """Open the field-editing menu for a task."""
    parts = query.data.split(":", 1)
    target_id = parts[1] if len(parts) > 1 else ""
    if not target_id:
        await query.answer()
        return
    task = await database.get_task(target_id)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()
    text = (
        f"✎ **Qaysi maydonni tahrirlash kerak?**\n\n"
        f"{_format_task_card(task)}"
    )
    try:
        await query.message.edit_text(text, parse_mode="Markdown", reply_markup=task_edit_menu(task))
    except TelegramBadRequest:
        await _safe_answer(query.message, text, parse_mode="Markdown", reply_markup=task_edit_menu(task))


@router.callback_query(F.data.startswith("editfield:"))
async def cb_edit_field(query: CallbackQuery, state: FSMContext) -> None:
    """User picked a field to edit. Show picker (for enums) or ask for free-text input."""
    _, tid, field = query.data.split(":", 2)
    task = await database.get_task(tid)
    if not task:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()

    if field == "priority":
        await query.message.edit_text(
            f"⚡ **Yangi prioritet?**\n\nHozir: `{task.get('priority', 'P2')}`",
            parse_mode="Markdown", reply_markup=priority_picker(tid),
        )
        return
    if field == "status":
        await query.message.edit_text(
            f"📊 **Yangi status?**\n\nHozir: `{_STATUS_LABEL_UZ.get(task.get('status', 'todo'), task.get('status'))}`",
            parse_mode="Markdown", reply_markup=status_picker(tid),
        )
        return
    if field == "deadline":
        current = task.get("deadline")
        current_label = _format_deadline_short(current)[0] if current else "yo'q"
        await query.message.edit_text(
            f"📅 **Yangi deadline?**\n\nHozir: {current_label}",
            parse_mode="Markdown", reply_markup=deadline_picker(tid),
        )
        return

    # Free-text fields → set FSM state, ask for input
    prompts = {
        "title": "📝 Yangi **sarlavha** yuboring (matn yoki ovoz):",
        "description": "📄 Yangi **tavsif** yuboring (yoki `-` deb yozsangiz tozalanadi):",
        "tags": "🏷 **Teglar** vergul bilan ajratib yuboring (masalan: `marketing, urgent`):",
        "category": "📁 **Kategoriya** yuboring (masalan: `Shartnomalar`; `-` = tozalash):",
    }
    prompt = prompts.get(field)
    if not prompt:
        await query.answer("Noto'g'ri maydon", show_alert=True)
        return

    await state.set_state(TaskEditFSM.awaiting_value)
    await state.update_data(task_id=tid, field=field)
    cancel_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✕ Bekor qilish", callback_data=f"edit_cancel:{tid}")
    ]])
    await query.message.edit_text(prompt, parse_mode="Markdown", reply_markup=cancel_kb)


@router.callback_query(F.data.startswith("setfield:"))
async def cb_set_field(query: CallbackQuery) -> None:
    """Apply an enum/preset value picked from picker."""
    _, tid, field, value = query.data.split(":", 3)
    if value == "none":
        update_val = None
    else:
        update_val = value
    if field == "status" and update_val == "done":
        await database.complete_task(tid)
    else:
        await database.update_task(tid, {field: update_val}, source="edit")
    # Toast — Uzbek labellarda. Foydalanuvchi raw kodlarni ko'rmasin.
    field_uz = {"priority": "Ustuvorlik", "status": "Holat",
                "deadline": "Muddat", "title": "Sarlavha",
                "description": "Tavsif", "tags": "Teglar"}.get(field, field)
    if update_val is None:
        value_uz = "tozalandi"
    elif field == "priority":
        value_uz = _PRIORITY_LABEL_UZ.get(update_val, update_val)
    elif field == "status":
        value_uz = _STATUS_LABEL_UZ.get(update_val, update_val)
    else:
        value_uz = update_val
    await query.answer(f"{field_uz} → {value_uz} ✅")
    task = await database.get_task(tid)
    if task:
        text = _format_task_card(task)
        try:
            await query.message.edit_text(
                text, parse_mode="Markdown",
                reply_markup=_task_card_kb_with_back(task),
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data.startswith("deadline_preset:"))
async def cb_deadline_preset(query: CallbackQuery) -> None:
    _, tid, preset = query.data.split(":", 2)
    now = datetime.now(database.TZ)
    new_dt = None
    if preset == "today":
        new_dt = now.replace(hour=17, minute=0, second=0, microsecond=0)
        if new_dt < now:
            # Round like every sibling preset — sub-second precision doesn't survive
            # an xlsx round-trip and a mismatch would re-arm the deadline reminder.
            new_dt = (now + timedelta(hours=2)).replace(second=0, microsecond=0)
    elif preset == "tomorrow":
        new_dt = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
    elif preset == "plus3":
        new_dt = (now + timedelta(days=3)).replace(hour=10, minute=0, second=0, microsecond=0)
    elif preset == "weekend":
        days_to_sunday = (6 - now.weekday()) % 7 or 7
        new_dt = (now + timedelta(days=days_to_sunday)).replace(hour=18, minute=0, second=0, microsecond=0)

    if new_dt:
        await database.update_task(tid, {"deadline": new_dt.isoformat()}, source="edit")
        await query.answer(f"Deadline: {new_dt.strftime('%d-%m %H:%M')} ✅")
        task = await database.get_task(tid)
        if task:
            try:
                await query.message.edit_text(
                    _format_task_card(task), parse_mode="Markdown",
                    reply_markup=_task_card_kb_with_back(task),
                )
            except TelegramBadRequest:
                pass


@router.callback_query(F.data.startswith("deadline_manual:"))
async def cb_deadline_manual(query: CallbackQuery, state: FSMContext) -> None:
    tid = query.data.split(":", 1)[1]
    await state.set_state(TaskEditFSM.awaiting_value)
    await state.update_data(task_id=tid, field="deadline")
    await query.answer()
    cancel_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✕ Bekor qilish", callback_data=f"edit_cancel:{tid}")
    ]])
    await query.message.edit_text(
        "📅 Deadline'ni quyidagi formatlardan birida yuboring:\n\n"
        "• `2026-05-25 14:30`\n"
        "• `25-05 14:30` (yil — bu yil)\n"
        "• `ertaga 09:00`\n"
        "• `juma 10:00`",
        parse_mode="Markdown",
        reply_markup=cancel_kb,
    )


@router.callback_query(F.data.startswith("edit_cancel:"))
async def cb_edit_cancel(query: CallbackQuery, state: FSMContext) -> None:
    tid = query.data.split(":", 1)[1]
    await state.clear()
    task = await database.get_task(tid)
    await query.answer("Bekor qilindi")
    if task:
        try:
            await query.message.edit_text(
                _format_task_card(task), parse_mode="Markdown",
                reply_markup=_task_card_kb_with_back(task),
            )
        except TelegramBadRequest:
            pass


@router.message(StateFilter(TaskEditFSM.awaiting_value), F.text | F.voice)
async def handle_edit_value(message: Message, state: FSMContext) -> None:
    _msg_text = await _get_text_or_transcribe(message, bot=message.bot)
    if _msg_text is None:
        return

    """User typed a new value while editing. Apply and exit FSM."""
    data = await state.get_data()
    tid = data.get("task_id")
    field = data.get("field")
    if not tid or not field:
        await state.clear()
        return
    raw = message.text.strip()
    past_note = ""

    if field == "title":
        await database.update_task(tid, {"title": raw[:200]}, source="edit")
    elif field == "description":
        await database.update_task(tid, {"description": None if raw == "-" else raw}, source="edit")
    elif field == "tags":
        tags = [t.strip() for t in raw.split(",") if t.strip()]
        await database.update_task(tid, {"tags": tags}, source="edit")
    elif field == "category":
        await database.update_task(tid, {"category": None if raw == "-" else raw[:60]}, source="edit")
    elif field == "deadline":
        # Try to parse via Claude (consistent with how user-input dates are parsed)
        parsed, reason = await _parse_deadline_natural(raw)
        if parsed:
            await database.update_task(tid, {"deadline": parsed}, source="edit")
            # Editing to a past date is a legitimate overdue-correction — save it,
            # but flag it so the user notices if it wasn't intended.
            if database.is_past_deadline(parsed):
                past_note = "⚠️ Diqqat: muddat o'tmishda.\n\n"
        else:
            await _safe_answer(message, _deadline_error_message(reason, kind="deadline"),
                               parse_mode="Markdown")
            return  # keep FSM open for retry
    await state.clear()

    task = await database.get_task(tid)
    if task:
        await _safe_answer(message, past_note + "✅ Saqlandi\n\n" + _format_task_card(task),
                           parse_mode="Markdown", reply_markup=_task_card_kb_with_back(task))


def _deadline_error_message(reason: str | None, *, kind: str = "deadline") -> str:
    """Map a _parse_deadline_natural() reason code to a specific, actionable
    message so different mistakes get different guidance — instead of one
    generic "tushunmadim" line for every kind of bad input.

    kind: 'deadline' (task due-date) or 'time' (reminder/edit time) — only
    changes the noun shown to the user.
    """
    noun = "Muddatni" if kind == "deadline" else "Vaqtni"
    if reason == "too_far":
        # The bot DID understand the input — it just exceeds the 7-day relative
        # cap. Saying "tushunmadim" here would be misleading, so be explicit.
        return ("⏳ Juda uzoq muddat — nisbiy vaqt eng ko'pi 7 kun bo'lishi mumkin.\n"
                "Aniq sana yuboring, masalan: `2026-06-10 15:00`.")
    if reason == "invalid":
        return ("📅 Bunday sana/vaqt mavjud emas (kun yoki oy noto'g'ri).\n"
                "Tekshirib qaytadan yuboring: `2026-06-10 15:00`.")
    # "unparsable" (or any unexpected reason) → generic-but-helpful with examples.
    return (f"❌ {noun} tushunmadim. Masalan: `ertaga 09:00`, `2 soat`, "
            f"yoki `2026-06-10 15:00`.")


async def _parse_deadline_natural(text: str) -> tuple[str | None, str | None]:
    """Lightweight natural-language → ISO 8601 in Asia/Tashkent.
    Handles common formats without invoking Claude.

    Returns (iso, reason):
      - (iso,  None)         — parsed successfully
      - (None, "too_far")    — relative offset exceeds the 7-day cap
      - (None, "invalid")    — matched a date/time shape but the value is impossible
      - (None, "unparsable") — nothing matched
    Pass the reason to _deadline_error_message() so each mistake gets its own
    fix-it message rather than one generic line.
    """
    import re
    text = text.strip().lower()
    now = datetime.now(database.TZ)

    # 0) Relative: "15 daqiqa", "2 soat", "2 soatdan keyin"
    # Cap at 7 days so absurd inputs like "999 daqiqa" don't quietly schedule
    # tasks 16 hours out (or worse, "99 soat" → 4 days). Anything bigger should
    # use an explicit date.
    MAX_RELATIVE_MINUTES = 7 * 24 * 60
    MAX_RELATIVE_HOURS = 7 * 24
    m = re.match(r"^(\d{1,4})\s*(daq|daqiqa|daqiqadan|min|minut)\b", text)
    if m:
        minutes = int(m.group(1))
        if 0 < minutes <= MAX_RELATIVE_MINUTES:
            return (now + timedelta(minutes=minutes)).replace(second=0, microsecond=0).isoformat(), None
        return None, "too_far"
    m = re.match(r"^(\d{1,3})\s*(soat|soatdan|hour|h)\b", text)
    if m:
        hours = int(m.group(1))
        if 0 < hours <= MAX_RELATIVE_HOURS:
            return (now + timedelta(hours=hours)).replace(second=0, microsecond=0).isoformat(), None
        return None, "too_far"

    # 1) ISO-ish: 2026-05-25 14:30 or 2026-05-25T14:30
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})[\s tT]+(\d{1,2}):(\d{2})$", text)
    if m:
        y, mo, d, hh, mm = map(int, m.groups())
        try:
            return database.TZ.localize(datetime(y, mo, d, hh, mm)).isoformat(), None
        except ValueError:
            return None, "invalid"

    # 2) Short: 25-05 14:30 (current year)
    m = re.match(r"^(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{2})$", text)
    if m:
        d, mo, hh, mm = map(int, m.groups())
        try:
            return database.TZ.localize(datetime(now.year, mo, d, hh, mm)).isoformat(), None
        except ValueError:
            return None, "invalid"

    # 3) Relative keywords
    weekday_map = {
        "dushanba": 0, "seshanba": 1, "chorshanba": 2, "payshanba": 3,
        "juma": 4, "shanba": 5, "yakshanba": 6,
    }
    m = re.match(r"^(ertaga|indin|bugun|dushanba|seshanba|chorshanba|payshanba|juma|shanba|yakshanba)\s+(?:soat\s+)?(\d{1,2})(?::(\d{2}))?$", text)
    if m:
        word = m.group(1)
        hh = int(m.group(2))
        mm = int(m.group(3) or 0)
        if word == "bugun":
            target = now
        elif word == "ertaga":
            target = now + timedelta(days=1)
        elif word == "indin":
            target = now + timedelta(days=2)
        else:
            target_wd = weekday_map[word]
            delta = (target_wd - now.weekday()) % 7 or 7
            target = now + timedelta(days=delta)
        try:
            return target.replace(hour=hh, minute=mm, second=0, microsecond=0).isoformat(), None
        except ValueError:
            return None, "invalid"
    return None, "unparsable"


def _task_card_kb_with_back(task: dict) -> InlineKeyboardMarkup:
    """Opened task card — direct actions, NO extra '⋯ Batafsil' step.
      Row 1: [✅ Bajarildi / ↺ Qaytarish] [✏️ Tahrir]
      Row 2 (active tasks only): one-tap reschedule [📅 Ertaga] [📅 +1 hafta] [📅 Dushanba]
      Row 3: [🗑 O'chirish] [⬅️ Ro'yxatga]  ← destructive sits away from ✅ (misclick guard)
    Field edits + 👤 Ijrochi live under ✏️ Tahrir. 🗑 still goes through a confirm step."""
    tid = task["id"]
    # A subtask's "back" returns to its parent's subtask view, not the flat list.
    back = (f"subview:{task['parent_id']}" if task.get("parent_id")
            else f"taskfilter:{_last_task_filter or 'active'}")
    is_done = task.get("status") == "done"
    if is_done:
        primary = InlineKeyboardButton(text="↺ Qaytarish", callback_data=f"reopen:{tid}")
    else:
        primary = InlineKeyboardButton(text="✅ Bajarildi", callback_data=f"complete:{tid}")
    rows: list[list[InlineKeyboardButton]] = [
        [primary, InlineKeyboardButton(text="✏️ Tahrir", callback_data=f"edit:{tid}")],
    ]
    # One-tap reschedule (snooze) — only meaningful for active tasks; a done task
    # has no live deadline to push.
    if not is_done:
        rows.append([
            InlineKeyboardButton(text="📅 Ertaga", callback_data=f"snooze:{tid}:tomorrow"),
            InlineKeyboardButton(text="📅 +1 hafta", callback_data=f"snooze:{tid}:week"),
            InlineKeyboardButton(text="📅 Dushanba", callback_data=f"snooze:{tid}:monday"),
        ])
    # Breakdown into real child tasks (top-level only — depth-1, no nesting).
    if not task.get("parent_id"):
        rows.append([InlineKeyboardButton(text="🌳 Sub-vazifalar", callback_data=f"subview:{tid}")])
    # Destructive action on its own row, away from ✅/✏️ (misclick guard). It still
    # routes through cb_task_del_confirm, which asks before deleting.
    rows.append([
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"task_del:{tid}"),
        InlineKeyboardButton(text="⬅️ Ro'yxatga", callback_data=back),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


class SubtaskAddFSM(StatesGroup):
    awaiting = State()


def _subtask_view_kb(parent_id: str, subs: list) -> InlineKeyboardMarkup:
    """Subtask view: each child opens its full task card; plus add + back-to-parent."""
    rows: list = []
    for s in subs:
        emoji = _STATUS_EMOJI.get(s.get("status", "todo"), "•")
        bits = [_truncate((s.get("title") or "—").strip(), 28)]
        if (s.get("assignee") or "").strip():
            bits.append(f"👤{s['assignee'].strip()[:12]}")
        rows.append([InlineKeyboardButton(text=f"{emoji} {' · '.join(bits)}",
                                          callback_data=f"taskopen:{s['id']}")])
    rows.append([InlineKeyboardButton(text="➕ Sub-vazifa", callback_data=f"subadd:{parent_id}")])
    rows.append([InlineKeyboardButton(text="⬅️ Vazifaga", callback_data=f"taskopen:{parent_id}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _render_subtask_view(query: CallbackQuery, parent: dict) -> None:
    subs = await database.list_subtasks(parent["id"])
    done = sum(1 for s in subs if s.get("status") in ("done", "cancelled"))
    head = f"🌳 **{(parent.get('title') or 'Vazifa').strip()}** — sub-vazifalar"
    body = f"{head}\n{done}/{len(subs)} bajarildi" if subs else f"{head}\n_Hali sub-vazifa yo'q._"
    kb = _subtask_view_kb(parent["id"], subs)
    try:
        await query.message.edit_text(body, parse_mode="Markdown", reply_markup=kb)
    except TelegramBadRequest:
        await _safe_answer(query.message, body, parse_mode="Markdown", reply_markup=kb)


@router.callback_query(F.data.startswith("subview:"))
async def cb_subtask_view(query: CallbackQuery) -> None:
    parent = await database.get_task(query.data.split(":", 1)[1])
    if not parent:
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await query.answer()
    await _render_subtask_view(query, parent)


@router.callback_query(F.data.startswith("subadd:"))
async def cb_subtask_add(query: CallbackQuery, state: FSMContext) -> None:
    pid = query.data.split(":", 1)[1]
    if not await database.get_task(pid):
        await query.answer("Vazifa topilmadi", show_alert=True)
        return
    await state.set_state(SubtaskAddFSM.awaiting)
    await state.update_data(pid=pid)
    await query.answer()
    await query.message.answer(
        "➕ Sub-vazifa nomini yozing (matn/ovoz). Ijrochi va muddatni keyin sub-vazifani "
        "ochib «✏️ Tahrir»dan qo'shasiz. Bir nechta — har qatorga bittadan.")


@router.message(StateFilter(SubtaskAddFSM.awaiting), F.text | F.voice)
async def handle_subtask_add(message: Message, state: FSMContext, bot: Bot) -> None:
    text = await _get_text_or_transcribe(message, bot=bot)
    if text is None:
        return
    data = await state.get_data()
    pid = data.get("pid")
    await state.clear()
    parent = await database.get_task(pid) if pid else None
    if not parent:
        await message.answer("Vazifa topilmadi.")
        return
    added = 0
    for line in (text or "").splitlines():
        s = line.strip(" -•\t·")
        if s:
            await database.create_task({
                "title": s[:200], "parent_id": pid,
                "priority": parent.get("priority", "P2"), "source": "subtask",
            })
            added += 1
    if not added:
        await message.answer("Bo'sh — bekor qilindi.")
        return
    subs = await database.list_subtasks(pid)
    await message.answer(
        f"✅ {added} ta sub-vazifa qo'shildi (jami {len(subs)}). Ijrochi/muddat uchun "
        f"sub-vazifani ochib «✏️ Tahrir».",
        reply_markup=_subtask_view_kb(pid, subs))


@router.callback_query()
async def cb_fallback(query: CallbackQuery) -> None:
    await query.answer()


# ─────────────────────── INLINE MODE ───────────────────────


@router.inline_query()
async def handle_inline_query(query: InlineQuery) -> None:
    """The SINGLE inline handler (principal only). Three query shapes:
      • 'proto:<meeting_id>' → that meeting's saved protocol (clean, sendable text)
      • 'txt:<token>'        → a cached polished text
      • anything else        → raw text + task/meeting quick-search

    All results use parse_mode=None so a protocol's stray markdown/emoji never makes
    answerInlineQuery fail. A 'proto:'/'txt:' miss shows "Topilmadi" — NEVER the raw
    "proto:<id>" string (that was the leaked-text bug).

    NOTE: inline share needs inline mode ENABLED in @BotFather (/setinline). When it's
    off, Telegram sends no inline_query and "@bot proto:<id>" would be sent as plain
    text — so if results never appear, check BotFather first."""
    uid = query.from_user.id if query.from_user else None
    if not _is_principal(uid):
        await query.answer(results=[], cache_time=1, is_personal=True)
        return

    q = (query.query or "").strip()
    results: list = []
    try:
        if q.startswith("proto:"):
            mid = q.split(":", 1)[1]
            m = await database.get_meeting(mid)
            if m:
                fu = m.get("follow_up_actions") or []
                text = (fu[0] if (isinstance(fu, list) and fu) else str(fu or "")).strip()
                if text:
                    results.append(InlineQueryResultArticle(
                        id=f"p-{mid}"[:64],
                        title="📄 Bayonnomani yuborish",
                        description=(m.get("title") or "Bayonnoma").strip()[:80],
                        input_message_content=InputTextMessageContent(
                            message_text=text.replace("**", "")[:4096], parse_mode=None),
                    ))
        elif q.startswith("txt:"):
            token = q.split(":", 1)[1]
            text = await database.get_share_text(token)
            if text:
                results.append(InlineQueryResultArticle(
                    id=f"t-{token}"[:64],
                    title="📤 Matnni yuborish",
                    description=text[:80],
                    input_message_content=InputTextMessageContent(
                        message_text=text[:4096], parse_mode=None),
                ))
        else:
            # Raw text as a quick scratchpad result.
            if q:
                results.append(InlineQueryResultArticle(
                    id="raw",
                    title=f"📝 «{q[:60]}»",
                    description="Bu matnni shu chatga jo'natish",
                    input_message_content=InputTextMessageContent(message_text=q, parse_mode=None),
                ))
            # Task/meeting quick-search.
            if len(q) >= 2:
                search_results = await database.search_all(q, limit=8)
                for t in search_results.get("tasks", [])[:5]:
                    badge = _task_badge(t)
                    deadline_label, _ = _format_deadline_short(t.get("deadline"))
                    results.append(InlineQueryResultArticle(
                        id=f"task:{t['id']}"[:64],
                        title=f"{badge} {_truncate(t['title'], 60)}",
                        description=f"Vazifa · {deadline_label}",
                        input_message_content=InputTextMessageContent(
                            message_text=f"{badge} {t['title']}\n📅 {deadline_label}", parse_mode=None),
                    ))
                for m in search_results.get("meetings", [])[:3]:
                    try:
                        dt = datetime.fromisoformat(m["datetime_start"]).astimezone(database.TZ)
                        time_str = dt.strftime("%d-%m %H:%M")
                    except (ValueError, TypeError):
                        time_str = "—"
                    participants = ", ".join(m.get("participants", [])[:2]) or "—"
                    results.append(InlineQueryResultArticle(
                        id=f"meeting:{m['id']}"[:64],
                        title=f"🤝 {_truncate(m['title'], 60)}",
                        description=f"Uchrashuv · {time_str} · {participants}",
                        input_message_content=InputTextMessageContent(
                            message_text=f"🤝 {m['title']}\n🕐 {time_str}\n👥 {participants}", parse_mode=None),
                    ))
    except Exception:
        logger.exception("Inline query handler failed")

    if not results:
        is_share = q.startswith(("proto:", "txt:"))
        results.append(InlineQueryResultArticle(
            id="empty",
            title="Topilmadi",
            description=("Bayonnoma topilmadi — avval saqlang." if is_share
                         else "Boshqa kalit so'z bilan urinib ko'ring"),
            input_message_content=InputTextMessageContent(
                message_text="(natija yo'q)", parse_mode=None),
        ))

    await query.answer(results=results, cache_time=1, is_personal=True)


# ─────────────────────── FALLBACK HANDLERS (last-resort) ───────────────────────
# These MUST be registered last — aiogram dispatches in registration order,
# so state-specific handlers win when they match. Anything that falls through
# (unexpected content type, message in a state with no matching handler,
# edited messages) lands here instead of silently disappearing.


@router.edited_message()
async def handle_edited_message(message: Message) -> None:
    """Telegram delivers message edits as `edited_message` updates. Without
    this handler they vanish — user thinks the bot saw their edit when it
    didn't. Reply with explicit guidance so they re-send."""
    await message.answer(
        "✏️ Tahrir qilingan xabarlar qabul qilinmaydi. "
        "Yangi xabar (matn yoki ovoz) yuboring."
    )


@router.message(F.voice)
async def handle_voice_fallback(message: Message, bot: Bot, state: FSMContext) -> None:
    """Catches voice messages in states whose specific handler is text-only.
    Transcribes, then re-emits the transcript as if the user typed it by
    re-dispatching through the router. Without this, voice in any non-
    default, non-meeting state vanishes."""
    transcript = await _get_text_or_transcribe(message, bot)
    if transcript is None:
        return
    # message.text was just patched in-place by the helper above. Now do the
    # equivalent of the default_state voice handler: send the voice-confirm
    # prompt only if we're in default_state. In other states, the handler
    # that would have caught this text directly didn't run because aiogram
    # already chose this fallback — so we either route through Claude or
    # echo a recovery hint. Default state → confirm; else → Claude direct.
    current = await state.get_state()
    if current is None:
        await _send_voice_confirm_prompt(message, state, transcript)
    else:
        # Voice arrived inside an FSM state that didn't claim it. Best UX:
        # treat the transcript as a free-form message to Claude (same as
        # text would in a section state's fall-through branch).
        await _process_and_reply(message, transcript, state=state)


@router.message(F.video | F.video_note | F.sticker | F.animation | F.audio)
async def handle_unsupported_attachment(message: Message) -> None:
    """Polite "not yet supported" instead of silent drop for media types the
    bot has no handler for. Documents and photos are handled by
    handle_incoming_file; this covers video/sticker/animation/audio."""
    kind = (
        "video" if message.video else
        "video-xabar" if message.video_note else
        "stiker" if message.sticker else
        "animatsiya" if message.animation else
        "audio fayl" if message.audio else
        "fayl"
    )
    await message.answer(
        f"📎 Hozircha {kind} qabul qila olmayman. "
        "Matn yoki ovoz xabari yuboring."
    )


@router.message()
async def handle_unmatched_message(message: Message) -> None:
    """Last-resort catch-all for anything that didn't match a more specific
    handler (unexpected content type, lingering FSM state with no matching
    filter, etc). Without this the message vanishes and the user has no
    idea what went wrong."""
    await message.answer(
        "🤔 Bu xabarni tushuna olmadim.\n\n"
        "• Matn yoki ovoz xabari yuboring\n"
        "• `/cancel` — joriy amalni bekor qilish\n"
        "• `/help` — qo'llanma\n"
        "• `/start` — boshidan boshlash",
        parse_mode="Markdown",
    )
